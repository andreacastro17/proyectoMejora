"""
Fase 7 — Valorización de Programas EAFIT en proceso de calidad académica.

Genera Programas_para_valorizacion_output.xlsx con dos secciones por programa y región:
  - MERCADO:     métricas del mercado regional (todas las IES de esa región)
  - REFERENTES:  métricas de las IES referentes filtradas por región

Fuente de datos: sabana_consolidada.parquet + agregados regionales (Fase 4)
Lista de programas: ref/backup/programas_para_valorizacion.xlsx
"""

from __future__ import annotations

import re as _re
import unicodedata as _ud
from functools import lru_cache
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd

from etl.config import (
    AÑO_FIN_DATOS,
    AÑO_INICIO_PRIMER_CURSO,
    OUTPUTS_DIR,
    RAW_HISTORIC_DIR,
    REF_DIR,
    TEMP_DIR,
    CHECKPOINT_BASE_MAESTRA,
    ESTUDIO_MERCADO_DIR,
    NIVELES_MERCADO,
)
from etl.pipeline_logger import log_info, log_warning
from etl.scoring import _SCORE_PARTICIPACION_PESO, apply_scoring
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

_PROG_IDS_PATH = REF_DIR / "backup" / "prog_ids.csv"
_CAT_IDS_PATH = REF_DIR / "backup" / "cat_ids.csv"

# Lazy: TF-IDF sobre categorías canónicas (cat_ids.csv)
_CATS_VECTORIZER: TfidfVectorizer | None = None
_CATS_MATRIX = None
_CATS_CANONICAS: list[str] | None = None

# Lazy: CAT_ID por CATEGORIA_FINAL (cat_ids.csv)
_CAT_IDS_MAP: dict[str, str] | None = None
_CAT_IDS_NORM_MAP: dict[str, str] | None = None
# Rango reservado EAFIT: 900001-900999.
_PROG_ID_BASE = 900001


def _get_or_assign_prog_id(programas: list[str]) -> dict[str, str]:
    """
    Lee/genera el registro permanente de IDs de programa EAFIT
    (ref/backup/prog_ids.csv). Una vez asignado, el ID no cambia.
    Los programas nuevos reciben el siguiente ID disponible.

    Cuando el MEN asigne un código SNIES real, actualizar prog_ids.csv
    manualmente reemplazando el EAFIT-9XXXXX por el código oficial.
    """
    registro: dict[str, str] = {}

    if _PROG_IDS_PATH.exists():
        try:
            df_reg = pd.read_csv(_PROG_IDS_PATH, dtype=str)
            for _, row in df_reg.iterrows():
                prog = str(row["PROGRAMA_EAFIT"]).strip().upper()
                registro[prog] = str(row["PROG_ID"]).strip()
        except Exception as e:
            log_warning(f"[PROG_ID] No se pudo leer {_PROG_IDS_PATH.name}: {e}")

    max_num = _PROG_ID_BASE - 1
    for pid in registro.values():
        try:
            max_num = max(max_num, int(pid.replace("EAFIT-", "")))
        except ValueError:
            pass

    nuevos: list[str] = []
    for prog in programas:
        prog_norm = str(prog).strip().upper()
        if prog_norm not in registro:
            nuevos.append(prog_norm)

    for prog_norm in nuevos:
        max_num += 1
        registro[prog_norm] = f"EAFIT-{max_num}"
        log_info(f"[PROG_ID] Nuevo programa registrado: {prog_norm} → EAFIT-{max_num}")

    if nuevos:
        _PROG_IDS_PATH.parent.mkdir(parents=True, exist_ok=True)
        df_out = pd.DataFrame(
            [(prog, pid) for prog, pid in sorted(registro.items(), key=lambda x: x[1])],
            columns=["PROGRAMA_EAFIT", "PROG_ID"],
        )
        df_out.to_csv(_PROG_IDS_PATH, index=False, encoding="utf-8-sig")
        log_info(f"[PROG_ID] Registro actualizado: {len(registro)} programas en {_PROG_IDS_PATH.name}")

    return registro


# TASA_CAPTURA_EAFIT — fracción del promedio regional de primer_curso que
# se estima capturará un programa nuevo de EAFIT en su primer año de operación.
TASA_CAPTURA_EAFIT: float = 0.35

# Horizonte fijo de proyección de primer curso (sin año de lanzamiento por programa).
_AÑO_PROYECCION_REGRESION: int = AÑO_FIN_DATOS + 2

# ── Definición de segmentos regionales (alineado con run_segmentos_regionales / nombres de parquet) ──
COL_DEPT = "DEPARTAMENTO_OFERTA_PROGRAMA"
COL_MOD = "MODALIDAD"

SEGMENTOS = ["Antioquia", "Bogota", "Eje_Cafetero", "Virtual"]
LABEL_REGION = {
    "Antioquia": "Antioquia",
    "Bogota": "Bogota",
    "Eje_Cafetero": "Eje Cafetero",
    "Virtual": "Virtual",
}

SEGMENTOS_FILTROS: dict[str, Callable] = {
    "Antioquia": lambda df: df[df[COL_DEPT] == "ANTIOQUIA"].copy() if COL_DEPT in df.columns else df.iloc[0:0],
    "Bogota": lambda df: df[df[COL_DEPT] == "BOGOTÁ D.C."].copy() if COL_DEPT in df.columns else df.iloc[0:0],
    "Eje_Cafetero": lambda df: df[df[COL_DEPT].isin(["CALDAS", "RISARALDA", "QUINDÍO"])].copy() if COL_DEPT in df.columns else df.iloc[0:0],
    "Virtual": lambda df: df[df[COL_MOD].astype(str).str.upper().str.strip() == "VIRTUAL"].copy() if COL_MOD in df.columns else df.iloc[0:0],
}

# Códigos SNIES exactos de las 13 IES referentes, verificados contra
# Instituciones.xlsx y la sábana consolidada. Se usan códigos numéricos
# para evitar errores de matching por variaciones de nombre.
CODIGOS_IES_REFERENTES: set[int] = {
    1701,  # PONTIFICIA UNIVERSIDAD JAVERIANA (Bogotá)
    1702,  # PONTIFICIA UNIVERSIDAD JAVERIANA (Cali)
    1710,  # UNIVERSIDAD PONTIFICIA BOLIVARIANA (Medellín)
    1711,  # UNIVERSIDAD DE LA SABANA
    1712,  # UNIVERSIDAD EAFIT
    1713,  # UNIVERSIDAD DEL NORTE
    1714,  # COLEGIO MAYOR DE NUESTRA SEÑORA DEL ROSARIO
    1723,  # UNIVERSIDAD PONTIFICIA BOLIVARIANA (Bucaramanga)
    1727,  # UNIVERSIDAD PONTIFICIA BOLIVARIANA (Montería)
    1729,  # UNIVERSIDAD EL BOSQUE
    1730,  # UNIVERSIDAD PONTIFICIA BOLIVARIANA (Palmira)
    1813,  # UNIVERSIDAD DE LOS ANDES
    1828,  # UNIVERSIDAD ICESI
    2704,  # COLEGIO DE ESTUDIOS SUPERIORES DE ADMINISTRACION-CESA
    2708,  # UNIVERSIDAD CES
    2727,  # FUNDACION UNIVERSITARIA-CEIPA
    2812,  # UNIVERSIDAD EAN
    2813,  # UNIVERSIDAD EIA
}


@lru_cache(maxsize=2048)
def _norm(s: str) -> str:
    """Mayúsculas y sin tildes (comparación de categorías / IES). Resultado cacheado."""
    import unicodedata

    t = str(s).upper().strip()
    return "".join(
        c for c in unicodedata.normalize("NFD", t)
        if unicodedata.category(c) != "Mn"
    )


def _get_archivo_valorizar() -> Path:
    """Busca el archivo de programas a valorizar en ref/backup/."""
    for nombre in ["programas_para_valorizacion.xlsx", "programas_para_valorizacion.csv"]:
        for carpeta in [REF_DIR / "backup", REF_DIR]:
            p = carpeta / nombre
            if p.exists():
                return p
    raise FileNotFoundError(
        "No se encontró programas_para_valorizacion.xlsx en ref/backup/.\n"
        "Coloca el archivo con columnas: Categoría, Nivel, "
        "'Programas en proceso Calidad académica', 'Tiene estudio de mercado'"
    )


def _leer_programas_eafit(log: Callable) -> pd.DataFrame:
    """Lee y normaliza la lista de programas EAFIT a valorizar."""
    ruta = _get_archivo_valorizar()
    if ruta.suffix == ".xlsx":
        # El archivo tiene dos filas de encabezado:
        # fila 0: bloques "Mercado" y "Referentes" (se ignora)
        # fila 1: nombres reales de columnas (Categoría, Nivel, etc.)
        # Solo necesitamos las primeras 5 columnas (identificación del programa)
        df = pd.read_excel(ruta, header=1, usecols=range(5))
    else:
        df = pd.read_csv(ruta, encoding="utf-8-sig")

    rename = {}
    for col in df.columns:
        c = str(col).strip().lower()
        if "categor" in c:
            rename[col] = "CATEGORIA_RAW"
        elif "nivel" in c:
            rename[col] = "NIVEL"
        elif "proceso" in c or "calidad" in c:
            rename[col] = "PROGRAMA_EAFIT"
        elif "estudio" in c or "mercado" in c:
            rename[col] = "TIENE_ESTUDIO_MERCADO"
    df = df.rename(columns=rename)

    for col in ["CATEGORIA_RAW", "NIVEL", "PROGRAMA_EAFIT"]:
        if col not in df.columns:
            raise ValueError(f"Columna '{col}' no encontrada en {ruta.name}")

    if "TIENE_ESTUDIO_MERCADO" not in df.columns:
        df["TIENE_ESTUDIO_MERCADO"] = "No"

    # Normalizar CATEGORIA_RAW antes del dedup para que variaciones de capitalización
    # no generen filas duplicadas (ej: "Inteligencia Artificial" == "INTELIGENCIA ARTIFICIAL")
    df["CATEGORIA_RAW"] = df["CATEGORIA_RAW"].astype(str).str.strip().str.upper()
    df["NIVEL"] = df["NIVEL"].astype(str).str.strip()
    df["PROGRAMA_EAFIT"] = df["PROGRAMA_EAFIT"].astype(str).str.strip()
    df = (
        df.dropna(subset=["PROGRAMA_EAFIT"])
        .drop_duplicates(subset=["PROGRAMA_EAFIT", "CATEGORIA_RAW"])
        .reset_index(drop=True)
    )
    log(f"  Programas EAFIT a valorizar: {len(df)}")
    _id_map = _get_or_assign_prog_id(df["PROGRAMA_EAFIT"].tolist())
    df["PROG_ID"] = (
        df["PROGRAMA_EAFIT"]
        .astype(str).str.strip().str.upper()
        .map(_id_map)
        .fillna("EAFIT-??????")
    )
    log(f"  PROG_IDs cargados. Rango: {df['PROG_ID'].min()} ... {df['PROG_ID'].max()}")
    return df[["PROG_ID", "CATEGORIA_RAW", "NIVEL", "PROGRAMA_EAFIT", "TIENE_ESTUDIO_MERCADO"]]


def _categorias_de_raw(categoria_raw: str) -> list[str]:
    """
    Parsea categorías que pueden venir separadas por guión o ' - '.
    Ejemplo: 'ANALITICA DE DATOS-INGENIERIA DE SOFTWARE' → ['ANALITICA DE DATOS', 'INGENIERIA DE SOFTWARE']
    Solo divide si cada parte tiene ≥ 2 palabras (para no confundir nombres de IES con guión).
    """
    raw = str(categoria_raw).strip()
    for sep in [" - ", "-"]:
        if sep in raw:
            partes = [p.strip() for p in raw.split(sep) if p.strip()]
            if all(len(p.split()) >= 2 for p in partes):
                return partes
    return [raw]


def _lookup_categoria(ag: pd.DataFrame | None, categorias: list[str]) -> dict:
    """
    Busca las métricas de una categoría en el DataFrame de agregado regional.
    Usa los nombres EXACTOS de columnas que produce run_fase4_desde_sabana().
    """
    VACIAS = {
        "CAT_ID": "",
        f"prom_matricula_por_programa_{AÑO_FIN_DATOS}": 0.0,
        f"prom_matricula_{AÑO_FIN_DATOS}": 0.0,
        f"participacion_{AÑO_FIN_DATOS}": 0.0,
        "AAGR_ROBUSTO": np.nan,
        "salario_promedio_smlmv": np.nan,
        f"pct_no_matriculados_{AÑO_FIN_DATOS}": np.nan,
        f"num_programas_{AÑO_FIN_DATOS}": 0,
        "distancia_costo_pct": np.nan,
        f"suma_matricula_{AÑO_FIN_DATOS}": 0.0,
        "programas_activos": 0,
        "programas_nuevos_3a": 0,
        "programas_inactivos": 0,
        "costo_promedio": np.nan,
        "pct_con_matricula": 0.0,
        "SEÑAL_TENDENCIA": np.nan,
        # Campos adicionales para scoring correcto en _score_y_calificacion
        "score_participacion": 1,        # pre-calculado sobre 288 cats; 1-fila colapsa a score 1
        "NIVEL_MAYORIT": "ESPECIALIZACIÓN",  # necesario para árbol de decisión AAGR ESP vs MAE
    }

    if ag is None or len(ag) == 0:
        return VACIAS.copy()

    # Columna de categoría en el parquet
    col_cat = next(
        (c for c in ag.columns if "CATEGORIA_FINAL" in str(c)),
        None,
    )
    if col_cat is None:
        return VACIAS.copy()

    col_cat_norm = "_cat_norm_cached"
    if col_cat_norm not in ag.columns:
        ag[col_cat_norm] = ag[col_cat].apply(lambda x: _norm(str(x)))

    resultados = []
    for cat in categorias:
        cat_n = _norm(cat)
        mask = ag[col_cat_norm] == cat_n
        sub = ag[mask]
        if len(sub) == 0:
            continue
        row = sub.iloc[0]

        def _get(col_name, default=np.nan):
            if col_name in row.index:
                v = row[col_name]
                return float(v) if pd.notna(v) else default
            return default

        _pmp = _get(f"prom_matricula_por_programa_{AÑO_FIN_DATOS}")
        if not pd.notna(_pmp):
            _pmp = _get(f"prom_primer_curso_{AÑO_FIN_DATOS}")  # fallback explícito si la columna alias no existe
        _pmp_val = _pmp if pd.notna(_pmp) else 0.0
        _cat_id_cell = row["CAT_ID"] if "CAT_ID" in row.index else ""
        _cat_id = str(_cat_id_cell).strip() if pd.notna(_cat_id_cell) and str(_cat_id_cell).strip() else ""
        met = {
            "CAT_ID": _cat_id,
            # prom_matricula_por_programa_2024 es el nombre que scoring.py busca PRIMERO (= primer_curso)
            f"prom_matricula_por_programa_{AÑO_FIN_DATOS}": _pmp_val,
            f"prom_matricula_{AÑO_FIN_DATOS}": _pmp_val,  # mantener para compatibilidad con columnas Excel
            f"participacion_{AÑO_FIN_DATOS}": _get(f"participacion_{AÑO_FIN_DATOS}", 0.0),
            "AAGR_ROBUSTO": _get("AAGR_ROBUSTO"),
            "salario_promedio_smlmv": _get("salario_promedio_smlmv"),
            f"pct_no_matriculados_{AÑO_FIN_DATOS}": _get(f"pct_no_matriculados_{AÑO_FIN_DATOS}"),
            f"num_programas_{AÑO_FIN_DATOS}": _get(f"num_programas_{AÑO_FIN_DATOS}", 0.0),
            "distancia_costo_pct": _get("distancia_costo_pct"),
            f"suma_matricula_{AÑO_FIN_DATOS}": _get(f"suma_matricula_{AÑO_FIN_DATOS}", 0.0),
            "programas_activos": _get("programas_activos", 0.0),
            "programas_nuevos_3a": _get("programas_nuevos_3a", 0.0),
            "programas_inactivos": _get("programas_inactivos", 0.0),
            "costo_promedio": _get("costo_promedio"),
            "pct_con_matricula": _get("pct_con_matricula", 0.0),
            # ── Campos para scoring contextual (no recomputables en 1 fila) ──
            # score_participacion: pre-calculado por apply_scoring() sobre el ag completo
            # (288 cats Colombia o referentes). Pasarlo evita el colapso de quintiles
            # que ocurre cuando _score_y_calificacion() llama apply_scoring() con 1 fila.
            "score_participacion": int(_get("score_participacion", 1)),
            # NIVEL_MAYORIT: nivel dominante de la categoría en el ag (ESP / MAE / UNIVERSITARIO).
            # Necesario para el árbol de decisión AAGR en apply_scoring() (scoring.py).
            # _get() solo maneja float; leer string directamente del row.
            "NIVEL_MAYORIT": (
                str(row["NIVEL_MAYORIT"]).strip()
                if "NIVEL_MAYORIT" in row.index and pd.notna(row["NIVEL_MAYORIT"])
                else "ESPECIALIZACIÓN"
            ),
            "SEÑAL_TENDENCIA": (
                str(row["SEÑAL_TENDENCIA"]).strip()
                if "SEÑAL_TENDENCIA" in row.index and pd.notna(row["SEÑAL_TENDENCIA"])
                else np.nan
            ),
        }
        resultados.append(met)

    if not resultados:
        return VACIAS.copy()
    if len(resultados) == 1:
        return resultados[0]

    # Promediar métricas de múltiples categorías (programas multi-categoría)
    prom: dict[str, float | str] = {}
    for k in resultados[0]:
        if k == "CAT_ID":
            cid = next(
                (str(m[k]).strip() for m in resultados if m.get(k) and str(m[k]).strip()),
                "",
            )
            prom[k] = cid
            continue
        # NIVEL_MAYORIT: tomar el nivel más frecuente (moda) entre las categorías
        if k == "NIVEL_MAYORIT":
            niveles = [m[k] for m in resultados if m.get(k)]
            prom[k] = max(set(niveles), key=niveles.count) if niveles else "ESPECIALIZACIÓN"
            continue
        if k == "SEÑAL_TENDENCIA":
            señales = [m[k] for m in resultados if isinstance(m.get(k), str) and m.get(k)]
            prom[k] = max(set(señales), key=señales.count) if señales else np.nan
            continue
        # score_participacion: redondear el promedio (es entero 1-5)
        vals = [m[k] for m in resultados if pd.notna(m.get(k))]
        if k == "score_participacion":
            prom[k] = int(round(float(np.mean(vals)))) if vals else 1
            continue
        prom[k] = float(np.mean(vals)) if vals else np.nan
    return prom


def _norm_cat(s: str) -> str:
    """Normaliza una cadena de categoría para comparación:
    mayúsculas, sin tildes, espacios colapsados."""
    s = str(s).strip().upper()
    s = _ud.normalize("NFD", s)
    s = "".join(c for c in s if _ud.category(c) != "Mn")
    return _re.sub(r"\s+", " ", s).strip()


def _ensure_cats_vectorizer() -> bool:
    """Construye vectorizador TF-IDF char_wb sobre CATEGORIA_FINAL de cat_ids.csv."""
    global _CATS_VECTORIZER, _CATS_MATRIX, _CATS_CANONICAS
    if _CATS_MATRIX is not None:
        return True
    if not _CAT_IDS_PATH.exists():
        return False
    try:
        df = pd.read_csv(_CAT_IDS_PATH, dtype=str)
        if "CATEGORIA_FINAL" not in df.columns:
            return False
        _CATS_CANONICAS = (
            df["CATEGORIA_FINAL"].dropna().astype(str).str.strip().tolist()
        )
        if not _CATS_CANONICAS:
            return False
        _CATS_VECTORIZER = TfidfVectorizer(
            ngram_range=(1, 3),
            max_features=50_000,
            sublinear_tf=True,
            analyzer="char_wb",
            min_df=1,
            strip_accents="unicode",
        )
        _CATS_MATRIX = _CATS_VECTORIZER.fit_transform(_CATS_CANONICAS)
        return True
    except Exception:
        return False


def _resolver_a_canonica(
    nombre: str,
    umbral: float = 0.55,
) -> tuple[str | None, float]:
    """
    Retorna (nombre_canonico, similitud) o (None, sim) si no hay match.

    1. Match exacto normalizado (_norm) contra cat_ids.csv
    2. Si no, TF-IDF char_wb coseno vs las categorías canónicas (umbral por defecto 0.55)
    """
    nombre = str(nombre).strip()
    if not nombre:
        return None, 0.0

    if not _ensure_cats_vectorizer():
        return None, 0.0

    assert _CATS_CANONICAS is not None
    norm_exact = {_norm(c): c for c in _CATS_CANONICAS}
    norm_key = _norm(nombre)
    if norm_key in norm_exact:
        return norm_exact[norm_key], 1.0

    assert _CATS_VECTORIZER is not None
    q_vec = _CATS_VECTORIZER.transform([nombre])
    sims = cosine_similarity(q_vec, _CATS_MATRIX).ravel()
    idx = int(sims.argmax())
    sim = float(sims[idx])
    if sim >= umbral:
        return _CATS_CANONICAS[idx], sim
    return None, sim


def _cargar_cat_ids_map() -> None:
    """Carga cat_ids.csv en mapas upper y _norm → CAT_ID."""
    global _CAT_IDS_MAP, _CAT_IDS_NORM_MAP
    if _CAT_IDS_MAP is not None:
        return
    _CAT_IDS_MAP = {}
    _CAT_IDS_NORM_MAP = {}
    if not _CAT_IDS_PATH.exists():
        return
    try:
        df = pd.read_csv(_CAT_IDS_PATH, dtype=str)
        if "CATEGORIA_FINAL" not in df.columns or "CAT_ID" not in df.columns:
            return
        for _, row in df.iterrows():
            cat = str(row["CATEGORIA_FINAL"]).strip()
            cid = str(row["CAT_ID"]).strip()
            if not cat or not cid:
                continue
            _CAT_IDS_MAP[cat.upper()] = cid
            _CAT_IDS_NORM_MAP[_norm(cat)] = cid
    except Exception:
        pass


def _cat_id_para(categoria_name: str) -> str:
    """CAT_ID estable desde cat_ids.csv (no depende del parquet regional)."""
    _cargar_cat_ids_map()
    assert _CAT_IDS_MAP is not None and _CAT_IDS_NORM_MAP is not None
    name = str(categoria_name).strip()
    if not name:
        return ""
    cid = _CAT_IDS_MAP.get(name.upper(), "")
    if cid:
        return cid
    cid = _CAT_IDS_NORM_MAP.get(_norm(name), "")
    if cid:
        return cid
    canon, _ = _resolver_a_canonica(name)
    if canon:
        cid = _CAT_IDS_MAP.get(canon.upper(), "")
        if cid:
            return cid
        return _CAT_IDS_NORM_MAP.get(_norm(canon), "")
    return ""


def _categorias_validas_desde_estudio() -> set[str]:
    """Categorías del estudio (hoja total) para validar splits compuestos."""
    ruta = ESTUDIO_MERCADO_DIR / "Estudio_Mercado_Colombia.xlsx"
    if not ruta.exists():
        return set()
    try:
        df_col = pd.read_excel(
            ruta, sheet_name="total", header=1, usecols=["CATEGORIA_FINAL"]
        )
        return set(df_col["CATEGORIA_FINAL"].dropna().astype(str).str.strip().unique())
    except Exception:
        return set()


def _expandir_categorias_compuestas(
    df: pd.DataFrame,
    cats_validas: set[str],
    col_cat: str = "CATEGORIA",
) -> pd.DataFrame:
    """
    Expande filas cuya columna col_cat contiene múltiples categorías
    unidas por '-' en filas independientes, una por categoría válida.

    Reglas:
    - Si la categoría completa ya coincide con una válida → sin cambio.
    - Si no coincide → dividir en '-' (con espacios opcionales), validar
      cada fragmento, crear una fila por fragmento válido.
    - Fragmentos sin match en cats_validas → se descartan con log_warning.
    - Si ningún fragmento es válido → conservar fila original sin cambio
      (para no perder datos) y emitir log_warning.
    - La función es idempotente: si ya está correctamente separada,
      no hace nada.

    Args:
        df: DataFrame con los programas a valorizar.
        cats_validas: conjunto de categorías normalizadas válidas.
        col_cat: nombre de la columna de categoría en df.

    Returns:
        DataFrame expandido (puede tener más filas que el original).
    """
    if not cats_validas:
        return df.copy()
    if col_cat not in df.columns:
        return df.copy()

    cats_norm_map: dict[str, str] = {_norm_cat(c): c for c in cats_validas}
    filas_out: list[pd.Series] = []

    for _, fila in df.iterrows():
        cat_original = str(fila.get(col_cat, "") or "").strip()

        # Caso 1: ya es una categoría válida → sin cambio
        if _norm_cat(cat_original) in cats_norm_map:
            filas_out.append(fila)
            continue

        # Caso 2: contiene '-' → intentar split
        if "-" in cat_original:
            partes = [p.strip() for p in _re.split(r"\s*-\s*", cat_original)]
            partes = [p for p in partes if p]  # quitar vacíos

            validas: list[str] = []
            invalidas: list[str] = []
            for parte in partes:
                key = _norm_cat(parte)
                if key in cats_norm_map:
                    validas.append(cats_norm_map[key])  # nombre canónico
                    continue
                canon, sim = _resolver_a_canonica(parte)
                if canon is not None:
                    canon_key = _norm_cat(canon)
                    if canon_key in cats_norm_map:
                        nombre_canon = cats_norm_map[canon_key]
                        validas.append(nombre_canon)
                        log_info(
                            f"[Valorización] Categoría resuelta automáticamente: "
                            f"'{parte}' → '{nombre_canon}' (sim={sim:.2f})"
                        )
                        continue
                invalidas.append(parte)

            if invalidas:
                prog = fila.get("PROGRAMA_EAFIT", fila.get(col_cat, "?"))
                log_warning(
                    f"[Valorización] '{prog}': fragmentos no encontrados "
                    f"en las 288 categorías → {invalidas}. Se descartan."
                )

            if validas:
                for cat_valida in validas:
                    nueva = fila.copy()
                    nueva[col_cat] = cat_valida
                    filas_out.append(nueva)
                continue  # ya procesado

        # Caso 3: no tiene '-' y no matchea → intentar resolución fuzzy
        canon, sim = _resolver_a_canonica(cat_original)
        if canon is not None:
            canon_key = _norm_cat(canon)
            if canon_key in cats_norm_map:
                nombre_canon = cats_norm_map[canon_key]
                log_info(
                    f"[Valorización] Categoría resuelta automáticamente: "
                    f"'{cat_original}' → '{nombre_canon}' (sim={sim:.2f})"
                )
                nueva = fila.copy()
                nueva[col_cat] = nombre_canon
                filas_out.append(nueva)
                continue

        prog = fila.get("PROGRAMA_EAFIT", cat_original)
        log_warning(
            f"[Valorización] Categoría '{cat_original}' (programa: '{prog}') "
            f"no se encontró en las 288 categorías válidas. "
            f"Se conserva sin modificar."
        )
        filas_out.append(fila)

    if not filas_out:
        return df.iloc[0:0].copy()  # DataFrame vacío con mismas columnas

    return pd.DataFrame(filas_out).reset_index(drop=True)


def _proyeccion_regresion_lineal(
    serie_pc: dict[int, float],
    año_objetivo: int,
    tasa_captura: float,
) -> float:
    """
    Proyecta estudiantes esperados para EAFIT usando regresión lineal
    sobre la serie histórica de primer_curso de la categoría × región.

    Args:
        serie_pc: {año: promedio_primer_curso} — solo años con datos válidos.
        año_objetivo: año al que se proyecta (horizonte fijo AÑO_FIN_DATOS + 2).
        tasa_captura: fracción del mercado que capturaría EAFIT (TASA_CAPTURA_EAFIT).

    Returns:
        Estudiantes proyectados (entero ≥ 1), o np.nan si datos insuficientes.
    """
    pts = [
        (yr, val)
        for yr, val in serie_pc.items()
        if pd.notna(val) and val > 0
    ]
    if len(pts) < 3:
        # Insuficientes puntos para regresión confiable
        return np.nan

    xs = np.array([p[0] for p in pts], dtype=float)
    ys = np.array([p[1] for p in pts], dtype=float)

    # Regresión lineal por mínimos cuadrados (grado 1)
    coef = np.polyfit(xs, ys, 1)          # [pendiente, intercepto]
    pred_mercado = np.polyval(coef, año_objetivo)

    if pred_mercado <= 0:
        return np.nan

    return max(1, round(pred_mercado * tasa_captura))


def _serie_primer_curso_sub(sub: pd.DataFrame) -> dict[int, float]:
    """Sumatoria anual de primer_curso_YYYY en el subconjunto (categoría × región).

    Usa sum() y no mean() porque la regresión proyecta el tamaño total del
    mercado (todos los programas de esa categoría × región), sobre el que luego
    se aplica TASA_CAPTURA_EAFIT para estimar los estudiantes de EAFIT.
    """
    _serie_pc: dict[int, float] = {}
    for _yr in range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1):
        _col_yr = f"primer_curso_{_yr}"
        if _col_yr in sub.columns:
            _val = float(sub[_col_yr].sum())
            if not np.isnan(_val) and _val > 0:
                _serie_pc[_yr] = _val
    return _serie_pc


def _subconjunto_por_categorias(df: pd.DataFrame, categorias: list[str]) -> pd.DataFrame:
    """Filtra filas de la sábana cuya CATEGORIA_FINAL coincide con alguna categoría."""
    if df is None or len(df) == 0 or "CATEGORIA_FINAL" not in df.columns:
        return pd.DataFrame()
    work = df
    col_norm = "_cat_norm_cached"
    if col_norm not in work.columns:
        work = work.copy()
        work[col_norm] = work["CATEGORIA_FINAL"].apply(lambda x: _norm(str(x)))
    mask = pd.Series(False, index=work.index)
    for cat in categorias:
        mask = mask | (work[col_norm] == _norm(cat))
    return work[mask]


def _agregar_metricas_categoria(
    df_region: pd.DataFrame,
    categorias: list[str],
) -> dict:
    """
    Filtra df_region por las categorías indicadas, agrega métricas y retorna
    un dict listo para apply_scoring. Si son varias categorías (programa multi-categoría),
    promedia las métricas de cada una.
    """
    col_norm = "_cat_norm_cached"
    if col_norm not in df_region.columns:
        df_region = df_region.copy()
        df_region[col_norm] = df_region["CATEGORIA_FINAL"].apply(lambda x: _norm(str(x)))

    resultados_por_cat = []
    for cat in categorias:
        cat_limpia = _norm(cat)
        mask = df_region[col_norm] == cat_limpia
        sub = df_region[mask]
        resultados_por_cat.append(_metricas_de_subconjunto(sub, df_region))

    if len(resultados_por_cat) == 1:
        return resultados_por_cat[0]

    # Promediar métricas de múltiples categorías
    met_prom: dict = {}
    for k in resultados_por_cat[0]:
        if k == "serie_primer_curso":
            met_prom[k] = resultados_por_cat[0].get(k, {})
            continue
        if k == "SEÑAL_TENDENCIA":
            señales = [
                m[k] for m in resultados_por_cat
                if isinstance(m.get(k), str) and m.get(k)
            ]
            met_prom[k] = max(set(señales), key=señales.count) if señales else np.nan
            continue
        vals = [m[k] for m in resultados_por_cat if pd.notna(m.get(k))]
        met_prom[k] = float(np.mean(vals)) if vals else np.nan
    return met_prom


def _metricas_de_subconjunto(sub: pd.DataFrame, df_region_completo: pd.DataFrame) -> dict:
    """
    Calcula las métricas de una categoría específica dentro de un DataFrame regional.
    Produce exactamente las variables que necesita apply_scoring().
    """
    if len(sub) == 0:
        return {
            f"prom_matricula_por_programa_{AÑO_FIN_DATOS}": 0.0,
            f"prom_matricula_{AÑO_FIN_DATOS}": 0.0,
            f"participacion_{AÑO_FIN_DATOS}": 0.0,
            "AAGR_ROBUSTO": np.nan,
            "salario_promedio_smlmv": np.nan,
            f"pct_no_matriculados_{AÑO_FIN_DATOS}": np.nan,
            f"num_programas_{AÑO_FIN_DATOS}": 0,
            "distancia_costo_pct": np.nan,
            # Extras para el Excel
            f"suma_primer_curso_{AÑO_FIN_DATOS}": 0.0,
            f"suma_matricula_{AÑO_FIN_DATOS}": 0,
            "programas_activos": 0,
            "programas_nuevos_3a": 0,
            "programas_inactivos": 0,
            "costo_promedio": np.nan,
            "pct_con_matricula": 0.0,
            "serie_primer_curso": {},
        }

    # Matrícula
    # Preferir primer_curso (flujo de nuevos) — mismo criterio que pipeline principal.
    # Fallback a matricula si no existe la columna en la sábana
    _pc_col = f"primer_curso_{AÑO_FIN_DATOS}" if f"primer_curso_{AÑO_FIN_DATOS}" in sub.columns else f"matricula_{AÑO_FIN_DATOS}"
    prom_mat = float(sub[_pc_col].mean()) if _pc_col in sub.columns else 0.0
    suma_mat = float(sub[_pc_col].sum()) if _pc_col in sub.columns else 0.0
    num_prog = int((sub[f"matricula_{AÑO_FIN_DATOS}"] > 0).sum()) if f"matricula_{AÑO_FIN_DATOS}" in sub.columns else 0

    # Participación sobre primer_curso (flujo), no sobre stock total — igual que pipeline principal
    _pc_col_reg = f"primer_curso_{AÑO_FIN_DATOS}" if f"primer_curso_{AÑO_FIN_DATOS}" in df_region_completo.columns else f"matricula_{AÑO_FIN_DATOS}"
    if _pc_col_reg in df_region_completo.columns and "CATEGORIA_FINAL" in df_region_completo.columns:
        todos_proms = df_region_completo.groupby("CATEGORIA_FINAL")[_pc_col_reg].mean()
        total_prom_sum = todos_proms.sum()
        participacion = prom_mat / total_prom_sum if total_prom_sum > 0 else 0.0
    else:
        participacion = 0.0

    # AAGR_ROBUSTO
    aagr = float(sub["AAGR_ROBUSTO"].mean()) if "AAGR_ROBUSTO" in sub.columns else np.nan

    # Salario en SMLMV (el campo ya viene en SMLMV de la sábana)
    salario_smlmv = float(sub["SALARIO_OLE"].mean()) if "SALARIO_OLE" in sub.columns else np.nan

    # Pct no matriculados 2024
    pct_no_mat = np.nan
    # Fórmula: (inscritos - primer_curso) / inscritos
    # NO comparar vs matricula_total (genera negativos porque acumula cohortes previas)
    _ins_col = next(
        (c for c in [f"inscritos_{AÑO_FIN_DATOS}", f"inscritos_{AÑO_FIN_DATOS}_suma"] if c in sub.columns),
        None,
    )
    _pc_col_pct = next((c for c in [f"primer_curso_{AÑO_FIN_DATOS}"] if c in sub.columns), None)
    if _ins_col and _pc_col_pct:
        ins = float(sub[_ins_col].sum())
        pc = float(sub[_pc_col_pct].sum())
        if ins > 0:
            pct_no_mat = float(np.clip((ins - pc) / ins, 0.0, 1.0))
        elif f"pct_no_matriculados_{AÑO_FIN_DATOS}" in sub.columns:
            pct_no_mat = float(sub[f"pct_no_matriculados_{AÑO_FIN_DATOS}"].mean())
    elif f"pct_no_matriculados_{AÑO_FIN_DATOS}" in sub.columns:
        pct_no_mat = float(sub[f"pct_no_matriculados_{AÑO_FIN_DATOS}"].mean())

    # Programas activos / inactivos / nuevos
    prog_activos = int(sub["es_activo"].sum()) if "es_activo" in sub.columns else len(sub)
    prog_inactivos = len(sub) - prog_activos
    prog_nuevos = int(sub["nuevo_en_snies_3a"].sum()) if "nuevo_en_snies_3a" in sub.columns else 0

    # Costo y distancia
    costo_col = "COSTO_MATRÍCULA_ESTUD_NUEVOS"
    costo = float(sub[costo_col].mean()) if costo_col in sub.columns else np.nan
    dist_costo = float(sub["_distancia_costo_prog"].mean()) if "_distancia_costo_prog" in sub.columns else np.nan

    pct_con_mat = num_prog / prog_activos if prog_activos > 0 else 0.0

    # ── Serie histórica primer_curso para regresión ─────────────────────
    _serie_pc = _serie_primer_curso_sub(sub)

    return {
        f"prom_matricula_por_programa_{AÑO_FIN_DATOS}": prom_mat,  # nombre primario para scoring.py
        f"prom_matricula_{AÑO_FIN_DATOS}": prom_mat,  # alias para columnas Excel
        f"participacion_{AÑO_FIN_DATOS}": participacion,
        "AAGR_ROBUSTO": aagr,
        "salario_promedio_smlmv": salario_smlmv,
        f"pct_no_matriculados_{AÑO_FIN_DATOS}": pct_no_mat,
        f"num_programas_{AÑO_FIN_DATOS}": num_prog,
        "distancia_costo_pct": dist_costo,
        # Extras para mostrar en el Excel
        f"suma_primer_curso_{AÑO_FIN_DATOS}": suma_mat,
        f"suma_matricula_{AÑO_FIN_DATOS}": suma_mat,
        "programas_activos": prog_activos,
        "programas_nuevos_3a": prog_nuevos,
        "programas_inactivos": prog_inactivos,
        "costo_promedio": costo,
        "pct_con_matricula": pct_con_mat,
        "serie_primer_curso": _serie_pc,
    }


def _score_y_calificacion(metricas: dict) -> dict:
    """Aplica scoring.py a las métricas y retorna el dict enriquecido con scores y calificacion_final."""
    # Valor de scoring de matrícula: preferir prom_matricula_por_programa_2024 (= primer_curso)
    # que scoring.py busca primero. Fallback a prom_matricula_2024 si no está.
    _pmp_scoring = metricas.get(
        f"prom_matricula_por_programa_{AÑO_FIN_DATOS}",
        metricas.get(f"prom_matricula_{AÑO_FIN_DATOS}", 0),
    )
    df_tmp = pd.DataFrame(
        [
            {
                f"prom_matricula_por_programa_{AÑO_FIN_DATOS}": _pmp_scoring,
                f"prom_matricula_{AÑO_FIN_DATOS}": _pmp_scoring,
                f"participacion_{AÑO_FIN_DATOS}": metricas.get(f"participacion_{AÑO_FIN_DATOS}", 0),
                "NIVEL_MAYORIT": metricas.get("NIVEL_MAYORIT", "ESPECIALIZACIÓN"),
                "AAGR_ROBUSTO": metricas.get("AAGR_ROBUSTO", np.nan),
                "salario_promedio_smlmv": metricas.get("salario_promedio_smlmv", np.nan),
                f"pct_no_matriculados_{AÑO_FIN_DATOS}": metricas.get(f"pct_no_matriculados_{AÑO_FIN_DATOS}", np.nan),
                f"num_programas_{AÑO_FIN_DATOS}": metricas.get(f"num_programas_{AÑO_FIN_DATOS}", 0),
                "distancia_costo_pct": metricas.get("distancia_costo_pct", np.nan),
            }
        ]
    )
    df_scored = apply_scoring(df_tmp, modo_local=False)
    row = df_scored.iloc[0]

    # apply_scoring con 1 fila colapsa score_participacion a 1; corregir calificacion_final.
    score_part_precomp = int(metricas.get("score_participacion", 1))
    score_part_colapso = int(row.get("score_participacion", 1))
    cal_final = float(row.get("calificacion_final", 1.0))
    cal_final_correcto = round(
        cal_final
        - score_part_colapso * _SCORE_PARTICIPACION_PESO
        + score_part_precomp * _SCORE_PARTICIPACION_PESO,
        4,
    )

    return {
        **metricas,
        "score_matricula":          row.get("score_matricula", 1),
        # Usar score_participacion pre-calculado del parquet (no el de apply_scoring con 1 fila).
        # apply_scoring con 1 fila siempre produce score 1 por colapso de quintiles dinámicos.
        "score_participacion":      metricas.get("score_participacion", row.get("score_participacion", 1)),
        # score_AAGR: ahora aplica árbol ESP/MAE porque NIVEL_MAYORIT está en df_tmp
        "score_AAGR":               row.get("score_AAGR", 1),
        "score_salario":            row.get("score_salario", 1),
        "score_pct_no_matriculados": row.get("score_pct_no_matriculados", 1),
        "score_num_programas":      row.get("score_num_programas", 1),
        "score_costo":              row.get("score_costo", 1),
        "calificacion_final":       cal_final_correcto,
    }


def _construir_tabla_crecimiento(sabana: pd.DataFrame) -> pd.DataFrame:
    """
    Construye tabla de crecimiento de mercado con 9 combinaciones de
    SEGMENTO × SECTOR × MODALIDAD, para ESP y MAE.

    Estructura: bloques apilados con cabecera de filtros + 3 filas de datos
    (ESPECIALIZACIÓN / MAESTRÍA / Total) por combinación.

    Fuente: archivos primer_curso_YYYY.xlsx en ref/backup/ (no la sábana).
    El parámetro sabana se conserva por compatibilidad de firma.
    """
    del sabana  # no se usa; datos desde backups SNIES

    CONTEXTOS = {
        "Colombia — Todo el mercado": (
            "Universo completo del mercado nacional de posgrado y pregrado.\n"
            "Incluye todas las metodologías (presencial, virtual, a distancia "
            "e híbrida) y todos los sectores de IES (pública y privada).\n"
            "Referencia de demanda bruta sin ningún filtro: útil para medir "
            "el tamaño real del mercado y las tendencias de largo plazo."
        ),
        "Colombia — No Virtual · Todos sectores": (
            "Excluye programas 100% virtuales.\n"
            "Muestra la evolución de la demanda presencial, a distancia "
            "e híbrida a nivel nacional.\n"
            "Contexto relevante para entender cuánto del crecimiento nacional "
            "corresponde a la virtualidad vs. la oferta tradicional."
        ),
        "Colombia — Presencial puro · Todos sectores": (
            "Solo programas presenciales puros, toda clase de IES.\n"
            "Es el segmento más cercano al modelo educativo de EAFIT y de "
            "sus competidores directos (Uniandes, ICESI, Rosario, Javeriana).\n"
            "Los CAGRs negativos post-2019 reflejan la migración de estudiantes "
            "hacia la virtualidad, especialmente en Especialización."
        ),
        "Colombia — Virtual · Todos sectores": (
            "Solo programas 100% virtuales a nivel nacional.\n"
            "Segmento de mayor crecimiento: compite por el perfil de estudiante "
            "trabajador que busca flexibilidad horaria.\n"
            "El salto de 2018-2019 (+93%) refleja la consolidación de plataformas "
            "virtuales; el de 2020-2021 (+19%) el efecto pandemia."
        ),
        "Colombia — Privada · Presencial puro": (
            "IES privadas, modalidad presencial exclusivamente.\n"
            "Competencia directa de EAFIT a nivel nacional: universidades privadas "
            "de calidad con campus físico.\n"
            "Benchmark para fijar participación de mercado esperada en nuevos "
            "programas presenciales."
        ),
        "Colombia — Privada · Virtual": (
            "IES privadas en modalidad 100% virtual.\n"
            "Muestra el crecimiento de la oferta digital de competidores privados "
            "directos de EAFIT.\n"
            "Relevante para programas de EAFIT que podrían tener versión virtual "
            "o enfrentar sustitución por parte de esta oferta."
        ),
        "Antioquia — Todo el mercado": (
            "Mercado regional de Antioquia, todas metodologías y sectores.\n"
            "Zona de influencia primaria de EAFIT: el mercado local más relevante "
            "para evaluar demanda de nuevos programas presenciales.\n"
            "El CAGR 2019-2024 negativo en Especialización refleja pérdida de "
            "estudiantes hacia Bogotá y plataformas virtuales nacionales."
        ),
        "Bogotá — Todo el mercado": (
            "Capital del país y mercado más grande (≈49% del mercado nacional "
            "en Especialización y Maestría).\n"
            "Refleja tendencias con antelación al resto del país.\n"
            "Relevante para programas EAFIT con modalidad virtual o planes de "
            "expansión fuera de Antioquia."
        ),
        "Eje Cafetero — Todo el mercado": (
            "Caldas, Risaralda y Quindío.\n"
            "Zona de influencia secundaria de EAFIT; mercado relevante para "
            "programas que pueden atraer estudiantes de Manizales, Pereira y "
            "Armenia.\n"
            "Tamaño pequeño pero con buena tasa de crecimiento reciente en "
            "Maestría (CAGR 2021-2024: +2.2%)."
        ),
    }

    años = list(range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1))
    yr_ini = años[0]
    yr_fin = años[-1]

    def _norm_sheet(name: object) -> str:
        s2 = _ud.normalize("NFD", str(name))
        return "".join(ch for ch in s2 if _ud.category(ch) != "Mn").upper()

    def _norm_sector(v: object) -> str:
        u = str(v).upper().strip()
        if u in ("PRIVADA", "PRIVADO"):
            return "PRIVADA"
        return u

    def _norm_dpto(v: object) -> str:
        """Unifica variantes SNIES: tildes, puntos finales, comas."""
        s = str(v).strip().upper().replace(",", "")
        s = "".join(
            c for c in _ud.normalize("NFD", s) if _ud.category(c) != "Mn"
        )
        return s.replace(".", "").strip()

    def _norm_modalidad(v: object) -> str:
        """PRESENCIAL | VIRTUAL | A DISTANCIA | HIBRIDA | OTROS"""
        u = str(v).upper().strip()
        _es_virtual_puro = (
            u == "VIRTUAL"
            or u in ("DISTANCIA (VIRTUAL)", "A DISTANCIA (VIRTUAL)")
            or (u.startswith("VIRTUAL") and "PRESENCIAL" not in u)
        )
        if _es_virtual_puro:
            return "VIRTUAL"
        if u == "PRESENCIAL":
            return "PRESENCIAL"
        if "DISTANCIA" in u or "TRADICIONAL" in u:
            return "A DISTANCIA"
        if "PRESENCIAL" in u and ("VIRTUAL" in u or "DUAL" in u):
            return "HIBRIDA"
        if "DUAL" in u or "HIBRIDA" in u or "HÍBRIDA" in u:
            return "HIBRIDA"
        return "OTROS"

    def _cargar_año(yr: int) -> pd.DataFrame | None:
        """Carga y normaliza primer_curso_YYYY → SECTOR, MOD, DPTO, NIVEL, PC."""
        candidatos = [
            REF_DIR / "backup" / "matriculas primer curso" / f"primer_curso_{yr}.xlsx",
            REF_DIR / "backup" / f"primer_curso_{yr}.xlsx",
            RAW_HISTORIC_DIR / f"primer_curso_{yr}.xlsx",
            REF_DIR / f"primer_curso_{yr}.xlsx",
        ]
        ruta = next((p for p in candidatos if p.exists()), None)
        if ruta is None:
            log_warning(f"[Crecimiento] {yr}: archivo no encontrado.")
            return None

        try:
            import openpyxl

            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            try:
                hoja = next(
                    (s for s in wb.sheetnames if "INDICE" not in _norm_sheet(s)),
                    wb.sheetnames[-1],
                )
            finally:
                wb.close()

            df_raw: pd.DataFrame | None = None
            for hdr in range(5, 13):
                try:
                    cand = pd.read_excel(ruta, sheet_name=hoja, header=hdr, dtype=str)
                except Exception:
                    continue
                cols_up = {str(c).upper().strip(): c for c in cand.columns}
                tiene_nivel = any("NIVEL" in k and "FORMAC" in k for k in cols_up)
                tiene_pc = any("PRIMER" in k for k in cols_up)
                if tiene_nivel and tiene_pc:
                    df_raw = cand
                    break

            if df_raw is None:
                preview = pd.read_excel(ruta, sheet_name=hoja, header=None, nrows=40)
                header_idx: int | None = None
                for idx in range(len(preview)):
                    vals = [
                        str(v).strip()
                        for v in preview.iloc[idx].tolist()
                        if pd.notna(v) and str(v).strip()
                    ]
                    if not vals or len(vals) < 5:
                        continue
                    first_n = _ud.normalize("NFD", vals[0])
                    first_n = "".join(
                        ch for ch in first_n if _ud.category(ch) != "Mn"
                    ).upper()
                    vals_n = [
                        "".join(
                            ch for ch in _ud.normalize("NFD", v)
                            if _ud.category(ch) != "Mn"
                        ).upper()
                        for v in vals
                    ]
                    if "CODIGO" in first_n and (
                        any("SNIES" in v for v in vals_n) or "INSTITUC" in first_n
                    ):
                        header_idx = idx
                        break
                if header_idx is not None:
                    cand = pd.read_excel(
                        ruta, sheet_name=hoja, header=header_idx, dtype=str
                    )
                    cols_up = {str(c).upper().strip(): c for c in cand.columns}
                    if any("NIVEL" in k and "FORMAC" in k for k in cols_up) and any(
                        "PRIMER" in k for k in cols_up
                    ):
                        df_raw = cand

            if df_raw is None:
                log_warning(f"[Crecimiento] {yr}: no se detectó encabezado válido.")
                return None

            cols_up = {str(c).upper().strip(): c for c in df_raw.columns}

            col_niv = next(
                (
                    orig
                    for key, orig in cols_up.items()
                    if "NIVEL" in key and "FORMAC" in key and not key.startswith("ID")
                ),
                None,
            )
            col_pc = next((orig for key, orig in cols_up.items() if "PRIMER" in key), None)
            col_sec = next(
                (
                    orig
                    for key, orig in cols_up.items()
                    if "SECTOR" in key and "IES" in key and not key.startswith("ID")
                ),
                None,
            )
            col_mod = next(
                (
                    orig
                    for key, orig in cols_up.items()
                    if key in ("METODOLOGÍA", "MODALIDAD", "METODOLOGIA")
                    and not key.startswith("ID")
                ),
                None,
            )
            if col_mod is None:
                col_mod = next(
                    (
                        orig
                        for key, orig in cols_up.items()
                        if "METODOL" in key and not key.startswith("ID")
                    ),
                    None,
                )
            col_dpto = next(
                (orig for key, orig in cols_up.items() if "DEPARTAMENTO" in key and "OFERTA" in key),
                None,
            )

            if not col_niv or not col_pc:
                log_warning(f"[Crecimiento] {yr}: columnas nivel/pc no encontradas.")
                return None

            out = pd.DataFrame()
            out["NIVEL"] = df_raw[col_niv].astype(str).str.strip().str.upper()
            out["PC"] = pd.to_numeric(df_raw[col_pc], errors="coerce").fillna(0)
            out["SECTOR"] = (
                df_raw[col_sec].apply(_norm_sector) if col_sec else "DESCONOCIDO"
            )
            out["MOD"] = (
                df_raw[col_mod].apply(_norm_modalidad) if col_mod else "PRESENCIAL"
            )
            out["DPTO"] = (
                df_raw[col_dpto].apply(_norm_dpto) if col_dpto else ""
            )
            mask = out["NIVEL"].str.contains("ESPECIALI|MAESTR", na=False)
            return out[mask].copy()

        except Exception as e:
            log_warning(f"[Crecimiento] {yr}: error al leer — {e}")
            return None

    datos: dict[int, pd.DataFrame | None] = {}
    for yr in años:
        datos[yr] = _cargar_año(yr)
        if datos[yr] is not None:
            log_info(f"[Crecimiento] {yr}: {len(datos[yr]):,} filas cargadas.")

    def _cagr(serie: dict[int, float], y0: int, y1: int) -> float | None:
        v0 = serie.get(y0, 0) or 0
        v1 = serie.get(y1, 0) or 0
        n = y1 - y0
        if v0 <= 0 or v1 <= 0 or n <= 0:
            return None
        return round((v1 / v0) ** (1 / n) - 1, 4)

    def _suma_serie(
        sec_filter: str | None,
        mod_filter: list[str],
        dpto_filter: list[str] | None,
        nivel_contains: str,
    ) -> dict[int, float]:
        serie: dict[int, float] = {}
        for yr in años:
            df_yr = datos.get(yr)
            if df_yr is None or df_yr.empty:
                serie[yr] = 0.0
                continue
            mask = df_yr["NIVEL"].str.contains(nivel_contains, na=False)
            if sec_filter:
                mask &= df_yr["SECTOR"] == sec_filter
            if mod_filter:
                mask &= df_yr["MOD"].isin(mod_filter)
            if dpto_filter:
                dptos_norm = [_norm_dpto(d) for d in dpto_filter]
                mask &= df_yr["DPTO"].isin(dptos_norm)
            serie[yr] = float(df_yr.loc[mask, "PC"].sum())
        return serie

    COMBOS = [
        (
            "Colombia — Todo el mercado",
            None,
            ["PRESENCIAL", "VIRTUAL", "A DISTANCIA", "HIBRIDA"],
            None,
        ),
        (
            "Colombia — No Virtual · Todos sectores",
            None,
            ["PRESENCIAL", "A DISTANCIA", "HIBRIDA"],
            None,
        ),
        (
            "Colombia — Presencial puro · Todos sectores",
            None,
            ["PRESENCIAL"],
            None,
        ),
        (
            "Colombia — Virtual · Todos sectores",
            None,
            ["VIRTUAL"],
            None,
        ),
        (
            "Colombia — Privada · Presencial puro",
            "PRIVADA",
            ["PRESENCIAL"],
            None,
        ),
        (
            "Colombia — Privada · Virtual",
            "PRIVADA",
            ["VIRTUAL"],
            None,
        ),
        (
            "Antioquia — Todo el mercado",
            None,
            ["PRESENCIAL", "VIRTUAL", "A DISTANCIA", "HIBRIDA"],
            ["ANTIOQUIA"],
        ),
        (
            "Bogotá — Todo el mercado",
            None,
            ["PRESENCIAL", "VIRTUAL", "A DISTANCIA", "HIBRIDA"],
            ["BOGOTÁ D.C.", "BOGOTA D.C."],
        ),
        (
            "Eje Cafetero — Todo el mercado",
            None,
            ["PRESENCIAL", "VIRTUAL", "A DISTANCIA", "HIBRIDA"],
            ["CALDAS", "RISARALDA", "QUINDÍO"],
        ),
    ]

    filas: list[dict] = []
    for nombre, sec_f, mod_f, dpto_f in COMBOS:
        serie_esp = _suma_serie(sec_f, mod_f, dpto_f, "ESPECIALI")
        serie_mae = _suma_serie(sec_f, mod_f, dpto_f, "MAESTR")
        serie_tot = {yr: serie_esp[yr] + serie_mae[yr] for yr in años}

        sector_txt = sec_f if sec_f else "(Todos)"
        mod_txt = " · ".join(mod_f) if mod_f else "(Todas)"
        dpto_txt = ", ".join(dpto_f) if dpto_f else "(Todos)"

        filas.append({
            "__tipo": "header",
            "__campo": "SEGMENTO",
            "__valor": nombre,
            "__contexto": CONTEXTOS.get(nombre, ""),
        })
        filas.append({"__tipo": "header", "__campo": "SECTOR IES", "__valor": sector_txt})
        filas.append({"__tipo": "header", "__campo": "METODOLOGÍA", "__valor": mod_txt})
        filas.append({"__tipo": "header", "__campo": "DEPARTAMENTO", "__valor": dpto_txt})
        filas.append({"__tipo": "blank"})

        for nivel_label, serie in [
            ("ESPECIALIZACIÓN", serie_esp),
            ("MAESTRÍA", serie_mae),
            ("Total", serie_tot),
        ]:
            fila: dict = {"__tipo": "data", "NIVEL": nivel_label}
            total_fila = 0.0
            for yr in años:
                v = serie.get(yr, 0) or 0
                fila[str(yr)] = v if v > 0 else None
                if v:
                    total_fila += v
            fila["Total general"] = total_fila if total_fila > 0 else None
            fila[f"CAGR {yr_ini}-{yr_fin}"] = _cagr(serie, yr_ini, yr_fin)
            fila["CAGR 2019-2024"] = _cagr(serie, 2019, yr_fin)
            fila["CAGR 2021-2024"] = _cagr(serie, 2021, yr_fin)
            filas.append(fila)

        filas.append({"__tipo": "blank"})

    return pd.DataFrame(filas)


def _construir_hoja_proyecciones(
    df_programas_val: pd.DataFrame,
    sabana: pd.DataFrame | None,
    log: Callable = print,
) -> pd.DataFrame:
    """
    Construye la hoja Proyecciones con horizonte 2027-2031.

    Una fila por programa EAFIT: P1 (regresión × participación) y
    P2 (análogos P25 / mediana / P75 o CAGR fallback).
    """
    from unidecode import unidecode

    AÑOS_DATOS = list(range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1))
    AÑOS_PROY = [2027, 2028, 2029, 2030, 2031]
    N_MIN = 2
    N_AÑOS_MIN = 3

    def _n(s: object) -> str:
        s = str(s).upper().strip()
        s = unidecode(s)
        s = _re.sub(r"[^A-Z0-9\s]", " ", s)
        return _re.sub(r"\s+", " ", s).strip()

    log("  [Proyecciones] Cargando series históricas primer_curso...")

    def _keys_sin_acento(keys: list[str]) -> dict[str, str]:
        return {
            k: "".join(
                c for c in _ud.normalize("NFD", k) if _ud.category(c) != "Mn"
            )
            for k in keys
        }

    def _detectar_cols_programa(columns: list[str]) -> tuple[str | None, str | None, str | None]:
        """IES, programa académico, primer curso — desde nombres de columna SNIES."""
        keys = [str(c).upper().replace("\n", " ").strip() for c in columns]
        kn = _keys_sin_acento(keys)

        col_ies = next(
            (
                keys[i]
                for i, k in enumerate(keys)
                if "INSTITUC" in kn[k]
                and "IES" in kn[k]
                and "PADRE" not in kn[k]
                and not k.startswith("ID")
            ),
            None,
        )
        col_prog = next(
            (
                keys[i]
                for i, k in enumerate(keys)
                if "PROGRAMA" in kn[k]
                and "ACAD" in kn[k]
                and "CODIGO" not in kn[k]
                and "ACRED" not in kn[k]
                and not k.startswith("ID")
            ),
            None,
        )
        if col_prog is None:
            col_prog = next(
                (
                    keys[i]
                    for i, k in enumerate(keys)
                    if "PROGRAMA" in kn[k]
                    and "CODIGO" not in kn[k]
                    and "ACRED" not in kn[k]
                    and not k.startswith("ID")
                ),
                None,
            )
        col_pc = next(
            (keys[i] for i, k in enumerate(keys) if "PRIMER" in kn[k]),
            None,
        )
        return col_ies, col_prog, col_pc

    filas_hist: list[pd.DataFrame] = []
    for yr in AÑOS_DATOS:
        candidatos = [
            REF_DIR / "backup" / "matriculas primer curso" / f"primer_curso_{yr}.xlsx",
            REF_DIR / "backup" / f"primer_curso_{yr}.xlsx",
            RAW_HISTORIC_DIR / f"primer_curso_{yr}.xlsx",
            REF_DIR / f"primer_curso_{yr}.xlsx",
        ]
        ruta = next((p for p in candidatos if p.exists()), None)
        if ruta is None:
            continue
        try:
            import openpyxl as _opxl

            _wb = _opxl.load_workbook(ruta, read_only=True, data_only=True)
            _hoja = next(
                (s for s in _wb.sheetnames if "NDICE" not in s.upper()),
                _wb.sheetnames[-1],
            )
            _wb.close()

            _df_raw: pd.DataFrame | None = None
            _hdr_ok: int | None = None
            for hdr in range(5, 13):
                try:
                    _cand = pd.read_excel(
                        ruta, sheet_name=_hoja, header=hdr, dtype=str, nrows=5
                    )
                    _col_ies, _col_prog, _col_pc = _detectar_cols_programa(
                        list(_cand.columns)
                    )
                    if all([_col_ies, _col_prog, _col_pc]):
                        _hdr_ok = hdr
                        break
                except Exception:
                    continue

            if _hdr_ok is None:
                continue

            _df_raw = pd.read_excel(
                ruta, sheet_name=_hoja, header=_hdr_ok, dtype=str
            )
            _col_ies, _col_prog, _col_pc = _detectar_cols_programa(list(_df_raw.columns))
            if not all([_col_ies, _col_prog, _col_pc]):
                continue

            _df_raw["_PC"] = pd.to_numeric(_df_raw[_col_pc], errors="coerce").fillna(0)
            _df_raw["_IES"] = _df_raw[_col_ies].apply(_n)
            _df_raw["_PROG"] = _df_raw[_col_prog].apply(_n)
            _df_raw["_AÑO"] = yr

            _grp = (
                _df_raw.groupby(["_IES", "_PROG", "_AÑO"])["_PC"]
                .sum()
                .reset_index()
            )
            filas_hist.append(_grp)
        except Exception as e:
            log_warning(f"  [Proyecciones] {yr}: lectura auxiliar fallida — {e}")
            continue

    if not filas_hist:
        log_warning("  [Proyecciones] No se cargaron datos históricos. Hoja vacía.")
        return pd.DataFrame()

    df_hist = pd.concat(filas_hist, ignore_index=True)
    df_hist.columns = ["IES", "PROGRAMA", "AÑO", "PC"]

    pv = (
        df_hist.groupby(["IES", "PROGRAMA", "AÑO"])["PC"]
        .sum()
        .reset_index()
        .pivot_table(
            index=["IES", "PROGRAMA"], columns="AÑO", values="PC", fill_value=0
        )
        .reset_index()
    )
    for yr in AÑOS_DATOS:
        if yr not in pv.columns:
            pv[yr] = 0.0

    pv["n_años"] = (pv[AÑOS_DATOS] > 0).sum(axis=1)
    pv["año_inicio"] = pv[AÑOS_DATOS].apply(
        lambda r: next((yr for yr in AÑOS_DATOS if r[yr] > 0), None), axis=1
    )

    mask_eafit = pv["IES"].str.contains("EAFIT", na=False)
    pv_eafit = pv[mask_eafit].copy()
    pv_otros = pv[~mask_eafit].copy()

    log(
        f"  [Proyecciones] EAFIT: {len(pv_eafit)} programas "
        f"({(pv_eafit['n_años'] >= N_AÑOS_MIN).sum()} con >={N_AÑOS_MIN} años) | "
        f"IES referentes: {len(pv_otros)} "
        f"({(pv_otros['n_años'] >= N_AÑOS_MIN).sum()} con >={N_AÑOS_MIN} años)"
    )

    df_v = df_programas_val[
        df_programas_val["PROGRAMA_EAFIT"].notna()
        & (df_programas_val["PROGRAMA_EAFIT"] != "Programa EAFIT")
    ].copy()
    df_v["_PROG_NORM"] = df_v["PROGRAMA_EAFIT"].apply(_n)
    df_v["_CAT"] = df_v["CATEGORIA"].astype(str).str.upper().str.strip()
    prog_a_cat = (
        df_v[["_PROG_NORM", "_CAT"]]
        .drop_duplicates()
        .set_index("_PROG_NORM")["_CAT"]
        .to_dict()
    )
    cat_a_progs_eafit: dict[str, list[str]] = {}
    for p, c in prog_a_cat.items():
        cat_a_progs_eafit.setdefault(c, []).append(p)

    _pares_path = REF_DIR / "backup" / "posParesPositivos.csv"
    if not _pares_path.exists():
        _pares_path = REF_DIR / "posParesPositivos.csv"

    pv_pares_cat: pd.DataFrame | None = None
    if _pares_path.exists():
        try:
            df_pares = pd.read_csv(_pares_path)
            df_pares["_PROG_PAR_N"] = df_pares["NOMBRE_DEL_PROGRAMA"].apply(_n)
            df_pares["_PROG_EAFIT_N"] = df_pares["NombrePrograma EAFIT"].apply(_n)
            df_pares["_CAMPO_EAFIT"] = df_pares["CAMPO_AMPLIO_EAFIT"].apply(_n)
            df_pares["_CATEGORIA"] = df_pares["_PROG_EAFIT_N"].map(prog_a_cat)
            pv_pares_cat = pv_otros.merge(
                df_pares[
                    ["_PROG_PAR_N", "_PROG_EAFIT_N", "_CAMPO_EAFIT", "_CATEGORIA"]
                ].drop_duplicates("_PROG_PAR_N"),
                left_on="PROGRAMA",
                right_on="_PROG_PAR_N",
                how="left",
            )
            log(
                f"  [Proyecciones] Pares mapeados: "
                f"{int(pv_pares_cat['_CATEGORIA'].notna().sum())} con categoría directa"
            )
        except Exception as e:
            log_warning(f"  [Proyecciones] No se pudo cargar posParesPositivos: {e}")

    serie_por_cat: dict[str, dict[int, float]] = {}
    if sabana is not None:
        try:
            pc_cols_sab = {
                yr: f"primer_curso_{yr}"
                for yr in AÑOS_DATOS
                if f"primer_curso_{yr}" in sabana.columns
            }
            cat_col = "CATEGORIA_FINAL"
            if cat_col in sabana.columns and pc_cols_sab:
                for cat, grp in sabana.groupby(cat_col):
                    serie_por_cat[str(cat).upper().strip()] = {
                        yr: float(grp[col].fillna(0).sum())
                        for yr, col in pc_cols_sab.items()
                    }
                log(
                    f"  [Proyecciones] Series de categoría desde sábana: "
                    f"{len(serie_por_cat)}"
                )
        except Exception as e:
            log_warning(f"  [Proyecciones] Error leyendo series de sábana: {e}")

    if not serie_por_cat:
        _total_nacional = df_hist.groupby("AÑO")["PC"].sum().to_dict()
        serie_por_cat = {cat: dict(_total_nacional) for cat in set(prog_a_cat.values())}
        log_warning(
            "  [Proyecciones] Serie de categoría desde total nacional (fallback)."
        )

    def _reg_lineal(serie: dict[int, float], año_target: int) -> float | None:
        xs = [yr for yr in AÑOS_DATOS if serie.get(yr, 0) > 0]
        ys = [serie[yr] for yr in xs]
        if len(xs) < 3:
            return None
        a, b = np.polyfit(np.array(xs, float), np.array(ys, float), 1)
        return max(0.0, round(a * año_target + b, 1))

    def _participacion_eafit(cat: str, prog_norm: str) -> float:
        serie_mkt = serie_por_cat.get(cat, {})
        eafit_rows = pv_eafit[pv_eafit["PROGRAMA"] == prog_norm]
        fracs = []
        for yr in AÑOS_DATOS[-4:]:
            pc_e = (
                float(eafit_rows[yr].sum())
                if (not eafit_rows.empty and yr in eafit_rows.columns)
                else 0.0
            )
            pc_m = float(serie_mkt.get(yr, 0))
            if pc_m > 0 and pc_e > 0:
                fracs.append(pc_e / pc_m)
        return float(np.mean(fracs)) if fracs else 0.01

    def _curva_norm(row: pd.Series, n: int = 5) -> list[float] | None:
        yr_ini = row.get("año_inicio")
        if yr_ini is None:
            return None
        base = float(row.get(yr_ini, 0))
        if base <= 0:
            return None
        curva = [float(row.get(yr_ini + i, 0)) / base for i in range(n)]
        return curva if len(curva) == n else None

    def _get_curvas(cat: str, prog_norm: str) -> tuple[list[list[float]], str]:
        curvas: list[list[float]] = []

        if pv_pares_cat is not None:
            mask_c1 = (pv_pares_cat["_CATEGORIA"] == cat) & (
                pv_pares_cat["n_años"] >= N_AÑOS_MIN
            )
            for _, r in pv_pares_cat[mask_c1].iterrows():
                c = _curva_norm(r)
                if c:
                    curvas.append(c)
        if len(curvas) >= N_MIN:
            return curvas, "IES_referentes_cat_directa"

        progs_cat = [p for p in cat_a_progs_eafit.get(cat, []) if p != prog_norm]
        mask_c2 = pv_eafit["PROGRAMA"].isin(progs_cat) & (
            pv_eafit["n_años"] >= N_AÑOS_MIN
        )
        for _, r in pv_eafit[mask_c2].iterrows():
            c = _curva_norm(r)
            if c:
                curvas.append(c)
        if len(curvas) >= N_MIN:
            return curvas, "EAFIT_cat_directa"

        palabras = [w for w in _re.sub(r"[^A-Z0-9\s]", " ", cat).split() if len(w) > 3]
        mask_c3e = (
            pv_eafit["PROGRAMA"].apply(lambda p: any(w in p for w in palabras))
            & (pv_eafit["n_años"] >= N_AÑOS_MIN)
            & ~pv_eafit["PROGRAMA"].isin(progs_cat + [prog_norm])
        )
        for _, r in pv_eafit[mask_c3e].iterrows():
            c = _curva_norm(r)
            if c:
                curvas.append(c)

        if pv_pares_cat is not None:
            mask_c3p = (
                pv_pares_cat["_CAMPO_EAFIT"].apply(
                    lambda x: any(w in _n(str(x)) for w in palabras)
                    if pd.notna(x)
                    else False
                )
                & (pv_pares_cat["n_años"] >= N_AÑOS_MIN)
                & (pv_pares_cat["_CATEGORIA"] != cat)
            )
            for _, r in pv_pares_cat[mask_c3p].iterrows():
                c = _curva_norm(r)
                if c:
                    curvas.append(c)

        if len(curvas) >= N_MIN:
            return curvas, "campo_amplio"

        return [], "CAGR_fallback"

    def _proy_analogos(
        curvas: list[list[float]],
        fuente: str,
        valor_base: float,
        serie_cat: dict[int, float],
    ) -> tuple[dict[int, dict[str, float | None]], bool]:
        resultado: dict[int, dict[str, float | None]] = {}
        fallback = fuente == "CAGR_fallback" or len(curvas) < N_MIN

        if not fallback and valor_base > 0:
            matriz = np.array(curvas[:50])
            for i, año_p in enumerate(AÑOS_PROY):
                col = matriz[:, i] if i < matriz.shape[1] else np.array([1.0])
                resultado[año_p] = {
                    "P25": max(
                        0.0, round(float(np.percentile(col, 25)) * valor_base, 1)
                    ),
                    "MEDIAN": max(
                        0.0, round(float(np.median(col)) * valor_base, 1)
                    ),
                    "P75": max(
                        0.0, round(float(np.percentile(col, 75)) * valor_base, 1)
                    ),
                }
        else:
            fallback = True
            vals = [
                (yr, serie_cat[yr])
                for yr in AÑOS_DATOS[-4:]
                if serie_cat.get(yr, 0) > 0
            ]
            if len(vals) >= 2 and valor_base > 0:
                v0, v1 = vals[0][1], vals[-1][1]
                n = vals[-1][0] - vals[0][0]
                cagr = (v1 / v0) ** (1 / n) - 1 if v0 > 0 and n > 0 else 0.0
                for i, año_p in enumerate(AÑOS_PROY):
                    central = valor_base * ((1 + cagr) ** (i + 1))
                    resultado[año_p] = {
                        "P25": max(0.0, round(central * 0.80, 1)),
                        "MEDIAN": max(0.0, round(central, 1)),
                        "P75": max(0.0, round(central * 1.20, 1)),
                    }
            else:
                for año_p in AÑOS_PROY:
                    resultado[año_p] = {"P25": None, "MEDIAN": None, "P75": None}

        return resultado, fallback

    progs_unicos = (
        df_v[
            [
                "PROGRAMA_EAFIT",
                "_PROG_NORM",
                "_CAT",
                "NIVEL",
                "PROG_ID",
            ]
        ]
        .drop_duplicates(subset=["_PROG_NORM"])
        .reset_index(drop=True)
    )

    filas_resultado: list[dict] = []
    for _, prow in progs_unicos.iterrows():
        prog_eafit = prow["PROGRAMA_EAFIT"]
        prog_norm = prow["_PROG_NORM"]
        cat = prow["_CAT"]
        nivel = prow.get("NIVEL", "")
        prog_id = prow.get("PROG_ID", "")

        serie_cat = serie_por_cat.get(cat, {yr: 0.0 for yr in AÑOS_DATOS})
        partic = _participacion_eafit(cat, prog_norm)
        proy_cat_27 = _reg_lineal(serie_cat, 2027)
        valor_base = proy_cat_27 * partic if proy_cat_27 and proy_cat_27 > 0 else 0.0

        fila: dict = {
            "ID Programa": prog_id,
            "Programa EAFIT": prog_eafit,
            "Nivel": nivel,
            "Categoría": cat,
            "Participación hist. EAFIT": round(partic, 4),
        }

        for año_p in AÑOS_PROY:
            proy_c = _reg_lineal(serie_cat, año_p)
            fila[f"P1_{año_p}"] = (
                round(proy_c * partic, 1) if proy_c else None
            )

        curvas, fuente = _get_curvas(cat, prog_norm)
        proy2, fb = _proy_analogos(curvas, fuente, valor_base, serie_cat)

        for año_p in AÑOS_PROY:
            v = proy2.get(año_p, {})
            fila[f"P2_P25_{año_p}"] = v.get("P25")
            fila[f"P2_MED_{año_p}"] = v.get("MEDIAN")
            fila[f"P2_P75_{año_p}"] = v.get("P75")

        fila["N° análogos"] = len(curvas)
        fila["Fuente análogos"] = fuente
        fila["¿Fallback?"] = "Sí" if fb else "No"

        filas_resultado.append(fila)

    df_result = pd.DataFrame(filas_resultado)
    log(
        f"  [Proyecciones] {len(df_result)} programas procesados "
        f"| con análogos: {(df_result['¿Fallback?'] == 'No').sum()} "
        f"| fallback: {(df_result['¿Fallback?'] == 'Sí').sum()}"
    )
    return df_result


def run_fase_valorizacion(log: Callable = print) -> Path:
    """
    Fase 7 — Genera Programas_para_valorizacion_output.xlsx.

    Flujo:
    1. Lee programas_para_valorizacion.xlsx
    2. Carga la sábana consolidada (base_maestra.parquet o sabana_consolidada.parquet)
    3. Pre-filtra la sábana en 4 segmentos × (todas_IES + solo_referentes)
    4. Por cada programa × región: calcula métricas MERCADO y REFERENTES, aplica scoring
    5. Exporta Excel con formato de doble sección
    """
    log("━━━ Fase 7 — Valorización de Programas EAFIT ━━━")

    # ── 1. Lista de programas ────────────────────────────────────────────────
    df_programas = _leer_programas_eafit(log)

    # ── 2. Cargar sábana ─────────────────────────────────────────────────────
    sabana_path = TEMP_DIR / "sabana_consolidada.parquet"
    if not sabana_path.exists():
        sabana_path = CHECKPOINT_BASE_MAESTRA
    if not sabana_path.exists():
        raise FileNotFoundError(
            "No se encontró sabana_consolidada.parquet ni base_maestra.parquet.\n"
            "Ejecuta primero el pipeline completo (Fases 1–3)."
        )
    log(f"  Cargando sábana: {sabana_path.name}...")
    sabana = pd.read_parquet(sabana_path)
    log(f"  Sábana: {len(sabana):,} programas")

    # Calcular distancia al costo si no existe en la sábana
    if "_distancia_costo_prog" not in sabana.columns:
        costo_col = "COSTO_MATRÍCULA_ESTUD_NUEVOS"
        if costo_col in sabana.columns and "NIVEL_DE_FORMACIÓN" in sabana.columns:
            try:
                from etl.config import get_benchmark_costo

                def _dist(row):
                    c = row.get(costo_col)
                    if pd.isna(c) or c == 0:
                        return np.nan
                    bench = get_benchmark_costo(str(row.get("NIVEL_DE_FORMACIÓN", "")))
                    return (float(c) - bench) / bench * 100

                sabana["_distancia_costo_prog"] = sabana.apply(_dist, axis=1)
            except Exception:
                sabana["_distancia_costo_prog"] = np.nan

    # Filtrar por código numérico de institución (más confiable que nombre)
    col_cod = "CÓDIGO_INSTITUCIÓN"
    if col_cod in sabana.columns:
        codigos_serie = pd.to_numeric(sabana[col_cod], errors="coerce")
        sabana["_es_referente"] = codigos_serie.isin(list(CODIGOS_IES_REFERENTES))
        n_ref = int(sabana["_es_referente"].sum())
        log(f"  Programas de IES referentes (todos los niveles): {n_ref:,} de {len(sabana):,}")
    else:
        sabana["_es_referente"] = False
        log_warning(f"  ⚠ Columna '{col_cod}' no encontrada en sábana. Referentes vacíos.")

    # Aplicar el mismo filtro de niveles que usa el estudio de mercado principal
    # (solo Especialización, Maestría y sus variantes) para que los referentes
    # sean comparables con el mercado que ya está filtrado en los parquets
    col_nivel = "NIVEL_DE_FORMACIÓN"
    sabana_ref = sabana[sabana["_es_referente"]].copy()
    if col_nivel in sabana_ref.columns and NIVELES_MERCADO:
        sabana_ref = sabana_ref[sabana_ref[col_nivel].isin(NIVELES_MERCADO)].copy()
        n_cat = sabana_ref["CATEGORIA_FINAL"].nunique() if "CATEGORIA_FINAL" in sabana_ref.columns else 0
        log(
            f"  Sábana referentes tras filtro NIVELES_MERCADO: {len(sabana_ref):,} programas | "
            f"{n_cat} categorías"
        )
    else:
        log(f"  Sábana referentes (sin filtro niveles): {len(sabana_ref):,} programas")

    # ── Expandir categorías compuestas ──────────────────────────────
    _col_cat_val = "CATEGORIA_FINAL"
    if _col_cat_val in sabana.columns:
        _cats_validas = set(
            sabana[_col_cat_val].dropna().astype(str).str.strip().unique()
        )
    else:
        _cats_validas = set()

    if _cats_validas:
        _n_antes = len(df_programas)
        df_programas = _expandir_categorias_compuestas(
            df_programas,
            _cats_validas,
            col_cat="CATEGORIA_RAW",
        )
        _n_despues = len(df_programas)
        if _n_despues != _n_antes:
            log(
                f"  Categorías compuestas expandidas: "
                f"{_n_antes} → {_n_despues} filas "
                f"(+{_n_despues - _n_antes} registros nuevos)"
            )
    else:
        log_warning("  No se pudo obtener lista de categorías válidas para expansión.")

    # ── 3. Agregados REFERENTES (Fase 4) + MERCADO desde parquet (TEMP_DIR) ───
    log("  Referentes por región (programas, tras niveles):")
    for seg in SEGMENTOS:
        log(f"    {LABEL_REGION[seg]}: {len(SEGMENTOS_FILTROS[seg](sabana_ref)):,} programas")

    # Referentes NACIONALES: un solo agregado sobre todas las IES referentes
    # sin filtro de región, igual para las 4 regiones de cada programa.
    # Esto replica el comportamiento del manual: las IES referentes compiten
    # a nivel nacional, no solo en una región específica.
    log("  Calculando agregado de referentes NACIONAL (sin filtro de región)...")
    from etl.mercado_pipeline import run_fase4_desde_sabana

    ag_ref_nacional: pd.DataFrame | None = None
    try:
        ag_ref_nacional = run_fase4_desde_sabana(sabana_ref)
        n_cats = len(ag_ref_nacional)
        n_aagr = (
            int(ag_ref_nacional["AAGR_ROBUSTO"].notna().sum())
            if "AAGR_ROBUSTO" in ag_ref_nacional.columns
            else 0
        )
        log(f"    ✓ Nacional: {n_cats} categorías referentes | AAGR disponible: {n_aagr}/{n_cats}")
    except Exception as e:
        log_warning(f"    ⚠ Error calculando referentes nacionales: {e}")
        ag_ref_nacional = None

    log("  Cargando agregados MERCADO regionales (parquet)...")
    agregados: dict[str, pd.DataFrame] = {}

    # Agregado nacional como fallback de AAGR cuando el regional es NaN.
    # Es fundamental para Eje Cafetero y Virtual donde muchas categorías
    # no tienen historia suficiente para calcular AAGR_ROBUSTO regional.
    ag_colombia: pd.DataFrame | None = None
    for _nombre_colombia in ["agregado_Colombia.parquet", "agregado_categorias.parquet"]:
        _path_col = TEMP_DIR / _nombre_colombia
        if _path_col.exists():
            try:
                ag_colombia = pd.read_parquet(_path_col)
                log(f"  ✓ Fallback AAGR nacional cargado: {_nombre_colombia} ({len(ag_colombia)} cats)")
            except Exception as _e:
                log_warning(f"  ⚠ No se pudo cargar {_nombre_colombia}: {_e}")
            break
    if ag_colombia is None:
        log_warning("  ⚠ No se encontró agregado nacional. AAGR faltantes quedarán como NaN.")
    for seg in SEGMENTOS:
        cache = TEMP_DIR / f"agregado_{seg}.parquet"
        if not cache.exists():
            log_warning(f"    ⚠ {seg}: no existe {cache.name}")
            agregados[seg] = pd.DataFrame()
            continue
        ag = pd.read_parquet(cache)
        # Verificar que tenga las columnas clave
        cols_clave = [
            "CATEGORIA_FINAL",
            "AAGR_ROBUSTO",
            f"prom_matricula_{AÑO_FIN_DATOS}",
            "calificacion_final",
            "salario_promedio_smlmv",
        ]
        cols_presentes = [c for c in cols_clave if c in ag.columns]
        cols_ausentes = [c for c in cols_clave if c not in ag.columns]
        agregados[seg] = ag
        log(f"    ✓ {seg}: {len(ag)} categorías | cols OK: {cols_presentes}")
        if cols_ausentes:
            log_warning(f"    ⚠ {seg}: columnas ausentes: {cols_ausentes}")

    # ── Diagnóstico: verificar cobertura de IES referentes ───────────────
    n_ref_total = int(sabana["_es_referente"].sum())
    if n_ref_total == 0:
        log_warning(
            "  ⚠ No se encontró ningún programa de IES referente. "
            f"Verificar que '{col_cod}' coincida con los códigos SNIES esperados."
        )
        if col_cod in sabana.columns:
            muestra = (
                pd.to_numeric(sabana[col_cod], errors="coerce")
                .dropna()
                .astype(int)
                .unique()[:15]
            )
            log(f"  Muestra de códigos de institución en sábana: {list(muestra)}")

    # ── Diagnóstico: verificar cobertura de categorías ───────────────────
    cats_sabana = set(sabana["CATEGORIA_FINAL"].apply(lambda x: _norm(str(x))).unique()) \
        if "CATEGORIA_FINAL" in sabana.columns else set()
    cats_sin_match = []
    for _, pr in df_programas.iterrows():
        cat_n = _norm(str(pr["CATEGORIA_RAW"]))
        if cat_n not in cats_sabana:
            cats_sin_match.append(str(pr["CATEGORIA_RAW"]))
    if cats_sin_match:
        log_warning(f"  ⚠ {len(cats_sin_match)} categorías del archivo no matchean en sábana:")
        for c in set(cats_sin_match):
            log_warning(f"    - {c!r}")
    else:
        log(f"  ✓ Todas las categorías del archivo encontradas en sábana")

    # ── 4. Calcular métricas por programa × región ────────────────────────────
    log("  Calculando métricas...")
    filas = []

    for _, prog_row in df_programas.iterrows():
        cat_raw = str(prog_row["CATEGORIA_RAW"]).strip()
        nivel = str(prog_row["NIVEL"]).strip()
        programa = str(prog_row["PROGRAMA_EAFIT"]).strip()
        tiene_em = str(prog_row["TIENE_ESTUDIO_MERCADO"]).strip()
        categorias = [cat_raw]

        for seg in SEGMENTOS:
            region = LABEL_REGION[seg]
            # MERCADO: agregado regional (parquet Fase 4 por segmento)
            met_m = _lookup_categoria(agregados.get(seg), categorias)

            # Si el parquet no trae primer_curso, calcular prom desde sábana regional (no matrícula total)
            _pmp_key = f"prom_matricula_por_programa_{AÑO_FIN_DATOS}"
            _pmp_parquet = float(met_m.get(_pmp_key, 0) or 0)
            if _pmp_parquet == 0.0:
                df_seg = SEGMENTOS_FILTROS[seg](sabana)
                if len(df_seg) > 0:
                    met_sabana = _agregar_metricas_categoria(df_seg, categorias)
                    _pmp_sab = float(met_sabana.get(_pmp_key, 0) or 0)
                    if _pmp_sab > 0:
                        met_m = {
                            **met_m,
                            _pmp_key: _pmp_sab,
                            f"prom_matricula_{AÑO_FIN_DATOS}": _pmp_sab,
                        }

            # Fallback AAGR: si el regional es NaN, usar el nacional (tendencia de largo plazo)
            if pd.isna(met_m.get("AAGR_ROBUSTO")) and ag_colombia is not None:
                met_nac_fallback = _lookup_categoria(ag_colombia, categorias)
                aagr_nac = met_nac_fallback.get("AAGR_ROBUSTO")
                if pd.notna(aagr_nac):
                    met_m = {**met_m, "AAGR_ROBUSTO": float(aagr_nac)}
                    log_info(
                        f"    AAGR fallback: {categorias[0]!r} en {LABEL_REGION[seg]} "
                        f"usa AAGR nacional = {float(aagr_nac):.1%}"
                    )

            met_m_s = _score_y_calificacion(met_m)

            # REFERENTES: agregado Fase 4 sobre IES referentes NACIONALES (sin filtro regional)
            met_r = _lookup_categoria(ag_ref_nacional, categorias)
            _sub_ref = _subconjunto_por_categorias(sabana_ref, categorias)
            met_r["serie_primer_curso"] = (
                _serie_primer_curso_sub(_sub_ref) if len(_sub_ref) else {}
            )
            met_r_s = _score_y_calificacion(met_r)

            # CAL_INTEGRADA = 0.4 × M + 0.6 × R
            # Peso 40% mercado general (demanda amplia) + 60% referentes nacionales
            # (benchmarks de calidad del mismo segmento que EAFIT).
            # Fórmula aritmética ponderada — más interpretable que la geométrica anterior.
            _m_cal = met_m_s.get("calificacion_final")
            _r_cal = met_r_s.get("calificacion_final")
            cal_integrada = (
                round(float(0.4 * _m_cal + 0.6 * _r_cal), 4)
                if (
                    pd.notna(_m_cal)
                    and pd.notna(_r_cal)
                    and _m_cal > 0
                    and _r_cal > 0
                )
                else np.nan
            )

            filas.append(
                {
                    # Identificación
                    "PROG_ID": str(prog_row.get("PROG_ID", "")),
                    "CAT_ID": _cat_id_para(cat_raw),
                    "CATEGORIA": cat_raw,
                    "NIVEL": nivel,
                    "PROGRAMA_EAFIT": programa,
                    "TIENE_ESTUDIO_MERCADO": tiene_em,
                    "REGION": region,
                    # ── SECCIÓN MERCADO ──────────────────────────────────────
                    "M_prom_matricula": met_m_s[f"prom_matricula_{AÑO_FIN_DATOS}"],
                    "M_score_matricula": met_m_s["score_matricula"],
                    "M_participacion": met_m_s[f"participacion_{AÑO_FIN_DATOS}"],
                    "M_score_participacion": met_m_s["score_participacion"],
                    "M_AAGR": met_m_s["AAGR_ROBUSTO"],
                    "M_score_AAGR": met_m_s["score_AAGR"],
                    "M_salario_smlmv": met_m_s["salario_promedio_smlmv"],
                    "M_score_salario": met_m_s["score_salario"],
                    "M_pct_no_matriculados": met_m_s[f"pct_no_matriculados_{AÑO_FIN_DATOS}"],
                    "M_score_no_mat": met_m_s["score_pct_no_matriculados"],
                    "M_num_programas": met_m_s[f"num_programas_{AÑO_FIN_DATOS}"],
                    "M_score_num_programas": met_m_s["score_num_programas"],
                    "M_pct_con_matricula": met_m_s["pct_con_matricula"],
                    "M_programas_activos": met_m_s["programas_activos"],
                    "M_programas_nuevos_3a": met_m_s["programas_nuevos_3a"],
                    "M_programas_inactivos": met_m_s["programas_inactivos"],
                    "M_costo_promedio": met_m_s["costo_promedio"],
                    "M_score_costo": met_m_s["score_costo"],
                    "M_calificacion": met_m_s["calificacion_final"],
                    "M_señal_tendencia": met_m_s.get("SEÑAL_TENDENCIA", np.nan),
                    # ── SECCIÓN REFERENTES ───────────────────────────────────
                    "R_prom_matricula": met_r_s[f"prom_matricula_{AÑO_FIN_DATOS}"],
                    "R_score_matricula": met_r_s["score_matricula"],
                    "R_participacion": met_r_s[f"participacion_{AÑO_FIN_DATOS}"],
                    "R_score_participacion": met_r_s["score_participacion"],
                    "R_AAGR": met_r_s["AAGR_ROBUSTO"],
                    "R_score_AAGR": met_r_s["score_AAGR"],
                    "R_salario_smlmv": met_r_s["salario_promedio_smlmv"],
                    "R_score_salario": met_r_s["score_salario"],
                    "R_pct_no_matriculados": met_r_s[f"pct_no_matriculados_{AÑO_FIN_DATOS}"],
                    "R_score_no_mat": met_r_s["score_pct_no_matriculados"],
                    "R_num_programas": met_r_s[f"num_programas_{AÑO_FIN_DATOS}"],
                    "R_score_num_programas": met_r_s["score_num_programas"],
                    "R_pct_con_matricula": met_r_s["pct_con_matricula"],
                    "R_programas_activos": met_r_s["programas_activos"],
                    "R_programas_nuevos_3a": met_r_s["programas_nuevos_3a"],
                    "R_programas_inactivos": met_r_s["programas_inactivos"],
                    "R_costo_promedio": met_r_s["costo_promedio"],
                    "R_score_costo": met_r_s["score_costo"],
                    "R_calificacion": met_r_s["calificacion_final"],
                    "R_señal_tendencia": met_r_s.get("SEÑAL_TENDENCIA", np.nan),
                    # ── CALIFICACIÓN INTEGRADA ────────────────────────────────
                    # 40% calificación mercado regional + 60% calificación referentes nacionales
                    "CAL_INTEGRADA": cal_integrada,
                    # ── VIABILIDAD_ESTUDIO — derivada de CAL_INTEGRADA ──────
                    "VIABILIDAD_ESTUDIO": (
                        "ALTA" if cal_integrada >= 3.5 else
                        "MEDIA" if cal_integrada >= 3.0 else
                        "BAJA" if cal_integrada >= 2.5 else
                        "MUY_BAJA"
                    ) if pd.notna(cal_integrada) else np.nan,
                    "VIABILIDAD_BINARIA": (
                        "Viable"
                        if pd.notna(cal_integrada) and cal_integrada >= 3.0
                        else "No viable"
                        if pd.notna(cal_integrada)
                        else np.nan
                    ),
                    "PROYECCION_REGRESION": (
                        _proyeccion_regresion_lineal(
                            serie_pc=met_r_s.get("serie_primer_curso", {}),
                            año_objetivo=_AÑO_PROYECCION_REGRESION,
                            tasa_captura=TASA_CAPTURA_EAFIT,
                        )
                        if pd.notna(cal_integrada)
                        else np.nan
                    ),
                }
            )

    df_out = pd.DataFrame(filas)
    log(f"  Total filas: {len(df_out)} ({len(df_programas)} programas × {len(SEGMENTOS)} regiones)")

    # ── 5. Exportar (misma carpeta que el estudio de mercado) ────────────────
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    ESTUDIO_MERCADO_DIR.mkdir(parents=True, exist_ok=True)
    ruta = ESTUDIO_MERCADO_DIR / "Programas_para_valorizacion_output.xlsx"

    df_out = _reordenar_columnas_valorizacion(df_out)

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Valorizacion", index=False)
        _formatear_hoja_valorizacion(writer, df_out)

        # ── Hoja 2: Crecimiento de Mercado ───────────────────────────
        df_crec = _construir_tabla_crecimiento(sabana_ref)
        ws_crec = writer.book.create_sheet("Crecimiento_Mercado")
        writer.sheets["Crecimiento_Mercado"] = ws_crec
        _formatear_hoja_crecimiento(writer, df_crec)

        # ── Hoja 3: Proyecciones 2027-2031 ───────────────────────────
        log("  Construyendo hoja Proyecciones...")
        try:
            df_proy = _construir_hoja_proyecciones(
                df_programas_val=df_out,
                sabana=sabana,
                log=log,
            )
            if not df_proy.empty:
                ws_proy = writer.book.create_sheet("Proyecciones")
                writer.sheets["Proyecciones"] = ws_proy
                try:
                    _formatear_hoja_proyecciones(writer, df_proy)
                except Exception as _fmt_e:
                    log_warning(
                        f"  Formato Proyecciones falló ({_fmt_e}); "
                        "escribiendo tabla sin formato."
                    )
                    from openpyxl.utils.dataframe import dataframe_to_rows

                    for r_i, row in enumerate(
                        dataframe_to_rows(df_proy, index=False, header=True), 1
                    ):
                        for c_i, val in enumerate(row, 1):
                            ws_proy.cell(row=r_i, column=c_i, value=val)
                log(f"  ✓ Hoja Proyecciones: {len(df_proy)} programas")
            else:
                log_warning("  ⚠ Hoja Proyecciones vacía — se omite.")
        except Exception as _e:
            log_warning(f"  ⚠ Error construyendo Proyecciones: {_e}")
            import traceback as _tb

            log_warning(_tb.format_exc())

    log(f"✓ Generado: {ruta}")
    return ruta


def _reordenar_columnas_valorizacion(df_out: pd.DataFrame) -> pd.DataFrame:
    """IDENTIFICACIÓN → CONCLUSIÓN → MERCADO → REFERENTES (orden intercalado valor|score)."""
    _cols_id = [
        "PROG_ID", "CAT_ID", "CATEGORIA", "NIVEL",
        "PROGRAMA_EAFIT", "TIENE_ESTUDIO_MERCADO", "REGION",
    ]
    _cols_concl = [
        "VIABILIDAD_ESTUDIO", "VIABILIDAD_BINARIA", "CAL_INTEGRADA", "PROYECCION_REGRESION",
    ]
    _cols_m = [
        "M_prom_matricula", "M_score_matricula",
        "M_participacion", "M_score_participacion",
        "M_AAGR", "M_score_AAGR",
        "M_salario_smlmv", "M_score_salario",
        "M_pct_no_matriculados", "M_score_no_mat",
        "M_num_programas", "M_score_num_programas",
        "M_pct_con_matricula",
        "M_programas_activos", "M_programas_nuevos_3a", "M_programas_inactivos",
        "M_costo_promedio", "M_score_costo",
        "M_calificacion", "M_señal_tendencia",
    ]
    _cols_r = [
        "R_prom_matricula", "R_score_matricula",
        "R_participacion", "R_score_participacion",
        "R_AAGR", "R_score_AAGR",
        "R_salario_smlmv", "R_score_salario",
        "R_pct_no_matriculados", "R_score_no_mat",
        "R_num_programas", "R_score_num_programas",
        "R_pct_con_matricula",
        "R_programas_activos", "R_programas_nuevos_3a", "R_programas_inactivos",
        "R_costo_promedio", "R_score_costo",
        "R_calificacion", "R_señal_tendencia",
    ]
    _orden = _cols_id + _cols_concl + _cols_m + _cols_r
    return df_out[[c for c in _orden if c in df_out.columns]].copy()


def _formatear_hoja_valorizacion(writer, df_out: pd.DataFrame) -> None:
    """Formato visual: encabezados de dos niveles, colores por sección, scores con escala de color."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    df_out = _reordenar_columnas_valorizacion(df_out)

    wb = writer.book
    ws = writer.sheets["Valorizacion"]
    cols = list(df_out.columns)
    _cols_m = [c for c in cols if c.startswith("M_")]

    AZUL_EAFIT = "000066"
    AZUL_MERC = "00A9E0"
    VERDE_REF = "1F7A3C"
    BLANCO = "FFFFFF"
    GRIS_ID = "F2F2F2"
    GRIS_ALT = "F9F9F9"

    thin = Side(style="thin", color="CCCCCC")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def font(color=BLANCO, bold=True, size=10):
        return Font(bold=bold, color=color, name="Arial", size=size)

    def score_fill(v):
        try:
            s = int(float(v))
        except (TypeError, ValueError):
            return fill("EEEEEE")
        return fill({1: "FFC7CE", 2: "FFD9B3", 3: "FFEB9C", 4: "C6EFCE", 5: "1F7A3C"}.get(s, "EEEEEE"))

    def score_font(v):
        try:
            s = int(float(v))
        except (TypeError, ValueError):
            return Font(name="Arial", size=9)
        color = "FFFFFF" if s == 5 else ("9C0006" if s == 1 else "1A1A1A")
        bold = s in (1, 5)
        return Font(bold=bold, color=color, name="Arial", size=9)

    # Insertar 2 filas de encabezado
    ws.insert_rows(1, 2)

    N_ID = 7  # PROG_ID, CAT_ID, Categoría, Nivel, Programa, ¿Tiene estudio?, Región
    N_CONCL = 4  # VIABILIDAD_ESTUDIO, VIABILIDAD_BINARIA, CAL_INTEGRADA, PROYECCION_REGRESION
    N_MET = len(_cols_m)  # columnas M_ (dinámico, actualmente 20)

    DORADO_CONCL = "7B5E00"  # dorado oscuro para sección conclusión

    # Fila 1: bloques de sección
    for c_ini, c_fin, titulo, color in [
        (1, N_ID, "IDENTIFICACIÓN", AZUL_EAFIT),
        (N_ID + 1, N_ID + N_CONCL, "CONCLUSIÓN", DORADO_CONCL),
        (N_ID + N_CONCL + 1, N_ID + N_CONCL + N_MET, "MERCADO", AZUL_MERC),
        (N_ID + N_CONCL + N_MET + 1, len(cols), "REFERENTES", VERDE_REF),
    ]:
        ws.merge_cells(start_row=1, start_column=c_ini, end_row=1, end_column=c_fin)
        cell = ws.cell(row=1, column=c_ini)
        cell.value = titulo
        cell.fill = fill(color)
        cell.font = font(BLANCO, bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[1].height = 22

    # Fila 2: nombres de columnas legibles
    NOMBRES = {
        "PROG_ID": "ID Programa",
        "CAT_ID": "ID Categoría",
        "CATEGORIA": "Categoría",
        "NIVEL": "Nivel",
        "PROGRAMA_EAFIT": "Programa EAFIT",
        "TIENE_ESTUDIO_MERCADO": "¿Tiene estudio?",
        "REGION": "Región",
        "M_prom_matricula": f"Prom. Primer Curso {AÑO_FIN_DATOS}",
        "M_score_matricula": "Score",
        "M_participacion": f"Participación {AÑO_FIN_DATOS}",
        "M_score_participacion": "Score",
        "M_AAGR": "AAGR Robusto",
        "M_score_AAGR": "Score",
        "M_salario_smlmv": "Salario (SMLMV)",
        "M_score_salario": "Score",
        "M_pct_no_matriculados": "% No Matriculados",
        "M_score_no_mat": "Score",
        "M_num_programas": "N° Programas",
        "M_score_num_programas": "Score",
        "M_pct_con_matricula": "% Con Matrícula",
        "M_programas_activos": "Activos",
        "M_programas_nuevos_3a": "Nuevos 3a",
        "M_programas_inactivos": "Inactivos",
        "M_costo_promedio": "Costo Promedio",
        "M_score_costo": "Score",
        "M_calificacion": "CALIFICACIÓN",
        "M_señal_tendencia": "Señal Tendencia",
        "R_prom_matricula": f"Prom. Primer Curso {AÑO_FIN_DATOS}",
        "R_score_matricula": "Score",
        "R_participacion": f"Participación {AÑO_FIN_DATOS}",
        "R_score_participacion": "Score",
        "R_AAGR": "AAGR Robusto",
        "R_score_AAGR": "Score",
        "R_salario_smlmv": "Salario (SMLMV)",
        "R_score_salario": "Score",
        "R_pct_no_matriculados": "% No Matriculados",
        "R_score_no_mat": "Score",
        "R_num_programas": "N° Programas",
        "R_score_num_programas": "Score",
        "R_pct_con_matricula": "% Con Matrícula",
        "R_programas_activos": "Activos",
        "R_programas_nuevos_3a": "Nuevos 3a",
        "R_programas_inactivos": "Inactivos",
        "R_costo_promedio": "Costo Promedio",
        "R_score_costo": "Score",
        "R_calificacion": "CALIFICACIÓN",
        "R_señal_tendencia": "Señal Tendencia",
        "CAL_INTEGRADA": "CAL. INTEGRADA (40%M + 60%R)",
        "VIABILIDAD_ESTUDIO": "Viabilidad (4 niveles)",
        "VIABILIDAD_BINARIA": "Viabilidad",
        "PROYECCION_REGRESION": f"Proyección Primer Curso {AÑO_FIN_DATOS + 2} (est.)",
    }
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci)
        cell.value = NOMBRES.get(col, col)
        if ci <= N_ID:
            cell.fill = fill(AZUL_EAFIT)
        elif ci <= N_ID + N_CONCL:
            cell.fill = fill(DORADO_CONCL)
        elif ci <= N_ID + N_CONCL + N_MET:
            cell.fill = fill(AZUL_MERC)
        else:
            cell.fill = fill(VERDE_REF)
        cell.font = font(BLANCO, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
        if col == "VIABILIDAD_ESTUDIO":
            cell.comment = Comment(
                "ALTA ≥ 3.5 · MEDIA ≥ 3.0 · BAJA ≥ 2.5 · MUY_BAJA < 2.5. "
                "Basado en CAL_INTEGRADA = 40%×Mercado + 60%×Referentes.",
                "SNIESManager",
            )
        elif col == "VIABILIDAD_BINARIA":
            cell.comment = Comment(
                "Replica el formato del manual de referencia. Viable si CAL_INTEGRADA ≥ 3.0.",
                "SNIESManager",
            )
        elif col == "PROYECCION_REGRESION":
            cell.comment = Comment(
                f"Estimación de primer_curso en {AÑO_FIN_DATOS + 2} usando regresión lineal "
                f"sobre IES referentes × tasa de captura EAFIT ({TASA_CAPTURA_EAFIT:.0%}).",
                "SNIESManager",
            )
    ws.row_dimensions[2].height = 32

    # Identificar columnas por tipo
    score_cols_m = {ci for ci, c in enumerate(cols, 1) if c.startswith("M_score")}
    score_cols_r = {ci for ci, c in enumerate(cols, 1) if c.startswith("R_score")}
    cal_cols = {ci for ci, c in enumerate(cols, 1) if c in (
        "M_calificacion", "R_calificacion", "CAL_INTEGRADA"
    )}
    pct_cols = {
        ci
        for ci, c in enumerate(cols, 1)
        if any(k in c for k in ("participacion", "AAGR", "no_matriculados", "pct_con"))
    }
    int_cols = {
        ci
        for ci, c in enumerate(cols, 1)
        if any(k in c for k in ("num_programas", "activos", "nuevos", "inactivos"))
        or c == "PROYECCION_REGRESION"
    }
    cost_cols = {ci for ci, c in enumerate(cols, 1) if "costo_promedio" in c}
    prom_cols = {ci for ci, c in enumerate(cols, 1) if "prom_matricula" in c}
    sal_cols = {ci for ci, c in enumerate(cols, 1) if "salario_smlmv" in c}

    # Formato de filas de datos
    for ri in range(3, 3 + len(df_out)):
        alt = ri % 2 == 0
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if ci in score_cols_m or ci in score_cols_r:
                cell.fill = score_fill(cell.value)
                cell.font = score_font(cell.value)

            elif col == "VIABILIDAD_ESTUDIO":
                viab_colores = {
                    "ALTA": "C6EFCE",
                    "MEDIA": "FFFDE7",
                    "BAJA": "FFD9B3",
                    "MUY_BAJA": "FFC7CE",
                }
                viab_texto = {
                    "ALTA": "1A6B2B",
                    "MEDIA": "7D6608",
                    "BAJA": "8A3A00",
                    "MUY_BAJA": "9C0006",
                }
                val_v = str(cell.value).strip().upper() if cell.value else ""
                bg = viab_colores.get(val_v, "F2F2F2")
                fg = viab_texto.get(val_v, "1A1A1A")
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(bold=True, color=fg, name="Arial", size=10)
                continue

            elif col == "VIABILIDAD_BINARIA":
                val_b = str(cell.value).strip() if cell.value else ""
                bg_b = "C6EFCE" if val_b == "Viable" else ("FFC7CE" if val_b == "No viable" else "F2F2F2")
                fg_b = "1A6B2B" if val_b == "Viable" else ("9C0006" if val_b == "No viable" else "1A1A1A")
                cell.fill = PatternFill("solid", fgColor=bg_b)
                cell.font = Font(bold=True, color=fg_b, name="Arial", size=10)
                continue

            elif "señal_tendencia" in col:
                tendencia_colores = {
                    "ACELERANDO": "C6EFCE",
                    "ESTABLE": "EBF9EE",
                    "DESACELERANDO": "FFFDE7",
                    "EN_DECLIVE": "FFD9B3",
                    "CONTRACCION": "FFC7CE",
                    "SIN_ACTIVIDAD": "F2F2F2",
                    "SIN_DATO": "EEEEEE",
                }
                val_t = str(cell.value).strip().upper() if cell.value else ""
                cell.fill = PatternFill("solid", fgColor=tendencia_colores.get(val_t, "F2F2F2"))
                cell.font = Font(name="Arial", size=9)
                continue

            elif col == "CAL_INTEGRADA":
                try:
                    v = float(cell.value) if cell.value is not None else None
                    if v is not None:
                        color_int = "EBF9EE" if v >= 4.0 else ("FFFDE7" if v >= 3.0 else "FFF0F0")
                        cell.fill = PatternFill("solid", fgColor=color_int)
                        cell.font = Font(bold=True, name="Arial", size=10)
                    cell.number_format = "0.00"
                except (TypeError, ValueError):
                    pass

            elif ci in cal_cols:
                try:
                    v = float(cell.value)
                    cell.fill = fill("EBF9EE" if v >= 4.0 else ("FFFDE7" if v >= 3.0 else "FFF0F0"))
                except (TypeError, ValueError):
                    pass
                cell.font = Font(bold=True, name="Arial", size=9)
                cell.number_format = "0.00"

            elif ci <= N_ID:
                cell.fill = fill(GRIS_ID if alt else BLANCO)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            elif ci <= N_ID + N_CONCL:
                if col not in ("VIABILIDAD_ESTUDIO", "VIABILIDAD_BINARIA", "CAL_INTEGRADA"):
                    cell.fill = fill("FFFDF0" if alt else "FFFFF8")
                    cell.font = Font(bold=False, name="Arial", size=9)

            else:
                cell.fill = fill(GRIS_ALT if alt else BLANCO)
                cell.font = Font(name="Arial", size=9)
                if ci in pct_cols:
                    cell.number_format = "0.0%"
                elif ci in cost_cols:
                    cell.number_format = "#,##0"
                elif ci in int_cols:
                    cell.number_format = "#,##0"
                elif ci in prom_cols:
                    cell.number_format = "0.0"
                elif ci in sal_cols:
                    cell.number_format = "0.00"

    # Anchos
    ANCHOS = {
        "PROG_ID": 14,
        "CAT_ID": 12,
        "CATEGORIA": 32,
        "NIVEL": 16,
        "PROGRAMA_EAFIT": 36,
        "TIENE_ESTUDIO_MERCADO": 12,
        "REGION": 13,
        "CAL_INTEGRADA": 16,
        "VIABILIDAD_ESTUDIO": 16,
        "VIABILIDAD_BINARIA": 13,
        "PROYECCION_REGRESION": 28,
        "M_señal_tendencia": 16,
        "R_señal_tendencia": 16,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = ANCHOS.get(col, 9 if "score" in col.lower() else 14)

    ws.freeze_panes = "H3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2 + len(df_out)}"


def _formatear_hoja_crecimiento(writer, df: pd.DataFrame) -> None:
    """
    Formatea la hoja Crecimiento_Mercado con bloques de cabecera + datos.
    Cada combo: 4 filas de cabecera + 1 blank + 3 datos + 1 blank.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets["Crecimiento_Mercado"]

    COLOR_HEADER_BG = "1C3557"
    COLOR_HEADER_FG = "FFFFFF"
    COLOR_TOTAL_BG = "D9E1F2"
    COLOR_ESP_BG = "EBF9EE"
    COLOR_MAE_BG = "FFFDE7"
    COLOR_COL_HEADER = "37474F"

    FMT_NUM = "#,##0"
    FMT_PCT = "0.0%"

    años_cols = [str(yr) for yr in range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1)]
    yr_ini = AÑO_INICIO_PRIMER_CURSO
    yr_fin = AÑO_FIN_DATOS
    all_cols = años_cols + [
        "Total general",
        f"CAGR {yr_ini}-{yr_fin}",
        "CAGR 2019-2024",
        "CAGR 2021-2024",
    ]
    last_col = len(all_cols) + 1
    CONTEXT_COL = last_col + 2

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    hdr_fill = PatternFill("solid", fgColor=COLOR_COL_HEADER)
    hdr_font = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=9)

    ws.cell(row=1, column=1, value="SEGMENTO / NIVEL")
    ws.cell(row=1, column=1).font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="left", vertical="center"
    )

    for j, col in enumerate(all_cols, start=2):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if "CAGR" in str(col):
            cell.number_format = FMT_PCT

    ctx_hdr = ws.cell(row=1, column=CONTEXT_COL, value="Descripción del segmento")
    ctx_hdr.font = hdr_font
    ctx_hdr.fill = hdr_fill
    ctx_hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B2"

    section_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    section_font_bold = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=9)
    section_font_val = Font(color=COLOR_HEADER_FG, name="Arial", size=9)

    ctx_border = Border(
        left=Side(style="medium", color="1C3557"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
    )

    current_row = 2
    context_start_row: int | None = None
    context_text: str | None = None

    for _, fila in df.iterrows():
        tipo = fila.get("__tipo", "data")

        if tipo == "header" and fila.get("__campo") == "SEGMENTO":
            context_start_row = current_row
            context_text = str(fila.get("__contexto", "") or "")

        if tipo == "blank":
            current_row += 1
            continue

        if tipo == "header":
            campo = fila.get("__campo", "")
            valor = fila.get("__valor", "")
            c1 = ws.cell(row=current_row, column=1, value=campo)
            c1.font = section_font_bold
            c1.fill = section_fill
            c2 = ws.cell(row=current_row, column=2, value=valor)
            c2.font = section_font_val
            c2.fill = section_fill
            ws.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=last_col,
            )
            ws.row_dimensions[current_row].height = 15
            current_row += 1
            continue

        if tipo == "data":
            nivel = fila.get("NIVEL", "")
            if nivel == "Total":
                bg = COLOR_TOTAL_BG
                bold = True
            elif "ESPECIALI" in str(nivel).upper():
                bg = COLOR_ESP_BG
                bold = False
            else:
                bg = COLOR_MAE_BG
                bold = False

            fill = PatternFill("solid", fgColor=bg)
            c0 = ws.cell(row=current_row, column=1, value=nivel)
            c0.font = Font(bold=bold, name="Arial", size=10)
            c0.fill = fill
            c0.alignment = Alignment(horizontal="left", vertical="center")

            for j, col in enumerate(all_cols, start=2):
                v = fila.get(col, None)
                cell = ws.cell(
                    row=current_row,
                    column=j,
                    value=v if v is not None and pd.notna(v) else None,
                )
                cell.fill = fill
                cell.font = Font(bold=bold, name="Arial", size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if "CAGR" in col:
                    cell.number_format = FMT_PCT
                else:
                    cell.number_format = FMT_NUM

            ws.row_dimensions[current_row].height = 16

            if str(nivel).strip() == "Total" and context_start_row is not None:
                context_end_row = current_row
                ctx_cell = ws.cell(
                    row=context_start_row,
                    column=CONTEXT_COL,
                    value=context_text,
                )
                ctx_cell.font = Font(
                    name="Arial", size=9, color="2C3E50", italic=False
                )
                ctx_cell.fill = PatternFill("solid", fgColor="EBF3FB")
                ctx_cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
                ctx_cell.border = ctx_border
                ws.merge_cells(
                    start_row=context_start_row,
                    start_column=CONTEXT_COL,
                    end_row=context_end_row,
                    end_column=CONTEXT_COL,
                )
                context_start_row = None
                context_text = None

            current_row += 1

    ws.column_dimensions["A"].width = 42
    for j in range(2, last_col + 1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = 10 if j <= len(años_cols) + 1 else 12
    ws.column_dimensions[get_column_letter(CONTEXT_COL)].width = 52

    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}1"


def _formatear_hoja_proyecciones(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
) -> None:
    """
    Escribe y formatea la hoja Proyecciones desde cero (3 filas de encabezado).
    Fila 1: bloques de sección | Fila 2: años (P2) | Fila 3: columnas | Datos: 4+
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets["Proyecciones"]
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    cols = list(df.columns)
    p1_cols = [c for c in cols if str(c).startswith("P1_")]
    p2_cols = [c for c in cols if str(c).startswith("P2_")]
    meta_cols = [c for c in cols if c in ("N° análogos", "Fuente análogos", "¿Fallback?")]
    id_cols = [c for c in cols if c not in p1_cols + p2_cols + meta_cols]

    n_id = len(id_cols)
    n_p1 = len(p1_cols)
    n_p2 = len(p2_cols)
    n_meta = len(meta_cols)
    n_cols = len(cols)

    AZUL = "000066"
    VERDE = "1F7A3C"
    NARANJA = "BF360C"
    GRIS = "37474F"
    BLANCO = "FFFFFF"
    AZUL_CLARO = "E8EEF7"
    VERDE_CLARO = "E8F5E9"
    NARANJA_CLARO = "FFF3E0"
    GRIS_CLARO = "ECEFF1"

    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _fill(c: str) -> PatternFill:
        return PatternFill("solid", fgColor=c)

    def _font(
        c: str = BLANCO, bold: bool = True, sz: int = 9
    ) -> Font:
        return Font(bold=bold, color=c, name="Calibri", size=sz)

    def _set_block(r: int, c1: int, c2: int, titulo: str, color: str, *, sz: int = 10) -> None:
        c2 = min(c2, n_cols)
        if c1 > c2 or c1 > n_cols or n_cols == 0:
            return
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=titulo)
        cell.fill = _fill(color)
        cell.font = _font(BLANCO, bold=True, sz=sz)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    # ── Fila 1: secciones principales ─────────────────────────────────────
    _set_block(1, 1, n_id, "IDENTIFICACIÓN DEL PROGRAMA", AZUL)
    _set_block(
        1,
        n_id + 1,
        n_id + n_p1,
        "PROYECCIÓN 1\nRegresión lineal del mercado × participación EAFIT\n(estudiantes / año)",
        VERDE,
    )
    _set_block(
        1,
        n_id + n_p1 + 1,
        n_id + n_p1 + n_p2,
        "PROYECCIÓN 2 — Escenarios por análogos históricos\nP25 (pesimista) · Mediana · P75 (optimista)",
        NARANJA,
    )
    _set_block(1, n_id + n_p1 + n_p2 + 1, n_cols, "METADATOS DEL MÉTODO", GRIS)
    ws.row_dimensions[1].height = 36

    # ── Fila 2: sub-encabezados por año (solo bloque P2) ──────────────────
    p2_start = n_id + n_p1 + 1
    años_p2: list[int] = []
    for c in p2_cols:
        yr_s = str(c).split("_")[-1]
        if yr_s.isdigit():
            años_p2.append(int(yr_s))
    años_p2 = sorted(set(años_p2))

    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=ci)
        cell.border = borde
        if ci < p2_start or ci >= p2_start + n_p2:
            cell.fill = _fill(
                AZUL if ci <= n_id else VERDE if ci <= n_id + n_p1 else GRIS
            )
        else:
            cell.fill = _fill(NARANJA)

    ci = p2_start
    for yr in años_p2:
        trio = [c for c in p2_cols if str(c).endswith(f"_{yr}")]
        if len(trio) != 3:
            continue
        c_fin = ci + len(trio) - 1
        ws.merge_cells(start_row=2, start_column=ci, end_row=2, end_column=c_fin)
        cell = ws.cell(row=2, column=ci, value=str(yr))
        cell.font = _font(BLANCO, bold=True, sz=10)
        cell.fill = _fill(NARANJA)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
        ci = c_fin + 1
    ws.row_dimensions[2].height = 18

    # ── Fila 3: nombres de columna ────────────────────────────────────────
    _HDR_ID = {
        "ID Programa": "ID",
        "Programa EAFIT": "Programa EAFIT",
        "Nivel": "Nivel",
        "Categoría": "Categoría mercado",
        "Participación hist. EAFIT": "Partic. EAFIT",
    }

    def _hdr_label(col: str) -> str:
        if col in _HDR_ID:
            return _HDR_ID[col]
        if col.startswith("P1_"):
            return str(col[3:])
        if col.startswith("P2_P25_"):
            return "P25"
        if col.startswith("P2_MED_"):
            return "Mediana"
        if col.startswith("P2_P75_"):
            return "P75"
        if col == "N° análogos":
            return "N análogos"
        if col == "Fuente análogos":
            return "Fuente"
        if col == "¿Fallback?":
            return "Fallback"
        return col

    def _bloque_col(ci: int) -> str:
        if ci <= n_id:
            return "id"
        if ci <= n_id + n_p1:
            return "p1"
        if ci <= n_id + n_p1 + n_p2:
            return "p2"
        return "meta"

    _HDR_FILL = {"id": AZUL, "p1": VERDE, "p2": NARANJA, "meta": GRIS}

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=ci, value=_hdr_label(col))
        cell.font = _font(BLANCO, bold=True, sz=9)
        cell.fill = _fill(_HDR_FILL[_bloque_col(ci)])
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = borde
    ws.row_dimensions[3].height = 32

    # ── Datos (fila 4+) ───────────────────────────────────────────────────
    FMT_NUM = "#,##0"
    FMT_PCT = "0.0%"

    def _valor_celda(col: str, val) -> object:
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        if "Participación" in col:
            return float(val)
        if col.startswith("P1_") or col.startswith("P2_"):
            num = pd.to_numeric(val, errors="coerce")
            return None if pd.isna(num) else float(num)
        return val

    _DATA_FILL = {
        "id": AZUL_CLARO,
        "p1": VERDE_CLARO,
        "p2": NARANJA_CLARO,
        "meta": GRIS_CLARO,
    }

    for ri, (_, row) in enumerate(df.iterrows(), start=4):
        zebra = ri % 2 == 0
        for ci, col in enumerate(cols, 1):
            blk = _bloque_col(ci)
            cell = ws.cell(row=ri, column=ci, value=_valor_celda(col, row.get(col)))
            base_fill = _DATA_FILL[blk]
            if zebra and blk == "id":
                cell.fill = _fill("F5F8FC")
            elif zebra:
                cell.fill = _fill("FAFAFA")
            else:
                cell.fill = _fill(base_fill)
            cell.border = borde
            cell.font = Font(name="Calibri", size=10)

            if col.startswith("P1_") or col.startswith("P2_"):
                cell.number_format = FMT_NUM
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif "Participación" in col:
                cell.number_format = FMT_PCT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == "Programa EAFIT":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            elif col in ("Categoría", "Fuente análogos"):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == "¿Fallback?":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value == "Sí":
                    cell.fill = _fill("FFEBEE")
                    cell.font = Font(name="Calibri", size=10, color="C62828")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[ri].height = 16

    # ── Anchos y vista ────────────────────────────────────────────────────
    _anchos = {
        "ID Programa": 11,
        "Programa EAFIT": 42,
        "Nivel": 14,
        "Categoría": 28,
        "Participación hist. EAFIT": 12,
        "N° análogos": 10,
        "Fuente análogos": 22,
        "¿Fallback?": 10,
    }
    for ci, col in enumerate(cols, 1):
        letter = get_column_letter(ci)
        if col.startswith("P1_") or col.startswith("P2_"):
            ws.column_dimensions[letter].width = 9
        else:
            ws.column_dimensions[letter].width = _anchos.get(col, 12)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{3 + len(df)}"
