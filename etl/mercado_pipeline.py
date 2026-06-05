"""
Orquestador del pipeline de mercado (Fases 1–6 + segmentos regionales/modales).

Fases: 1-Base maestra (ML) · 2-Scrapers · 3-Sábana · 4-Scoring ·
       5-Exportación nacional · 6-EAFIT vs Mercado ·
       run_segmentos_regionales — Bogotá / Antioquia / Eje Cafetero / Virtual.
"""

from __future__ import annotations

from pathlib import Path
import threading
from typing import Callable
import re
import unicodedata

import joblib
import numpy as np
import pandas as pd
from sklearn.calibration import CalibratedClassifierCV
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, f1_score
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline

from etl.config import (
    AÑO_FIN_DATOS,
    AÑO_INICIO_HISTORICO,
    AÑO_INICIO_PRIMER_CURSO,
    ARCHIVO_PROGRAMAS,
    ARCHIVO_CATALOGO_EAFIT,
    ARCHIVO_REFERENTE_CATEGORIAS,
    BENCHMARK_COSTO,
    CHECKPOINT_BASE_MAESTRA,
    HISTORICO_ESTUDIO_MERCADO_DIR,
    HOJA_PROGRAMAS,
    HOJA_REFERENTE_CATEGORIAS,
    MODELO_CLASIFICADOR_MERCADO,
    MODELS_DIR,
    NIVELES_MERCADO,
    NIVELES_POSGRADO,
    NIVELES_PREGRADO,
    RAW_HISTORIC_DIR,
    REF_DIR,
    UMBRAL_REGIONAL_PRIMER_CURSO,
    get_benchmark_costo,
    get_smlmv_sesion,
)
from etl.exceptions_helpers import leer_excel_con_reintentos
from etl.normalizacion import limpiar_texto
from etl.pipeline_logger import (
    log_error,
    log_etapa_completada,
    log_etapa_iniciada,
    log_info,
    log_resultado,
    log_warning,
)
from etl.scraper_matriculas import SNIESMatriculasScraper
from etl.scraper_ole import OLEScraper
from etl.scoring import apply_scoring

_CAT_IDS_PATH = REF_DIR / "backup" / "cat_ids.csv"
_SCORING_PY_PATH = Path(__file__).resolve().parent / "scoring.py"


def _cache_invalida(parquet_path: Path) -> bool:
    """
    Invalida un parquet de caché si etl/scoring.py es más reciente que él.
    Retorna True si el archivo se eliminó (o no existía), False si sigue válido.

    Se usa para garantizar que cualquier recalibración del scoring (o cambio de
    SCORING_CONFIG / thresholds) se propague forzando recalculación, en lugar de
    leer caché stale.
    """
    if not parquet_path.exists():
        return True
    try:
        if (
            _SCORING_PY_PATH.exists()
            and _SCORING_PY_PATH.stat().st_mtime > parquet_path.stat().st_mtime
        ):
            parquet_path.unlink()
            log_info(f"[Cache] Invalidado {parquet_path.name} (scoring.py más reciente)")
            return True
    except OSError as e:
        log_warning(f"[Cache] No se pudo invalidar {parquet_path.name}: {e}")
    return False


def _get_or_assign_cat_id(categorias: list[str]) -> dict[str, str]:
    """
    Lee el registro permanente de IDs de categoría (ref/backup/cat_ids.csv).
    Las categorías existentes conservan su ID original.
    Las categorías nuevas reciben el siguiente ID disponible y se persisten.

    Formato del CSV: CATEGORIA_FINAL,CAT_ID

    Retorna: dict {CATEGORIA_FINAL_upper → CAT_ID}
    """
    registro: dict[str, str] = {}

    if _CAT_IDS_PATH.exists():
        try:
            df_reg = pd.read_csv(_CAT_IDS_PATH, dtype=str)
            registro.update(
                zip(
                    df_reg["CATEGORIA_FINAL"].astype(str).str.strip().str.upper(),
                    df_reg["CAT_ID"].astype(str).str.strip(),
                )
            )
        except Exception as e:
            log_warning(f"[CAT_ID] No se pudo leer {_CAT_IDS_PATH.name}: {e}")

    max_num = 0
    for cid in registro.values():
        try:
            max_num = max(max_num, int(cid.replace("CAT-", "")))
        except ValueError:
            pass

    nuevas: list[str] = []
    for cat in categorias:
        cat_norm = str(cat).strip().upper()
        if cat_norm not in registro:
            nuevas.append(cat_norm)

    for cat_norm in nuevas:
        max_num += 1
        registro[cat_norm] = f"CAT-{max_num:04d}"
        log_info(f"[CAT_ID] Nueva categoría registrada: {cat_norm} → CAT-{max_num:04d}")

    if nuevas:
        _CAT_IDS_PATH.parent.mkdir(parents=True, exist_ok=True)
        df_out = pd.DataFrame(
            [(cat, cid) for cat, cid in sorted(registro.items(), key=lambda x: x[1])],
            columns=["CATEGORIA_FINAL", "CAT_ID"],
        )
        df_out.to_csv(_CAT_IDS_PATH, index=False, encoding="utf-8-sig")
        log_info(f"[CAT_ID] Registro actualizado: {len(registro)} categorías en {_CAT_IDS_PATH.name}")

    return registro


def _normalizar_codigo_snies(serie: pd.Series) -> pd.Series:
    """Convierte códigos SNIES a string y elimina sufijo '.0'."""
    return (
        serie.astype(str)
        .str.strip()
        .str.upper()
        .str.replace(r"\.0$", "", regex=True)
    )


def _leer_primer_curso_anual(year: int, ref_dir: Path) -> pd.Series:
    """
    Lee primer_curso_{year}.xlsx desde ref/backup/ y retorna una Series
    {snies_norm -> total_anual (S1+S2)} indexada por código SNIES normalizado
    (misma lógica que _normalizar_codigo_snies).
    Retorna Series vacía si el archivo no existe o hay error.

    Cobertura soportada: AÑO_INICIO_PRIMER_CURSO..AÑO_FIN_DATOS (ref/backup/matriculas primer curso/).
    Nombre esperado: `primer_curso_{year}.xlsx`.
    El layout exacto del header se detecta dinámicamente buscando la fila
    que contenga 'CODIGO' + 'SNIES' (header=5 sigue siendo la primera apuesta
    rápida para archivos 2023/2024).
    """
    pc_dir = ref_dir / "backup" / "matriculas primer curso"
    candidatos = [
        pc_dir / f"primer_curso_{year}.xlsx",
        ref_dir / "backup" / f"primer_curso_{year}.xlsx",
        ref_dir / "backup" / "primer_curso" / f"primer_curso_{year}.xlsx",
        ref_dir / f"primer_curso_{year}.xlsx",
    ]
    ruta = next((p for p in candidatos if p.exists()), None)
    if ruta is None:
        log_warning(
            f"[Fase 1] primer_curso_{year}.xlsx no encontrado en ref/backup/. "
            "Columna quedará vacía."
        )
        return pd.Series(dtype=float, name=f"PRIMER_CURSO_{year}")

    try:
        import openpyxl
        import unicodedata as _ud

        def _norm_up(s: str) -> str:
            s2 = _ud.normalize("NFD", str(s))
            s2 = "".join(ch for ch in s2 if _ud.category(ch) != "Mn")
            return s2.upper().strip()

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        try:
            hoja = next(
                (
                    s
                    for s in wb.sheetnames
                    if "INDICE" not in str(s).upper() and "ÍNDICE" not in str(s).upper()
                ),
                wb.sheetnames[-1],
            )
        finally:
            wb.close()

        # Detección de columnas (tolerante a layouts viejos): primero probamos
        # header=5 (rápido, confirmado para 2023/2024); si las columnas clave
        # no aparecen, escaneamos las primeras 30 filas igual que el scraper.
        def _detectar_columnas(df: pd.DataFrame) -> tuple[str | None, str | None, str | None]:
            df.columns = [str(c).strip() for c in df.columns]
            col_snies = next(
                (c for c in df.columns if _norm_up(c).startswith("CODIGO") and "SNIES" in _norm_up(c)),
                None,
            ) or next(
                (c for c in df.columns if "SNIES" in _norm_up(c) and "PROGRAMA" in _norm_up(c)),
                None,
            )
            col_pc = next(
                (c for c in df.columns if "MATRICULADOS" in _norm_up(c) and "PRIMER" in _norm_up(c)),
                None,
            ) or next(
                (c for c in df.columns if "PRIMER" in _norm_up(c) and "CURSO" in _norm_up(c)),
                None,
            )
            col_sem = next((c for c in df.columns if _norm_up(c) == "SEMESTRE"), None)
            return col_snies, col_pc, col_sem

        df_pc = pd.read_excel(ruta, sheet_name=hoja, header=5, dtype=str, engine="openpyxl")
        col_snies, col_pc, col_sem = _detectar_columnas(df_pc)

        if not col_snies or not col_pc:
            # Fallback: detectar header dinámicamente
            preview = pd.read_excel(
                ruta, sheet_name=hoja, header=None, nrows=30, dtype=str, engine="openpyxl"
            )
            header_idx: int | None = None
            for idx in range(len(preview)):
                vals = [str(v) for v in preview.iloc[idx].tolist() if pd.notna(v) and str(v).strip()]
                if not vals:
                    continue
                if "CODIGO" in _norm_up(vals[0]) and any("SNIES" in _norm_up(v) for v in vals):
                    header_idx = idx
                    break
            if header_idx is not None and header_idx != 5:
                df_pc = pd.read_excel(
                    ruta, sheet_name=hoja, header=header_idx, dtype=str, engine="openpyxl"
                )
                col_snies, col_pc, col_sem = _detectar_columnas(df_pc)

        if not col_snies or not col_pc:
            log_warning(
                f"[Fase 1] primer_curso_{year}: columnas no detectadas "
                f"(buscadas: 'CÓDIGO SNIES DEL PROGRAMA', 'MATRICULADOS PRIMER CURSO'). "
                f"Columna quedará vacía."
            )
            return pd.Series(dtype=float, name=f"PRIMER_CURSO_{year}")

        use_cols = [col_snies, col_pc] + ([col_sem] if col_sem else [])
        df_pc = df_pc[use_cols].copy()
        df_pc.columns = ["SNIES", "PC"] + (["SEM"] if col_sem else [])
        if col_sem:
            df_pc = df_pc[df_pc["SEM"].astype(str).str.strip().isin(("1", "2"))].copy()

        df_pc["SNIES"] = _normalizar_codigo_snies(df_pc["SNIES"])
        df_pc["PC"] = pd.to_numeric(df_pc["PC"], errors="coerce").fillna(0)
        df_pc = df_pc[df_pc["SNIES"].str.match(r"^\d+$", na=False)]

        serie = df_pc.groupby("SNIES", sort=False)["PC"].sum()
        serie.name = f"PRIMER_CURSO_{year}"

        log_info(
            f"[Fase 1] primer_curso_{year}: {len(serie):,} programas, "
            f"{float(serie.sum()):,.0f} matriculados totales (S1+S2)."
        )
        return serie

    except Exception as exc:
        log_warning(f"[Fase 1] Error leyendo primer_curso_{year}: {exc}. Columna quedará vacía.")
        return pd.Series(dtype=float, name=f"PRIMER_CURSO_{year}")


def _build_texto_ml(df: pd.DataFrame) -> pd.Series:
    """
    Construye el texto de entrenamiento/predicción combinando campos descriptivos.

    Orden de prioridad: nombre > título > campo detallado > campo específico > área conocimiento.
    NO incluye CINE_F_2013_AC_CAMPO_AMPLIO ni NIVEL_DE_FORMACIÓN para evitar ruido.
    """
    cols = [
        "NOMBRE_DEL_PROGRAMA",
        "TITULO_OTORGADO",
        "CINE_F_2013_AC_CAMPO_DETALLADO",
        "CINE_F_2013_AC_CAMPO_ESPECÍFIC",
        "ÁREA_DE_CONOCIMIENTO",
    ]
    texto = pd.Series("", index=df.index, dtype="string")
    for col in cols:
        if col in df.columns:
            texto = texto + " " + df[col].fillna("").astype(str)
    return texto.astype(str).str.strip()


def _normalizar_nombre_programa(nombre: str) -> str:
    """
    Elimina prefijos de nivel académico del nombre para obtener el tema puro.
    Ejemplos:
      "ESPECIALIZACIÓN EN DERECHO PENAL" → "DERECHO PENAL"
      "LICENCIATURA EN MATEMÁTICAS"      → "MATEMÁTICAS"
      "TECNOLOGÍA EN SISTEMAS"           → "SISTEMAS"
      "DOCTORADO EN BIOLOGÍA"            → "BIOLOGÍA"
      "MAESTRÍA EN GERENCIA FINANCIERA"  → "GERENCIA FINANCIERA"
    """
    if not nombre or pd.isna(nombre):
        return ""
    n = str(nombre).strip()

    PREFIJOS = [
        r"^ESPECIALIZACI[OÓ]N\s+M[EÉ]DICO[- ]QUIR[ÚU]RGICA\s+EN\s+",
        r"^ESPECIALIZACI[OÓ]N\s+M[EÉ]DICO[- ]QUIR[ÚU]RGICA\s+",
        r"^ESPECIALIZACI[ÓO]N\s+EN\s+",
        r"^ESPECIALIZACION\s+EN\s+",
        r"^ESPECIALIZACIÓN\s+EN\s+",
        r"^ESPECIALIZACI[ÓO]N\s+",
        r"^ESPECIALIZACION\s+",
        r"^MAESTR[ÍI]A\s+EN\s+",
        r"^MAESTRIA\s+EN\s+",
        r"^MAESTR[ÍI]A\s+",
        r"^DOCTORADO\s+EN\s+",
        r"^DOCTORADO\s+",
        r"^LICENCIATURA\s+EN\s+",
        r"^LICENCIATURA\s+",
        r"^TECNOLOG[ÍI]A\s+EN\s+",
        r"^TECNOLOG[ÍI]A\s+",
        r"^T[ÉE]CNICA\s+PROFESIONAL\s+EN\s+",
        r"^TECNICO\s+PROFESIONAL\s+EN\s+",
        r"^T[ÉE]CNICA\s+EN\s+",
        r"^T[ÉE]CNICO\s+EN\s+",
        r"^PROGRAMA\s+DE\s+",
        r"^INGENIER[ÍI]A\s+EN\s+",
        r"^INGENIER[ÍI]A\s+",
    ]

    for p in PREFIJOS:
        n2 = re.sub(p, "", n, flags=re.IGNORECASE).strip()
        if n2 != n:
            return n2.upper()

    return n.upper()


# Mapeo verificado: nombre exacto del programa EAFIT → CATEGORIA_FINAL del mercado
# Programas sin entrada en el mapeo se reportan en log como warnings (Fase 6 no bloqueante).
MAPEO_PROGRAMAS_EAFIT: dict[str, str] = {
    "Especialización en Administración de Negocios": "ADMINISTRACION DE EMPRESAS",
    "Especialización en Agricultura Inteligente": "CIENCIAS AGROPECUARIAS",
    "Especialización en Agroquímica": "CIENCIAS AGROPECUARIAS",
    "Especialización en Ciberseguridad": "SEGURIDAD INFORMATICA Y DIGITAL",
    "Especialización en Ciencia de los Datos y Analítica": "ANALITICA DE DATOS",
    "Especialización en Derecho Humanos": "DERECHOS HUMANOS",
    "Especialización en Derecho y Sostenibilidad": "DERECHO AMBIENTAL",
    "Especialización en Gerencia Integral para Agronegocios": "AGRONEGOCIOS",
    "Especialización en Gerencia Social": "GERENCIA SOCIAL",
    "Especialización en Gerencia Tecnológica": "GERENCIA DE TECNOLOGIA",
    "Especialización en Gestión Tributaria Internacional": "GESTION TRIBUTARIA Y FISCAL",
    "Especialización en Gestión del Riesgo de Desastres y Cambio Global": "GESTION DEL RIESGO DE EMERGENCIAS Y DESASTRES",
    "Especialización en Ingeniería de Datos": "ANALITICA DE DATOS",
    "Especialización en Inteligencia Artificial": "INTELIGENCIA ARTIFICIAL",
    "Especialización en Marketing Digital": "MERCADEO DIGITAL",
    "Especialización en Matematicas e Inteligencia Artificial": "INTELIGENCIA ARTIFICIAL",
    "Especialización en Optimización": "MATEMATICAS",
    "Especialización en Propiedad Intelectual": "PROPIEDAD INTELECTUAL",
    "Especialización en Servicio y Experiencia": "DESARROLLO DE NEGOCIOS",
    "Especialización en Ética Digital": "ETICA",
    "Maestría en Biotecnología": "BIOTECNOLOGIA",
    "Maestría en Dirección de Marketing Digital": "MERCADEO DIGITAL",
    "Maestría en Energía Sostenible": "ENERGIA Y RECURSOS ENERGETICOS",
    "Maestría en Futuros": "PROSPECTIVA",
    "Maestría en Gestión del Riesgo de Desastres y Cambio Global": "GESTION DEL RIESGO DE EMERGENCIAS Y DESASTRES",
    "Maestría en Inteligencia Artificial": "INTELIGENCIA ARTIFICIAL",
    "Maestría en Liderazgo Educativo": "LIDERAZGO",
    "Maestría en Marca y Publicidad": "MERCADEO Y PUBLICIDAD",
    # Pregrados legacy (nombres con prefijo; ya no están todos en catálogo activo)
    "Pregrado en Ingeniería en Energía": "ENERGIA Y RECURSOS ENERGETICOS",
    "Pregrado en Inteligencia de Negocios": "INTELIGENCIA DE NEGOCIOS",
    # ── Pregrados EAFIT (claves = Nombre Programa EAFIT en catálogo, minúsculas) ──
    "administracion de negocios": "ADMINISTRACION DE EMPRESAS",
    "contaduria publica": "CONTABILIDAD Y AUDITORIA",
    "mercadeo": "MERCADEO Y PUBLICIDAD",
    "finanzas": "ADMINISTRACION FINANCIERA",
    "negocios internacionales": "NEGOCIOS INTERNACIONALES",
    "derecho": "DERECHO Y CIENCIAS JURIDICAS",
    "economia": "ECONOMIA",
    "psicologia": "PSICOLOGIA GENERAL",
    "ciencias politicas": "CIENCIA POLITICA",
    "comunicacion social": "COMUNICACION Y PERIODISMO",
    "ingenieria de sistemas": "INGENIERIA DE SISTEMAS Y SOFTWARE",
    "ingenieria civil": "INGENIERIA CIVIL",
    "ingenieria mecanica": "INGENIERIA MECANICA E INDUSTRIAL",
    "ingenieria de produccion": "INGENIERÍA DE PRODUCCIÓN Y PROCESOS",
    "ingenieria de procesos": "INGENIERIA QUIMICA Y PROCESOS",
    "ingenieria de diseno de producto": "INGENIERÍA DE PRODUCCIÓN Y PROCESOS",
    "ingenieria fisica": "FISICA",
    "ingenieria matematica": "MATEMATICAS",
    "ingenieria agronomica": "CIENCIAS AGROPECUARIAS",
    "biologia": "CIENCIAS BIOLOGICAS",
    "geologia": "SUELOS Y RECURSOS MINERALES",
    "musica": "MUSICA",
    "literatura": "LINGUISTICA Y LITERATURA",
    "diseno interactivo": "DISEÑO DIGITAL Y GRAFICO",
    "diseno urbano y gestion del habitat": "ARQUITECTURA",
}


def run_fase1() -> pd.DataFrame:
    """
    Fase 1: Base maestra con categorías (ML).

    - Lee Programas.xlsx y Referente_Categorias.xlsx.
    - Cruza por SNIES (left join).
    - Entrena TF-IDF + Logistic Regression sobre el referente.
    - Predice CATEGORIA_FINAL para programas sin cruce.
    - Guarda checkpoint en outputs/temp/base_maestra.parquet.

    Returns:
        DataFrame con todas las columnas de Programas + CATEGORIA_FINAL +
        FUENTE_CATEGORIA + TASA_COTIZANTES + SALARIO_OLE + INSCRITOS_2023 + INSCRITOS_2024
        + PRIMER_CURSO_{AÑO_INICIO_PRIMER_CURSO}..PRIMER_CURSO_{AÑO_FIN_DATOS} (desde ref/backup/matriculas primer curso/).
    """
    log_etapa_iniciada("Fase 1: Base maestra con categorías (ML)")

    # 1.1 Leer Programas.xlsx
    if not ARCHIVO_PROGRAMAS.exists():
        msg = f"No se encontró {ARCHIVO_PROGRAMAS}. Ejecute primero el pipeline principal."
        log_error(msg)
        raise FileNotFoundError(msg)
    df_programas = leer_excel_con_reintentos(ARCHIVO_PROGRAMAS, sheet_name=HOJA_PROGRAMAS)
    log_info(f"Programas.xlsx cargado: {len(df_programas)} filas")

    # 1.2 Leer Referente_Categorias.xlsx y deduplicar por SNIES (primera ocurrencia)
    if not ARCHIVO_REFERENTE_CATEGORIAS.exists():
        msg = f"No se encontró {ARCHIVO_REFERENTE_CATEGORIAS}."
        log_error(msg)
        raise FileNotFoundError(msg)
    df_referente = leer_excel_con_reintentos(
        ARCHIVO_REFERENTE_CATEGORIAS,
        sheet_name=HOJA_REFERENTE_CATEGORIAS,
    )
    if "SNIES" not in df_referente.columns:
        msg = "Referente_Categorias debe tener columna SNIES."
        log_error(msg)
        raise ValueError(msg)
    df_referente = df_referente.drop_duplicates(subset=["SNIES"], keep="first").reset_index(drop=True)
    log_info(f"Referente_Categorias cargado y deduplicado: {len(df_referente)} filas")

    # 1.2b Human-in-the-Loop: aplicar correcciones manuales desde ref/feedback_manual.csv (no se modifica el Excel del referente)
    feedback_path = REF_DIR / "feedback_manual.csv"
    if feedback_path.exists():
        try:
            df_feedback = pd.read_csv(feedback_path, dtype={"SNIES": str}, encoding="utf-8-sig")
            if "SNIES" in df_feedback.columns and "CATEGORIA_FINAL" in df_feedback.columns:
                df_fb = (
                    df_feedback[["SNIES", "CATEGORIA_FINAL"]]
                    .dropna(subset=["SNIES"])
                    .drop_duplicates(subset=["SNIES"], keep="last")
                )
                df_fb["SNIES"] = df_fb["SNIES"].astype(str).str.strip()
                if len(df_fb) > 0:
                    ref_idx = df_referente.set_index("SNIES")
                    fb_idx = df_fb.set_index("SNIES")
                    ref_idx.update(fb_idx)
                    df_referente = ref_idx.reset_index()
                    log_info(f"Referente actualizado con {len(df_fb)} correcciones desde feedback_manual.csv")
        except Exception as e:
            log_warning(f"No se pudo aplicar feedback_manual.csv: {e}. Se continúa con el referente original.")

    # 1.3 Normalizar llaves de cruce y texto para ML
    clave_prog = "CÓDIGO_SNIES_DEL_PROGRAMA"
    if clave_prog not in df_programas.columns:
        msg = f"Programas.xlsx debe tener columna {clave_prog}."
        log_error(msg)
        raise ValueError(msg)
    df_programas["_snies_norm"] = _normalizar_codigo_snies(df_programas[clave_prog])
    df_referente["_snies_norm"] = _normalizar_codigo_snies(df_referente["SNIES"])

    df_programas["_texto_ml_norm"] = _build_texto_ml(df_programas).apply(limpiar_texto)
    df_referente["_texto_ml_norm"] = _build_texto_ml(df_referente).apply(limpiar_texto)

    if "CATEGORIA_FINAL" not in df_referente.columns:
        msg = "Referente_Categorias debe tener columna CATEGORIA_FINAL."
        log_error(msg)
        raise ValueError(msg)

    # 1.4 Left join Programas sobre Referente por SNIES
    columnas_extra_ref = [
        "CATEGORIA_FINAL",
        "TASA_COTIZANTES",
        "SALARIO_OLE",
        f"INSCRITOS_{AÑO_FIN_DATOS - 1}",
        f"INSCRITOS_{AÑO_FIN_DATOS}",
    ]
    columnas_traer = [c for c in columnas_extra_ref if c in df_referente.columns]
    ref_merge = df_referente[["_snies_norm"] + columnas_traer].copy()

    df_base = df_programas.merge(
        ref_merge,
        on="_snies_norm",
        how="left",
        suffixes=("", "_ref"),
    )

    # Asegurar columnas de referente (nulas donde no hay cruce)
    for c in columnas_extra_ref:
        if c not in df_base.columns:
            df_base[c] = pd.NA

    # FUENTE_CATEGORIA: CRUCE_SNIES donde sí hubo cruce, resto se llenará con MATCH_NOMBRE / KNN_TFIDF
    df_base["FUENTE_CATEGORIA"] = pd.NA
    mask_cruce = df_base["CATEGORIA_FINAL"].notna()
    df_base.loc[mask_cruce, "FUENTE_CATEGORIA"] = "CRUCE_SNIES"
    log_info(f"Registros con categoría por cruce SNIES: {mask_cruce.sum()}")
    log_info(f"Registros sin categoría (a predecir): {(~mask_cruce).sum()}")

    # 1.5 Cascada de clasificación: SNIES (ya hecho) → Match nombre → KNN char_wb
    if "PROBABILIDAD" not in df_base.columns:
        df_base["PROBABILIDAD"] = pd.NA
    if "REQUIERE_REVISION" not in df_base.columns:
        df_base["REQUIERE_REVISION"] = False
    df_base["REQUIERE_REVISION"] = df_base["REQUIERE_REVISION"].fillna(False)

    # Capa 2 — Match exacto por nombre normalizado (100% certeza)
    df_referente["_nombre_base"] = df_referente["NOMBRE_DEL_PROGRAMA"].apply(_normalizar_nombre_programa)
    df_nombre_candidates = df_referente.dropna(subset=["_nombre_base", "CATEGORIA_FINAL"])
    df_nombre_candidates = df_nombre_candidates[
        df_nombre_candidates["_nombre_base"].astype(str).str.len() > 0
    ]
    df_nombre_mode = (
        df_nombre_candidates.groupby("_nombre_base")["CATEGORIA_FINAL"]
        .agg(lambda x: x.mode().iloc[0] if not x.mode().empty else x.iloc[0])
    )
    _dict_nombre_base = df_nombre_mode.to_dict()

    mask_sin_cat = df_base["FUENTE_CATEGORIA"].isna()
    if mask_sin_cat.any():
        df_base.loc[mask_sin_cat, "_nombre_base"] = df_base.loc[mask_sin_cat, "NOMBRE_DEL_PROGRAMA"].apply(_normalizar_nombre_programa)
        df_base.loc[mask_sin_cat, "_cat_nombre_base"] = df_base.loc[mask_sin_cat, "_nombre_base"].map(_dict_nombre_base)

        mask_match = mask_sin_cat & df_base["_cat_nombre_base"].notna()
        if mask_match.any():
            df_base.loc[mask_match, "CATEGORIA_FINAL"] = df_base.loc[mask_match, "_cat_nombre_base"]
            df_base.loc[mask_match, "FUENTE_CATEGORIA"] = "MATCH_NOMBRE"
            df_base.loc[mask_match, "PROBABILIDAD"] = 1.0
            df_base.loc[mask_match, "REQUIERE_REVISION"] = False

    # Alias explícitos para casos verificados donde el nombre base incluye sufijos
    _ALIAS_CATEGORIA: dict[str, str] = {
        "MEDICINA C": "MEDICINA",
    }

    # Capa 1.5 — Match directo nombre_base → categoría exacta del referente
    # Cubre el caso: nombre del programa == nombre de categoría, pero el programa
    # no está en el referente. El referente ya incluye programas universitarios (pregrado).
    # La Capa 2 busca en *programas* del referente; esta capa busca en *categorías*.
    _cats_validas = set(
        df_referente["CATEGORIA_FINAL"]
        .dropna()
        .astype(str)
        .str.strip()
        .str.upper()
        .unique()
    )
    mask_sin_cat_15 = df_base["FUENTE_CATEGORIA"].isna()
    if mask_sin_cat_15.any():
        if "_nombre_base" not in df_base.columns:
            df_base["_nombre_base"] = df_base["NOMBRE_DEL_PROGRAMA"].apply(
                _normalizar_nombre_programa
            )
        # Reutilizar _nombre_base; calcular solo para filas aún faltantes.
        mask_sin_nombre_base = mask_sin_cat_15 & df_base["_nombre_base"].isna()
        if mask_sin_nombre_base.any():
            df_base.loc[mask_sin_nombre_base, "_nombre_base"] = (
                df_base.loc[mask_sin_nombre_base, "NOMBRE_DEL_PROGRAMA"]
                .apply(_normalizar_nombre_programa)
            )

        _nombre_base_15 = (
            df_base.loc[mask_sin_cat_15, "_nombre_base"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        _hit_15 = _nombre_base_15.isin(_cats_validas)
        mask_match_15 = mask_sin_cat_15.copy()
        mask_match_15.loc[mask_sin_cat_15] = _hit_15.values

        if mask_match_15.any():
            df_base.loc[mask_match_15, "CATEGORIA_FINAL"] = (
                _nombre_base_15[_hit_15].values
            )
            df_base.loc[mask_match_15, "FUENTE_CATEGORIA"] = "MATCH_CATEGORIA"
            df_base.loc[mask_match_15, "PROBABILIDAD"] = 1.0
            df_base.loc[mask_match_15, "REQUIERE_REVISION"] = False
            log_info(
                f"Capa 1.5 — MATCH_CATEGORIA: {int(mask_match_15.sum())} programas "
                f"clasificados por nombre exacto de categoría."
            )

    # Capa 1.5b — Alias explícitos para nombres con sufijos conocidos
    mask_sin_cat_15b = df_base["FUENTE_CATEGORIA"].isna()
    if mask_sin_cat_15b.any():
        if "_nombre_base" not in df_base.columns:
            df_base["_nombre_base"] = df_base["NOMBRE_DEL_PROGRAMA"].apply(
                _normalizar_nombre_programa
            )
        _nb_15b = (
            df_base.loc[mask_sin_cat_15b, "_nombre_base"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        _hit_alias = _nb_15b.map(_ALIAS_CATEGORIA)
        mask_alias = mask_sin_cat_15b.copy()
        mask_alias.loc[mask_sin_cat_15b] = _hit_alias.notna().values

        if mask_alias.any():
            df_base.loc[mask_alias, "CATEGORIA_FINAL"] = _hit_alias[_hit_alias.notna()].values
            df_base.loc[mask_alias, "FUENTE_CATEGORIA"] = "MATCH_CATEGORIA"
            df_base.loc[mask_alias, "PROBABILIDAD"] = 1.0
            df_base.loc[mask_alias, "REQUIERE_REVISION"] = False
            log_info(
                f"Capa 1.5b — alias: {int(mask_alias.sum())} programa(s) "
                f"corregidos por alias explícito."
            )

    # Capa 3 — KNN con TF-IDF de caracteres (robusta a prefijos/ortografía)
    mask_sin_cat_final = df_base["FUENTE_CATEGORIA"].isna()
    if mask_sin_cat_final.any():
        area_col = "ÁREA_DE_CONOCIMIENTO"
        if area_col not in df_referente.columns:
            df_referente[area_col] = ""
        if area_col not in df_base.columns:
            df_base[area_col] = ""

        # Entrenar KNN solo con clases válidas.
        # Si CATEGORIA_FINAL viene como NaN, al hacer astype(str) se convierte en la etiqueta literal "nan"
        # y contaminaría el entrenamiento.
        df_referente_knn = df_referente.dropna(subset=["CATEGORIA_FINAL"]).copy()
        if "_nombre_base" not in df_referente_knn.columns:
            df_referente_knn["_nombre_base"] = df_referente_knn["NOMBRE_DEL_PROGRAMA"].apply(_normalizar_nombre_programa)
        df_referente_knn["_nombre_base"] = (
            df_referente_knn["_nombre_base"].fillna("").astype(str).str.strip()
        )
        df_referente_knn["CATEGORIA_FINAL"] = (
            df_referente_knn["CATEGORIA_FINAL"].astype(str).str.strip()
        )
        mask_label_ok = (
            df_referente_knn["CATEGORIA_FINAL"].str.len() > 0
            & (df_referente_knn["CATEGORIA_FINAL"].str.lower() != "nan")
        )
        df_referente_knn = df_referente_knn.loc[mask_label_ok]
        if df_referente_knn.empty:
            msg = "[Fase 1] KNN: no hay registros válidos en df_referente para entrenar (CATEGORIA_FINAL vacío/NaN)."
            log_error(msg)
            raise ValueError(msg)

        ref_name_base = df_referente_knn["_nombre_base"].fillna("").astype(str).str.strip()
        ref_area = df_referente_knn[area_col].fillna("").astype(str).str.upper().str.strip()
        ref_text = (ref_name_base + " " + ref_area).str.strip()

        vectorizer_knn = TfidfVectorizer(
            ngram_range=(1, 3),
            max_features=50_000,
            sublinear_tf=True,
            analyzer="char_wb",
            min_df=1,
            strip_accents="unicode",
        )
        X_ref_knn = vectorizer_knn.fit_transform(ref_text.astype(str))
        y_ref_knn = df_referente_knn["CATEGORIA_FINAL"].values

        knn_clf = KNeighborsClassifier(
            n_neighbors=5,
            metric="cosine",
            algorithm="brute",
            weights="distance",
        )
        knn_clf.fit(X_ref_knn, y_ref_knn)

        # Persistir el clasificador para compatibilidad con tests/downstream.
        # Aunque la predicción en cascada no requiera reutilizar el modelo,
        # otras partes del sistema esperan que exista el artefacto.
        try:
            MODELS_DIR.mkdir(parents=True, exist_ok=True)
            joblib.dump(
                {
                    "vectorizer": vectorizer_knn,
                    "knn": knn_clf,
                    "version": "cascada_knn_char_wb_v1",
                },
                MODELO_CLASIFICADOR_MERCADO,
            )
            log_info(f"Modelo KNN (cascada) guardado en {MODELO_CLASIFICADOR_MERCADO}")
        except Exception as e:
            log_warning(f"No se pudo persistir MODELO_CLASIFICADOR_MERCADO: {e}")

        base_subset = df_base.loc[mask_sin_cat_final]
        base_name_base = base_subset["NOMBRE_DEL_PROGRAMA"].apply(_normalizar_nombre_programa).astype(str).str.strip()
        base_area = base_subset[area_col].fillna("").astype(str).str.upper().str.strip()
        base_text = (base_name_base + " " + base_area).str.strip()

        X_pred = vectorizer_knn.transform(base_text.astype(str))
        proba_matrix = knn_clf.predict_proba(X_pred)
        preds = knn_clf.predict(X_pred)
        max_proba = proba_matrix.max(axis=1)

        clases = knn_clf.classes_
        if proba_matrix.shape[1] >= 2:
            top2_idx = np.argsort(proba_matrix, axis=1)[:, -2:][:, ::-1]
            segunda_cat = clases[top2_idx[:, 1]]
            segunda_proba = proba_matrix[np.arange(len(proba_matrix)), top2_idx[:, 1]]
        else:
            segunda_cat = np.array([""] * len(preds))
            segunda_proba = np.zeros(len(preds))

        df_base.loc[mask_sin_cat_final, "CATEGORIA_FINAL"] = preds
        df_base.loc[mask_sin_cat_final, "FUENTE_CATEGORIA"] = "KNN_TFIDF"
        df_base.loc[mask_sin_cat_final, "PROBABILIDAD"] = pd.Series(
            max_proba, index=df_base.index[mask_sin_cat_final]
        ).round(4)
        df_base.loc[mask_sin_cat_final, "REQUIERE_REVISION"] = pd.Series(
            (max_proba < 0.50),
            index=df_base.index[mask_sin_cat_final],
        ).astype(bool)
        df_base.loc[mask_sin_cat_final, "CATEGORIA_ALTERNATIVA"] = pd.Series(
            segunda_cat, index=df_base.index[mask_sin_cat_final]
        )
        df_base.loc[mask_sin_cat_final, "PROBABILIDAD_ALTERNATIVA"] = pd.Series(
            segunda_proba.round(4), index=df_base.index[mask_sin_cat_final]
        )

        n_total_pred = int(mask_sin_cat_final.sum())
        n_revision = int((max_proba < 0.50).sum())
        n_confiables = n_total_pred - n_revision
        pct_revision = (n_revision / n_total_pred * 100) if n_total_pred > 0 else 0

        log_info(f"KNN_TFIDF: {n_total_pred:,} programas predichos")
        log_info(f"  Confianza >= 0.50 (OK):               {n_confiables:,} ({100 - pct_revision:.1f}%)")
        log_info(f"  Confianza < 0.50 (REQUIERE_REVISION): {n_revision:,} ({pct_revision:.1f}%)")

        # Distribución de confianza
        bins = [0, 0.50, 0.60, 0.70, 0.80, 0.90, 1.01]
        labels = ["<0.50", "0.50-0.60", "0.60-0.70", "0.70-0.80", "0.80-0.90", ">=0.90"]
        proba_series = pd.Series(max_proba)
        counts = (
            pd.cut(proba_series, bins=bins, labels=labels, right=False)
            .value_counts()
            .sort_index()
        )
        log_info("Distribución de confianza en KNN_TFIDF:")
        for rango, count in counts.items():
            pct = (count / len(proba_series) * 100) if len(proba_series) else 0
            marker = " ← REVISIÓN REQUERIDA" if str(rango) == "<0.50" else ""
            log_info(f"  {rango}: {int(count):,} programas ({pct:.1f}%){marker}")
    else:
        # Si no hay programas sin categoría, igual aseguramos la existencia del artefacto.
        try:
            MODELS_DIR.mkdir(parents=True, exist_ok=True)
            joblib.dump({"version": "cascada_knn_char_wb_v1", "note": "no_layer3_needed"}, MODELO_CLASIFICADOR_MERCADO)
            log_info(f"Modelo stub (capa 3 no necesaria) guardado en {MODELO_CLASIFICADOR_MERCADO}")
        except Exception as e:
            log_warning(f"No se pudo persistir MODELO_CLASIFICADOR_MERCADO (stub): {e}")

    # Para programas que cruzaron por SNIES: probabilidad = 1.0, no requieren revisión
    mask_cruce_snies = df_base["FUENTE_CATEGORIA"] == "CRUCE_SNIES"
    if mask_cruce_snies.any():
        df_base.loc[mask_cruce_snies, "PROBABILIDAD"] = 1.0
        df_base.loc[mask_cruce_snies, "REQUIERE_REVISION"] = False

    # ── Primer curso AÑO_INICIO_PRIMER_CURSO..AÑO_FIN_DATOS ──────────────────
    # Fuente: ref/backup/matriculas primer curso/primer_curso_{year}.xlsx.
    # Agrega S1+S2 por CÓDIGO_SNIES_DEL_PROGRAMA mediante _leer_primer_curso_anual,
    # que detecta el header dinámicamente y tolera el archivo 2015 sin guión bajo.
    # Idempotente: borra la columna si ya existe antes de re-añadirla.
    # Si un archivo no existe se loggea warning y la columna queda con NaN.
    _SNIES_COL = "CÓDIGO_SNIES_DEL_PROGRAMA"
    _snies_norm_base = _normalizar_codigo_snies(df_base[_SNIES_COL])
    for _pc_year in range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1):
        _col_out = f"PRIMER_CURSO_{_pc_year}"
        if _col_out in df_base.columns:
            df_base.drop(columns=[_col_out], inplace=True)
        _serie_pc = _leer_primer_curso_anual(_pc_year, REF_DIR)
        df_base[_col_out] = _snies_norm_base.map(_serie_pc)
        _n_con_dato = int(df_base[_col_out].notna().sum())
        _n_base = len(df_base)
        _pct = (_n_con_dato / _n_base * 100) if _n_base else 0.0
        log_info(
            f"[Fase 1] {_col_out}: {_n_con_dato:,} / {_n_base:,} programas "
            f"con dato ({_pct:.1f}% cobertura)."
        )
    # ── Fin primer curso ──────────────────────────────────────────────────────

    # Reservar 'MANUAL' para correcciones futuras (no se asigna aquí)
    df_base = df_base.drop(
        columns=[
            "_snies_norm",
            "_nombre_norm",
            "_texto_ml_norm",
            "_nombre_base",
            "_cat_nombre_base",
        ],
        errors="ignore",
    )

    # 1.7 Guardar checkpoint parquet
    assert "PROBABILIDAD" in df_base.columns, "Falta columna PROBABILIDAD"
    assert "REQUIERE_REVISION" in df_base.columns, "Falta columna REQUIERE_REVISION"
    CHECKPOINT_BASE_MAESTRA.parent.mkdir(parents=True, exist_ok=True)
    df_base.to_parquet(CHECKPOINT_BASE_MAESTRA, index=False)
    log_info(f"Checkpoint guardado: {CHECKPOINT_BASE_MAESTRA} ({len(df_base)} filas)")
    log_etapa_completada("Fase 1: Base maestra con categorías (ML)", f"{len(df_base)} filas")

    return df_base


def validar_archivos_entrada() -> tuple[bool, list[str]]:
    """
    Verifica que los archivos necesarios para el pipeline existan y tengan
    el formato mínimo esperado. Retorna (ok, lista_de_errores).
    Llamar antes de run_fase2() para dar feedback temprano al usuario.
    """
    from etl.config import ARCHIVO_PROGRAMAS, ARCHIVO_REFERENTE_CATEGORIAS, REF_DIR

    errores: list[str] = []
    advertencias: list[str] = []

    # 1. Programas.xlsx (Fase 1)
    if not ARCHIVO_PROGRAMAS.exists():
        errores.append(f"No se encuentra Programas.xlsx en {ARCHIVO_PROGRAMAS.parent}")
    else:
        try:
            xl = pd.ExcelFile(ARCHIVO_PROGRAMAS)
            if not xl.sheet_names:
                errores.append("Programas.xlsx no tiene hojas válidas.")
        except Exception as e:
            errores.append(f"No se pudo abrir Programas.xlsx: {e}")

    # 2. Archivos de matrículas (Fase 2)
    matriculas_dir = REF_DIR / "backup" / "matriculas"
    if not matriculas_dir.exists():
        errores.append(f"Carpeta de matrículas no existe: {matriculas_dir}")
    else:
        años_encontrados: list[int] = []
        for f in matriculas_dir.glob("*.xlsx"):
            for año in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 2):
                if str(año) in f.name:
                    años_encontrados.append(año)
                    break
        años_encontrados = sorted(set(años_encontrados))
        años_esperados = list(range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1))
        años_faltantes = [a for a in años_esperados if a not in años_encontrados]
        if not años_encontrados:
            errores.append(
                f"No hay archivos de matrículas en {matriculas_dir}. "
                "Coloque los Excels con el año en el nombre (ej. matriculados_2024.xlsx)."
            )
        elif años_faltantes:
            advertencias.append(
                f"Años de matrículas faltantes: {años_faltantes}. "
                "Las categorías de esos años quedarán con matrícula 0."
            )

    # 3. OLE (opcional pero advertir si falta)
    ole_candidates = [
        REF_DIR / "backup" / "ole_indicadores.csv",
        REF_DIR / "backup" / "ole_indicadores.xlsx",
    ]
    if not any(p.exists() for p in ole_candidates):
        advertencias.append(
            "No se encuentra ole_indicadores.csv/.xlsx en ref/backup/. "
            "El salario promedio quedará como NaN (score 1) para todas las categorías."
        )

    # 4. Referente de categorías (Fase 1)
    if not ARCHIVO_REFERENTE_CATEGORIAS.exists():
        errores.append(f"No se encuentra el referente de categorías: {ARCHIVO_REFERENTE_CATEGORIAS}")

    # Log resultados
    for adv in advertencias:
        log_warning(f"[Validación] ⚠ {adv}")
    for err in errores:
        log_error(f"[Validación] ✗ {err}")
    if not errores and not advertencias:
        log_info("[Validación] ✓ Todos los archivos de entrada verificados correctamente.")
    elif not errores:
        log_info(f"[Validación] ✓ Archivos OK con {len(advertencias)} advertencia(s). Continuando.")

    return len(errores) == 0, errores


def run_fase2() -> None:
    """
    Fase 2: Descarga de datos faltantes (matrículas históricas SNIES e indicadores OLE).
    Si los scrapers no están implementados o fallan, se registra warning y se continúa
    para que Fase 3 pueda trabajar con imputación.
    """
    log_etapa_iniciada("Fase 2: Descarga de datos faltantes (Selenium)")
    scraper_mat = SNIESMatriculasScraper()
    scraper_ole = OLEScraper()

    any_matriculas = False
    any_ole = False

    # Scraper A: matrículas por semestre + inscritos anual (S1+S2) + primer_curso/graduados por semestre
    for year in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        try:
            df_ins = scraper_mat.download_inscritos(year)
            if df_ins is not None and len(df_ins) > 0:
                any_matriculas = True
        except Exception as e:
            log_warning(f"Inscritos {year}: {e}. Continuando.")
        for semestre in (1, 2):
            try:
                df_mat = scraper_mat.download_matriculados(year, semestre)
                if df_mat is not None and len(df_mat) > 0:
                    any_matriculas = True
            except Exception as e:
                log_warning(f"Matriculados {year}-{semestre}: {e}. Continuando.")
            try:
                df_pc = scraper_mat.download_primer_curso(year, semestre)
                if df_pc is not None and len(df_pc) > 0:
                    any_matriculas = True
            except Exception as e:
                log_warning(f"Primer curso {year}-{semestre}: {e}. Continuando.")
            try:
                df_grad = scraper_mat.download_graduados(year, semestre)
                if df_grad is not None and len(df_grad) > 0:
                    any_matriculas = True
            except Exception as e:
                log_warning(f"Graduados {year}-{semestre}: {e}. Continuando.")

    # Scraper B: indicadores OLE (lista SNIES desde checkpoint Fase 1)
    snies_list = []
    if CHECKPOINT_BASE_MAESTRA.exists():
        try:
            df_base = pd.read_parquet(CHECKPOINT_BASE_MAESTRA)
            col = "CÓDIGO_SNIES_DEL_PROGRAMA"
            if col in df_base.columns:
                snies_list = df_base[col].dropna().astype(str).str.strip().unique().tolist()
        except Exception as e:
            log_warning(f"No se pudo leer base maestra para lista SNIES: {e}")
    try:
        df_ole = scraper_ole.download_indicadores(snies_list)
        if df_ole is not None and len(df_ole) > 0:
            any_ole = True
    except Exception as e:
        log_warning(f"Descarga OLE: {e}. Continuando.")

    if not any_matriculas and not any_ole:
        log_error(
            "Fase 2: ambos scrapers fallaron o no devolvieron datos. "
            "Fase 3 continuará con imputación."
        )
    else:
        log_info("Fase 2: al menos un scraper aportó datos.")
    log_etapa_completada("Fase 2: Descarga de datos faltantes", "")


def _cargar_csv_raw(raw_dir: Path, nombre: str) -> pd.DataFrame:
    """Carga un CSV de raw si existe; si no, retorna DataFrame vacío con columnas esperadas."""
    path = raw_dir / nombre
    if not path.exists():
        return pd.DataFrame()
    try:
        df = pd.read_csv(
            path,
            dtype={"CÓDIGO_SNIES_DEL_PROGRAMA": str, "CODIGO_SNIES_DEL_PROGRAMA": str},
            encoding="utf-8-sig",
        )
        return df if df is not None and len(df) > 0 else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _cargar_ole_indicadores(raw_dir: Path, ref_dir: Path) -> tuple[pd.DataFrame, str]:
    """
    Carga ole_indicadores con cascada de fuentes:
      1) raw_dir/ole_indicadores.csv           → scraper reciente (prioridad máxima)
      2) ref_dir/backup/ole_indicadores.csv    → datos estáticos (backup)
      3) vacío                                 → Fase 3 imputará por mediana

    Retorna: (df_ole, fuente_label) donde fuente_label ∈ {"SCRAPER", "BACKUP", "NONE"}.
    El df retornado queda estandarizado con columnas:
      - CÓDIGO_SNIES_DEL_PROGRAMA
      - TASA_COTIZANTES
      - SALARIO_OLE
    """
    COLS_OUT = ["CÓDIGO_SNIES_DEL_PROGRAMA", "TASA_COTIZANTES", "SALARIO_OLE"]

    def _normalizar_cols(df: pd.DataFrame) -> pd.DataFrame:
        # Unificar variantes de nombre de columna SNIES
        rename_map = {}
        for c in df.columns:
            c_str = str(c).strip()
            if c_str.upper() in {"CODIGO_SNIES_DEL_PROGRAMA", "CÓDIGO_SNIES_DEL_PROGRAMA"}:
                rename_map[c] = "CÓDIGO_SNIES_DEL_PROGRAMA"
        if rename_map:
            df = df.rename(columns=rename_map)
        return df

    def _intentar_cargar(ruta: Path, label: str) -> pd.DataFrame | None:
        if not ruta.exists():
            return None
        try:
            df = pd.read_csv(ruta, dtype={"CÓDIGO_SNIES_DEL_PROGRAMA": str, "CODIGO_SNIES_DEL_PROGRAMA": str}, encoding="utf-8-sig")
            df = _normalizar_cols(df)
            if "CÓDIGO_SNIES_DEL_PROGRAMA" not in df.columns:
                log_warning(f"[Fase 3] OLE desde {label}: falta columna CÓDIGO_SNIES_DEL_PROGRAMA/CODIGO_SNIES_DEL_PROGRAMA.")
                return None
            if "TASA_COTIZANTES" not in df.columns or "SALARIO_OLE" not in df.columns:
                log_warning(
                    f"[Fase 3] OLE desde {label}: columnas insuficientes. "
                    f"Encontradas: {set(df.columns)}"
                )
                return None
            df["CÓDIGO_SNIES_DEL_PROGRAMA"] = df["CÓDIGO_SNIES_DEL_PROGRAMA"].astype(str).str.strip()
            df["TASA_COTIZANTES"] = pd.to_numeric(df["TASA_COTIZANTES"], errors="coerce")
            df["SALARIO_OLE"] = pd.to_numeric(df["SALARIO_OLE"], errors="coerce")
            tasa_max = df["TASA_COTIZANTES"].max(skipna=True)
            if pd.notna(tasa_max) and float(tasa_max) > 1.5:
                log_warning(
                    f"[Fase 3] OLE desde {label}: TASA_COTIZANTES parece venir en porcentaje (max={tasa_max:.1f}). "
                    "Convirtiendo a decimal (/100)."
                )
                df["TASA_COTIZANTES"] = df["TASA_COTIZANTES"] / 100.0
            log_info(
                f"[Fase 3] OLE cargado desde {label}: {len(df):,} programas | "
                f"con tasa: {int(df['TASA_COTIZANTES'].notna().sum()):,} | "
                f"con salario: {int(df['SALARIO_OLE'].notna().sum()):,}"
            )
            return df[COLS_OUT].copy()
        except Exception as e:
            log_warning(f"[Fase 3] Error leyendo OLE desde {label}: {e}")
            return None

    fuentes = [
        (raw_dir / "ole_indicadores.csv", "RAW/scraper", "SCRAPER"),
        (ref_dir / "backup" / "ole_indicadores.csv", "ref/backup", "BACKUP"),
    ]
    for ruta, label, out_label in fuentes:
        df = _intentar_cargar(ruta, label)
        if df is not None and len(df) > 0:
            return df, out_label

    log_warning(
        "[Fase 3] No se encontró ole_indicadores.csv en ninguna fuente. "
        "TASA_COTIZANTES y SALARIO_OLE quedarán sin dato (FUENTE_OLE=SIN_DATO)."
    )
    return pd.DataFrame(columns=COLS_OUT), "NONE"


def _limpiar_raw_csv(raw_dir: Path) -> None:
    """
    Elimina los CSVs intermedios de matrículas, inscritos y OLE
    que Fase 2 genera en outputs/historico/raw/.
    Estos archivos ya quedaron incorporados en sabana_consolidada.parquet
    y no tienen uso posterior; Fase 2 los regenera en la próxima ejecución.
    """
    patrones = [
        "matriculados_*.csv",
        "inscritos_*.csv",
        "ole_indicadores.csv",
        "primer_curso_*.csv",
        "graduados_*.csv",
    ]
    eliminados = 0
    for patron in patrones:
        for f in raw_dir.glob(patron):
            try:
                f.unlink()
                eliminados += 1
            except Exception as e:
                log_warning(f"[Fase 3] No se pudo eliminar CSV temporal {f.name}: {e}")
    if eliminados:
        log_info(f"[Fase 3] Limpieza: {eliminados} CSVs intermedios eliminados de raw/")


def run_fase3() -> None:
    """
    Fase 3: Consolidación en sábana única.
    Incorpora matrículas históricas, OLE en cascada, costo de matrícula y columnas derivadas.
    """
    log_etapa_iniciada("Fase 3: Consolidación en sábana única")
    raw_dir = RAW_HISTORIC_DIR
    base = pd.read_parquet(CHECKPOINT_BASE_MAESTRA)
    codigo_col = "CÓDIGO_SNIES_DEL_PROGRAMA"

    # Guard idempotencia: eliminar columnas previas que serán recalculadas (evita mezclar ejecuciones)
    cols_to_refresh = (
        [f"matricula_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [f"matricula_{y}_1" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [f"matricula_{y}_2" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [f"inscritos_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [f"primer_curso_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [f"graduados_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
    )
    cols_existentes = [c for c in cols_to_refresh if c in base.columns]
    if cols_existentes:
        base = base.drop(columns=cols_existentes)
        log_info(
            f"[Fase 3] Idempotencia: eliminadas {len(cols_existentes)} columnas previas "
            f"de matrículas/inscritos/primer_curso/graduados para recalcular."
        )

    # Normalizar código para joins
    base["_codigo_norm"] = _normalizar_codigo_snies(base[codigo_col])

    # 3.1 Matrículas e inscritos históricos (2019-2024, semestre 1+2)
    matricula_cols = [f"matricula_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
    sem_cols = [f"matricula_{y}_{s}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1) for s in (1, 2)]
    inscritos_cols = [f"inscritos_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
    primer_curso_cols = [f"primer_curso_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
    graduados_cols = [f"graduados_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]

    for year in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        m1 = _cargar_csv_raw(raw_dir, f"matriculados_{year}_1.csv")
        m2 = _cargar_csv_raw(raw_dir, f"matriculados_{year}_2.csv")
        if codigo_col in m1.columns:
            m1[codigo_col] = _normalizar_codigo_snies(m1[codigo_col])
        if codigo_col in m2.columns:
            m2[codigo_col] = _normalizar_codigo_snies(m2[codigo_col])

        # Columnas semestrales individuales (matricula_{year}_1, matricula_{year}_2)
        col_sem1 = f"matricula_{year}_1"
        col_sem2 = f"matricula_{year}_2"
        if len(m1) > 0 and "MATRICULADOS" in m1.columns:
            val_m1_sem = m1.groupby(codigo_col, as_index=False)["MATRICULADOS"].sum()
            val_m1_sem = val_m1_sem.rename(columns={"MATRICULADOS": col_sem1})
            val_m1_sem["_codigo_norm"] = _normalizar_codigo_snies(val_m1_sem[codigo_col])
            if col_sem1 in base.columns:
                base = base.drop(columns=[col_sem1])
            base = base.merge(
                val_m1_sem[["_codigo_norm", col_sem1]],
                on="_codigo_norm",
                how="left",
            )
        else:
            base[col_sem1] = 0
        if len(m2) > 0 and "MATRICULADOS" in m2.columns:
            val_m2_sem = m2.groupby(codigo_col, as_index=False)["MATRICULADOS"].sum()
            val_m2_sem = val_m2_sem.rename(columns={"MATRICULADOS": col_sem2})
            val_m2_sem["_codigo_norm"] = _normalizar_codigo_snies(val_m2_sem[codigo_col])
            if col_sem2 in base.columns:
                base = base.drop(columns=[col_sem2])
            base = base.merge(
                val_m2_sem[["_codigo_norm", col_sem2]],
                on="_codigo_norm",
                how="left",
            )
        else:
            base[col_sem2] = 0

        # Suma anual (matricula_{year} = sem1 + sem2)
        val_m1 = m1.groupby(codigo_col, as_index=False)["MATRICULADOS"].sum() if len(m1) > 0 and "MATRICULADOS" in m1.columns else pd.DataFrame(columns=[codigo_col, "MATRICULADOS"])
        val_m2 = m2.groupby(codigo_col, as_index=False)["MATRICULADOS"].sum() if len(m2) > 0 and "MATRICULADOS" in m2.columns else pd.DataFrame(columns=[codigo_col, "MATRICULADOS"])
        merge_m = val_m1.merge(val_m2, on=codigo_col, how="outer", suffixes=("", "_2"))
        merge_m["matricula"] = merge_m["MATRICULADOS"].fillna(0) + (merge_m["MATRICULADOS_2"].fillna(0) if "MATRICULADOS_2" in merge_m.columns else 0)
        merge_m["_codigo_norm"] = _normalizar_codigo_snies(merge_m[codigo_col])
        col_name_mat = f"matricula_{year}"
        if col_name_mat in base.columns:
            base = base.drop(columns=[col_name_mat])
        base = base.merge(merge_m[["_codigo_norm", "matricula"]].rename(columns={"matricula": col_name_mat}), on="_codigo_norm", how="left")

        # Inscritos: anual S1+S2 (preferido). Fallback a esquema viejo por semestres.
        i_anual = _cargar_csv_raw(raw_dir, f"inscritos_{year}.csv")
        col_name_ins = f"inscritos_{year}"
        if len(i_anual) > 0 and "INSCRITOS" in i_anual.columns:
            if codigo_col in i_anual.columns:
                i_anual[codigo_col] = _normalizar_codigo_snies(i_anual[codigo_col])
            val_i = (
                i_anual.groupby(codigo_col, as_index=False)["INSCRITOS"].sum()
                if codigo_col in i_anual.columns
                else pd.DataFrame(columns=[codigo_col, "INSCRITOS"])
            )
            val_i["_codigo_norm"] = _normalizar_codigo_snies(val_i[codigo_col])
            if col_name_ins in base.columns:
                base = base.drop(columns=[col_name_ins])
            base = base.merge(
                val_i[["_codigo_norm", "INSCRITOS"]].rename(columns={"INSCRITOS": col_name_ins}),
                on="_codigo_norm",
                how="left",
            )
        else:
            i1 = _cargar_csv_raw(raw_dir, f"inscritos_{year}_1.csv")
            i2 = _cargar_csv_raw(raw_dir, f"inscritos_{year}_2.csv")
            if codigo_col in i1.columns:
                i1[codigo_col] = _normalizar_codigo_snies(i1[codigo_col])
            if codigo_col in i2.columns:
                i2[codigo_col] = _normalizar_codigo_snies(i2[codigo_col])
            val_i1 = (
                i1.groupby(codigo_col, as_index=False)["INSCRITOS"].sum()
                if len(i1) > 0 and "INSCRITOS" in i1.columns
                else pd.DataFrame(columns=[codigo_col, "INSCRITOS"])
            )
            val_i2 = (
                i2.groupby(codigo_col, as_index=False)["INSCRITOS"].sum()
                if len(i2) > 0 and "INSCRITOS" in i2.columns
                else pd.DataFrame(columns=[codigo_col, "INSCRITOS"])
            )
            merge_i = val_i1.merge(val_i2, on=codigo_col, how="outer", suffixes=("", "_2"))
            merge_i["inscritos"] = merge_i["INSCRITOS"].fillna(0) + (
                merge_i["INSCRITOS_2"].fillna(0) if "INSCRITOS_2" in merge_i.columns else 0
            )
            merge_i["_codigo_norm"] = _normalizar_codigo_snies(merge_i[codigo_col])
            if col_name_ins in base.columns:
                base = base.drop(columns=[col_name_ins])
            base = base.merge(
                merge_i[["_codigo_norm", "inscritos"]].rename(columns={"inscritos": col_name_ins}),
                on="_codigo_norm",
                how="left",
            )

    # 3.2b Primer curso por año y semestre (2019-2024)
    for year in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        pc1 = _cargar_csv_raw(raw_dir, f"primer_curso_{year}_1.csv")
        pc2 = _cargar_csv_raw(raw_dir, f"primer_curso_{year}_2.csv")
        for df_pc in (pc1, pc2):
            if (
                len(df_pc) > 0
                and "PRIMER_CURSO" in df_pc.columns
                and codigo_col in df_pc.columns
            ):
                df_pc["_codigo_norm"] = _normalizar_codigo_snies(df_pc[codigo_col])
        val_pc1 = (
            pc1.groupby("_codigo_norm", as_index=False)["PRIMER_CURSO"].sum()
            if len(pc1) > 0 and "PRIMER_CURSO" in pc1.columns and "_codigo_norm" in pc1.columns
            else pd.DataFrame(columns=["_codigo_norm", "PRIMER_CURSO"])
        )
        val_pc2 = (
            pc2.groupby("_codigo_norm", as_index=False)["PRIMER_CURSO"].sum()
            if len(pc2) > 0 and "PRIMER_CURSO" in pc2.columns and "_codigo_norm" in pc2.columns
            else pd.DataFrame(columns=["_codigo_norm", "PRIMER_CURSO"])
        )
        merge_pc = val_pc1.merge(val_pc2, on="_codigo_norm", how="outer", suffixes=("", "_2"))
        merge_pc[f"primer_curso_{year}"] = merge_pc["PRIMER_CURSO"].fillna(0) + (
            merge_pc["PRIMER_CURSO_2"].fillna(0) if "PRIMER_CURSO_2" in merge_pc.columns else 0
        )
        col_pc = f"primer_curso_{year}"
        if col_pc in base.columns:
            base = base.drop(columns=[col_pc])
        base = base.merge(merge_pc[["_codigo_norm", col_pc]], on="_codigo_norm", how="left")

    # 3.2c Graduados por año y semestre (2019-2024)
    for year in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        g1 = _cargar_csv_raw(raw_dir, f"graduados_{year}_1.csv")
        g2 = _cargar_csv_raw(raw_dir, f"graduados_{year}_2.csv")
        for df_g in (g1, g2):
            if len(df_g) > 0 and "GRADUADOS" in df_g.columns and codigo_col in df_g.columns:
                df_g["_codigo_norm"] = _normalizar_codigo_snies(df_g[codigo_col])
        val_g1 = (
            g1.groupby("_codigo_norm", as_index=False)["GRADUADOS"].sum()
            if len(g1) > 0 and "GRADUADOS" in g1.columns and "_codigo_norm" in g1.columns
            else pd.DataFrame(columns=["_codigo_norm", "GRADUADOS"])
        )
        val_g2 = (
            g2.groupby("_codigo_norm", as_index=False)["GRADUADOS"].sum()
            if len(g2) > 0 and "GRADUADOS" in g2.columns and "_codigo_norm" in g2.columns
            else pd.DataFrame(columns=["_codigo_norm", "GRADUADOS"])
        )
        merge_g = val_g1.merge(val_g2, on="_codigo_norm", how="outer", suffixes=("", "_2"))
        merge_g[f"graduados_{year}"] = merge_g["GRADUADOS"].fillna(0) + (
            merge_g["GRADUADOS_2"].fillna(0) if "GRADUADOS_2" in merge_g.columns else 0
        )
        col_g = f"graduados_{year}"
        if col_g in base.columns:
            base = base.drop(columns=[col_g])
        base = base.merge(merge_g[["_codigo_norm", col_g]], on="_codigo_norm", how="left")

    # 3.2d IES — añadir ACREDITADA_ALTA_CALIDAD (no CARÁCTER_ACADÉMICO ni SECTOR)
    ies_path = REF_DIR / "backup" / "ies" / "Instituciones.xlsx"
    if ies_path.exists():
        try:
            df_ies = pd.read_excel(ies_path, sheet_name="Instituciones", dtype=str)
            df_ies.columns = [str(c).strip() for c in df_ies.columns]

            def _norm_header(h: str) -> str:
                s = unicodedata.normalize("NFD", str(h))
                s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
                return s.upper()

            col_cod_ies = None
            for c in df_ies.columns:
                cn = _norm_header(c)
                if "CODIGO" in cn and "INSTITUCION" in cn:
                    col_cod_ies = c
                    break
            if col_cod_ies and "ACREDITADA_ALTA_CALIDAD" in df_ies.columns:
                df_ies["_cod_ies_norm"] = (
                    df_ies[col_cod_ies].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                )
                df_ies_merge = df_ies[["_cod_ies_norm", "ACREDITADA_ALTA_CALIDAD"]].drop_duplicates(
                    subset=["_cod_ies_norm"]
                )
                col_ies_sabana = None
                for c in base.columns:
                    cu = _norm_header(c)
                    if (
                        "CODIGO" in cu
                        and "INSTITUCION" in cu
                        and "SNIES" not in cu
                        and "PADRE" not in cu
                    ):
                        col_ies_sabana = c
                        break
                if col_ies_sabana:
                    if "ACREDITADA_ALTA_CALIDAD" in base.columns:
                        base = base.drop(columns=["ACREDITADA_ALTA_CALIDAD"])
                    base["_cod_ies_norm"] = (
                        base[col_ies_sabana].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                    )
                    base = base.merge(df_ies_merge, on="_cod_ies_norm", how="left")
                    base = base.drop(columns=["_cod_ies_norm"])
                    n_acred = base["ACREDITADA_ALTA_CALIDAD"].notna().sum()
                    log_info(f"[Fase 3] IES: ACREDITADA_ALTA_CALIDAD añadida ({n_acred:,} programas con dato)")
                else:
                    log_warning("[Fase 3] IES: no se encontró CÓDIGO_INSTITUCIÓN en la sábana.")
            else:
                log_warning(
                    f"[Fase 3] IES: columnas clave no encontradas. Cols: {list(df_ies.columns)[:10]}"
                )
        except Exception as e:
            log_warning(f"[Fase 3] No se pudo leer Instituciones.xlsx: {e}")

    # Rellenar nulos de matrícula (anual + semestral) e inscritos con 0
    for col in matricula_cols + sem_cols + inscritos_cols:
        if col in base.columns:
            base[col] = base[col].fillna(0)

    # Rellenar nulos de primer_curso y graduados con 0
    for col in primer_curso_cols + graduados_cols:
        if col in base.columns:
            base[col] = base[col].fillna(0)

    # Cobertura de inscritos (fuente primaria: SNIES por código; los CSVs vienen de Fase 2)
    for year in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        col_ins = f"inscritos_{year}"
        if col_ins in base.columns:
            n_con_datos = int((base[col_ins] > 0).sum())
            n_total = len(base)
            pct = (n_con_datos / n_total) if n_total else 0
            log_info(f"[Fase 3] {col_ins}: {n_con_datos:,} programas con dato real de {n_total:,} ({pct:.0%})")

    # Cobertura de primer curso y graduados (últimos 2 años)
    for col in (
        [f"primer_curso_{y}" for y in (AÑO_FIN_DATOS - 1, AÑO_FIN_DATOS)]
        + [f"graduados_{y}" for y in (AÑO_FIN_DATOS - 2, AÑO_FIN_DATOS - 1)]
    ):
        if col in base.columns:
            n_con_datos = int((base[col] > 0).sum())
            n_total = len(base)
            pct = (n_con_datos / n_total) if n_total else 0
            log_info(f"[Fase 3] {col}: {n_con_datos:,} programas con dato real de {n_total:,} ({pct:.0%})")

    # 3.2 OLE en cascada: REFERENTE → SCRAPER/BACKUP → SIN_DATO (sin imputación)
    if "FUENTE_OLE" not in base.columns:
        base["FUENTE_OLE"] = pd.NA
    tiene_ref = base["TASA_COTIZANTES"].notna() | base["SALARIO_OLE"].notna()
    base.loc[tiene_ref, "FUENTE_OLE"] = "REFERENTE"

    df_ole, fuente_ole = _cargar_ole_indicadores(raw_dir=raw_dir, ref_dir=REF_DIR)
    if len(df_ole) > 0 and "CÓDIGO_SNIES_DEL_PROGRAMA" in df_ole.columns and "TASA_COTIZANTES" in df_ole.columns:
        df_ole["_codigo_norm"] = _normalizar_codigo_snies(df_ole["CÓDIGO_SNIES_DEL_PROGRAMA"])
        ole_agg = df_ole.drop_duplicates(subset=["_codigo_norm"], keep="first")[["_codigo_norm", "TASA_COTIZANTES", "SALARIO_OLE"]]
        ole_agg = ole_agg.rename(columns={"TASA_COTIZANTES": "_ole_tasa", "SALARIO_OLE": "_ole_salario"})
        base = base.merge(ole_agg, on="_codigo_norm", how="left")
        mask_sin_ole = base["FUENTE_OLE"].isna()
        base.loc[mask_sin_ole, "TASA_COTIZANTES"] = base.loc[mask_sin_ole, "TASA_COTIZANTES"].fillna(base.loc[mask_sin_ole, "_ole_tasa"])
        base.loc[mask_sin_ole, "SALARIO_OLE"] = base.loc[mask_sin_ole, "SALARIO_OLE"].fillna(base.loc[mask_sin_ole, "_ole_salario"])
        base.loc[
            mask_sin_ole & (base["TASA_COTIZANTES"].notna() | base["SALARIO_OLE"].notna()),
            "FUENTE_OLE",
        ] = fuente_ole if fuente_ole in ("SCRAPER", "BACKUP") else "SCRAPER"
        base = base.drop(columns=["_ole_tasa", "_ole_salario"], errors="ignore")

    # Sin imputación: si no hay dato OLE, el salario queda en NaN → score_salario = 1.
    # Los pesos ponderados absorben correctamente la ausencia del dato.
    # Solo marcar la fuente para trazabilidad.
    mask_sin_ole_final = base["FUENTE_OLE"].isna()
    if mask_sin_ole_final.any():
        base.loc[mask_sin_ole_final, "FUENTE_OLE"] = "SIN_DATO"

    # Logs detallados de OLE
    total_prog = len(base)
    if total_prog:
        fuente_counts = base["FUENTE_OLE"].value_counts(dropna=False)
        n_ref = int(fuente_counts.get("REFERENTE", 0))
        n_scr = int(fuente_counts.get("SCRAPER", 0))
        n_bak = int(fuente_counts.get("BACKUP", 0))
        n_sin = int(fuente_counts.get("SIN_DATO", 0))
        log_info(
            f"OLE — Referente: {n_ref} | Scraper: {n_scr} | Backup: {n_bak} | Sin dato: {n_sin} "
            f"(total: {total_prog} programas)"
        )
        if "CATEGORIA_FINAL" in base.columns and n_sin:
            df_sin_ole = base[base["FUENTE_OLE"] == "SIN_DATO"]
            if not df_sin_ole.empty:
                sin_por_cat = df_sin_ole.groupby("CATEGORIA_FINAL")["FUENTE_OLE"].count()
                total_por_cat = base.groupby("CATEGORIA_FINAL")["FUENTE_OLE"].count()
                top_sin = sin_por_cat.sort_values(ascending=False).head(5)
                log_info("Top 5 categorías con más programas sin dato OLE:")
                for cat, cnt in top_sin.items():
                    total_cat = int(total_por_cat.get(cat, 0))
                    pct = (cnt / total_cat * 100) if total_cat else 0
                    log_info(f"  · {cat}: {cnt} sin dato de {total_cat} ({pct:.1f}%)")

    # 3.3 Costo de matrícula: base → Cobertura (Principal) → mediana por categoría
    costo_col = "COSTO_MATRÍCULA_ESTUD_NUEVOS"
    if costo_col not in base.columns:
        base[costo_col] = pd.NA
    # Bandera para identificar costos imputados por mediana de categoría
    if "COSTO_IMPUTADO_MEDIANA" not in base.columns:
        base["COSTO_IMPUTADO_MEDIANA"] = False
    try:
        df_cob = leer_excel_con_reintentos(ARCHIVO_PROGRAMAS, sheet_name="Cobertura")
        if df_cob is not None and len(df_cob) > 0 and "TIPO_CUBRIMIENTO" in df_cob.columns:
            cob_principal = df_cob[df_cob["TIPO_CUBRIMIENTO"].astype(str).str.strip().str.upper() == "PRINCIPAL"].copy()
            if "VALOR_MATRICULA" in cob_principal.columns and codigo_col in cob_principal.columns:
                cob_principal["_codigo_norm"] = _normalizar_codigo_snies(cob_principal[codigo_col])
                cob_principal = cob_principal.groupby("_codigo_norm", as_index=False)["VALOR_MATRICULA"].first()
                base = base.merge(cob_principal.rename(columns={"VALOR_MATRICULA": "_costo_cob"}), on="_codigo_norm", how="left")
                base[costo_col] = base[costo_col].fillna(base["_costo_cob"])
                base = base.drop(columns=["_costo_cob"], errors="ignore")
    except Exception as e:
        log_warning(f"No se pudo cargar hoja Cobertura: {e}")
    mask_costo_nulo = base[costo_col].isna()
    if mask_costo_nulo.any() and "CATEGORIA_FINAL" in base.columns:
        medianas_costo = base.groupby("CATEGORIA_FINAL")[costo_col].transform("median")
        mask_imputar_costo = mask_costo_nulo & medianas_costo.notna()
        base.loc[mask_imputar_costo, costo_col] = base.loc[mask_imputar_costo, costo_col].fillna(
            medianas_costo[mask_imputar_costo]
        )
        base.loc[mask_imputar_costo, "COSTO_IMPUTADO_MEDIANA"] = True

    # Logs detallados de costos
    n_total_costos = len(base)
    if n_total_costos:
        n_imputados_costo = int(base["COSTO_IMPUTADO_MEDIANA"].sum())
        n_con_dato = int(base[costo_col].notna().sum())
        log_info(
            f"Costos — Con dato real: {n_con_dato} | Imputados por mediana: {n_imputados_costo}"
        )
        if "CATEGORIA_FINAL" in base.columns and n_imputados_costo:
            df_imp_costo = base[base["COSTO_IMPUTADO_MEDIANA"]]
            if not df_imp_costo.empty:
                imp_costo_cat = df_imp_costo.groupby("CATEGORIA_FINAL")["COSTO_IMPUTADO_MEDIANA"].count()
                total_costo_cat = base.groupby("CATEGORIA_FINAL")[costo_col].count()
                top_costo = imp_costo_cat.sort_values(ascending=False).head(5)
                log_info("Top 5 categorías con más costos imputados:")
                for cat, cnt in top_costo.items():
                    total_cat = int(total_costo_cat.get(cat, 0))
                    pct = (cnt / total_cat * 100) if total_cat else 0
                    log_info(f"  · {cat}: {cnt} imputados de {total_cat} ({pct:.1f}%)")
                    if total_cat and (cnt / total_cat) > 0.80:
                        log_warning(
                            f"⚠️ Categoría '{cat}' tiene {pct:.1f}% de costos imputados — baja confiabilidad."
                        )

    # 3.4 Columnas derivadas
    # Regla de negocio: ACTIVO = (matricula_último_año > 0) OR (ESTADO_PROGRAMA == 'activo')
    # Corrige programas marcados "inactivo" pero con matrícula real.
    ultimo_y = next((y for y in range(AÑO_FIN_DATOS + 1, AÑO_INICIO_HISTORICO - 1, -1) if f"matricula_{y}" in base.columns), None)
    mat_ultimo = base.get(f"matricula_{ultimo_y}", pd.Series(0, index=base.index)).fillna(0) if ultimo_y else pd.Series(0, index=base.index)
    estado_activo = (
        base["ESTADO_PROGRAMA"].astype(str).str.strip().str.lower() == "activo"
        if "ESTADO_PROGRAMA" in base.columns
        else pd.Series(False, index=base.index)
    )
    base["es_activo"] = ((mat_ultimo > 0) | estado_activo).astype(bool)
    try:
        log_info(
            f"[Fase 3] es_activo: {int(base['es_activo'].sum()):,} activos | "
            f"{int((mat_ultimo > 0).sum()):,} por matrícula ({ultimo_y}) | "
            f"{int(estado_activo.sum()):,} por ESTADO_PROGRAMA"
        )
    except Exception:
        pass
    # nuevo_por_primera_matricula: programa que entró al mercado recientemente.
    # Criterio: sin actividad en el año base (AÑO_INICIO_HISTORICO) pero con matrícula en
    # alguno de los últimos 3 años disponibles.
    # Más robusto que FECHA_DE_REGISTRO_EN_SNIES, que incluye programas sin actividad real.
    try:
        _mat_base = base.get(f"matricula_{AÑO_INICIO_HISTORICO}", pd.Series(0, index=base.index)).fillna(0)
        _años_recientes = [
            y for y in range(AÑO_FIN_DATOS - 2, AÑO_FIN_DATOS + 1)
            if f"matricula_{y}" in base.columns
        ]
        if _años_recientes:
            _mat_reciente = base[[f"matricula_{y}" for y in _años_recientes]].fillna(0).max(axis=1)
            base["nuevo_en_snies_3a"] = (_mat_base == 0) & (_mat_reciente > 0)
        else:
            # Fallback a fecha de registro si no hay años recientes de matrícula
            _fechas = pd.to_datetime(base.get("FECHA_DE_REGISTRO_EN_SNIES", pd.Series()), errors="coerce")
            base["nuevo_en_snies_3a"] = _fechas >= (pd.Timestamp.today().normalize() - pd.DateOffset(years=3))
        log_info(
            f"[Fase 3] programas_nuevos: criterio primera_matricula | "
            f"nuevos={int(base['nuevo_en_snies_3a'].sum()):,} "
            f"({base['nuevo_en_snies_3a'].mean()*100:.1f}% del universo)"
        )
    except Exception as _e:
        log_warning(f"[Fase 3] no se pudo calcular nuevo_en_snies_3a: {_e}")
        base["nuevo_en_snies_3a"] = False
    base["nuevo_en_snies_3a"] = base["nuevo_en_snies_3a"].fillna(False)
    _col_mat_fin = f"matricula_{AÑO_FIN_DATOS}"
    mat_fin = base.get(_col_mat_fin, pd.Series(0, index=base.index)).fillna(0)
    base[f"tiene_matricula_{AÑO_FIN_DATOS}"] = (mat_fin > 0).astype(bool)

    base = base.drop(columns=["_codigo_norm"], errors="ignore")

    # Validación post-merge: matrículas anuales y semestrales
    n_total = len(base)
    log_info("[Fase 3] === VALIDACIÓN DE MATRÍCULAS ===")
    for year in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        col = f"matricula_{year}"
        if col in base.columns:
            nonzero = int((base[col].fillna(0) > 0).sum())
            pct = (nonzero / n_total * 100) if n_total else 0
            flag = "✓" if pct > 5 else "⚠ ALERTA"
            log_info(
                f"[Fase 3] {flag} {col}: {nonzero:,} programas con matrícula > 0 ({pct:.1f}%)"
            )
        col_s1 = f"matricula_{year}_1"
        col_s2 = f"matricula_{year}_2"
        if col_s1 in base.columns and col_s2 in base.columns:
            nz1 = int((base[col_s1].fillna(0) > 0).sum())
            nz2 = int((base[col_s2].fillna(0) > 0).sum())
            log_info(f"[Fase 3]   sem1={nz1:,}  sem2={nz2:,}")

    # Alerta: si la mayoría de programas tiene matrícula en 0 (falla masiva)
    for col in [f"matricula_{y}" for y in range(AÑO_FIN_DATOS - 2, AÑO_FIN_DATOS + 1)]:
        if col not in base.columns or n_total == 0:
            continue
        con_dato = (base[col].fillna(0) > 0).sum()
        pct = con_dato / n_total
        if pct < 0.05:
            log_warning(
                f"[Fase 3] ALERTA: Falla masiva de datos en {col} "
                f"({pct * 100:.1f}% programas con valor > 0). Coloca los archivos en ref/backup/matriculas/."
            )

    # AAGR_ROBUSTO por programa — requerido por valorizacion_pipeline._metricas_de_subconjunto()
    def _aagr_robusto_por_programa(df_prog: pd.DataFrame) -> pd.Series:
        col_nivel = "NIVEL_DE_FORMACIÓN"
        pc_ini = df_prog.get(
            f"primer_curso_{AÑO_INICIO_HISTORICO}",
            pd.Series(0, index=df_prog.index),
        ).fillna(0)
        pc_fin = df_prog.get(
            f"primer_curso_{AÑO_FIN_DATOS}",
            pd.Series(0, index=df_prog.index),
        ).fillna(0)
        nivel = df_prog.get(
            col_nivel,
            pd.Series("ESPECIALIZACIÓN", index=df_prog.index),
        ).fillna("ESPECIALIZACIÓN")

        def _umbral(n: str) -> int:
            u = str(n).strip().upper()
            if "MAEST" in u:
                return 15
            if "UNIVER" in u:
                return 100
            return 30

        umbral = nivel.apply(_umbral)

        vars_pc = []
        for _y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
            c_c = f"primer_curso_{_y}"
            c_p = f"primer_curso_{_y - 1}"
            if c_c in df_prog.columns and c_p in df_prog.columns:
                den = df_prog[c_p].replace(0, np.nan)
                vars_pc.append((df_prog[c_c] - df_prog[c_p]) / den)
        aagr = (
            pd.concat(vars_pc, axis=1).mean(axis=1)
            if vars_pc
            else pd.Series(np.nan, index=df_prog.index)
        )

        n_years = max(AÑO_FIN_DATOS - AÑO_INICIO_HISTORICO, 1)
        mask_pos = (pc_ini > 0) & (pc_fin > 0)
        cagr = pd.Series(np.nan, index=df_prog.index)
        cagr[mask_pos] = (pc_fin[mask_pos] / pc_ini[mask_pos]) ** (1 / n_years) - 1

        result = aagr.copy()
        mask_bp = (pc_ini > 0) & (pc_ini < umbral) & cagr.notna()
        result[mask_bp] = cagr[mask_bp]
        result[(pc_fin == 0) & (pc_ini > 0)] = -1.0
        result[(pc_fin == 0) & (pc_ini == 0)] = np.nan
        return result.rename("AAGR_ROBUSTO")

    if "AAGR_ROBUSTO" not in base.columns:
        base["AAGR_ROBUSTO"] = _aagr_robusto_por_programa(base)
        log_info(
            f"[Fase 3] AAGR_ROBUSTO por programa: {base['AAGR_ROBUSTO'].notna().sum():,} con dato"
        )

    # 3.5 Guardar sábana y log
    sabana_path = CHECKPOINT_BASE_MAESTRA.parent / "sabana_consolidada.parquet"
    SCHEMA_VERSION = "v4"
    base["schema_version"] = SCHEMA_VERSION
    base.to_parquet(sabana_path, index=False)
    n = len(base)
    pct_mat24 = (base[f"tiene_matricula_{AÑO_FIN_DATOS}"].sum() / n * 100) if n else 0
    ole_reales = (base["FUENTE_OLE"].isin(["REFERENTE", "SCRAPER", "BACKUP"])).sum()
    pct_ole = (ole_reales / n * 100) if n else 0
    tiene_costo = base[costo_col].notna().sum()
    pct_costo = (tiene_costo / n * 100) if n else 0
    log_info(f"Sábana consolidada: {sabana_path} ({n} filas)")
    log_resultado(f"Total filas: {n}")
    log_resultado(f"% programas con matricula_2024 > 0: {pct_mat24:.1f}%")
    log_resultado(f"% programas con datos OLE reales (no imputados): {pct_ole:.1f}%")
    log_resultado(f"% programas con costo de matrícula disponible: {pct_costo:.1f}%")

    # Limpiar CSVs intermedios (ya incorporados en sabana_consolidada.parquet)
    _limpiar_raw_csv(raw_dir)

    log_etapa_completada("Fase 3: Consolidación en sábana única", f"{n} filas")


def run_fase4_desde_sabana(
    df: pd.DataFrame,
    modo_local: bool = False,
    niveles: frozenset[str] | None = None,
    universo: str = "posgrado",
) -> pd.DataFrame:
    """
    Ejecuta la lógica de agregación y scoring de la Fase 4 a partir de un DataFrame de sábana ya cargado.

    Args:
        df: DataFrame sábana consolidada.
        modo_local: True → score_matricula con quintiles dinámicos (segmentos regionales).
        niveles:    subconjunto de NIVELES_MERCADO a incluir. Si None → NIVELES_POSGRADO.
        universo:   "posgrado" o "pregrado" — pasado a apply_scoring para seleccionar
                    SCORING_CONFIG y thresholds de matrícula.

    No lee ni escribe archivos; retorna únicamente el DataFrame agregado por CATEGORIA_FINAL.
    """
    if "CATEGORIA_FINAL" not in df.columns:
        raise ValueError("La sábana no tiene columna CATEGORIA_FINAL.")

    dup_cols = [c for c in df.columns if str(c).endswith("_x") or str(c).endswith("_y")]
    if dup_cols:
        raise ValueError(
            f"La sábana tiene columnas duplicadas (_x/_y) como {dup_cols[:5]}. "
            "Elimine el archivo 'outputs/temp/sabana_consolidada.parquet' y vuelva a ejecutar la Fase 3 para limpiar los datos."
        )

    # ── Filtro de programas por nivel de formación ────────────────────────────
    # Filtra PROGRAMAS, no categorías. El groupby posterior produce tantas categorías
    # como categorías únicas tengan programas del nivel solicitado.
    # - NIVELES_POSGRADO → 288 categorías (todas las del referente tienen ≥1 ESP/MAE)
    # - NIVELES_PREGRADO → 144 categorías (las que tienen ≥1 UNIVERSITARIO)
    col_nivel = "NIVEL_DE_FORMACIÓN"
    _niveles_activos = niveles if niveles is not None else NIVELES_POSGRADO
    if col_nivel in df.columns and _niveles_activos:
        n_antes = len(df)
        df = df[df[col_nivel].isin(_niveles_activos)].copy()
        n_despues = len(df)
        log_info(
            f"[Fase 4 / {universo}] Filtro de programas: {n_antes:,} → {n_despues:,} "
            f"({n_antes - n_despues:,} programas de otros niveles excluidos)."
        )
        if n_despues == 0:
            raise ValueError(
                f"El filtro de niveles excluyó todos los programas. "
                f"universo='{universo}', niveles={_niveles_activos}"
            )
    else:
        log_info(
            f"[Fase 4 / {universo}] Columna NIVEL_DE_FORMACIÓN no encontrada — procesando todos."
        )

    years = list(range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1))
    grouped = df.groupby("CATEGORIA_FINAL", as_index=True)

    def _count_false(s: pd.Series) -> int:
        return int((~s.fillna(True).astype(bool)).sum())

    simple_agg = {}
    for y in years:
        c = f"matricula_{y}"
        if c in df.columns:
            simple_agg[f"suma_matricula_{y}"] = pd.NamedAgg(column=c, aggfunc="sum")
            simple_agg[f"prom_matricula_{y}"] = pd.NamedAgg(column=c, aggfunc="mean")
    for y in years:
        for s in (1, 2):
            c_sem = f"matricula_{y}_{s}"
            if c_sem in df.columns:
                simple_agg[f"suma_matricula_{y}_{s}"] = pd.NamedAgg(column=c_sem, aggfunc="sum")
    for y in (AÑO_FIN_DATOS - 1, AÑO_FIN_DATOS):
        c = f"inscritos_{y}"
        if c in df.columns:
            simple_agg[f"inscritos_{y}_suma"] = pd.NamedAgg(column=c, aggfunc="sum")
            simple_agg[f"inscritos_{y}_prom"] = pd.NamedAgg(column=c, aggfunc="mean")

    for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        c = f"primer_curso_{y}"
        if c in df.columns:
            simple_agg[f"suma_primer_curso_{y}"] = pd.NamedAgg(column=c, aggfunc="sum")
            # FIX 24: denominador = solo registros con primer_curso > 0 ese año.
            # Antes: mean sobre todos los registros (incluye PC=0) → denominador fijo
            #        → var_prom = var_suma algebraicamente → AAGR_prom = AAGR_suma siempre.
            # Ahora: mean sobre registros activos → denominador variable → AAGR_prom ≠ AAGR_suma
            #        cuando programas entran o salen de la categoría entre años.
            # Interpretación: "promedio de nuevos matriculados por programa ACTIVO ese año".
            simple_agg[f"prom_primer_curso_{y}"] = pd.NamedAgg(
                column=c,
                aggfunc=lambda x: x[x > 0].mean(),
            )

    for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        c = f"graduados_{y}"
        if c in df.columns:
            simple_agg[f"graduados_{y}_suma"] = pd.NamedAgg(column=c, aggfunc="sum")

    if "SALARIO_OLE" in df.columns:
        simple_agg["salario_promedio"] = pd.NamedAgg(column="SALARIO_OLE", aggfunc="mean")
    if "COSTO_MATRÍCULA_ESTUD_NUEVOS" in df.columns:
        simple_agg["costo_promedio"] = pd.NamedAgg(column="COSTO_MATRÍCULA_ESTUD_NUEVOS", aggfunc="mean")
    if "es_activo" in df.columns:
        simple_agg["programas_activos"] = pd.NamedAgg(column="es_activo", aggfunc="sum")
        simple_agg["programas_inactivos"] = pd.NamedAgg(column="es_activo", aggfunc=_count_false)
    if "nuevo_en_snies_3a" in df.columns:
        simple_agg["programas_nuevos_3a"] = pd.NamedAgg(column="nuevo_en_snies_3a", aggfunc="sum")
    if "nuevo_vs_snapshot_anterior" in df.columns:
        simple_agg["nuevos_vs_snapshot"] = pd.NamedAgg(column="nuevo_vs_snapshot_anterior", aggfunc="sum")

    ag = grouped.agg(**simple_agg)

    for y in years:
        mat_col = f"matricula_{y}"
        if mat_col in df.columns:
            ag[f"num_programas_{y}"] = grouped[mat_col].apply(lambda s: (s > 0).sum())
    if f"num_programas_{AÑO_INICIO_HISTORICO}" not in ag.columns:
        ag[f"num_programas_{AÑO_INICIO_HISTORICO}"] = 0
    if f"num_programas_{AÑO_FIN_DATOS}" not in ag.columns:
        ag[f"num_programas_{AÑO_FIN_DATOS}"] = 0

    ag = ag.reset_index()

    # NIVEL_MAYORIT — nivel de formación dominante de la categoría (para AAGR por universo)
    _nivel_counts = (
        df.groupby(["CATEGORIA_FINAL", "NIVEL_DE_FORMACIÓN"])
        .size()
        .reset_index(name="_cnt")
    )
    _nivel_mayorit = (
        _nivel_counts.sort_values("_cnt", ascending=False)
        .drop_duplicates(subset=["CATEGORIA_FINAL"])
        .set_index("CATEGORIA_FINAL")["NIVEL_DE_FORMACIÓN"]
    )
    if "CATEGORIA_FINAL" in ag.columns:
        ag["NIVEL_MAYORIT"] = ag["CATEGORIA_FINAL"].map(_nivel_mayorit)
    else:
        ag["NIVEL_MAYORIT"] = "ESPECIALIZACIÓN"

    for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
        c_curr = f"suma_primer_curso_{y}"
        c_prev = f"suma_primer_curso_{y-1}"
        if c_curr in ag.columns and c_prev in ag.columns:
            den = ag[c_prev].replace(0, np.nan)
            ag[f"var_primer_curso_{y}"] = (ag[c_curr] - ag[c_prev]) / den

    for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
        c_curr_p = f"prom_primer_curso_{y}"
        c_prev_p = f"prom_primer_curso_{y-1}"
        if c_curr_p in ag.columns and c_prev_p in ag.columns:
            den_p = ag[c_prev_p].replace(0, np.nan)
            ag[f"var_prom_primer_curso_{y}"] = (ag[c_curr_p] - ag[c_prev_p]) / den_p

    var_pc_cols = [
        f"var_primer_curso_{y}" for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1) if f"var_primer_curso_{y}" in ag.columns
    ]
    if var_pc_cols:
        ag["AAGR_primer_curso"] = ag[var_pc_cols].mean(axis=1)
    else:
        ag["AAGR_primer_curso"] = np.nan

    if f"graduados_{AÑO_FIN_DATOS - 1}_suma" in ag.columns and f"suma_matricula_{AÑO_FIN_DATOS - 2}" in ag.columns:
        den_grad = ag[f"suma_matricula_{AÑO_FIN_DATOS - 2}"].replace(0, np.nan)
        ag["tasa_graduacion"] = (ag[f"graduados_{AÑO_FIN_DATOS - 1}_suma"] / den_grad).clip(0, 2)

    if f"suma_primer_curso_{AÑO_FIN_DATOS}" in ag.columns:
        ag["tiene_primer_curso_real"] = ag[f"suma_primer_curso_{AÑO_FIN_DATOS}"] > 0

    # var_suma y var_prom (2020-2024)
    for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
        s_curr = ag.get(f"suma_matricula_{y}", pd.Series(dtype=float))
        s_prev = ag.get(f"suma_matricula_{y-1}", pd.Series(dtype=float))
        if s_curr is not None and s_prev is not None and len(s_prev) == len(ag):
            den = s_prev.replace(0, np.nan)
            ag[f"var_suma_{y}"] = (s_curr - s_prev) / den
        p_curr = ag.get(f"prom_matricula_{y}", pd.Series(dtype=float))
        p_prev = ag.get(f"prom_matricula_{y-1}", pd.Series(dtype=float))
        if p_curr is not None and p_prev is not None and len(p_prev) == len(ag):
            den_p = p_prev.replace(0, np.nan)
            ag[f"var_prom_{y}"] = (p_curr - p_prev) / den_p

    # participacion_2019 — sobre primer_curso (consistente con participacion_2024)
    _col_pc_ini = f"suma_primer_curso_{AÑO_INICIO_HISTORICO}"
    if _col_pc_ini in ag.columns:
        _total_pc_ini = ag[_col_pc_ini].sum()
        if _total_pc_ini and _total_pc_ini != 0:
            ag[f"participacion_{AÑO_INICIO_HISTORICO}"] = ag[_col_pc_ini] / _total_pc_ini
        else:
            ag[f"participacion_{AÑO_INICIO_HISTORICO}"] = np.nan
            log_warning(
                f"[Fase 4] participacion_{AÑO_INICIO_HISTORICO} = NaN para todas las categorías. "
                f"Verificar que primer_curso_{AÑO_INICIO_HISTORICO}.xlsx exista en ref/backup/."
            )
    else:
        ag[f"participacion_{AÑO_INICIO_HISTORICO}"] = np.nan
        log_warning(
            f"[Fase 4] No existe columna suma_primer_curso_{AÑO_INICIO_HISTORICO}. "
            f"participacion_{AÑO_INICIO_HISTORICO} quedará vacía. "
            f"Verificar que primer_curso_{AÑO_INICIO_HISTORICO}.xlsx exista en ref/backup/."
        )

    # participacion_2024 — sobre primer_curso (peso relativo del flujo de nuevos)
    if f"prom_primer_curso_{AÑO_FIN_DATOS}" in ag.columns:
        total_pc_2024 = ag[f"prom_primer_curso_{AÑO_FIN_DATOS}"].sum()
        if total_pc_2024 and total_pc_2024 != 0:
            ag[f"participacion_{AÑO_FIN_DATOS}"] = ag[f"prom_primer_curso_{AÑO_FIN_DATOS}"] / total_pc_2024
        else:
            ag[f"participacion_{AÑO_FIN_DATOS}"] = np.nan
    else:
        total_prom_2024 = ag[f"prom_matricula_{AÑO_FIN_DATOS}"].sum() if f"prom_matricula_{AÑO_FIN_DATOS}" in ag.columns else 0
        ag[f"participacion_{AÑO_FIN_DATOS}"] = (
            ag[f"prom_matricula_{AÑO_FIN_DATOS}"] / total_prom_2024
            if total_prom_2024 != 0 else np.nan
        )

    # AAGR_suma — promedio de variaciones anuales de PRIMER CURSO (no matrícula total)
    var_pc_sum_cols = [
        f"var_primer_curso_{y}"
        for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1)
        if f"var_primer_curso_{y}" in ag.columns
    ]
    ag["AAGR_suma"] = ag[var_pc_sum_cols].mean(axis=1) if var_pc_sum_cols else np.nan
    _var_prom_pc_cols = [
        f"var_prom_primer_curso_{y}"
        for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1)
        if f"var_prom_primer_curso_{y}" in ag.columns
    ]
    ag["AAGR_prom"] = ag[_var_prom_pc_cols].mean(axis=1) if _var_prom_pc_cols else np.nan

    # CAGR_suma — tasa compuesta sobre primer_curso (demanda nueva)
    ag["CAGR_suma"] = np.nan
    _n_years_cagr = max(AÑO_FIN_DATOS - AÑO_INICIO_HISTORICO, 1)
    _col_pc_ini = f"suma_primer_curso_{AÑO_INICIO_HISTORICO}"
    _col_pc_fin = f"suma_primer_curso_{AÑO_FIN_DATOS}"
    if _col_pc_ini in ag.columns and _col_pc_fin in ag.columns:
        _pc_ini = ag[_col_pc_ini]
        _pc_fin = ag[_col_pc_fin]
        _mask_cagr = (_pc_ini > 0) & (_pc_fin > 0)
        ag.loc[_mask_cagr, "CAGR_suma"] = (
            _pc_fin[_mask_cagr] / _pc_ini[_mask_cagr]
        ) ** (1 / _n_years_cagr) - 1

    # CAGR_primer_curso — idéntico a CAGR_suma; se mantienen ambos nombres por compatibilidad
    ag["CAGR_primer_curso"] = np.nan
    if f"suma_primer_curso_{AÑO_INICIO_HISTORICO}" in ag.columns and f"suma_primer_curso_{AÑO_FIN_DATOS}" in ag.columns:
        pc_2019 = ag[f"suma_primer_curso_{AÑO_INICIO_HISTORICO}"]
        pc_2024 = ag[f"suma_primer_curso_{AÑO_FIN_DATOS}"]
        mask_cagr_pc = (pc_2019 > 0) & (pc_2024 > 0)
        ag.loc[mask_cagr_pc, "CAGR_primer_curso"] = (
            pc_2024[mask_cagr_pc] / pc_2019[mask_cagr_pc]
        ) ** (1 / _n_years_cagr) - 1

    # AAGR_ROBUSTO — árbol de decisión con UMBRAL_BASE diferenciado por universo.
    # El umbral determina cuándo la base histórica (suma_primer_curso_2019) es suficientemente
    # grande para confiar en AAGR en vez de CAGR.
    # Valores calibrados sobre distribución real Colombia:
    #   ESP: P10 de suma_PC_2019 = 13 → umbral 30 clasifica 69% como NORMAL ✓
    #   MAE: P10 de suma_PC_2019 = 12 → umbral 15 clasifica 73% como NORMAL ✓
    #   PRE: P10 de suma_PC_2019 = 54 → umbral 100 clasifica 65% como NORMAL ✓
    _ESP_NIVELES = {
        "ESPECIALIZACIÓN", "ESPECIALIZACIÓN MÉDICO QUIRÚRGICA",
        "ESPECIALIZACIÓN TECNOLÓGICA", "ESPECIALIZACIÓN TÉCNICO PROFESIONAL",
    }
    _MAE_NIVELES = {"MAESTRÍA"}
    _PRE_NIVELES = {"UNIVERSITARIO"}

    def _umbral_base(nivel: str) -> int:
        nivel_u = str(nivel).strip().upper()
        if nivel_u in _PRE_NIVELES:
            return 100    # Pregrado: base histórica mínima alta (mayor volumen)
        if nivel_u in _MAE_NIVELES:
            return 15     # Maestría: base histórica pequeña → umbral bajo
        return 30         # Especialización (default): umbral intermedio

    # Calcular UMBRAL_BASE por fila usando NIVEL_MAYORIT
    if "NIVEL_MAYORIT" in ag.columns:
        _umbral_series = ag["NIVEL_MAYORIT"].apply(_umbral_base)
    else:
        _umbral_series = pd.Series(30, index=ag.index)  # fallback global
    if f"suma_primer_curso_{AÑO_INICIO_HISTORICO}" in ag.columns and f"suma_primer_curso_{AÑO_FIN_DATOS}" in ag.columns:
        m19 = ag[f"suma_primer_curso_{AÑO_INICIO_HISTORICO}"].fillna(0)
        m24 = ag[f"suma_primer_curso_{AÑO_FIN_DATOS}"].fillna(0)

        cond_normal = m19 >= _umbral_series
        cond_pequena = (m19 > 0) & (m19 < _umbral_series)
        cond_nueva = (m19 == 0) & (m24 > 0)
        cond_extinta = (m24 == 0) & (m19 > 0)
        cond_sin_act = (m19 == 0) & (m24 == 0)

        tipo = pd.Series("NORMAL", index=ag.index, dtype=object)
        tipo[cond_pequena] = "BASE_PEQUENA"
        tipo[cond_nueva] = "CATEGORIA_NUEVA"
        tipo[cond_extinta] = "EXTINTA"
        tipo[cond_sin_act] = "SIN_ACTIVIDAD"
        ag["TIPO_CRECIMIENTO"] = tipo

        aagr_r = (
            ag["AAGR_primer_curso"].copy()
            if "AAGR_primer_curso" in ag.columns
            else pd.Series(np.nan, index=ag.index)
        )
        if "CAGR_primer_curso" in ag.columns:
            mask_cagr_ok = cond_pequena & ag["CAGR_primer_curso"].notna()
            aagr_r[mask_cagr_ok] = ag.loc[mask_cagr_ok, "CAGR_primer_curso"]

        # CATEGORIA_NUEVA: AAGR_primer_curso solo si hay >= 3 años con primer_curso > 0
        MIN_AÑOS_NUEVA = 3
        if "AAGR_primer_curso" in ag.columns:
            _cols_pc = [
                f"suma_primer_curso_{y}"
                for y in range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1)
                if f"suma_primer_curso_{y}" in ag.columns
            ]
            if _cols_pc:
                _años_con_dato = (ag[_cols_pc].fillna(0) > 0).sum(axis=1)
            else:
                _años_con_dato = pd.Series(0, index=ag.index)

            mask_nueva_con_dato = (
                cond_nueva
                & (_años_con_dato >= MIN_AÑOS_NUEVA)
                & ag["AAGR_primer_curso"].notna()
            )
            mask_nueva_sin_dato = cond_nueva & ~mask_nueva_con_dato

            aagr_r[mask_nueva_sin_dato] = np.nan
            log_info(
                f"CATEGORIA_NUEVA: {int(cond_nueva.sum())} total | "
                f"{int(mask_nueva_con_dato.sum())} con AAGR_primer_curso "
                f"(>={MIN_AÑOS_NUEVA} años) | "
                f"{int(mask_nueva_sin_dato.sum())} con NaN (<{MIN_AÑOS_NUEVA} años)"
            )
        else:
            aagr_r[cond_nueva] = np.nan
        aagr_r[cond_extinta] = -1.0
        aagr_r[cond_sin_act] = np.nan
        ag["AAGR_ROBUSTO"] = aagr_r

        _tipos_str = ", ".join(
            f"{t}={int((tipo == t).sum())}"
            for t in ["NORMAL", "BASE_PEQUENA", "CATEGORIA_NUEVA", "EXTINTA", "SIN_ACTIVIDAD"]
            if (tipo == t).any()
        )
        if "NIVEL_MAYORIT" in ag.columns:
            for _univ in ["ESPECIALIZACIÓN", "MAESTRÍA", "UNIVERSITARIO"]:
                _mask_univ = ag["NIVEL_MAYORIT"].astype(str).str.upper() == _univ.upper()
                _n_normal = int((tipo[_mask_univ] == "NORMAL").sum())
                _n_total = int(_mask_univ.sum())
                log_info(
                    f"  AAGR {_univ[:3]}: NORMAL={_n_normal}/{_n_total} "
                    f"({_n_normal / max(_n_total, 1) * 100:.0f}%)"
                )
        log_info(f"AAGR_ROBUSTO calculado. {_tipos_str}")
    else:
        ag["AAGR_ROBUSTO"] = ag.get("AAGR_primer_curso", pd.Series(np.nan, index=ag.index))
        ag["TIPO_CRECIMIENTO"] = "SIN_DATOS"
        log_info("AAGR_ROBUSTO: no se encontraron columnas suma_primer_curso_2019/2024.")

    # ── MOMENTUM YoY: crecimiento del último año vs. tendencia histórica ──────────
    # var_yoy_2024: variación real del último año (2023→2024)
    # diferencial_tendencia: cuánto se desvía el YoY del AAGR histórico
    # SEÑAL_TENDENCIA: etiqueta legible para el analista
    if f"suma_primer_curso_{AÑO_FIN_DATOS - 1}" in ag.columns and f"suma_primer_curso_{AÑO_FIN_DATOS}" in ag.columns:
        m23 = pd.to_numeric(ag[f"suma_primer_curso_{AÑO_FIN_DATOS - 1}"], errors="coerce").fillna(0)
        m24 = pd.to_numeric(ag[f"suma_primer_curso_{AÑO_FIN_DATOS}"], errors="coerce").fillna(0)
        den_yoy = m23.replace(0, np.nan)
        ag[f"var_yoy_{AÑO_FIN_DATOS}"] = (m24 - m23) / den_yoy

        if "AAGR_ROBUSTO" in ag.columns:
            ag["diferencial_tendencia"] = ag[f"var_yoy_{AÑO_FIN_DATOS}"] - pd.to_numeric(ag["AAGR_ROBUSTO"], errors="coerce")
        else:
            ag["diferencial_tendencia"] = np.nan

        # SEÑAL_TENDENCIA: combina YoY y diferencial para una etiqueta accionable
        def _señal(row: pd.Series) -> str:
            yoy = row.get(f"var_yoy_{AÑO_FIN_DATOS}", np.nan)
            dif = row.get("diferencial_tendencia", np.nan)
            if pd.isna(yoy):
                return "SIN_DATO"
            if yoy >= 0.10 and (pd.isna(dif) or dif >= -0.05):
                return "ACELERANDO"     # Crece bien y no hay desaceleración significativa
            if yoy >= 0.00 and not pd.isna(dif) and dif < -0.10:
                return "DESACELERANDO"  # Sigue creciendo pero mucho menos que antes
            if yoy >= 0.00:
                return "ESTABLE"        # Crecimiento positivo o nulo sin señal de alarma
            if yoy < 0.00 and not pd.isna(dif) and dif < -0.10:
                return "EN_DECLIVE"     # Cae y peor que su promedio histórico
            return "CONTRACCION"        # Cae pero dentro de lo esperable por su historia

        ag["SEÑAL_TENDENCIA"] = ag.apply(_señal, axis=1)

        # Sobrescribir SEÑAL_TENDENCIA para categorías extintas y sin actividad.
        if "TIPO_CRECIMIENTO" in ag.columns:
            mask_extinta = ag["TIPO_CRECIMIENTO"] == "EXTINTA"
            mask_sin_act = ag["TIPO_CRECIMIENTO"] == "SIN_ACTIVIDAD"
            ag.loc[mask_extinta, "SEÑAL_TENDENCIA"] = "SIN_ACTIVIDAD"
            ag.loc[mask_sin_act, "SEÑAL_TENDENCIA"] = "SIN_ACTIVIDAD"
            n_sin_act_total = int((mask_extinta | mask_sin_act).sum())
            if n_sin_act_total:
                log_info(
                    f"SEÑAL_TENDENCIA: {n_sin_act_total} categorías marcadas como SIN_ACTIVIDAD "
                    f"(EXTINTA={int(mask_extinta.sum())}, "
                    f"SIN_ACTIVIDAD_tipo={int(mask_sin_act.sum())})"
                )

        log_info(
            "Momentum YoY calculado. Señales: "
            + ", ".join(
                f"{s}={int((ag['SEÑAL_TENDENCIA'] == s).sum())}"
                for s in [
                    "ACELERANDO", "ESTABLE", "DESACELERANDO", "EN_DECLIVE",
                    "CONTRACCION", "SIN_ACTIVIDAD", "SIN_DATO",
                ]
                if (ag["SEÑAL_TENDENCIA"] == s).any()
            )
        )
    else:
        ag[f"var_yoy_{AÑO_FIN_DATOS}"] = np.nan
        ag["diferencial_tendencia"] = np.nan
        ag["SEÑAL_TENDENCIA"] = "SIN_DATO"
        log_info("Momentum YoY: no se encontraron columnas suma_primer_curso_2023/2024.")

    # Bloque B: pct_no_matriculados y var_inscritos
    # Fórmula: pct = (inscritos - primer_curso) / inscritos
    # Comparar inscritos vs primer_curso (no vs matricula_total):
    #   inscritos      = personas que aplicaron ese año
    #   primer_curso   = personas que se matricularon por primera vez
    #   pct_no_mat     = tasa de rechazo/abandono en la transición inscripción→matrícula
    # Comparar con matricula_total era incorrecto: matrícula acumula múltiples cohortes
    # y siempre supera a inscritos (ratio típico: ins/mat ≈ 0.3–0.7).

    def _calc_pct_vs_pc(ins_col: str, pc_col: str) -> "pd.Series":
        """
        pct = (inscritos - primer_curso) / inscritos, rango [0, 1].
        Si inscritos < primer_curso → NaN (dato SNIES incoherente, no se imputa).
        Si inscritos = 0 → NaN (sin denominador).
        """
        if ins_col not in ag.columns or pc_col not in ag.columns:
            return pd.Series(np.nan, index=ag.index)
        ins = pd.to_numeric(ag[ins_col], errors="coerce")
        pc = pd.to_numeric(ag[pc_col], errors="coerce")
        den = ins.replace(0, np.nan)
        raw = (ins - pc) / den
        mask_invalido = ins.notna() & pc.notna() & (ins < pc)
        raw = raw.where(~mask_invalido, other=np.nan)
        return raw.clip(lower=0, upper=1)

    # pct_no_matriculados_2023
    ag[f"pct_no_matriculados_{AÑO_FIN_DATOS - 1}"] = _calc_pct_vs_pc(
        f"inscritos_{AÑO_FIN_DATOS - 1}_suma", f"suma_primer_curso_{AÑO_FIN_DATOS - 1}"
    )

    # pct_no_matriculados_2024 — prioridad 2024, fallback 2023 (solo si falta dato, no si incoherente)
    pct_2024 = _calc_pct_vs_pc(f"inscritos_{AÑO_FIN_DATOS}_suma", f"suma_primer_curso_{AÑO_FIN_DATOS}")
    pct_2023 = _calc_pct_vs_pc(f"inscritos_{AÑO_FIN_DATOS - 1}_suma", f"suma_primer_curso_{AÑO_FIN_DATOS - 1}")

    _col_ins_24 = f"inscritos_{AÑO_FIN_DATOS}_suma"
    _col_pc_24 = f"suma_primer_curso_{AÑO_FIN_DATOS}"
    _ins_24 = pd.to_numeric(ag.get(_col_ins_24, pd.Series(np.nan, index=ag.index)), errors="coerce")
    _pc_24 = pd.to_numeric(ag.get(_col_pc_24, pd.Series(np.nan, index=ag.index)), errors="coerce")
    _mask_inv_2024 = _ins_24.notna() & _pc_24.notna() & (_ins_24 < _pc_24)
    _pct_combined = pct_2024.combine_first(pct_2023)
    _pct_combined = _pct_combined.where(~_mask_inv_2024, other=np.nan)
    ag[f"pct_no_matriculados_{AÑO_FIN_DATOS}"] = _pct_combined
    if _mask_inv_2024.any():
        log_warning(
            f"[Consistencia] {_mask_inv_2024.sum()} categorías con "
            f"inscritos_{AÑO_FIN_DATOS} < primer_curso_{AÑO_FIN_DATOS} — "
            f"pct_no_matriculados_{AÑO_FIN_DATOS} = NaN (no se usa fallback 2023)."
        )

    # FUENTE: basado en si inscritos_suma > 0 (no en coherencia ins>=mat)
    ins24_ok = ag.get(f"inscritos_{AÑO_FIN_DATOS}_suma", pd.Series(0, index=ag.index))
    ins24_ok = pd.to_numeric(ins24_ok, errors="coerce").fillna(0) > 0
    ins23_ok = ag.get(f"inscritos_{AÑO_FIN_DATOS - 1}_suma", pd.Series(0, index=ag.index))
    ins23_ok = pd.to_numeric(ins23_ok, errors="coerce").fillna(0) > 0

    ag["FUENTE_PCT_NO_MAT"] = "SIN_DATOS"
    ag.loc[ins24_ok, "FUENTE_PCT_NO_MAT"] = f"INSCRITOS_{AÑO_FIN_DATOS}"
    ag.loc[~ins24_ok & ins23_ok, "FUENTE_PCT_NO_MAT"] = f"INSCRITOS_{AÑO_FIN_DATOS - 1}_FALLBACK"

    ag["tiene_inscritos_reales"] = ins24_ok | ins23_ok

    # Inscritos promedio por programa (columna informativa — equivale a "Inscritos Prom" del manual)
    # = inscritos_suma / num_programas_{AÑO_FIN_DATOS} (programas con matrícula activa)
    for _yr in (AÑO_FIN_DATOS - 1, AÑO_FIN_DATOS):
        suma_col = f"inscritos_{_yr}_suma"
        if suma_col in ag.columns and f"num_programas_{AÑO_FIN_DATOS}" in ag.columns:
            _den = pd.to_numeric(ag[f"num_programas_{AÑO_FIN_DATOS}"], errors="coerce").replace(0, np.nan)
            _suma = pd.to_numeric(ag[suma_col], errors="coerce")
            ag[f"inscritos_{_yr}_prom_por_programa"] = (_suma / _den).round(1)
        else:
            ag[f"inscritos_{_yr}_prom_por_programa"] = np.nan

    # Variación del promedio por programa de inscritos (penúltimo → último año)
    _prom_prev_col = f"inscritos_{AÑO_FIN_DATOS - 1}_prom_por_programa"
    _prom_last_col = f"inscritos_{AÑO_FIN_DATOS}_prom_por_programa"
    if _prom_prev_col in ag.columns and _prom_last_col in ag.columns:
        _prom23 = pd.to_numeric(ag[_prom_prev_col], errors="coerce").replace(0, np.nan)
        _prom24 = pd.to_numeric(ag[_prom_last_col], errors="coerce")
        ag["var_inscritos_prom"] = ((_prom24 - _prom23) / _prom23).clip(-1.0, 3.0)
    else:
        ag["var_inscritos_prom"] = np.nan

    # var_inscritos: variación anual de inscripciones
    if f"inscritos_{AÑO_FIN_DATOS - 1}_suma" in ag.columns and f"inscritos_{AÑO_FIN_DATOS}_suma" in ag.columns:
        ins23 = pd.to_numeric(ag[f"inscritos_{AÑO_FIN_DATOS - 1}_suma"], errors="coerce")
        ins24 = pd.to_numeric(ag[f"inscritos_{AÑO_FIN_DATOS}_suma"], errors="coerce")
        den_i = ins23.replace(0, np.nan)
        ag["var_inscritos"] = ((ins24 - ins23) / den_i).clip(-1.0, 3.0)
    else:
        ag["var_inscritos"] = np.nan

    # Bloque C: var_programas, pct_con_matricula, prom_matricula_por_programa_2024
    if f"num_programas_{AÑO_INICIO_HISTORICO}" in ag.columns and f"num_programas_{AÑO_FIN_DATOS}" in ag.columns:
        den_p = ag[f"num_programas_{AÑO_INICIO_HISTORICO}"].replace(0, np.nan)
        ag["var_programas"] = (
            (ag[f"num_programas_{AÑO_FIN_DATOS}"] - ag[f"num_programas_{AÑO_INICIO_HISTORICO}"]) / den_p
        ).clip(-1.0, 3.0)  # Acotar a [-100%, +300%] — elimina outliers extremos
        # Nota: fórmula idéntica al archivo manual de referencia.
        # El clip evita que 5 categorías con crecimiento >200% distorsionen el promedio.
    else:
        ag["var_programas"] = np.nan
    if "programas_activos" in ag.columns and "programas_inactivos" in ag.columns:
        _total_registros = ag["programas_activos"] + ag["programas_inactivos"]
        ag["pct_con_matricula"] = (
            ag["programas_activos"] / _total_registros.replace(0, np.nan)
        )
        ag["pct_con_matricula"] = ag["pct_con_matricula"].clip(0, 1)
    else:
        ag["pct_con_matricula"] = np.nan
    # NOTA: prom_matricula_por_programa_XXXX es un alias de prom_primer_curso_XXXX.
    # El nombre se mantiene por compatibilidad con scoring.py y valorizacion_pipeline.py.
    # Semánticamente representa "nuevos matriculados promedio por registro SNIES",
    # no "matrícula total promedio". Ver Fix 11 y Fix 20.
    if f"prom_primer_curso_{AÑO_FIN_DATOS}" in ag.columns:
        ag[f"prom_matricula_por_programa_{AÑO_FIN_DATOS}"] = ag[f"prom_primer_curso_{AÑO_FIN_DATOS}"]
    elif f"num_programas_{AÑO_FIN_DATOS}" in ag.columns and f"suma_matricula_{AÑO_FIN_DATOS}" in ag.columns:
        ag[f"prom_matricula_por_programa_{AÑO_FIN_DATOS}"] = (
            ag[f"suma_matricula_{AÑO_FIN_DATOS}"] / ag[f"num_programas_{AÑO_FIN_DATOS}"].replace(0, np.nan)
        )
    else:
        ag[f"prom_matricula_por_programa_{AÑO_FIN_DATOS}"] = np.nan

    # Bloque D: distancia_costo_pct por nivel de formación
    # Se calcula a nivel de programa (antes de agregar) para usar el benchmark correcto por nivel
    col_costo_prog = "COSTO_MATRÍCULA_ESTUD_NUEVOS"
    col_nivel_prog = "NIVEL_DE_FORMACIÓN"
    if col_costo_prog in df.columns:
        def _distancia_costo(row):
            costo = row.get(col_costo_prog)
            nivel = row.get(col_nivel_prog, "")
            if pd.isna(costo) or costo == 0:
                return np.nan
            bench = get_benchmark_costo(str(nivel))
            return (float(costo) - bench) / bench * 100

        df["_distancia_costo_prog"] = df.apply(_distancia_costo, axis=1)
        distancia_por_cat = df.groupby("CATEGORIA_FINAL")["_distancia_costo_prog"].mean()
        ag = ag.merge(
            distancia_por_cat.rename("distancia_costo_pct").reset_index(),
            on="CATEGORIA_FINAL",
            how="left",
        )
        df = df.drop(columns=["_distancia_costo_prog"], errors="ignore")
    elif "costo_promedio" in ag.columns:
        # Fallback: si no hay costo por programa, usar costo_promedio con benchmark general
        ag["distancia_costo_pct"] = (ag["costo_promedio"] - BENCHMARK_COSTO) / BENCHMARK_COSTO * 100
    else:
        ag["distancia_costo_pct"] = np.nan

    # salario_promedio ya viene expresado en SMLMV (ej. 3.5). No dividir por SMLMV en pesos.
    # NaN en salario_promedio (sin OLE) se propaga sin fillna → score_salario = 1.
    if "salario_promedio" in ag.columns:
        ag["salario_promedio_smlmv"] = ag["salario_promedio"]
        smlmv = float(get_smlmv_sesion())
        ag["salario_proyectado_pesos_hoy"] = ag["salario_promedio"] * smlmv if smlmv else np.nan
    else:
        ag["salario_promedio_smlmv"] = np.nan
        ag["salario_proyectado_pesos_hoy"] = np.nan

    # Reemplazar inf por nan
    ag = ag.replace([np.inf, -np.inf], np.nan)

    # Bloque E: scoring
    ag = apply_scoring(ag, modo_local=modo_local, universo=universo)

    # Restaurar NaN en export para inscritos < primer_curso (scoring usa fillna 0.25 interno)
    _col_pct_nm = f"pct_no_matriculados_{AÑO_FIN_DATOS}"
    if _col_pct_nm in ag.columns and _mask_inv_2024.any():
        ag.loc[_mask_inv_2024, _col_pct_nm] = np.nan

    # CAT_ID — identificador estable de categoría desde registro permanente
    if "CATEGORIA_FINAL" in ag.columns:
        _id_map = _get_or_assign_cat_id(ag["CATEGORIA_FINAL"].tolist())
        ag["CAT_ID"] = (
            ag["CATEGORIA_FINAL"]
            .astype(str).str.strip().str.upper()
            .map(_id_map)
            .fillna("CAT-????")
        )
    else:
        ag["CAT_ID"] = "CAT-????"

    return ag


def run_fase4() -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
    """
    Fase 4: Agregación por CATEGORIA_FINAL y scoring ponderado.
    Retorna (ag_posgrado, ag_pregrado).

    ag_posgrado → hoja 'total'           (288 categorías, solo programas ESP+MAE)
    ag_pregrado → hoja 'total_pregrado'  (144 categorías, solo programas UNIVERSITARIO)
    """
    log_etapa_iniciada("Fase 4: Agregación por categoría")
    out_path = CHECKPOINT_BASE_MAESTRA.parent / "agregado_categorias.parquet"
    out_path_pre = CHECKPOINT_BASE_MAESTRA.parent / "agregado_categorias_pregrado.parquet"
    sabana_path = CHECKPOINT_BASE_MAESTRA.parent / "sabana_consolidada.parquet"

    if not sabana_path.exists():
        log_error("No existe sábana consolidada. Ejecutar Fase 3 antes.")
        return None, None
    df = pd.read_parquet(sabana_path)
    SCHEMA_VERSION = "v4"
    if "schema_version" in df.columns:
        sv = str(df["schema_version"].iloc[0]) if len(df) else ""
        if sv and sv != SCHEMA_VERSION:
            log_warning(
                f"[Fase 4] ALERTA: sabana_consolidada.parquet tiene schema_version='{sv}' "
                f"pero se esperaba '{SCHEMA_VERSION}'. Elimine 'sabana_consolidada.parquet' y re-ejecute Fase 3."
            )
    else:
        log_warning(
            f"[Fase 4] ALERTA: sabana_consolidada.parquet no tiene schema_version. "
            "Si ve columnas mezcladas, elimine el parquet y re-ejecute Fase 3."
        )
    if "CATEGORIA_FINAL" not in df.columns:
        log_error("Sábana sin columna CATEGORIA_FINAL.")
        return None, None

    smlmv_actual = get_smlmv_sesion()
    log_info(f"SMLMV usado en scoring: ${smlmv_actual:,.0f}")

    # ── Invalidar cachés si scoring.py es más reciente ────────────────────────
    _cache_invalida(out_path)
    _cache_invalida(out_path_pre)
    for _seg in ["Antioquia", "Bogota", "Eje_Cafetero", "Virtual", "Colombia"]:
        _cache_invalida(CHECKPOINT_BASE_MAESTRA.parent / f"agregado_{_seg}.parquet")

    # ── Universo posgrado (ESP + MAE, 288 categorías) ─────────────────────────
    ag_pos = run_fase4_desde_sabana(
        df, modo_local=False, niveles=NIVELES_POSGRADO, universo="posgrado"
    )
    if ag_pos is None:
        log_error("Fase 4: la agregación de posgrado no produjo datos.")
        return None, None

    out_path.parent.mkdir(parents=True, exist_ok=True)
    ag_pos.to_parquet(out_path, index=False)
    log_info(
        f"[Fase 4] Agregado posgrado guardado: {out_path.name} "
        f"({len(ag_pos)} categorías, solo ESP+MAE)"
    )

    # ── Universo pregrado (UNIVERSITARIO, 144 categorías) ─────────────────────
    ag_pre: pd.DataFrame | None = None
    try:
        ag_pre = run_fase4_desde_sabana(
            df, modo_local=False, niveles=NIVELES_PREGRADO, universo="pregrado"
        )
        if ag_pre is not None:
            ag_pre.to_parquet(out_path_pre, index=False)
            log_info(
                f"[Fase 4] Agregado pregrado guardado: {out_path_pre.name} "
                f"({len(ag_pre)} categorías, solo UNIVERSITARIO)"
            )
        else:
            log_warning(
                "[Fase 4] Pregrado: agregación vacía — hoja total_pregrado no se generará."
            )
    except Exception as _e:
        log_warning(
            f"[Fase 4] Pregrado: error en agregación ({_e}) — hoja total_pregrado omitida."
        )
        ag_pre = None

    log_etapa_completada(
        "Fase 4: Agregación por categoría",
        f"{len(ag_pos)} cats posgrado | "
        f"{len(ag_pre) if ag_pre is not None else 0} cats pregrado",
    )
    return ag_pos, ag_pre


# Bloques para hoja "total" (encabezado fila 1)
_BLOQUES_TOTAL = [
    ("CATEGORÍA", [
        "CAT_ID", "CATEGORIA_FINAL", "FUENTE_CATEGORIA", "NIVEL_MAYORIT",
    ]),
    ("DEMANDA NUEVA — PRIMER CURSO",
        [f"suma_primer_curso_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [f"prom_primer_curso_{y}" for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)]
        + [
            f"var_primer_curso_{y}"
            for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1)
        ]
        + [
            f"var_prom_primer_curso_{y}"
            for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1)
        ],
    ),
    ("PARTICIPACIÓN Y CRECIMIENTO", [
        f"participacion_{AÑO_INICIO_HISTORICO}", f"participacion_{AÑO_FIN_DATOS}",
        "AAGR_suma", "AAGR_prom", "CAGR_suma",
        "AAGR_ROBUSTO", "TIPO_CRECIMIENTO", "SEÑAL_TENDENCIA",
    ]),
    ("INSCRITOS", [
        f"inscritos_{AÑO_FIN_DATOS - 1}_suma", f"inscritos_{AÑO_FIN_DATOS}_suma",
        f"inscritos_{AÑO_FIN_DATOS - 1}_prom_por_programa",
        f"inscritos_{AÑO_FIN_DATOS}_prom_por_programa",
        f"pct_no_matriculados_{AÑO_FIN_DATOS - 1}", f"pct_no_matriculados_{AÑO_FIN_DATOS}",
        "FUENTE_PCT_NO_MAT", "var_inscritos", "var_inscritos_prom",
    ]),
    ("SALARIO OLE", [
        "salario_promedio", "salario_proyectado_pesos_hoy",
    ]),
    ("OFERTA DE PROGRAMAS", [
        f"num_programas_{AÑO_INICIO_HISTORICO}", f"num_programas_{AÑO_FIN_DATOS}",
        "programas_activos", "programas_inactivos",
        "programas_nuevos_3a", "nuevos_vs_snapshot",
        "var_programas", "pct_con_matricula",
    ]),
    ("COSTO", [
        "costo_promedio", "distancia_costo_pct",
    ]),
    ("SCORING — valor | puntuación", [
        f"prom_primer_curso_{AÑO_FIN_DATOS}",   "score_matricula",
        f"participacion_{AÑO_FIN_DATOS}",        "score_participacion",
        "AAGR_ROBUSTO",                           "score_AAGR",
        "salario_promedio",                       "score_salario",
        f"pct_no_matriculados_{AÑO_FIN_DATOS}",  "score_pct_no_matriculados",
        f"num_programas_{AÑO_FIN_DATOS}",         "score_num_programas",
        "distancia_costo_pct",                    "score_costo",
    ]),
    ("CALIFICACIÓN FINAL", [
        "calificacion_final",
    ]),
]

HEADERS_CONTEXTO_NACIONAL = {
    "CATEGORIA_FINAL": "Categoría de mercado",
    "calificacion_final": "Calificación en este segmento (1-5)",
    "calificacion_nacional": "Calificación Colombia nacional (1-5)",
    "AAGR_ROBUSTO": "AAGR primer curso en segmento",
    "AAGR_ROBUSTO_nacional": "AAGR primer curso nacional",
    f"participacion_{AÑO_FIN_DATOS}": "Participación del segmento en mercado nacional",
    f"num_programas_{AÑO_FIN_DATOS}": "N° programas en el segmento",
}

HEADERS_CONTEXTO_NACIONAL.update({
    f"suma_primer_curso_regional_{AÑO_FIN_DATOS}": (
        f"Primer curso total segmento {AÑO_FIN_DATOS}"
    ),
    f"suma_primer_curso_{AÑO_FIN_DATOS}": f"Primer curso total nacional {AÑO_FIN_DATOS}",
})

COL_ANCHOS_PROGRAMAS = {
    "CÓDIGO_SNIES_DEL_PROGRAMA": 18,
    "NOMBRE_DEL_PROGRAMA": 50,
    "NOMBRE_INSTITUCIÓN": 40,
}

VERDE = "C6EFCE"
AMARILLO = "FFEB9C"
ROJO = "FFC7CE"


def _escribir_hoja_estandar(writer: pd.ExcelWriter) -> None:
    """
    Crea la hoja 'estandar_calificacion' con la tabla de umbrales del scoring.

    Propósito: que el analista pueda entender qué significa cada score (1-5)
    para cada métrica, sin necesidad de abrir scoring.py.
    Se ubica entre resumen_ejecutivo y total para dar contexto antes de los datos.
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    ws = writer.book.create_sheet("estandar_calificacion")

    AZUL_OSC = "000066"
    BLANCO = "FFFFFF"
    GRIS_CLR = "F5F5F5"
    SCORE_COLORES = {1: "FFCDD2", 2: "FFE0B2", 3: "FFF9C4", 4: "DCEDC8", 5: "C8E6C9"}
    SCORE_FUENTE = {1: "B71C1C", 2: "E65100", 3: "827717", 4: "1B5E20", 5: "1B5E20"}

    thin = Side(style="thin", color="BDBDBD")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(ws, row, col, value, bold=False, bg=None, fg="000000",
              halign="center", wrap=False, size=10):
        c = ws.cell(row=row, column=col)
        c.value = str(value) if value is not None else ""
        c.data_type = "s"
        c.font = Font(bold=bold, color=fg, name="Arial", size=size)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
        c.border = borde
        return c

    ws.merge_cells("A1:N1")
    _cell(ws, 1, 1,
          "ESTÁNDAR DE CALIFICACIÓN — Cómo se evalúa cada categoría de mercado",
          bold=True, bg=AZUL_OSC, fg=BLANCO, size=12)
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:N2")
    _cell(ws, 2, 1,
          "Cada categoría recibe un score de 1 (peor) a 5 (mejor) en 7 métricas. "
          "La Calificación Final = suma ponderada de los 7 scores. "
          "Verde ≥ 4.0 · Amarillo 3.0–3.9 · Rojo < 3.0",
          bold=False, bg="E3F2FD", fg="0D47A1", halign="left", wrap=True, size=10)
    ws.row_dimensions[2].height = 30

    fila = 4
    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1, "PESOS DE CADA MÉTRICA EN LA CALIFICACIÓN FINAL",
          bold=True, bg="37474F", fg=BLANCO, size=10)
    ws.row_dimensions[fila].height = 20
    fila += 1

    _yr_fin = AÑO_FIN_DATOS
    _yr_ant = AÑO_FIN_DATOS - 1
    pesos_data = [
        (f"Prom. primer curso {_yr_fin}", "S. Primer curso", "30%",
         f"Nuevos matriculados promedio por registro SNIES en {_yr_fin}. "
         "Mide el tamaño de la demanda real del mercado."),
        ("Participación en mercado", "S. Participación", "15%",
         "Fracción del mercado nacional que representa esta categoría. "
         "Quintiles dinámicos del segmento actual."),
        ("AAGR primer curso (robusto)", "S. AAGR", "20%",
         f"Crecimiento anual promedio histórico del primer_curso "
         f"{AÑO_INICIO_HISTORICO}-{_yr_fin}. "
         f"Usa CAGR para categorías con base pequeña (< 30 estudiantes en {AÑO_INICIO_HISTORICO})."),
        ("Salario promedio (SMLMV)", "S. Salario", "15%",
         "Salario promedio de egresados en SMLMV, según OLE. "
         "Indica el retorno laboral del área de conocimiento."),
        ("% Inscritos no matriculados", "S. No matr.", "10%",
         "Fracción de inscritos que no se matricularon. "
         "Inverso: menos es mejor. Indica conversión de demanda en matrícula."),
        ("N° programas en mercado", "S. N° Programas", "5%",
         "Número de programas activos en la categoría. "
         "Inverso: menos programas = menor competencia = mejor oportunidad."),
        ("Distancia vs benchmark EAFIT (%)", "S. Costo", "5%",
         "Cuánto cuesta el mercado vs lo que cobra EAFIT. "
         "Positivo: el mercado es más caro (EAFIT puede competir). "
         "Negativo: el mercado es más barato (EAFIT debe revisar precio)."),
    ]

    for ci, h in enumerate(["Métrica", "Columna score", "Peso", "Descripción"], 1):
        _cell(ws, fila, ci, h, bold=True, bg="546E7A", fg=BLANCO, size=9)
    fila += 1

    for i, (metrica, col_score, peso, desc) in enumerate(pesos_data):
        bg = GRIS_CLR if i % 2 == 0 else BLANCO
        _cell(ws, fila, 1, metrica, bold=True, bg=bg, halign="left")
        _cell(ws, fila, 2, col_score, bg=bg)
        _cell(ws, fila, 3, peso, bold=True, bg=bg)
        _cell(ws, fila, 4, desc, bg=bg, halign="left", wrap=True)
        ws.row_dimensions[fila].height = 32
        fila += 1

    fila += 1

    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1,
          "UMBRALES DE SCORE POR MÉTRICA — Posgrado (Especialización y Maestría)",
          bold=True, bg="37474F", fg=BLANCO, size=10)
    ws.row_dimensions[fila].height = 20
    fila += 1

    METRICAS_HEADERS = [
        "Score", "Peso",
        f"Prom. primer\ncurso {_yr_fin}\n(S. Primer curso, 30%)",
        "AAGR primer\ncurso robusto\n(S. AAGR, 20%) — ESP",
        "AAGR primer\ncurso robusto\n(S. AAGR, 20%) — MAE",
        "Salario\n(SMLMV)\n(S. Salario, 15%)",
        f"% Inscritos\nno matr. {_yr_fin}\n(S. No matr., 10%)",
        "N° programas\nen mercado\n(S. N° Prog., 5%)",
        "Distancia vs\nbenchmark EAFIT\n(S. Costo, 5%)",
    ]
    for ci, h in enumerate(METRICAS_HEADERS, 1):
        _cell(ws, fila, ci, h, bold=True, bg="455A64", fg=BLANCO, wrap=True, size=9)
        ws.row_dimensions[fila].height = 44
    fila += 1

    UMBRALES = [
        (1, "0-1.3", "≤ 3.0", "≤ 0.8%", "≤ -5.1%", "≤ 2 SMLMV", "> 50%", "> 32", "< -60%"),
        (2, "1.3-2.3", "≤ 5.4", "≤ 5.8%", "≤ -0.6%", "≤ 3 SMLMV", "≤ 50%", "≤ 32", "< -40%"),
        (3, "2.3-3.3", "≤ 8.5", "≤ 10.3%", "≤ 3.3%", "≤ 5 SMLMV", "≤ 30%", "≤ 18", "< -15%"),
        (4, "3.3-4.3", "≤ 13.8", "≤ 18.0%", "≤ 8.0%", "≤ 8 SMLMV", "≤ 20%", "≤ 10", "< +20%"),
        (5, "4.3-5.0", "> 13.8", "> 18.0%", "> 8.0%", "> 8 SMLMV", "≤ 10%", "≤ 4", "≥ +20%"),
    ]
    ETIQUETAS = {
        1: "1 — Bajo", 2: "2 — Bajo-Medio", 3: "3 — Medio",
        4: "4 — Medio-Alto", 5: "5 — Alto",
    }

    for score, peso_rango, mat, aagr_e, aagr_m, sal, nomat, nprog, costo in UMBRALES:
        bg = SCORE_COLORES[score]
        fg = SCORE_FUENTE[score]
        _cell(ws, fila, 1, ETIQUETAS[score], bold=True, bg=bg, fg=fg)
        _cell(ws, fila, 2, peso_rango, bg=bg, fg=fg)
        for ci, val in enumerate([mat, aagr_e, aagr_m, sal, nomat, nprog, costo], 3):
            _cell(ws, fila, ci, val, bg=bg, fg=fg)
        ws.row_dimensions[fila].height = 20
        fila += 1

    fila += 1
    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1,
          "(*) El AAGR usa umbrales distintos para Especialización (ESP) y Maestría (MAE) "
          "porque el mercado de maestrías tiene menor volatilidad (63% de categorías MAE "
          "tienen AAGR < 3.3%). La participación usa quintiles dinámicos del segmento actual.",
          bg="FFF8E1", fg="5D4037", halign="left", wrap=True, size=9)
    ws.row_dimensions[fila].height = 28
    fila += 1

    fila += 1
    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1, "TIPO DE MERCADO — Qué significa cada clasificación",
          bold=True, bg="2E7D32", fg=BLANCO, size=10)
    ws.row_dimensions[fila].height = 20
    fila += 1

    _yr_ini = AÑO_INICIO_HISTORICO
    TIPOS_MERCADO = [
        ("NORMAL",
         f"El mercado tiene datos desde {_yr_ini} con base suficiente (≥ 30 estudiantes en ESP, "
         f"≥ 15 en MAE). El AAGR se calcula como promedio de las 5 variaciones anuales."),
        ("BASE_PEQUEÑA",
         f"El mercado existe desde {_yr_ini} pero con pocos estudiantes (< 30 ESP / < 15 MAE). "
         "El AAGR usa el CAGR (tasa compuesta) en lugar del promedio de variaciones, "
         "para reducir el ruido estadístico de mercados pequeños."),
        ("CATEGORIA_NUEVA",
         f"El mercado no existía en {_yr_ini} (primer_curso_{_yr_ini} = 0) pero sí en años recientes. "
         "El AAGR se calcula desde el primer año con dato disponible."),
        ("EXTINTA",
         f"El mercado existía en {_yr_ini} pero no tiene primer_curso en {_yr_fin}. "
         "AAGR_ROBUSTO = -1.0 (penalización máxima). Score AAGR = 1."),
        ("SIN_ACTIVIDAD",
         "El mercado no tiene primer_curso en ningún año del período. "
         "AAGR_ROBUSTO = NaN. Score AAGR = 1 (fallback conservador)."),
    ]

    for ci, h in enumerate(["Tipo de mercado", "Significado"], 1):
        _cell(ws, fila, ci, h, bold=True, bg="388E3C", fg=BLANCO, size=9)
    ws.row_dimensions[fila].height = 18
    fila += 1

    TIPO_BG = {
        "NORMAL": "F1F8E9", "BASE_PEQUEÑA": "FFF8E1", "CATEGORIA_NUEVA": "E3F2FD",
        "EXTINTA": "FFEBEE", "SIN_ACTIVIDAD": "F5F5F5",
    }
    for tipo, desc in TIPOS_MERCADO:
        bg = TIPO_BG.get(tipo, "FFFFFF")
        _cell(ws, fila, 1, tipo, bold=True, bg=bg, halign="left")
        _cell(ws, fila, 2, desc, bg=bg, halign="left", wrap=True)
        ws.row_dimensions[fila].height = 36
        fila += 1

    fila += 1
    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1,
          f"SEÑAL DE TENDENCIA — Momento actual del mercado "
          f"(variación primer_curso {_yr_ant}→{_yr_fin})",
          bold=True, bg="1565C0", fg=BLANCO, size=10)
    ws.row_dimensions[fila].height = 20
    fila += 1

    SEÑALES = [
        ("▲ ACELERANDO", "C8E6C9", "1B5E20",
         f"Var. {_yr_ant}→{_yr_fin} > +10%. El mercado crece fuerte en el último año."),
        ("→ ESTABLE", "FFF9C4", "827717",
         f"Var. {_yr_ant}→{_yr_fin} entre −10% y +10%. Crecimiento moderado y consistente."),
        ("▼ DESACELERANDO", "FFE0B2", "E65100",
         f"Var. {_yr_ant}→{_yr_fin} negativa pero superior a −20%. Crecimiento que se frena."),
        ("↓ EN DECLIVE", "FFCDD2", "B71C1C",
         f"Var. {_yr_ant}→{_yr_fin} entre −20% y −50%. El mercado pierde estudiantes."),
        ("↓↓ CONTRACCION", "EF9A9A", "B71C1C",
         f"Var. {_yr_ant}→{_yr_fin} < −50%. Caída severa de la demanda."),
        ("— SIN DATO", "EEEEEE", "757575",
         f"No hay dato de primer_curso en {_yr_ant} o {_yr_fin} para calcular la variación."),
    ]

    for ci, h in enumerate(["Señal", "Descripción"], 1):
        _cell(ws, fila, ci, h, bold=True, bg="1976D2", fg=BLANCO, size=9)
    ws.row_dimensions[fila].height = 18
    fila += 1

    for señal, bg, fg, desc in SEÑALES:
        _cell(ws, fila, 1, señal, bold=True, bg=bg, fg=fg)
        _cell(ws, fila, 2, desc, bg=bg, fg=fg, halign="left", wrap=True)
        ws.row_dimensions[fila].height = 28
        fila += 1

    fila += 1
    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1,
          "DIFERENCIA ENTRE AAGR SUMA Y AAGR PROM POR PROGRAMA",
          bold=True, bg="5D4037", fg=BLANCO, size=10)
    ws.row_dimensions[fila].height = 18
    fila += 1

    ws.merge_cells(f"A{fila}:N{fila}")
    _cell(ws, fila, 1,
          "AAGR suma del mercado: promedio de las variaciones interanuales de la SUMA total "
          "de primer_curso. Responde a: ¿el mercado como un todo crece o decrece en nuevos "
          "estudiantes? Es el insumo de AAGR_ROBUSTO (y del scoring S. AAGR). "
          "||  "
          "AAGR prom. por programa: promedio de las variaciones del PROMEDIO por registro SNIES. "
          "Responde a: ¿cada programa individualmente captura más estudiantes? "
          "Métrica informativa, no entra al scoring. "
          "Si AAGR_suma sube pero AAGR_prom baja, el mercado crece porque hay más programas, "
          "no porque los existentes mejoren.",
          bg="EFEBE9", fg="3E2723", halign="left", wrap=True, size=10)
    ws.row_dimensions[fila].height = 60

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 18
    ws.sheet_view.showGridLines = False
    ws.sheet_state = "visible"


def _escribir_resumen_ejecutivo(
    writer: pd.ExcelWriter,
    sabana: pd.DataFrame,
    ag: pd.DataFrame,
    _sem_verde: int | None = None,
    _sem_amarillo: int | None = None,
    _sem_rojo: int | None = None,
    titulo: str = "COLOMBIA",
) -> None:
    """
    Genera hoja "resumen_ejecutivo" (primera) con KPIs globales, rankings y calidad.
    No debe romper el export: cualquier excepción se captura y se continúa.
    """
    try:
        from openpyxl.styles import Border, Font, PatternFill, Side
        from openpyxl.styles import Alignment as XLAlignment

        # Estilos (según especificación)
        AZUL_EAFIT = "000066"
        VERDE_FILL = "C6EFCE"
        AMARILLO_FILL = "FFEB9C"
        ROJO_FILL = "FFC7CE"
        GRIS_HEADER = "F2F2F2"
        DATA_ALT = "F9F9F9"

        wb = writer.book

        # Reposicionar la hoja al inicio
        if "resumen_ejecutivo" in wb.sheetnames:
            ws_old = wb["resumen_ejecutivo"]
            wb.remove(ws_old)
        ws = wb.create_sheet("resumen_ejecutivo", 0)

        # Evitar que la sheet por defecto quede intercalada si está presente
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1 and wb["Sheet"].max_column <= 1:
            try:
                wb.remove(wb["Sheet"])
            except Exception:
                pass

        ws.freeze_panes = "A3"

        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_font = Font(size=16, bold=True, color=AZUL_EAFIT)
        subtitle_font = Font(size=11, color="666666")
        header_fill = PatternFill(start_color=AZUL_EAFIT, end_color=AZUL_EAFIT, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_fill_alt = PatternFill(start_color=GRIS_HEADER, end_color=GRIS_HEADER, fill_type="solid")
        data_fill = PatternFill(start_color=DATA_ALT, end_color=DATA_ALT, fill_type="solid")

        def set_cell(
            r: int,
            c: int,
            value,
            *,
            font: Font | None = None,
            fill: PatternFill | None = None,
            align: XLAlignment | None = None,
            number_format: str | None = None,
            bold: bool | None = None,
            border_on: bool = False,
        ) -> None:
            cell = ws.cell(row=r, column=c)
            if value is None or (isinstance(value, float) and np.isnan(value)):
                cell.value = None
            else:
                cell.value = value
            if font is not None:
                cell.font = font
            if bold is not None:
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=bold,
                    color=cell.font.color,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                )
            if fill is not None:
                cell.fill = fill
            if align is not None:
                cell.alignment = align
            if border_on:
                cell.border = border
            if number_format is not None:
                cell.number_format = number_format

        def _valor_excel(value, number_fmt: str | None):
            """Coerce a tipo numérico nativo para que Excel aplique number_format."""
            if number_fmt is None or value in ("", None):
                return value
            if isinstance(value, float) and np.isnan(value):
                return None
            if number_fmt == "0.0%":
                return float(pd.to_numeric(value, errors="coerce"))
            if number_fmt in ("#,##0", "0.00"):
                num = pd.to_numeric(value, errors="coerce")
                return None if pd.isna(num) else float(num)
            return value

        def merge_and_set_title(
            r: int, c1: int, c2: int, value: str, *, fill: PatternFill, font: Font
        ) -> None:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            set_cell(
                r,
                c1,
                value,
                fill=fill,
                font=font,
                align=XLAlignment(horizontal="left", vertical="center"),
                border_on=False,
            )

        # Anchos de columna
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15

        # Bloque 1 — KPIs globales
        import datetime

        mat_cols = {
            str(AÑO_INICIO_HISTORICO): f"matricula_{AÑO_INICIO_HISTORICO}",
            str(AÑO_FIN_DATOS): f"matricula_{AÑO_FIN_DATOS}",
        }
        def _sum_col(df: pd.DataFrame, col: str) -> float:
            if df is None or col not in df.columns:
                return 0.0
            return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())

        mat19 = _sum_col(sabana, mat_cols[str(AÑO_INICIO_HISTORICO)])
        mat24 = _sum_col(sabana, mat_cols[str(AÑO_FIN_DATOS)])
        crecimiento_global_pct = ((mat24 - mat19) / mat19) if mat19 else 0.0

        total_programas = int(len(sabana)) if sabana is not None else 0
        total_categorias = int(len(ag)) if ag is not None else 0

        es_activo_sum = int(
            pd.to_numeric(sabana.get("es_activo", False), errors="coerce").fillna(0).astype(int).sum()
        ) if isinstance(sabana, pd.DataFrame) else 0
        # tiene_matricula_2024 suele existir en ag; en caso contrario usamos matricula_2024>0
        if isinstance(sabana, pd.DataFrame) and f"tiene_matricula_{AÑO_FIN_DATOS}" in sabana.columns:
            tiene_matricula_2024_sum = int(sabana[f"tiene_matricula_{AÑO_FIN_DATOS}"].fillna(False).astype(bool).sum())
        else:
            tiene_matricula_2024_sum = int(pd.to_numeric(sabana.get(f"matricula_{AÑO_FIN_DATOS}", 0), errors="coerce").fillna(0).gt(0).sum())

        calif = pd.to_numeric(ag.get("calificacion_final", np.nan), errors="coerce") if isinstance(ag, pd.DataFrame) else pd.Series(dtype=float)
        categorias_verdes   = _sem_verde    if _sem_verde    is not None else int((calif >= 4.0).sum())
        categorias_amarillo = _sem_amarillo if _sem_amarillo is not None else int(((calif >= 3.0) & (calif < 4.0)).sum())
        categorias_rojas    = _sem_rojo     if _sem_rojo     is not None else int((calif < 3.0).sum())
        calif_promedio = float(calif.mean()) if len(calif) else 0.0

        fuente = sabana.get("FUENTE_CATEGORIA", pd.Series([], dtype=object))
        cruce_snies = int(fuente.astype(str).str.upper().str.strip().eq("CRUCE_SNIES").sum())
        match_nombre = int(fuente.astype(str).str.upper().str.strip().eq("MATCH_NOMBRE").sum())
        knn_tfidf = int(fuente.astype(str).str.upper().str.strip().eq("KNN_TFIDF").sum())

        total_confianza_100 = (
            ((cruce_snies + match_nombre) / total_programas * 100.0) if total_programas else 0.0
        )

        req_revision = 0
        if "REQUIERE_REVISION" in sabana.columns:
            req_revision = int(sabana["REQUIERE_REVISION"].fillna(False).astype(bool).sum())

        ws["A1"].value = f"ESTUDIO DE MERCADO — {titulo.upper()}"
        ws["A1"].font = title_font

        generado_dt = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A2"].value = f"Generado el {generado_dt}"
        ws["A2"].font = subtitle_font

        kpis = [
            ("Total registros SNIES analizados", total_programas, "#,##0", False),
            ("Total categorías", total_categorias, "#,##0", False),
            (f"Matrícula total {AÑO_FIN_DATOS}", mat24, "#,##0", False),
            (f"Matrícula total {AÑO_INICIO_HISTORICO}", mat19, "#,##0", False),
            (f"Crecimiento global {AÑO_INICIO_HISTORICO}→{AÑO_FIN_DATOS}", crecimiento_global_pct, "0.0%", True),
            ("Registros con matrícula activa", es_activo_sum, "#,##0", False),
            (f"Programas únicos con dato {AÑO_FIN_DATOS}", tiene_matricula_2024_sum, "#,##0", False),
            ("Categorías VERDES (calif. ≥ 4.0)", categorias_verdes, "#,##0", False),
            ("Categorías AMARILLO (3.0-3.9)", categorias_amarillo, "#,##0", False),
            ("Categorías ROJAS (calif. < 3.0)", categorias_rojas, "#,##0", False),
            ("Calificación promedio", calif_promedio, "0.00", False),
            ("Certeza clasificación 100%", total_confianza_100 / 100.0, "0.0%", True),
            ("Requieren revisión manual", req_revision, "#,##0", False),
        ]

        start_row_kpi = 4
        for i, (label, value, fmt, _is_percent) in enumerate(kpis):
            r = start_row_kpi + i
            set_cell(r, 1, label, font=Font(bold=True), align=XLAlignment(horizontal="left"), border_on=False)
            set_cell(
                r,
                2,
                _valor_excel(value, fmt),
                number_format=fmt,
                align=XLAlignment(horizontal="right"),
                border_on=False,
            )

            # Semáforo en la celda de conteo (opcional, suave) — solo fill, preserva number_format
            cell_val = ws.cell(row=r, column=2)
            if "VERDES" in label:
                cell_val.fill = PatternFill(start_color=VERDE_FILL, end_color=VERDE_FILL, fill_type="solid")
            elif "AMARILLO" in label:
                cell_val.fill = PatternFill(start_color=AMARILLO_FILL, end_color=AMARILLO_FILL, fill_type="solid")
            elif "ROJAS" in label:
                cell_val.fill = PatternFill(start_color=ROJO_FILL, end_color=ROJO_FILL, fill_type="solid")

        # Bloque 2 — Rankings por dimensión
        start_row = 17

        def write_top_table(
            df: pd.DataFrame,
            title: str,
            start_r: int,
            start_c: int,
            cols: list[str],
            data_cols: list[str],
            n: int,
            sort_by: str,
            ascending: bool = False,
        ) -> None:
            # title row (merged)
            merge_and_set_title(start_r, start_c, start_c + len(cols) - 1, title, fill=header_fill, font=header_font)

            # column headers
            header_r = start_r + 1
            for j, h in enumerate(cols):
                set_cell(
                    header_r,
                    start_c + j,
                    h,
                    fill=header_fill_alt,
                    font=Font(bold=True, color="000000"),
                    align=XLAlignment(horizontal="left", vertical="center"),
                    border_on=True,
                )

            top = df.sort_values(sort_by, ascending=ascending).head(n) if sort_by in df.columns else df.head(n)
            data_start_r = header_r + 1
            _PCT_COLS = {"AAGR_suma", "AAGR_ROBUSTO", "AAGR_primer_curso"}
            for k, (_, row) in enumerate(top.iterrows()):
                r = data_start_r + k
                for j, dc in enumerate(data_cols):
                    v = row.get(dc, "")
                    number_fmt = None
                    if dc in {
                        f"suma_matricula_{AÑO_FIN_DATOS}",
                        f"suma_primer_curso_{AÑO_FIN_DATOS}",
                    }:
                        number_fmt = "#,##0"
                    elif dc in _PCT_COLS:
                        number_fmt = "0.0%"
                    elif dc in {"calificacion_final"}:
                        number_fmt = "0.00"
                    elif dc in {"salario_promedio"}:
                        number_fmt = "0.00"
                    elif dc in {"salario_proyectado_pesos_hoy"}:
                        number_fmt = "#,##0"
                    set_cell(
                        r,
                        start_c + j,
                        _valor_excel(v, number_fmt),
                        fill=data_fill if k % 2 == 1 else None,
                        align=XLAlignment(horizontal="left", vertical="center"),
                        number_format=number_fmt,
                        border_on=False,
                    )

        # Table A: Top 5 mayor primer curso (último año) (A-D)
        _col_pc_top = (
            f"suma_primer_curso_{AÑO_FIN_DATOS}"
            if f"suma_primer_curso_{AÑO_FIN_DATOS}" in ag.columns
            else f"suma_matricula_{AÑO_FIN_DATOS}"
        )
        write_top_table(
            ag,
            f"🎓 TOP 5 — MAYOR PRIMER CURSO {AÑO_FIN_DATOS}",
            start_row,
            1,
            ["Categoría", f"Primer curso {AÑO_FIN_DATOS}", "Calificación", "Crecimiento AAGR"],
            ["CATEGORIA_FINAL", _col_pc_top, "calificacion_final", "AAGR_ROBUSTO"],
            5,
            _col_pc_top,
            ascending=False,
        )

        # Table B: Top 5 mayor crecimiento AAGR (F-I)
        write_top_table(
            ag,
            "📈 TOP 5 — MAYOR CRECIMIENTO",
            start_row,
            6,
            ["Categoría", "AAGR", f"Matrícula {AÑO_FIN_DATOS}", "Calificación"],
            ["CATEGORIA_FINAL", "AAGR_ROBUSTO", f"suma_matricula_{AÑO_FIN_DATOS}", "calificacion_final"],
            5,
            "AAGR_ROBUSTO",
            ascending=False,
        )

        # Table C: Top 5 mejor salario (A-D, desplazada 9 filas)
        start_row_c = start_row + 9
        write_top_table(
            ag,
            "💰 TOP 5 — MEJOR SALARIO (SMLMV)",
            start_row_c,
            1,
            ["Categoría", "Salario SMLMV", "Salario pesos hoy", "Calificación"],
            ["CATEGORIA_FINAL", "salario_promedio", "salario_proyectado_pesos_hoy", "calificacion_final"],
            5,
            "salario_promedio",
            ascending=False,
        )

        # Table D: Top 5 peor crecimiento (F-I, misma fila que C)
        write_top_table(
            ag,
            "📉 TOP 5 — MENOR CRECIMIENTO",
            start_row_c,
            6,
            ["Categoría", "AAGR", f"Matrícula {AÑO_FIN_DATOS}", "Calificación"],
            ["CATEGORIA_FINAL", "AAGR_ROBUSTO", f"suma_matricula_{AÑO_FIN_DATOS}", "calificacion_final"],
            5,
            "AAGR_ROBUSTO",
            ascending=True,
        )

        # Bloque 3 — Calidad de clasificación (al final)
        row_q = start_row_c + 9
        merge_and_set_title(row_q, 1, 4, "✅ CALIDAD DE CLASIFICACIÓN", fill=header_fill, font=header_font)

        header_r = row_q + 1
        q_cols = ["Fuente", "Programas", "% del total", "Confianza"]
        for j, h in enumerate(q_cols):
            set_cell(
                header_r,
                1 + j,
                h,
                fill=header_fill_alt,
                font=Font(bold=True, color="000000"),
                align=XLAlignment(horizontal="left", vertical="center"),
                border_on=True,
            )

        q_rows = [
            ("CRUCE_SNIES", cruce_snies, "100% — cruce exacto por código SNIES"),
            ("MATCH_NOMBRE", match_nombre, "100% — match exacto por nombre"),
            ("KNN_TFIDF", knn_tfidf, "Variable (mediana 79.3%)"),
            ("Requieren revisión", req_revision, "—"),
        ]

        for i, (source, count, confianza) in enumerate(q_rows):
            r = header_r + 1 + i
            pct = (count / total_programas) if total_programas else 0.0
            set_cell(r, 1, source, fill=data_fill if i % 2 else None)
            set_cell(r, 2, int(count), number_format="#,##0", fill=data_fill if i % 2 else None)
            set_cell(
                r,
                3,
                _valor_excel(pct, "0.0%"),
                number_format="0.0%",
                fill=data_fill if i % 2 else None,
            )
            set_cell(r, 4, confianza, fill=data_fill if i % 2 else None)

    except Exception as e:
        log_warning(f"[Fase 5] No se pudo generar hoja resumen_ejecutivo: {e}. Se continúa sin esa hoja.")


# Agrupación de departamentos en regiones geográficas de Colombia
_REGION_MAP: dict[str, str] = {
    "BOGOTÁ D.C.":                    "Bogotá",
    "ANTIOQUIA":                       "Eje Andino",
    "CALDAS":                          "Eje Andino",
    "RISARALDA":                       "Eje Andino",
    "QUINDÍO":                         "Eje Andino",
    "VALLE DEL CAUCA":                 "Pacífico",
    "CAUCA":                           "Pacífico",
    "NARIÑO":                          "Pacífico",
    "CHOCÓ":                           "Pacífico",
    "SANTANDER":                       "Nororiente",
    "NORTE DE SANTANDER":              "Nororiente",
    "BOYACÁ":                          "Nororiente",
    "CUNDINAMARCA":                    "Nororiente",
    "ATLÁNTICO":                       "Caribe",
    "BOLÍVAR":                         "Caribe",
    "MAGDALENA":                       "Caribe",
    "CESAR":                           "Caribe",
    "CÓRDOBA":                         "Caribe",
    "SUCRE":                           "Caribe",
    "LA GUAJIRA":                      "Caribe",
    "ARCHIPIÉLAGO DE SAN ANDRÉS, PROVIDENCIA Y SANTA CATALINA": "Caribe",
    "TOLIMA":                          "Centro-Sur",
    "HUILA":                           "Centro-Sur",
    "META":                            "Centro-Sur",
    "CASANARE":                        "Llanos",
    "ARAUCA":                          "Llanos",
    "VICHADA":                         "Llanos",
    "AMAZONAS":                        "Amazonía",
    "CAQUETÁ":                         "Amazonía",
    "PUTUMAYO":                        "Amazonía",
    "GUAINÍA":                         "Amazonía",
    "VAUPÉS":                          "Amazonía",
}

# Orden en que aparecen las regiones en el Excel (de mayor a menor mercado)
_REGION_ORDEN: list[str] = [
    "Bogotá", "Eje Andino", "Pacífico", "Nororiente",
    "Caribe", "Centro-Sur", "Llanos", "Amazonía",
]


def run_analisis_regional(
    sabana: pd.DataFrame,
    ag_nacional: pd.DataFrame,
) -> pd.DataFrame | None:
    """
    Análisis regional integrado (Opción B).

    Para cada par (CATEGORIA_FINAL, DEPARTAMENTO_OFERTA_PROGRAMA) calcula las
    mismas métricas de la Fase 4 restringidas a ese departamento y añade las
    métricas nacionales como referencia para comparación directa.

    Celdas con menos de UMBRAL_REGIONAL_PRIMER_CURSO nuevos matriculados (primer_curso)
    en 2024 se incluyen pero con DATOS_INSUFICIENTES=True y sin métricas de crecimiento.
    Retorna None si la columna de departamento no existe en la sábana.
    """
    COL_DEPT = "DEPARTAMENTO_OFERTA_PROGRAMA"
    if COL_DEPT not in sabana.columns:
        log_warning("[Regional] Sin columna DEPARTAMENTO_OFERTA_PROGRAMA — hoja regional omitida.")
        return None

    col_nivel = "NIVEL_DE_FORMACIÓN"
    # Hojas regionales solo para posgrado (los segmentos no exportan total_pregrado).
    if col_nivel in sabana.columns and NIVELES_POSGRADO:
        sabana = sabana[sabana[col_nivel].isin(NIVELES_POSGRADO)].copy()

    years = list(range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1))
    registros: list[dict] = []

    for (cat, dept), grupo in sabana.groupby(["CATEGORIA_FINAL", COL_DEPT]):
        _col_pc_fin = f"primer_curso_{AÑO_FIN_DATOS}"
        pc_2024_regional = (
            pd.to_numeric(grupo.get(_col_pc_fin, pd.Series(dtype=float)), errors="coerce")
            .fillna(0)
            .sum()
        )

        fila: dict = {
            "REGION": _REGION_MAP.get(dept, "Otra"),
            "DEPARTAMENTO": dept,
            "CATEGORIA_FINAL": cat,
            f"num_programas_regional_{AÑO_FIN_DATOS}": int(
                (pd.to_numeric(grupo.get(_col_pc_fin, pd.Series(dtype=float)), errors="coerce")
                 .fillna(0) > 0).sum()
            ),
            f"suma_primer_curso_regional_{AÑO_FIN_DATOS}": pc_2024_regional,
            "DATOS_INSUFICIENTES": pc_2024_regional < UMBRAL_REGIONAL_PRIMER_CURSO,
        }

        if pc_2024_regional < UMBRAL_REGIONAL_PRIMER_CURSO:
            registros.append(fila)
            continue

        # Primer curso anual (fuente correcta para AAGR regional)
        sumas_pc: dict[int, float] = {}
        for y in years:
            col_pc = f"primer_curso_{y}"
            if col_pc in grupo.columns:
                sumas_pc[y] = pd.to_numeric(grupo[col_pc], errors="coerce").fillna(0).sum()
                fila[f"suma_primer_curso_regional_{y}"] = sumas_pc[y]
            else:
                sumas_pc[y] = 0.0

        # Matrícula total solo como referencia (para umbral y contexto)
        sumas_mat: dict[int, float] = {}
        for y in years:
            col_m = f"matricula_{y}"
            if col_m in grupo.columns:
                sumas_mat[y] = pd.to_numeric(grupo[col_m], errors="coerce").fillna(0).sum()
                fila[f"suma_matricula_regional_{y}"] = sumas_mat[y]

        # Variaciones y AAGR regional — sobre primer_curso
        vars_pc_reg: list[float] = []
        for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
            if y in sumas_pc and (y - 1) in sumas_pc and sumas_pc[y - 1] > 0:
                v = (sumas_pc[y] - sumas_pc[y - 1]) / sumas_pc[y - 1]
                fila[f"var_primer_curso_regional_{y}"] = v
                vars_pc_reg.append(v)
            else:
                fila[f"var_primer_curso_regional_{y}"] = np.nan

        fila["AAGR_regional"] = float(np.mean(vars_pc_reg)) if vars_pc_reg else np.nan

        # CAGR regional — sobre primer_curso
        pc19_reg = sumas_pc.get(AÑO_INICIO_HISTORICO, 0)
        pc24_reg = sumas_pc.get(AÑO_FIN_DATOS, 0)
        _n_years_cagr = max(AÑO_FIN_DATOS - AÑO_INICIO_HISTORICO, 1)
        fila["CAGR_regional"] = (
            float((pc24_reg / pc19_reg) ** (1 / _n_years_cagr) - 1)
            if pc19_reg > 0 and pc24_reg > 0 else np.nan
        )

        # Participación regional sobre el total nacional de primer_curso
        total_nacional_2024 = (
            ag_nacional[f"suma_primer_curso_{AÑO_FIN_DATOS}"].sum()
            if f"suma_primer_curso_{AÑO_FIN_DATOS}" in ag_nacional.columns else 0
        )
        fila[f"participacion_regional_{AÑO_FIN_DATOS}"] = (
            (pc24_reg / total_nacional_2024) if total_nacional_2024 > 0 else np.nan
        )

        # Referencia nacional para comparación directa
        fila_nac = ag_nacional[ag_nacional["CATEGORIA_FINAL"] == cat]
        if len(fila_nac) == 1:
            nac = fila_nac.iloc[0]
            fila[f"suma_primer_curso_nacional_{AÑO_FIN_DATOS}"] = nac.get(
                f"suma_primer_curso_{AÑO_FIN_DATOS}", np.nan
            )
            fila["AAGR_nacional"] = nac.get("AAGR_suma", np.nan)
            fila["calificacion_nacional"] = nac.get("calificacion_final", np.nan)
            pc_nac = nac.get(f"suma_primer_curso_{AÑO_FIN_DATOS}", 0) or 0
            fila["pct_mercado_regional"] = (pc24_reg / pc_nac) if pc_nac > 0 else np.nan
        else:
            fila[f"suma_primer_curso_nacional_{AÑO_FIN_DATOS}"] = np.nan
            fila["AAGR_nacional"] = np.nan
            fila["calificacion_nacional"] = np.nan
            fila["pct_mercado_regional"] = np.nan

        # Salario y costo promedios del grupo regional
        if "SALARIO_OLE" in grupo.columns:
            sal = pd.to_numeric(grupo["SALARIO_OLE"], errors="coerce")
            fila["salario_promedio_regional"] = float(sal.mean()) if sal.notna().any() else np.nan
        if "COSTO_MATRÍCULA_ESTUD_NUEVOS" in grupo.columns:
            costo = pd.to_numeric(grupo["COSTO_MATRÍCULA_ESTUD_NUEVOS"], errors="coerce")
            fila["costo_promedio_regional"] = float(costo.mean()) if costo.notna().any() else np.nan

        registros.append(fila)

    if not registros:
        log_warning("[Regional] No se generaron registros para el análisis regional.")
        return None

    df_regional = pd.DataFrame(registros)

    # Ordenar: región → departamento → calificación nacional desc
    region_order = {r: i for i, r in enumerate(_REGION_ORDEN)}
    df_regional["_region_ord"] = df_regional["REGION"].map(region_order).fillna(99)
    df_regional = df_regional.sort_values(
        ["_region_ord", "DATOS_INSUFICIENTES", "DEPARTAMENTO", "calificacion_nacional"],
        ascending=[True, True, True, False],
    ).drop(columns=["_region_ord"]).reset_index(drop=True)

    log_info(
        f"[Regional] {len(df_regional):,} celdas generadas "
        f"({df_regional['DATOS_INSUFICIENTES'].sum():,} con datos insuficientes, "
        f"umbral primer_curso={UMBRAL_REGIONAL_PRIMER_CURSO})."
    )
    return df_regional


def _exportar_estudio_segmento(
    etiqueta: str,
    sabana_seg: pd.DataFrame,
    ag_seg: pd.DataFrame,
    ag_nacional: pd.DataFrame,
    ruta: Path,
) -> None:
    """
    Exporta un Excel de estudio de mercado para un segmento geográfico o modal.
    Misma estructura que run_fase5 pero:
      - Sin merge_incremental (siempre fresco).
      - Sin análisis regional (aplica solo al nacional).
      - Añade hoja 'contexto_nacional' con comparación segmento vs. país.
    """
    ruta.parent.mkdir(parents=True, exist_ok=True)

    NAC_COLS = {
        "calificacion_final": "calificacion_nacional",
        "AAGR_ROBUSTO": "AAGR_ROBUSTO_nacional",
        f"suma_primer_curso_{AÑO_FIN_DATOS}": f"suma_primer_curso_{AÑO_FIN_DATOS}",
    }
    keys = [k for k in NAC_COLS if k in ag_nacional.columns]
    if not keys:
        log_warning(
            f"[Segmento {etiqueta}] ag_nacional sin columnas de comparación — "
            "contexto_nacional quedará limitado."
        )
    nac_cols_use = ["CATEGORIA_FINAL"] + keys
    nac_merge = ag_nacional[nac_cols_use].rename(columns={k: NAC_COLS[k] for k in keys})

    _col_pc_seg = f"suma_primer_curso_{AÑO_FIN_DATOS}"
    _col_pc_reg = f"suma_primer_curso_regional_{AÑO_FIN_DATOS}"
    ag_seg_ctx = ag_seg.copy()
    if _col_pc_seg in ag_seg_ctx.columns:
        ag_seg_ctx = ag_seg_ctx.rename(columns={_col_pc_seg: _col_pc_reg})
    ag_con_nac = ag_seg_ctx.merge(nac_merge, on="CATEGORIA_FINAL", how="left")

    import shutil
    from datetime import date, datetime

    if ruta.exists():
        snap_dir = HISTORICO_ESTUDIO_MERCADO_DIR / "snapshots"
        snap_dir.mkdir(parents=True, exist_ok=True)
        fecha = date.today().isoformat()
        nombre_base = ruta.stem
        dest = snap_dir / f"{nombre_base}_{fecha}.xlsx"
        if dest.exists():
            hora = datetime.now().strftime("%H%M")
            dest = snap_dir / f"{nombre_base}_{fecha}_{hora}.xlsx"
        shutil.copy2(ruta, dest)
        log_info(f"[Segmento] Snapshot guardado: {dest.name}")

    while True:
        try:
            sabana_seg = sabana_seg.copy()
            ag_seg = ag_seg.copy()

            _sf_col_seg = pd.to_numeric(
                ag_seg["calificacion_final"]
                if "calificacion_final" in ag_seg.columns
                else pd.Series(dtype=float),
                errors="coerce",
            ).fillna(0.0)
            _sem_verde_seg = int((_sf_col_seg >= 4.0).sum())
            _sem_amarillo_seg = int(((_sf_col_seg >= 3.0) & (_sf_col_seg < 4.0)).sum())
            _sem_rojo_seg = int((_sf_col_seg < 3.0).sum())

            with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
                # Filtrar por niveles de POSGRADO antes de calcular KPIs del resumen,
                # para que matrículas y certeza sean consistentes con programas_detalle
                # del segmento (los Excels regionales solo cubren ESP+MAE).
                col_nivel = "NIVEL_DE_FORMACIÓN"
                sabana_export = sabana_seg.copy()
                if col_nivel in sabana_export.columns and NIVELES_POSGRADO:
                    sabana_export = sabana_export[sabana_export[col_nivel].isin(NIVELES_POSGRADO)]

                try:
                    _escribir_resumen_ejecutivo(
                        writer, sabana_export, ag_seg,
                        _sem_verde_seg, _sem_amarillo_seg, _sem_rojo_seg,
                        titulo=etiqueta,
                    )
                except Exception as e:
                    log_warning(f"[Segmento {etiqueta}] Resumen ejecutivo: {e}")

                # Eliminar columnas de tracking del pipeline global que no aplican
                # a los segmentos regionales para mantener consistencia de estructura.
                COLS_TRACKING_GLOBAL = [
                    "ACTIVO_PIPELINE",
                    "FECHA_PRIMERA_VEZ",
                    "FECHA_ULTIMO_ACTIVO",
                    "nuevo_vs_snapshot_anterior",
                ]
                cols_a_drop = [c for c in COLS_TRACKING_GLOBAL if c in sabana_export.columns]
                if cols_a_drop:
                    sabana_export = sabana_export.drop(columns=cols_a_drop)

                sabana_export.to_excel(writer, sheet_name="programas_detalle", index=False)

                col_order = _escribir_hoja_total(writer, ag_seg)

                ctx_cols = [
                    "CATEGORIA_FINAL",
                    "calificacion_final",
                    "calificacion_nacional",
                    "AAGR_ROBUSTO",
                    "AAGR_ROBUSTO_nacional",
                    f"suma_primer_curso_regional_{AÑO_FIN_DATOS}",
                    f"suma_primer_curso_{AÑO_FIN_DATOS}",
                    f"participacion_{AÑO_FIN_DATOS}",
                    f"num_programas_{AÑO_FIN_DATOS}",
                ]
                ctx_export = ag_con_nac[[c for c in ctx_cols if c in ag_con_nac.columns]].copy()
                _ctx_rename = {
                    k: v for k, v in HEADERS_CONTEXTO_NACIONAL.items() if k in ctx_export.columns
                }
                ctx_export = ctx_export.rename(columns=_ctx_rename)
                _col_sort = HEADERS_CONTEXTO_NACIONAL.get("calificacion_final")
                if _col_sort and _col_sort in ctx_export.columns:
                    ctx_export = ctx_export.sort_values(_col_sort, ascending=False)
                elif "CATEGORIA_FINAL" in ctx_export.columns:
                    ctx_export = ctx_export.sort_values("CATEGORIA_FINAL")
                ctx_export.to_excel(writer, sheet_name="contexto_nacional", index=False)

                wb = writer.book
                ws_det = wb["programas_detalle"]
                ws_det.freeze_panes = "A2"
                ws_det.auto_filter.ref = ws_det.dimensions
                from openpyxl.utils import get_column_letter

                for col_name, width in COL_ANCHOS_PROGRAMAS.items():
                    if col_name in sabana_export.columns:
                        idx = list(sabana_export.columns).index(col_name) + 1
                        ws_det.column_dimensions[get_column_letter(idx)].width = width
                ws_total = wb["total"]
                if col_order:
                    _aplicar_formato_total(ws_total, col_order)

            log_info(f"[Segmento {etiqueta}] Excel exportado: {ruta}")
            break

        except PermissionError:
            log_warning(
                f"[Segmento {etiqueta}] No se pudo escribir '{ruta.name}' — "
                "el archivo está abierto en Excel. Ciérrelo e intente de nuevo desde la UI."
            )
            break
        except Exception as e:
            log_error(f"[Segmento {etiqueta}] Error exportando: {e}")
            break


def run_segmentos_regionales(
    sabana: pd.DataFrame,
    ag_nacional: pd.DataFrame,
    cancel_event: threading.Event | None = None,
    force_recalc: bool = False,
) -> dict[str, pd.DataFrame]:
    """
    Genera un Excel independiente para cada segmento geográfico/modal.

    Cada Excel recalcula Fase 4 completa (scoring, AAGR, participación, semáforo)
    solo con los programas del segmento, sin tocar los datos nacionales.

    Segmentos:
      - Bogotá      → DEPARTAMENTO_OFERTA_PROGRAMA == "BOGOTÁ D.C."
      - Antioquia   → DEPARTAMENTO_OFERTA_PROGRAMA == "ANTIOQUIA"
      - Eje Cafetero → DEPARTAMENTO_OFERTA_PROGRAMA in {CALDAS, RISARALDA, QUINDÍO}
      - Virtual     → MODALIDAD normalizada == "VIRTUAL"

    Retorna dict {nombre_segmento: DataFrame_agregado} con los resultados.
    """
    from etl.config import OUTPUTS_DIR, TEMP_DIR, ESTUDIO_MERCADO_DIR

    COL_DEPT = "DEPARTAMENTO_OFERTA_PROGRAMA"
    COL_MOD = "MODALIDAD"

    SEGMENTOS: list[dict] = [
        {
            "nombre": "Bogota",
            "etiqueta": "Bogotá D.C.",
            "filtro": (
                lambda df: df[df[COL_DEPT] == "BOGOTÁ D.C."].copy()
                if COL_DEPT in df.columns
                else df.iloc[0:0]
            ),
        },
        {
            "nombre": "Antioquia",
            "etiqueta": "Antioquia",
            "filtro": (
                lambda df: df[df[COL_DEPT] == "ANTIOQUIA"].copy()
                if COL_DEPT in df.columns
                else df.iloc[0:0]
            ),
        },
        {
            "nombre": "Eje_Cafetero",
            "etiqueta": "Eje Cafetero",
            "filtro": (
                lambda df: df[
                    df[COL_DEPT].isin(["CALDAS", "RISARALDA", "QUINDÍO"])
                ].copy()
                if COL_DEPT in df.columns
                else df.iloc[0:0]
            ),
        },
        {
            "nombre": "Virtual",
            "etiqueta": "Virtual (Colombia)",
            "filtro": (
                lambda df: df[
                    df[COL_MOD].astype(str).str.upper().str.strip() == "VIRTUAL"
                ].copy()
                if COL_MOD in df.columns
                else df.iloc[0:0]
            ),
        },
    ]

    resultados: dict[str, pd.DataFrame] = {}

    log_etapa_iniciada("Segmentos regionales/modales")

    for _seg_name in ["Antioquia", "Bogota", "Eje_Cafetero", "Virtual"]:
        _cache_invalida(TEMP_DIR / f"agregado_{_seg_name}.parquet")

    for seg in SEGMENTOS:
        if cancel_event is not None and cancel_event.is_set():
            log_warning("[Segmentos] Cancelado por el usuario antes de continuar.")
            break

        nombre = seg["nombre"]
        etiqueta = seg["etiqueta"]

        try:
            df_seg = seg["filtro"](sabana)

            if len(df_seg) == 0:
                log_warning(f"[Segmento {nombre}] Sin programas tras filtro — omitido.")
                continue

            log_info(
                f"[Segmento {nombre}] {len(df_seg):,} programas "
                f"→ recalculando Fase 4..."
            )

            cache_path = TEMP_DIR / f"agregado_{nombre}.parquet"
            sabana_path_check = TEMP_DIR / "sabana_consolidada.parquet"

            usar_cache = False
            ag_seg: pd.DataFrame | None = None
            if (
                not force_recalc
                and cache_path.exists()
                and sabana_path_check.exists()
                and cache_path.stat().st_mtime >= sabana_path_check.stat().st_mtime
            ):
                try:
                    ag_seg = pd.read_parquet(cache_path)
                    log_info(f"[Segmento {nombre}] Cargado desde caché (sábana sin cambios).")
                    usar_cache = True
                except Exception:
                    ag_seg = None
                    usar_cache = False

            if not usar_cache:
                try:
                    ag_seg = run_fase4_desde_sabana(
                        df_seg,
                        modo_local=True,
                        niveles=NIVELES_POSGRADO,
                        universo="posgrado",
                    )
                except Exception as e:
                    log_error(f"[Segmento {nombre}] Fase 4 falló: {e}")
                    continue
                try:
                    ag_seg.to_parquet(cache_path, index=False)
                    log_info(f"[Segmento {nombre}] Caché guardado: {cache_path.name}")
                except Exception as e:
                    log_warning(f"[Segmento {nombre}] No se pudo guardar caché: {e}")

            if ag_seg is None or len(ag_seg) == 0:
                log_warning(f"[Segmento {nombre}] Fase 4 sin resultados — omitido.")
                continue

            log_info(
                f"[Segmento {nombre}] {len(ag_seg)} categorías. "
                f"Verde(>=4): {(ag_seg['calificacion_final'] >= 4.0).sum()}, "
                f"Amarillo: {((ag_seg['calificacion_final'] >= 3.0) & (ag_seg['calificacion_final'] < 4.0)).sum()}, "
                f"Rojo(<3): {(ag_seg['calificacion_final'] < 3.0).sum()}"
            )

            ruta = ESTUDIO_MERCADO_DIR / f"Estudio_Mercado_{nombre}.xlsx"
            _exportar_estudio_segmento(etiqueta, df_seg, ag_seg, ag_nacional, ruta)

            resultados[nombre] = ag_seg

        except Exception as e:
            log_error(f"[Segmento {nombre}] Error inesperado: {e}")
            continue

    log_etapa_completada(
        "Segmentos regionales/modales",
        f"{len(resultados)}/{len(SEGMENTOS)} segmentos exportados",
    )
    return resultados


def _escribir_hoja_delta(
    writer: pd.ExcelWriter,
    ag_nuevo: pd.DataFrame,
) -> None:
    """
    Genera hoja 'cambios_vs_anterior' comparando la calificación_final actual
    contra la última ejecución guardada en agregado_categorias_anterior.parquet
    bajo TEMP_DIR. Si no existe anterior, crea el snapshot y omite la hoja.
    """
    from etl.config import TEMP_DIR
    from openpyxl.styles import Alignment, Font, PatternFill

    snapshot_path = TEMP_DIR / "agregado_categorias_anterior.parquet"
    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    cols_snap = [
        "CATEGORIA_FINAL",
        "calificacion_final",
        f"suma_matricula_{AÑO_FIN_DATOS}",
        "AAGR_ROBUSTO",
        f"num_programas_{AÑO_FIN_DATOS}",
    ]
    cols_pres = [c for c in cols_snap if c in ag_nuevo.columns]
    if "CATEGORIA_FINAL" not in cols_pres:
        log_warning("[Delta] ag_nuevo sin CATEGORIA_FINAL — omitiendo hoja de cambios.")
        return

    if not snapshot_path.exists():
        try:
            ag_nuevo[cols_pres].to_parquet(snapshot_path, index=False)
            log_info("[Delta] Snapshot inicial guardado. La hoja de cambios aparecerá en la próxima ejecución.")
        except Exception as e:
            log_warning(f"[Delta] No se pudo guardar snapshot: {e}")
        return

    try:
        ag_ant = pd.read_parquet(snapshot_path)
    except Exception as e:
        log_warning(f"[Delta] No se pudo leer snapshot anterior: {e}. Omitiendo hoja de cambios.")
        return

    try:
        df_nuevo = ag_nuevo[[c for c in cols_snap if c in ag_nuevo.columns]].copy()
        ant_cols = [c for c in cols_snap if c in ag_ant.columns]
        if "CATEGORIA_FINAL" not in ant_cols:
            log_warning("[Delta] Snapshot sin CATEGORIA_FINAL — omitiendo hoja.")
            return
        df_ant = ag_ant[ant_cols].copy()
        df_ant = df_ant.rename(
            columns={
                "calificacion_final": "calif_anterior",
                f"suma_matricula_{AÑO_FIN_DATOS}": "matricula_anterior",
                "AAGR_ROBUSTO": "aagr_anterior",
                f"num_programas_{AÑO_FIN_DATOS}": "programas_anterior",
            }
        )

        merged = df_nuevo.merge(df_ant, on="CATEGORIA_FINAL", how="outer", indicator=True)

        def _semaforo(v):
            if pd.isna(v):
                return "SIN DATO"
            return "VERDE" if v >= 4.0 else ("AMARILLO" if v >= 3.0 else "ROJO")

        merged["semaforo_nuevo"] = merged["calificacion_final"].apply(_semaforo)
        merged["semaforo_anterior"] = merged["calif_anterior"].apply(_semaforo)
        merged["delta_calif"] = merged["calificacion_final"] - merged["calif_anterior"]
        merged["delta_matricula"] = merged[f"suma_matricula_{AÑO_FIN_DATOS}"] - merged["matricula_anterior"]
        merged["cambio_semaforo"] = merged["semaforo_nuevo"] != merged["semaforo_anterior"]

        def _tipo(row):
            if row["_merge"] == "left_only":
                return "CATEGORÍA NUEVA"
            if row["_merge"] == "right_only":
                return "CATEGORÍA ELIMINADA"
            if row["cambio_semaforo"]:
                if row["semaforo_anterior"] in ("ROJO", "AMARILLO") and row["semaforo_nuevo"] == "VERDE":
                    return "SUBIÓ A VERDE ▲"
                if row["semaforo_anterior"] == "VERDE":
                    return "BAJÓ DE VERDE ▼"
                if row["semaforo_anterior"] == "ROJO" and row["semaforo_nuevo"] == "AMARILLO":
                    return "MEJORÓ ▲"
                if row["semaforo_anterior"] == "AMARILLO" and row["semaforo_nuevo"] == "ROJO":
                    return "EMPEORÓ ▼"
            dc = row.get("delta_calif")
            if pd.notna(dc) and abs(float(dc)) >= 0.3:
                return "CAMBIO SIGNIFICATIVO"
            return "SIN CAMBIO RELEVANTE"

        merged["tipo_cambio"] = merged.apply(_tipo, axis=1)

        orden = {
            "SUBIÓ A VERDE ▲": 0,
            "BAJÓ DE VERDE ▼": 1,
            "CATEGORÍA NUEVA": 2,
            "EMPEORÓ ▼": 3,
            "MEJORÓ ▲": 4,
            "CAMBIO SIGNIFICATIVO": 5,
            "CATEGORÍA ELIMINADA": 6,
            "SIN CAMBIO RELEVANTE": 7,
        }
        merged["_ord"] = merged["tipo_cambio"].map(orden).fillna(99)
        merged = merged.sort_values(["_ord", "delta_calif"], ascending=[True, True])
        merged_stats = merged.copy()
        merged = merged.drop(columns=["_merge", "_ord", "cambio_semaforo"])

        cols_export = [
            "CATEGORIA_FINAL",
            "tipo_cambio",
            "semaforo_nuevo",
            "calificacion_final",
            "semaforo_anterior",
            "calif_anterior",
            "delta_calif",
            f"suma_matricula_{AÑO_FIN_DATOS}",
            "matricula_anterior",
            "delta_matricula",
            f"num_programas_{AÑO_FIN_DATOS}",
            "programas_anterior",
        ]
        cols_export = [c for c in cols_export if c in merged.columns]
        merged[cols_export].to_excel(writer, sheet_name="cambios_vs_anterior", index=False)

        wb = writer.book
        ws = wb["cambios_vs_anterior"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        AZUL = "000066"
        VERDE_F = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        AMAR_F = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ROJO_F = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        GRIS_F = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        sem_nuevo_idx = cols_export.index("semaforo_nuevo") + 1 if "semaforo_nuevo" in cols_export else None
        sem_ant_idx = cols_export.index("semaforo_anterior") + 1 if "semaforo_anterior" in cols_export else None

        n_cols = len(cols_export)
        for ri in range(2, ws.max_row + 1):
            zebra = GRIS_F if ri % 2 == 0 else None
            for ci in range(1, n_cols + 1):
                c = ws.cell(row=ri, column=ci)
                if zebra:
                    c.fill = zebra
            if sem_nuevo_idx:
                v = ws.cell(row=ri, column=sem_nuevo_idx).value
                c = ws.cell(row=ri, column=sem_nuevo_idx)
                if v == "VERDE":
                    c.fill = VERDE_F
                elif v == "AMARILLO":
                    c.fill = AMAR_F
                elif v == "ROJO":
                    c.fill = ROJO_F
            if sem_ant_idx:
                v = ws.cell(row=ri, column=sem_ant_idx).value
                c = ws.cell(row=ri, column=sem_ant_idx)
                if v == "VERDE":
                    c.fill = VERDE_F
                elif v == "AMARILLO":
                    c.fill = AMAR_F
                elif v == "ROJO":
                    c.fill = ROJO_F

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 26

        n_subio = int((merged_stats["tipo_cambio"] == "SUBIÓ A VERDE ▲").sum())
        n_bajo = int((merged_stats["tipo_cambio"] == "BAJÓ DE VERDE ▼").sum())
        n_nuevas = int((merged_stats["tipo_cambio"] == "CATEGORÍA NUEVA").sum())
        n_elim = int((merged_stats["tipo_cambio"] == "CATEGORÍA ELIMINADA").sum())
        ws.cell(row=1, column=n_cols + 2).value = (
            f"Resumen: +verde={n_subio} | -verde={n_bajo} | nuevas={n_nuevas} | eliminadas={n_elim}"
        )

        ag_nuevo[cols_pres].to_parquet(snapshot_path, index=False)

        log_info(
            f"[Delta] Hoja 'cambios_vs_anterior' generada: "
            f"subió_verde={n_subio}, bajó_verde={n_bajo}, nuevas={n_nuevas}, eliminadas={n_elim}"
        )

    except Exception as e:
        log_warning(f"[Delta] No se pudo generar hoja de cambios: {e}")


def run_fase5(
    agregado_df: pd.DataFrame | None,
    ag_pre: pd.DataFrame | None = None,
) -> None:
    """
    Fase 5: Exportación formateada a Estudio_Mercado_Colombia.xlsx.
    Hoja programas_detalle: sábana Fase 3 (freeze, filtros, anchos), solo posgrado.
    Hoja total: agregado posgrado Fase 4 con dos filas de encabezado y formato por bloque.
    Hoja total_pregrado (opcional): agregado pregrado, mismo formato.
    """
    from etl.config import ARCHIVO_ESTUDIO_MERCADO

    log_etapa_iniciada("Fase 5: Exportación formateada")
    sabana_path = CHECKPOINT_BASE_MAESTRA.parent / "sabana_consolidada.parquet"
    if not sabana_path.exists():
        log_error("No existe sábana consolidada. Ejecutar Fase 3 antes.")
        return
    sabana = pd.read_parquet(sabana_path)
    # ── Filtro de niveles ────────────────────────────────────────────────────
    # programas_detalle solo incluye programas de POSGRADO (ESP+MAE).
    # Los pregrados se reportan exclusivamente en la hoja agregada total_pregrado.
    col_nivel = "NIVEL_DE_FORMACIÓN"
    if col_nivel in sabana.columns and NIVELES_POSGRADO:
        n_antes = len(sabana)
        sabana = sabana[sabana[col_nivel].isin(NIVELES_POSGRADO)].copy()
        log_info(
            f"[Fase 5] programas_detalle: filtrado a solo posgrado "
            f"({len(sabana):,} de {n_antes:,} programas)."
        )
    if agregado_df is None or len(agregado_df) == 0:
        log_error("No hay DataFrame agregado. Ejecutar Fase 4 antes.")
        return

    out_path = ARCHIVO_ESTUDIO_MERCADO
    out_path.parent.mkdir(parents=True, exist_ok=True)

    while True:
        try:
            # Merge incremental: mantiene histórico, respeta manuales y guarda snapshot antes de modificar
            try:
                from etl.merge_incremental import ESTUDIO_PATH, merge_incremental

                merged = merge_incremental(nuevo=sabana, nuevo_total=agregado_df)
                sabana_final = merged.get("programas_detalle") if isinstance(merged, dict) else sabana
                total_final = merged.get("total") if isinstance(merged, dict) else agregado_df
                if sabana_final is None or len(sabana_final) == 0:
                    sabana_final = sabana
                if total_final is None or len(total_final) == 0:
                    total_final = agregado_df
                out_path = ESTUDIO_PATH
            except Exception as e:
                log_warning(f"[Fase 5] Merge incremental falló ({e}). Exportando modo clásico (sobrescritura).")
                sabana_final = sabana
                total_final = agregado_df

            sabana_final = sabana_final.copy()
            total_final = total_final.copy()

            # Semáforos calculados desde total_final ya procesado — antes de entrar al writer
            # para garantizar que el resumen y la hoja total muestren exactamente lo mismo.
            _sf_col = pd.to_numeric(
                total_final["calificacion_final"]
                if "calificacion_final" in total_final.columns
                else pd.Series(dtype=float),
                errors="coerce",
            ).fillna(0.0)
            _sem_verde = int((_sf_col >= 4.0).sum())
            _sem_amarillo = int(((_sf_col >= 3.0) & (_sf_col < 4.0)).sum())
            _sem_rojo = int((_sf_col < 3.0).sum())

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                # Generar hoja de resumen ejecutivo (primera hoja)
                _escribir_resumen_ejecutivo(
                    writer, sabana_final, total_final,
                    _sem_verde, _sem_amarillo, _sem_rojo,
                    titulo="COLOMBIA",
                )
                _escribir_hoja_estandar(writer)
                log_info("✓ Hoja 'estandar_calificacion' creada.")
                sabana_final.to_excel(writer, sheet_name="programas_detalle", index=False)
                col_order = _escribir_hoja_total(writer, total_final)
                try:
                    _wb_sheets = writer.book._sheets
                    _ws_est = writer.book["estandar_calificacion"]
                    _ws_res = writer.book["resumen_ejecutivo"]
                    _idx_res = _wb_sheets.index(_ws_res)
                    _wb_sheets.remove(_ws_est)
                    _wb_sheets.insert(_idx_res + 1, _ws_est)
                except Exception as _e_ord:
                    log_warning(
                        f"[Fase 5] No se pudo reordenar estandar_calificacion: {_e_ord}"
                    )
                wb = writer.book
                ws_detalle = wb["programas_detalle"]
                ws_detalle.freeze_panes = "A2"
                ws_detalle.auto_filter.ref = ws_detalle.dimensions
                from openpyxl.utils import get_column_letter
                for col_name, width in COL_ANCHOS_PROGRAMAS.items():
                    if col_name in sabana_final.columns:
                        idx = list(sabana_final.columns).index(col_name) + 1
                        ws_detalle.column_dimensions[get_column_letter(idx)].width = width
                _aplicar_formato_total(wb["total"], col_order)

                # ── Hoja total_pregrado (144 categorías, solo UNIVERSITARIO) ──
                # Sin merge incremental: no hay baseline anterior para esta hoja
                # (es nueva tras la separación posgrado/pregrado).
                if ag_pre is not None and len(ag_pre) > 0:
                    try:
                        ag_pre_export = ag_pre.copy()
                        col_order_pre = _escribir_hoja_total(
                            writer, ag_pre_export, sheet_name="total_pregrado"
                        )
                        _aplicar_formato_total(wb["total_pregrado"], col_order_pre)
                        # Reposicionar para que aparezca justo después de `total`
                        try:
                            _ws_pre = wb["total_pregrado"]
                            _wb_sheets = wb._sheets  # API interna estable de openpyxl
                            _wb_sheets.remove(_ws_pre)
                            _idx_total = _wb_sheets.index(wb["total"])
                            _wb_sheets.insert(_idx_total + 1, _ws_pre)
                        except Exception as _e_ord:
                            log_warning(
                                f"[Fase 5] No se pudo reordenar total_pregrado: {_e_ord}"
                            )
                        log_info(
                            f"✓ Hoja 'total_pregrado' escrita: "
                            f"{len(ag_pre)} categorías (solo programas UNIVERSITARIO)."
                        )
                    except Exception as _e_pre:
                        log_warning(
                            f"[Fase 5] No se pudo escribir total_pregrado: {_e_pre}"
                        )
                else:
                    log_info(
                        "[Fase 5] ag_pre vacío o no provisto — hoja total_pregrado omitida."
                    )

                try:
                    _escribir_hoja_delta(writer, total_final)
                except Exception as e:
                    log_warning(f"[Delta] Hoja de cambios omitida: {e}")

                # ── Gap de Oportunidad: Océanos Azules (opcional, no bloqueante) ──
                try:
                    df_gap = run_gap_oportunidades(total_final, log_info)
                    if df_gap is not None and len(df_gap) > 0:
                        df_gap.to_excel(
                            writer,
                            sheet_name="oportunidades_expansion",
                            index=False,
                        )
                        _formatear_hoja_gap(writer, df_gap)
                        log_info(f"✓ Hoja 'oportunidades_expansion' añadida: {len(df_gap)} categorías.")
                    else:
                        log_warning("⚠ Gap omitido — sin oportunidades disponibles.")
                except Exception as e:
                    log_warning(f"[Gap] No se pudo generar hoja oportunidades_expansion: {e}")

                # ── Hojas Regional_* desactivadas ─────────────────────────────
                # Los Excel regionales separados (Estudio_Mercado_<Region>.xlsx) ya
                # cubren este análisis; las hojas Regional_* dentro del Excel Colombia
                # eran redundantes y pesadas. `run_analisis_regional()` sigue definida
                # más arriba por si se necesita reactivar este bloque a futuro.

                # Hoja informativa: programas con baja confianza del ML (REQUIERE_REVISION)
                if "REQUIERE_REVISION" in sabana_final.columns:
                    df_revision = sabana_final[sabana_final["REQUIERE_REVISION"]].copy()
                else:
                    df_revision = pd.DataFrame()

                if df_revision is not None and len(df_revision) > 0:
                    cols_revision = [
                        "CÓDIGO_SNIES_DEL_PROGRAMA",
                        "NOMBRE_DEL_PROGRAMA",
                        "NIVEL_DE_FORMACIÓN",
                        "ÁREA_DE_CONOCIMIENTO",
                        "CINE_F_2013_AC_CAMPO_DETALLADO",
                        "CATEGORIA_FINAL",
                        "PROBABILIDAD",
                        "CATEGORIA_ALTERNATIVA",
                        "PROBABILIDAD_ALTERNATIVA",
                        "NOMBRE_INSTITUCIÓN",
                        "DEPARTAMENTO_OFERTA_PROGRAMA",
                        "ESTADO_PROGRAMA",
                    ]
                    cols_existentes = [c for c in cols_revision if c in df_revision.columns]
                    df_revision_export = df_revision[cols_existentes].copy()
                    if "PROBABILIDAD" in df_revision_export.columns:
                        df_revision_export = df_revision_export.sort_values("PROBABILIDAD")

                    df_revision_export.to_excel(writer, sheet_name="revision_requerida", index=False)
                    ws_rev = wb["revision_requerida"]
                    ws_rev.freeze_panes = "A2"
                    ws_rev.auto_filter.ref = ws_rev.dimensions

                    # Resaltar filas con probabilidad < 0.50 en amarillo
                    try:
                        from openpyxl.styles import PatternFill

                        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        if "PROBABILIDAD" in df_revision_export.columns:
                            prob_col_idx = list(df_revision_export.columns).index("PROBABILIDAD") + 1
                            for r in range(2, ws_rev.max_row + 1):
                                v = ws_rev.cell(row=r, column=prob_col_idx).value
                                try:
                                    pv = float(v) if v is not None else 1.0
                                except (TypeError, ValueError):
                                    pv = 1.0
                                if pv < 0.50:
                                    for c in range(1, ws_rev.max_column + 1):
                                        ws_rev.cell(row=r, column=c).fill = yellow

                        for prob_col_name in ["PROBABILIDAD", "PROBABILIDAD_ALTERNATIVA"]:
                            if prob_col_name in df_revision_export.columns:
                                col_idx = list(df_revision_export.columns).index(prob_col_name) + 1
                                for r in range(2, ws_rev.max_row + 1):
                                    ws_rev.cell(row=r, column=col_idx).number_format = "0.0%"
                    except Exception:
                        pass

                    log_info(
                        "Hoja 'revision_requerida' exportada: "
                        f"{len(df_revision_export):,} programas con confianza ML < 0.50."
                    )
                else:
                    log_info("No hay programas que requieran revisión (todos >= 0.50 de confianza).")
            break
        except PermissionError:
            try:
                from tkinter import messagebox
                reintentar = messagebox.askretrycancel(
                    "Archivo en uso",
                    "El archivo Estudio_Mercado_Colombia.xlsx está abierto en otro programa (por ejemplo Excel).\n\n"
                    "Cierre el archivo y presione 'Reintentar' para guardar, o 'Cancelar' para abortar la exportación.",
                    parent=None,
                )
            except Exception:
                reintentar = False
            if not reintentar:
                log_error("Exportación cancelada: el usuario eligió no reintentar (archivo Excel en uso).")
                raise RuntimeError(
                    "Exportación cancelada por el usuario: cierre Estudio_Mercado_Colombia.xlsx e intente de nuevo."
                ) from None

    log_info(f"Exportado: {out_path}")
    log_etapa_completada("Fase 5: Exportación formateada", str(out_path))


def run_gap_oportunidades(ag: pd.DataFrame, log) -> pd.DataFrame:
    """
    Gap de Oportunidad — Océanos Azules.

    Identifica categorías de mercado atractivas donde EAFIT no tiene presencia.
    Criterio de inclusión: calificacion_final >= 3.0 y categoría no cubierta
    por MAPEO_PROGRAMAS_EAFIT.

    Columnas de salida (en orden de lectura para el analista):
    IDENTIFICACIÓN → DECISIÓN → MERCADO → TENDENCIA → RETORNO → OFERTA → COSTO

    Thresholds calibrados sobre distribución real de las 129 categorías de
    oportunidad Colombia 2024 (pipeline v4.5+).
    """
    log("━━━ Gap de Oportunidad — Océanos Azules ━━━")

    if ag is None or len(ag) == 0 or "CATEGORIA_FINAL" not in ag.columns:
        log("⚠ Gap: ag vacío o sin CATEGORIA_FINAL. Se omite hoja.")
        return pd.DataFrame()

    col_cal = "calificacion_final"
    if col_cal not in ag.columns:
        log("⚠ Gap: columna calificacion_final no encontrada. Se omite hoja.")
        return pd.DataFrame()

    cats_eafit = {str(v).strip().upper() for v in MAPEO_PROGRAMAS_EAFIT.values() if v}
    ag_work = ag.copy()
    ag_work["_cat_norm"] = ag_work["CATEGORIA_FINAL"].astype(str).str.strip().str.upper()
    ag_work[col_cal] = pd.to_numeric(ag_work[col_cal], errors="coerce")

    mask_sin_eafit = ~ag_work["_cat_norm"].isin(cats_eafit)
    mask_cal_min = ag_work[col_cal] >= 3.0
    df_gap = ag_work[mask_sin_eafit & mask_cal_min].drop(columns=["_cat_norm"],
                                                          errors="ignore").copy()

    if len(df_gap) == 0:
        log("⚠ Gap: no se encontraron oportunidades con calificación ≥ 3.0.")
        return pd.DataFrame()

    _num_cols = [
        col_cal,
        "AAGR_ROBUSTO",
        f"suma_primer_curso_{AÑO_FIN_DATOS}",
        "SEÑAL_TENDENCIA",
        "TIPO_CRECIMIENTO",
        f"pct_no_matriculados_{AÑO_FIN_DATOS}",
        "salario_promedio",
        f"num_programas_{AÑO_FIN_DATOS}",
        "programas_nuevos_3a",
        "costo_promedio",
        "distancia_costo_pct",
        "NIVEL_MAYORIT",
        f"var_yoy_{AÑO_FIN_DATOS}",
        "CAGR_suma",
        "tasa_graduacion",
    ]
    for c in _num_cols:
        if c in df_gap.columns and c not in ("SEÑAL_TENDENCIA", "TIPO_CRECIMIENTO",
                                              "NIVEL_MAYORIT"):
            df_gap[c] = pd.to_numeric(df_gap[c], errors="coerce")

    df_gap["SEMAFORO"] = df_gap[col_cal].apply(
        lambda c: "VERDE" if c >= 4.0 else ("AMARILLO" if c >= 3.0 else "ROJO")
    )

    def _pts_cal(c):
        if pd.isna(c):
            return 0.0
        return max(0.0, min(40.0, (float(c) - 3.0) / 2.0 * 40.0))

    def _pts_aagr(a):
        if pd.isna(a):
            return 0.0
        a = float(a)
        if a < 0:
            return 0.0
        if a < 0.05:
            return round(a / 0.05 * 6, 2)
        if a < 0.15:
            return round(6 + (a - 0.05) / 0.10 * 10, 2)
        if a < 0.25:
            return round(16 + (a - 0.15) / 0.10 * 6, 2)
        return 25.0

    def _pts_pc(v):
        if pd.isna(v):
            return 0.0
        v = float(v)
        if v < 68:
            return round(v / 68 * 4, 2)
        if v < 151:
            return round(4 + (v - 68) / 83 * 4, 2)
        if v < 420:
            return round(8 + (v - 151) / 269 * 4, 2)
        if v < 1150:
            return round(12 + (v - 420) / 730 * 4, 2)
        if v < 2000:
            return round(16 + (v - 1150) / 850 * 4, 2)
        return 20.0

    def _pts_sal(s):
        if pd.isna(s):
            return 0.0
        s = float(s)
        if s < 3:
            return 0.0
        if s < 5.1:
            return round((s - 3) / 2.1 * 5, 2)
        if s < 6.0:
            return round(5 + (s - 5.1) / 0.9 * 4, 2)
        if s < 7.2:
            return round(9 + (s - 6.0) / 1.2 * 3, 2)
        if s < 9.0:
            return round(12 + (s - 7.2) / 1.8 * 3, 2)
        return 15.0

    df_gap["_pts_cal"] = df_gap[col_cal].apply(_pts_cal)
    df_gap["_pts_aagr"] = df_gap["AAGR_ROBUSTO"].apply(_pts_aagr)
    df_gap["_pts_pc"] = df_gap[f"suma_primer_curso_{AÑO_FIN_DATOS}"].apply(_pts_pc)
    df_gap["_pts_sal"] = df_gap["salario_promedio"].apply(_pts_sal)

    df_gap["PUNTUACION_OPORTUNIDAD"] = (
        df_gap["_pts_cal"] + df_gap["_pts_aagr"] +
        df_gap["_pts_pc"] + df_gap["_pts_sal"]
    ).round(1)
    df_gap.drop(columns=["_pts_cal", "_pts_aagr", "_pts_pc", "_pts_sal"],
                inplace=True, errors="ignore")

    def _senal_favorable(s: str) -> bool:
        s = str(s).strip().upper()
        return s in ("ACELERANDO", "ESTABLE") or s.startswith("▲") or s.startswith("→")

    def _nivel_oportunidad(row) -> str:
        pts = row.get("PUNTUACION_OPORTUNIDAD", 0) or 0
        senal = str(row.get("SEÑAL_TENDENCIA", "")).strip()
        pc = row.get(f"suma_primer_curso_{AÑO_FIN_DATOS}", 0) or 0

        if pc < 68:
            return "🟡 Baja-Media"

        if pts >= 75:
            if _senal_favorable(senal):
                return "🟢 Alta"
            return "🟡 Alta con Cautela"
        if pts >= 55:
            if _senal_favorable(senal):
                return "🟢 Media-Alta"
            return "🟡 Media"
        if pts >= 35:
            return "🟡 Media"
        return "🔴 Baja"

    df_gap["NIVEL_OPORTUNIDAD"] = df_gap.apply(_nivel_oportunidad, axis=1)

    _SENAL_EMOJI = {
        "ACELERANDO": "▲ ACELERANDO",
        "ESTABLE": "→ ESTABLE",
        "DESACELERANDO": "▼ DESACELERANDO",
        "EN_DECLIVE": "↓ EN DECLIVE",
        "CONTRACCION": "↓↓ CONTRACCION",
        "SIN_DATO": "— SIN DATO",
    }
    if "SEÑAL_TENDENCIA" in df_gap.columns:
        df_gap["SEÑAL_TENDENCIA"] = df_gap["SEÑAL_TENDENCIA"].apply(
            lambda v: _SENAL_EMOJI.get(str(v).strip(), str(v).strip())
            if pd.notna(v) else "— SIN DATO"
        )

    COLS_SALIDA_ORDEN = [
        "CATEGORIA_FINAL",
        "NIVEL_MAYORIT",
        "SEMAFORO",
        "NIVEL_OPORTUNIDAD",
        "PUNTUACION_OPORTUNIDAD",
        col_cal,
        f"suma_primer_curso_{AÑO_FIN_DATOS}",
        "AAGR_ROBUSTO",
        "CAGR_suma",
        "SEÑAL_TENDENCIA",
        "TIPO_CRECIMIENTO",
        f"var_yoy_{AÑO_FIN_DATOS}",
        f"pct_no_matriculados_{AÑO_FIN_DATOS}",
        "salario_promedio",
        "tasa_graduacion",
        f"num_programas_{AÑO_FIN_DATOS}",
        "programas_nuevos_3a",
        "costo_promedio",
        "distancia_costo_pct",
    ]

    RENAME_COLS = {
        "CATEGORIA_FINAL": "Categoría de mercado",
        "NIVEL_MAYORIT": "Nivel predominante",
        "SEMAFORO": "Semáforo",
        "NIVEL_OPORTUNIDAD": "Nivel de oportunidad",
        "PUNTUACION_OPORTUNIDAD": "Puntuación (0-100)",
        col_cal: "Calificación (1-5)",
        f"suma_primer_curso_{AÑO_FIN_DATOS}": f"Primer curso {AÑO_FIN_DATOS}",
        "AAGR_ROBUSTO": "AAGR (% anual)",
        "CAGR_suma": "CAGR 2019-2024 (%)",
        "SEÑAL_TENDENCIA": "Señal de tendencia",
        "TIPO_CRECIMIENTO": "Tipo de mercado",
        f"var_yoy_{AÑO_FIN_DATOS}": "Var. último año (%)",
        f"pct_no_matriculados_{AÑO_FIN_DATOS}": "% Inscritos no matriculados",
        "salario_promedio": "Salario prom. (SMLMV)",
        "tasa_graduacion": "Tasa de graduación",
        f"num_programas_{AÑO_FIN_DATOS}": "N° programas en mercado",
        "programas_nuevos_3a": "Prog. nuevos (3 años)",
        "costo_promedio": "Costo prom. mercado ($)",
        "distancia_costo_pct": "Distancia vs EAFIT (%)",
    }

    cols_ok = [c for c in COLS_SALIDA_ORDEN if c in df_gap.columns]
    df_gap = df_gap[cols_ok].rename(columns=RENAME_COLS)

    sort_cols = []
    if "Puntuación (0-100)" in df_gap.columns:
        sort_cols.append(("Puntuación (0-100)", False))
    if "Calificación (1-5)" in df_gap.columns:
        sort_cols.append(("Calificación (1-5)", False))
    if sort_cols:
        df_gap = df_gap.sort_values(
            [c for c, _ in sort_cols],
            ascending=[a for _, a in sort_cols],
        ).reset_index(drop=True)

    alta = (df_gap["Nivel de oportunidad"].str.contains("Alta", na=False)).sum() \
        if "Nivel de oportunidad" in df_gap.columns else 0
    med_alt = (df_gap["Nivel de oportunidad"].str.contains("Media-Alta", na=False)).sum() \
        if "Nivel de oportunidad" in df_gap.columns else 0
    if "Puntuación (0-100)" in df_gap.columns:
        log(
            f"✓ Gap completado: {len(df_gap)} oportunidades | "
            f"🟢 Alta: {alta} | 🟢 Media-Alta: {med_alt} | "
            f"Puntuación máxima: {df_gap['Puntuación (0-100)'].max():.0f} | "
            f"mínima: {df_gap['Puntuación (0-100)'].min():.0f}"
        )
    else:
        log(f"✓ Gap completado: {len(df_gap)} oportunidades")
    return df_gap


def _formatear_hoja_gap(writer: pd.ExcelWriter, df_gap: pd.DataFrame) -> None:
    """Formato visual de la hoja oportunidades_expansion."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets.get("oportunidades_expansion")
    if ws is None:
        return

    AZUL_OSC = "000066"
    BLANCO = "FFFFFF"
    GRIS_ALT = "F7F7F7"

    FILA_FILLS = {"VERDE": "EBF9EE", "AMARILLO": "FFFDE7", "ROJO": "FFF0F0"}

    NIVEL_FILLS = {
        "🟢 Alta": ("1F7A3C", "FFFFFF"),
        "🟢 Media-Alta": ("C6EFCE", "1A5C2A"),
        "🟡 Alta con Cautela": ("FFD966", "7D4800"),
        "🟡 Media": ("FFFDE7", "7D4800"),
        "🟡 Baja-Media": ("FFF0CC", "7D4800"),
        "🔴 Baja": ("FFC7CE", "9C0006"),
    }

    SENAL_FILLS = {
        "▲ ACELERANDO": ("1F7A3C", "FFFFFF"),
        "→ ESTABLE": ("C6EFCE", "1A5C2A"),
        "▼ DESACELERANDO": ("FFD966", "7D4800"),
        "↓ EN DECLIVE": ("FF7043", "FFFFFF"),
        "↓↓ CONTRACCION": ("C62828", "FFFFFF"),
        "— SIN DATO": ("EEEEEE", "888888"),
    }

    thin = Side(style="thin", color="D9D9D9")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [cell.value for cell in ws[1]]

    _col_pc_hdr = f"Primer curso {AÑO_FIN_DATOS}"
    BLOQUES = {
        "Categoría de mercado": ("IDENTIFICACIÓN", "37474F"),
        "Nivel predominante": ("IDENTIFICACIÓN", "37474F"),
        "Semáforo": ("DECISIÓN", AZUL_OSC),
        "Nivel de oportunidad": ("DECISIÓN", AZUL_OSC),
        "Puntuación (0-100)": ("DECISIÓN", AZUL_OSC),
        "Calificación (1-5)": ("DECISIÓN", AZUL_OSC),
        _col_pc_hdr: ("MERCADO", "2E7D32"),
        "AAGR (% anual)": ("MERCADO", "2E7D32"),
        "CAGR 2019-2024 (%)": ("MERCADO", "2E7D32"),
        "Señal de tendencia": ("MERCADO", "2E7D32"),
        "Tipo de mercado": ("MERCADO", "2E7D32"),
        "Var. último año (%)": ("MERCADO", "2E7D32"),
        "% Inscritos no matriculados": ("DEMANDA", "E65100"),
        "Salario prom. (SMLMV)": ("RETORNO", "5D4037"),
        "Tasa de graduación": ("RETORNO", "5D4037"),
        "N° programas en mercado": ("OFERTA", "2E4057"),
        "Prog. nuevos (3 años)": ("OFERTA", "2E4057"),
        "Costo prom. mercado ($)": ("COSTO", "6A1B9A"),
        "Distancia vs EAFIT (%)": ("COSTO", "6A1B9A"),
    }

    ws.insert_rows(1)
    current_blk = None
    blk_start = 1
    prev_color = "455A64"
    for ci, col_name in enumerate(cols, start=1):
        blk_info = BLOQUES.get(str(col_name), ("OTRO", "455A64"))
        blk_name, blk_color = blk_info
        if blk_name != current_blk:
            if current_blk is not None:
                ws.merge_cells(start_row=1, start_column=blk_start,
                               end_row=1, end_column=ci - 1)
                cell = ws.cell(row=1, column=blk_start)
                cell.value = current_blk
                cell.fill = PatternFill("solid", fgColor=prev_color)
                cell.font = Font(bold=True, color=BLANCO, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            current_blk = blk_name
            prev_color = blk_color
            blk_start = ci
        cell2 = ws.cell(row=2, column=ci)
        cell2.fill = PatternFill("solid", fgColor=blk_color)
        cell2.font = Font(bold=True, color=BLANCO, name="Arial", size=9)
        cell2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell2.border = borde

    if current_blk:
        ws.merge_cells(start_row=1, start_column=blk_start,
                       end_row=1, end_column=len(cols))
        cell = ws.cell(row=1, column=blk_start)
        cell.value = current_blk
        cell.fill = PatternFill("solid", fgColor=prev_color)
        cell.font = Font(bold=True, color=BLANCO, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32

    sem_col_idx = cols.index("Semáforo") + 1 if "Semáforo" in cols else None
    niv_col_idx = cols.index("Nivel de oportunidad") + 1 if "Nivel de oportunidad" in cols else None

    for ri in range(3, ws.max_row + 1):
        zebra = ri % 2 == 0
        sem_raw = str(ws.cell(row=ri, column=sem_col_idx).value or "") if sem_col_idx else ""
        sem_key = sem_raw.split()[-1] if sem_raw else ""
        base_fill = PatternFill(
            "solid",
            fgColor=FILA_FILLS.get(sem_key, GRIS_ALT) if not zebra else "FAFAFA",
        )

        for ci, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci)
            col_s = str(col_name)
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = base_fill

            if col_s == "Semáforo":
                sem_raw_v = str(cell.value or "")
                sk = sem_raw_v.split()[-1] if sem_raw_v else ""
                color_map = {"VERDE": "1F7A3C", "AMARILLO": "B8860B", "ROJO": "C62828"}
                emoji_map = {"VERDE": "🟢 VERDE", "AMARILLO": "🟡 AMARILLO", "ROJO": "🔴 ROJO"}
                cell.value = emoji_map.get(sk, sem_raw_v)
                cell.font = Font(bold=True, color=color_map.get(sk, "000000"),
                                 name="Arial", size=10)

            elif col_s == "Nivel de oportunidad":
                val = str(cell.value or "")
                fg, fc = NIVEL_FILLS.get(val, ("EEEEEE", "555555"))
                cell.fill = PatternFill("solid", fgColor=fg)
                cell.font = Font(bold=True, color=fc, name="Arial", size=10)

            elif col_s == "Puntuación (0-100)":
                cell.number_format = "0.0"
                try:
                    v = float(cell.value)
                    fc = "1A5C2A" if v >= 75 else ("7D4800" if v >= 55 else
                                                    ("9C0006" if v < 35 else "000000"))
                    cell.font = Font(bold=True, color=fc, name="Arial", size=10)
                except (TypeError, ValueError):
                    pass

            elif col_s == "Calificación (1-5)":
                cell.number_format = "0.00"
                try:
                    v = float(cell.value)
                    fc = "1A5C2A" if v >= 4.0 else ("7D4800" if v >= 3.0 else "9C0006")
                    cell.font = Font(bold=True, color=fc, name="Arial", size=10)
                except (TypeError, ValueError):
                    pass

            elif col_s == "Señal de tendencia":
                val = str(cell.value or "")
                fg, fc = SENAL_FILLS.get(val, ("EEEEEE", "888888"))
                cell.fill = PatternFill("solid", fgColor=fg)
                cell.font = Font(bold=True, color=fc, name="Arial", size=10)

            elif col_s in ("AAGR (% anual)", "CAGR 2019-2024 (%)", "Var. último año (%)"):
                cell.number_format = "0.0%"
                cell.font = Font(name="Arial", size=10)

            elif col_s == "% Inscritos no matriculados":
                cell.number_format = "0.0%"
                cell.font = Font(name="Arial", size=10)

            elif col_s == _col_pc_hdr:
                cell.number_format = "#,##0"
                cell.font = Font(name="Arial", size=10)

            elif col_s == "Salario prom. (SMLMV)":
                cell.number_format = "0.0"
                cell.font = Font(name="Arial", size=10)

            elif col_s == "Costo prom. mercado ($)":
                cell.number_format = "#,##0"
                cell.font = Font(name="Arial", size=10)

            elif col_s == "Distancia vs EAFIT (%)":
                cell.number_format = '0.0"%"'
                cell.font = Font(name="Arial", size=10)

            elif col_s in ("N° programas en mercado", "Prog. nuevos (3 años)"):
                cell.number_format = "#,##0"
                cell.font = Font(name="Arial", size=10)

            elif col_s == "Tasa de graduación":
                cell.number_format = "0.0%"
                cell.font = Font(name="Arial", size=10)

            else:
                cell.font = Font(name="Arial", size=10)

    ANCHOS = {
        "Categoría de mercado": 36,
        "Nivel predominante": 18,
        "Semáforo": 12,
        "Nivel de oportunidad": 20,
        "Puntuación (0-100)": 14,
        "Calificación (1-5)": 14,
        _col_pc_hdr: 14,
        "AAGR (% anual)": 13,
        "CAGR 2019-2024 (%)": 13,
        "Señal de tendencia": 20,
        "Tipo de mercado": 16,
        "Var. último año (%)": 14,
        "% Inscritos no matriculados": 22,
        "Salario prom. (SMLMV)": 18,
        "Tasa de graduación": 16,
        "N° programas en mercado": 20,
        "Prog. nuevos (3 años)": 18,
        "Costo prom. mercado ($)": 20,
        "Distancia vs EAFIT (%)": 18,
    }
    for ci, col_name in enumerate(cols, start=1):
        w = ANCHOS.get(str(col_name))
        if w:
            ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"


def _escribir_hoja_total(
    writer: pd.ExcelWriter,
    ag: pd.DataFrame,
    sheet_name: str = "total",
) -> list[str]:
    """Escribe una hoja con estructura idéntica a `total` (fila 1 = bloques, fila 2 = nombres
    de columnas). El parámetro `sheet_name` permite reutilizar la lógica para `total_pregrado`.
    Retorna el orden de columnas escritas."""
    from openpyxl.styles import Alignment as _Al
    from openpyxl.styles import Font as _Ft
    from openpyxl.styles import PatternFill as _PF

    NOMBRES_LEGIBLES = {
        "CAT_ID": "ID Categoría",
        "CATEGORIA_FINAL": "Categoría de mercado",
        "FUENTE_CATEGORIA": "Fuente de categoría",
        "NIVEL_MAYORIT": "Nivel predominante",
        f"suma_matricula_{AÑO_INICIO_HISTORICO}": f"Matr. total {AÑO_INICIO_HISTORICO}",
        f"suma_matricula_{AÑO_FIN_DATOS}": f"Matr. total {AÑO_FIN_DATOS}",
        f"prom_matricula_{AÑO_INICIO_HISTORICO}": f"Prom. matr. {AÑO_INICIO_HISTORICO}",
        f"prom_matricula_{AÑO_FIN_DATOS}": f"Prom. matr. {AÑO_FIN_DATOS}",
        f"participacion_{AÑO_INICIO_HISTORICO}": f"Part. primer curso {AÑO_INICIO_HISTORICO}",
        f"participacion_{AÑO_FIN_DATOS}": f"Part. primer curso {AÑO_FIN_DATOS}",
        "AAGR_suma": "AAGR primer curso — suma del mercado",
        "AAGR_prom": "AAGR primer curso — prom. por programa",
        "CAGR_suma": "CAGR primer curso",
        "AAGR_ROBUSTO": "AAGR primer curso (robusto)",
        "TIPO_CRECIMIENTO": "Tipo de mercado",
        "SEÑAL_TENDENCIA": "Señal tendencia actual",
        f"suma_primer_curso_{AÑO_INICIO_HISTORICO}": f"Primer curso {AÑO_INICIO_HISTORICO}",
        f"suma_primer_curso_{AÑO_FIN_DATOS}": f"Primer curso {AÑO_FIN_DATOS}",
        f"prom_primer_curso_{AÑO_INICIO_HISTORICO}": (
            f"Prom. primer curso {AÑO_INICIO_HISTORICO} (por registro)"
        ),
        f"prom_primer_curso_{AÑO_FIN_DATOS}": (
            f"Prom. primer curso {AÑO_FIN_DATOS} (por registro)"
        ),
        "AAGR_primer_curso": "AAGR primer curso (histórico)",
        f"inscritos_{AÑO_FIN_DATOS - 1}_suma": f"Inscritos suma {AÑO_FIN_DATOS - 1}",
        f"inscritos_{AÑO_FIN_DATOS}_suma": f"Inscritos suma {AÑO_FIN_DATOS}",
        f"inscritos_{AÑO_FIN_DATOS - 1}_prom_por_programa": (
            f"Inscritos prom/prog {AÑO_FIN_DATOS - 1}"
        ),
        f"inscritos_{AÑO_FIN_DATOS}_prom_por_programa": (
            f"Inscritos prom/prog {AÑO_FIN_DATOS}"
        ),
        f"pct_no_matriculados_{AÑO_FIN_DATOS - 1}": (
            f"% No matriculados {AÑO_FIN_DATOS - 1}"
        ),
        f"pct_no_matriculados_{AÑO_FIN_DATOS}": f"% No matriculados {AÑO_FIN_DATOS}",
        "FUENTE_PCT_NO_MAT": "Fuente % no matr.",
        "var_inscritos": f"Var. inscritos {AÑO_FIN_DATOS - 1}→{AÑO_FIN_DATOS}",
        "var_inscritos_prom": (
            f"Var. inscritos prom {AÑO_FIN_DATOS - 1}→{AÑO_FIN_DATOS}"
        ),
        "salario_promedio": "Salario promedio (SMLMV)",
        "salario_proyectado_pesos_hoy": "Salario promedio (pesos hoy)",
        f"num_programas_{AÑO_INICIO_HISTORICO}": (
            f"Programas con dato {AÑO_INICIO_HISTORICO}"
        ),
        f"num_programas_{AÑO_FIN_DATOS}": f"Programas con dato {AÑO_FIN_DATOS}",
        "programas_activos": "Registros con matrículas",
        "programas_inactivos": "Registros sin matrículas",
        "programas_nuevos_3a": "Prog. nuevos (últimos 3 años)",
        "nuevos_vs_snapshot": "Nuevos vs año anterior",
        "var_programas": (
            f"Var. N° programas {AÑO_FIN_DATOS - 1}→{AÑO_FIN_DATOS}"
        ),
        "pct_con_matricula": "% Registros con matrícula",
        "costo_promedio": "Costo prom. matrícula",
        "distancia_costo_pct": "Distancia vs benchmark EAFIT (%)",
        "score_matricula": "S. Primer curso",
        "score_participacion": "S. Participación",
        "score_AAGR": "S. AAGR",
        "score_salario": "S. Salario",
        "score_pct_no_matriculados": "S. No matriculados",
        "score_num_programas": "S. N° Programas",
        "score_costo": "S. Costo",
        "calificacion_final": "Calificación final",
    }
    for _y in range(AÑO_INICIO_PRIMER_CURSO, AÑO_FIN_DATOS + 1):
        NOMBRES_LEGIBLES.setdefault(f"suma_primer_curso_{_y}", f"Primer curso {_y}")

    for _y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        NOMBRES_LEGIBLES.setdefault(
            f"prom_primer_curso_{_y}",
            f"Prom. primer curso {_y} (por registro)",
        )

    for _y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
        NOMBRES_LEGIBLES.setdefault(
            f"var_primer_curso_{_y}",
            f"Var. suma primer curso {_y - 1}→{_y}",
        )
        NOMBRES_LEGIBLES.setdefault(
            f"var_prom_primer_curso_{_y}",
            f"Var. prom. por prog. {_y - 1}→{_y}",
        )

    NOMBRES_LEGIBLES["AAGR_suma"] = "AAGR primer curso — suma del mercado"
    NOMBRES_LEGIBLES["AAGR_prom"] = "AAGR primer curso — prom. por programa"

    for _y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        NOMBRES_LEGIBLES.setdefault(f"inscritos_{_y}_suma", f"Inscritos suma {_y}")
        NOMBRES_LEGIBLES.setdefault(f"inscritos_{_y}_prom_por_programa", f"Inscritos prom/prog {_y}")
        NOMBRES_LEGIBLES.setdefault(f"suma_matricula_{_y}", f"Matr. total {_y}")
        NOMBRES_LEGIBLES.setdefault(f"prom_matricula_{_y}", f"Prom. matr. {_y}")

    for _y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1):
        for _s in (1, 2):
            NOMBRES_LEGIBLES.setdefault(f"suma_matricula_{_y}_{_s}", f"Matr. total {_y} S{_s}")

    for _y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1):
        NOMBRES_LEGIBLES.setdefault(f"var_suma_{_y}", f"Var. matr. total suma {_y}")
        NOMBRES_LEGIBLES.setdefault(f"var_prom_{_y}", f"Var. matr. total prom {_y}")

    COLORES_BLOQUES = {
        "CATEGORÍA": "37474F",
        "DEMANDA NUEVA — PRIMER CURSO": "2E7D32",
        "PARTICIPACIÓN Y CRECIMIENTO": "1A5276",
        "INSCRITOS": "E65100",
        "SALARIO OLE": "5D4037",
        "OFERTA DE PROGRAMAS": "2E4057",
        "COSTO": "6A1B9A",
        "SCORING — valor | puntuación": "000066",
        "CALIFICACIÓN FINAL": "000066",
    }
    wb = writer.book
    # Posición canónica de la hoja: `total` va en índice 1 (justo después de
    # resumen_ejecutivo); `total_pregrado` se añade al final y luego se reordena.
    if sheet_name == "total":
        ws = wb.create_sheet(sheet_name, 1)
    else:
        ws = wb.create_sheet(sheet_name)
    col_order: list[str] = []
    col_idx = 1
    for block_name, cols in _BLOQUES_TOTAL:
        present = [c for c in cols if c in ag.columns]
        if not present:
            continue
        start = col_idx
        for c in present:
            nombre_header = NOMBRES_LEGIBLES.get(c, c)
            ws.cell(row=2, column=col_idx, value=nombre_header)
            col_order.append(c)
            col_idx += 1
        ws.cell(row=1, column=start, value=block_name)
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col_idx - 1)
        bg = COLORES_BLOQUES.get(block_name, "455A64")
        for col_b in range(start, col_idx):
            cell_b = ws.cell(row=1, column=col_b)
            cell_b.fill = _PF("solid", fgColor=bg)
            cell_b.font = _Ft(bold=True, color="FFFFFF", name="Arial", size=9)
            cell_b.alignment = _Al(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 28

    for c in range(1, ws.max_column + 1):
        cell_h = ws.cell(row=2, column=c)
        cell_h.font = _Ft(bold=True, name="Arial", size=9)
        cell_h.alignment = _Al(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(ag.itertuples(index=False), start=3):
        c_idx = 1
        for _, cols in _BLOQUES_TOTAL:
            for c in cols:
                if c in ag.columns:
                    val = getattr(row, c, None)
                    ws.cell(row=r_idx, column=c_idx, value=val)
                    c_idx += 1
    return col_order


def _aplicar_formato_total(ws, col_order: list[str]) -> None:
    """Aplica formatos de número y color de fila según calificacion_final. col_order es el orden de columnas en la hoja."""
    from openpyxl.styles import Alignment, PatternFill
    pct_fmt = "0.0%"
    moneda_fmt = "#,##0"
    score_fmt = "0"
    calif_fmt = "0.00"
    col_calif = None
    for j, col in enumerate(col_order):
        if col == "calificacion_final":
            col_calif = j
            break
    SCORE_FILLS = {
        1: PatternFill("solid", fgColor="FFC7CE"),  # rojo
        2: PatternFill("solid", fgColor="FFD9B3"),  # naranja
        3: PatternFill("solid", fgColor="FFEB9C"),  # amarillo
        4: PatternFill("solid", fgColor="C6EFCE"),  # verde claro
        5: PatternFill("solid", fgColor="4CAF50"),  # verde fuerte
    }
    SCORE_COLS_SET = {
        "score_matricula",
        "score_participacion",
        "score_AAGR",
        "score_salario",
        "score_pct_no_matriculados",
        "score_num_programas",
        "score_costo",
    }

    # Paleta de colores para SEÑAL_TENDENCIA
    SENAL_FILLS = {
        "ACELERANDO":    PatternFill("solid", fgColor="1F7A3C"),  # verde oscuro
        "ESTABLE":       PatternFill("solid", fgColor="C6EFCE"),  # verde suave
        "DESACELERANDO": PatternFill("solid", fgColor="FFD966"),  # amarillo ámbar
        "EN_DECLIVE":    PatternFill("solid", fgColor="FF7043"),  # naranja-rojo
        "CONTRACCION":   PatternFill("solid", fgColor="C62828"),  # rojo oscuro
        "SIN_DATO":      PatternFill("solid", fgColor="EEEEEE"),  # gris neutro
        "SIN_ACTIVIDAD": PatternFill("solid", fgColor="BDBDBD"),  # gris oscuro — mercado extinto
    }
    SENAL_FONTS = {
        "ACELERANDO":    {"bold": True,  "color": "FFFFFF"},
        "ESTABLE":       {"bold": False, "color": "1A5C2A"},
        "DESACELERANDO": {"bold": True,  "color": "7D4800"},
        "EN_DECLIVE":    {"bold": True,  "color": "FFFFFF"},
        "CONTRACCION":   {"bold": True,  "color": "FFFFFF"},
        "SIN_DATO":      {"bold": False, "color": "888888"},
        "SIN_ACTIVIDAD": {"bold": False, "color": "424242"},
    }
    # Texto con emoji para que sea legible sin necesidad de leyenda
    SENAL_LABELS = {
        "ACELERANDO":    "▲ ACELERANDO",
        "ESTABLE":       "→ ESTABLE",
        "DESACELERANDO": "▼ DESACELERANDO",
        "EN_DECLIVE":    "↓ EN DECLIVE",
        "CONTRACCION":   "↓↓ CONTRACCION",
        "SIN_DATO":      "— SIN DATO",
        "SIN_ACTIVIDAD": "✕ SIN ACTIVIDAD",
    }

    for r in range(3, ws.max_row + 1):
        calif = ws.cell(row=r, column=col_calif + 1).value if col_calif is not None else None
        try:
            calif_f = float(calif) if calif is not None else 0.0
        except (TypeError, ValueError):
            calif_f = 0.0
        if calif_f >= 4.0:
            fill = PatternFill("solid", fgColor="EBF9EE")  # verde muy suave
        elif calif_f >= 3.0:
            fill = PatternFill("solid", fgColor="FFFDE7")  # amarillo muy suave
        else:
            fill = PatternFill("solid", fgColor="FFF0F0")  # rojo muy suave
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            col_name = col_order[c - 1] if c - 1 < len(col_order) else None
            if col_name in ("CAT_ID", "CATEGORIA_FINAL", "FUENTE_CATEGORIA", "NIVEL_MAYORIT"):
                continue
            if col_name in SCORE_COLS_SET:
                try:
                    sv = int(float(cell.value)) if cell.value is not None else 0
                except (TypeError, ValueError):
                    sv = 0
                if sv in SCORE_FILLS:
                    cell.fill = SCORE_FILLS[sv]
                    from openpyxl.styles import Font as _Font
                    if sv == 5:
                        cell.font = _Font(bold=True, color="FFFFFF", name="Arial", size=10)
                    elif sv == 1:
                        cell.font = _Font(bold=True, color="9C0006", name="Arial", size=10)
                    else:
                        cell.font = _Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Promedios de estudiantes (personas): enteros sin decimales
            if col_name and (
                "prom_primer_curso" in col_name
                and not col_name.startswith("var_")
                or col_name == f"prom_matricula_por_programa_{AÑO_FIN_DATOS}"
                or ("inscritos" in col_name and "prom" in col_name)
            ):
                cell.number_format = "#,##0"
            # Sumas de primer_curso y matrícula: enteros sin decimales
            elif col_name and (
                "suma_primer_curso" in col_name
                or "suma_matricula" in col_name
                or "num_programas" in col_name
                or ("inscritos" in col_name and "suma" in col_name)
            ):
                cell.number_format = "#,##0"
            # Salario SMLMV: 1 decimal (ej. 3.1, 4.5)
            elif col_name == "salario_promedio":
                cell.number_format = "0.0"
            # Costo en pesos: sin centavos
            elif col_name == "costo_promedio":
                cell.number_format = "#,##0"
            elif col_name == "salario_proyectado_pesos_hoy":
                cell.number_format = "#,##0"
            elif col_name == "distancia_costo_pct":
                cell.number_format = '0.0"%"'
            elif col_name and (
                "var_primer_curso" in col_name or "var_prom_primer_curso" in col_name
            ):
                cell.number_format = pct_fmt
            # Porcentajes y tasas
            elif col_name and (
                "participacion" in col_name
                or "pct_" in col_name
                or "var_" in col_name
                or "AAGR" in col_name
                or "CAGR" in col_name
            ) and not col_name.startswith("score_"):
                cell.number_format = pct_fmt
            elif col_name and col_name.startswith("score_"):
                cell.number_format = score_fmt
            elif col_name == "calificacion_final":
                cell.number_format = calif_fmt
            elif col_name == "SEÑAL_TENDENCIA":
                raw_val = str(cell.value).strip() if cell.value is not None else "SIN_DATO"
                senal = raw_val if raw_val in SENAL_FILLS else "SIN_DATO"
                cell.value = SENAL_LABELS.get(senal, senal)
                cell.fill = SENAL_FILLS[senal]
                fnt_cfg = SENAL_FONTS[senal]
                from openpyxl.styles import Font as _Font
                cell.font = _Font(
                    bold=fnt_cfg["bold"],
                    color=fnt_cfg["color"],
                    name="Arial",
                    size=10,
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _ANCHOS_TOTAL = {
        "CAT_ID": 8,
        "CATEGORIA_FINAL": 42,
        "FUENTE_CATEGORIA": 14,
        "NIVEL_MAYORIT": 22,
        **{f"suma_primer_curso_{y}": 13
           for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)},
        **{f"prom_primer_curso_{y}": 18
           for y in range(AÑO_INICIO_HISTORICO, AÑO_FIN_DATOS + 1)},
        **{f"var_primer_curso_{y}": 16
           for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1)},
        **{f"var_prom_primer_curso_{y}": 18
           for y in range(AÑO_INICIO_HISTORICO + 1, AÑO_FIN_DATOS + 1)},
        f"participacion_{AÑO_INICIO_HISTORICO}": 16,
        f"participacion_{AÑO_FIN_DATOS}": 16,
        "AAGR_suma": 18,
        "AAGR_prom": 18,
        "CAGR_suma": 16,
        "AAGR_ROBUSTO": 18,
        "TIPO_CRECIMIENTO": 18,
        "SEÑAL_TENDENCIA": 22,
        f"inscritos_{AÑO_FIN_DATOS - 1}_suma": 16,
        f"inscritos_{AÑO_FIN_DATOS}_suma": 16,
        f"inscritos_{AÑO_FIN_DATOS - 1}_prom_por_programa": 18,
        f"inscritos_{AÑO_FIN_DATOS}_prom_por_programa": 18,
        f"pct_no_matriculados_{AÑO_FIN_DATOS - 1}": 18,
        f"pct_no_matriculados_{AÑO_FIN_DATOS}": 18,
        "FUENTE_PCT_NO_MAT": 22,
        "var_inscritos": 16,
        "var_inscritos_prom": 18,
        "salario_promedio": 18,
        "salario_proyectado_pesos_hoy": 20,
        f"num_programas_{AÑO_INICIO_HISTORICO}": 18,
        f"num_programas_{AÑO_FIN_DATOS}": 18,
        "programas_activos": 20,
        "programas_inactivos": 20,
        "programas_nuevos_3a": 22,
        "nuevos_vs_snapshot": 18,
        "var_programas": 18,
        "pct_con_matricula": 22,
        "costo_promedio": 18,
        "distancia_costo_pct": 24,
        "score_matricula": 9,
        "score_participacion": 9,
        "score_AAGR": 9,
        "score_salario": 9,
        "score_pct_no_matriculados": 9,
        "score_num_programas": 9,
        "score_costo": 9,
        "calificacion_final": 16,
    }

    from openpyxl.utils import get_column_letter as _gcl2

    for _ci, _col in enumerate(col_order, start=1):
        _w = _ANCHOS_TOTAL.get(_col)
        if _w:
            ws.column_dimensions[_gcl2(_ci)].width = _w

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = ws.dimensions


def exportar_base_maestra_excel(ruta_salida: Path | None = None) -> Path:
    """
    Exporta un Excel formateado con los resultados de la Fase 1 (base_maestra.parquet).

    Filtra los niveles activos en NIVELES_MERCADO (incluye UNIVERSITARIO si está configurado).
    Genera dos hojas:
      - 'Programas_Categorizados': todos los programas filtrados con sus categorías.
      - 'Revision_Requerida': solo los que tienen REQUIERE_REVISION=True (confianza KNN < 50%).

    Aplica formato visual:
      - Verde  : CRUCE_SNIES / MATCH_NOMBRE / MATCH_CATEGORIA  (certeza 100%)
      - Amarillo: KNN_TFIDF con confianza >= 0.50
      - Rojo   : KNN_TFIDF con confianza < 0.50 (requiere revisión)

    Retorna la ruta del archivo generado.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not CHECKPOINT_BASE_MAESTRA.exists():
        raise FileNotFoundError(
            "No existe base_maestra.parquet. Ejecuta primero la Fase 1 del pipeline de mercado."
        )

    log_info("[Exportar F1] Leyendo base_maestra.parquet...")
    df = pd.read_parquet(CHECKPOINT_BASE_MAESTRA)

    # ── Filtro de niveles ────────────────────────────────────────────────────
    col_nivel = "NIVEL_DE_FORMACIÓN"
    n_total = len(df)
    if col_nivel in df.columns and NIVELES_MERCADO:
        df = df[df[col_nivel].isin(NIVELES_MERCADO)].copy()
        log_info(f"[Exportar F1] Filtro niveles: {n_total:,} → {len(df):,} programas. Niveles activos: {NIVELES_MERCADO}")
    else:
        log_info("[Exportar F1] No se aplicó filtro de niveles (columna no disponible).")

    if len(df) == 0:
        raise ValueError("No hay programas de Especialización/Maestría en la base maestra.")

    # ── Columnas a exportar (orden lógico) ───────────────────────────────────
    COLS_ORDEN = [
        "CÓDIGO_SNIES_DEL_PROGRAMA",
        "NOMBRE_DEL_PROGRAMA",
        "TITULO_OTORGADO",
        "NIVEL_DE_FORMACIÓN",
        "MODALIDAD",
        "NOMBRE_INSTITUCIÓN",
        "CARÁCTER_ACADÉMICO",
        "SECTOR",
        "DEPARTAMENTO_OFERTA_PROGRAMA",
        "MUNICIPIO_OFERTA_PROGRAMA",
        "ÁREA_DE_CONOCIMIENTO",
        "NÚCLEO_BÁSICO_DEL_CONOCIMIENTO",
        "ESTADO_PROGRAMA",
        # Resultado de la clasificación
        "CATEGORIA_FINAL",
        "FUENTE_CATEGORIA",
        "PROBABILIDAD",
        "CATEGORIA_ALTERNATIVA",
        "PROBABILIDAD_ALTERNATIVA",
        "REQUIERE_REVISION",
    ]
    cols_export = [c for c in COLS_ORDEN if c in df.columns]
    # Agregar columnas que no están en la lista pero sí en el df (al final)
    extra = [c for c in df.columns if c not in cols_export and not c.startswith("_")]
    df_export = df[cols_export + extra].copy()

    # Limpiar columnas internas
    cols_internas = [c for c in df_export.columns if c.startswith("_") or c == "schema_version"]
    df_export = df_export.drop(columns=cols_internas, errors="ignore")

    # ── Ruta de salida ───────────────────────────────────────────────────────
    if ruta_salida is None:
        from etl.config import OUTPUTS_DIR

        ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
        ruta_salida = OUTPUTS_DIR / f"Base_Programas_Categoria_F1_{ts}.xlsx"
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    # ── Hoja de revisión requerida ───────────────────────────────────────────
    col_rev = "REQUIERE_REVISION"
    if col_rev in df_export.columns:
        df_revision = df_export[
            df_export[col_rev].astype(str).str.lower().isin(["true", "1", "yes", "sí"])
        ].copy()
    else:
        df_revision = pd.DataFrame()

    # ── Estilos ──────────────────────────────────────────────────────────────
    AZUL = "000066"
    VERDE_F = "C6EFCE"
    AMARI_F = "FFF2CC"
    ROJO_F = "FFC7CE"
    BLANCO = "FFFFFF"
    GRIS_ALT = "F5F5F5"

    thin = Side(style="thin", color="CCCCCC")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _estilo_header(cell):
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    def _color_fila(fuente: str, requiere: bool, _prob: float | None) -> str:
        fuente_up = str(fuente).upper().strip()
        if fuente_up in ("CRUCE_SNIES", "MATCH_NOMBRE", "MATCH_CATEGORIA"):
            return VERDE_F
        if requiere:
            return ROJO_F
        if fuente_up == "KNN_TFIDF":
            return AMARI_F
        return BLANCO

    def _escribir_hoja(ws, df_h: pd.DataFrame, titulo: str) -> None:
        """Escribe un DataFrame en la hoja ws con encabezados y colores."""
        # Fila 1: título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_h.columns))
        cell_t = ws.cell(row=1, column=1, value=titulo)
        cell_t.font = Font(bold=True, color="FFFFFF", size=12)
        cell_t.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
        cell_t.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Fila 2: encabezados de columna
        for ci, col in enumerate(df_h.columns, start=1):
            cell = ws.cell(row=2, column=ci, value=col)
            _estilo_header(cell)
        ws.row_dimensions[2].height = 30

        # Anchos de columna
        ANCHOS = {
            "CÓDIGO_SNIES_DEL_PROGRAMA": 18,
            "NOMBRE_DEL_PROGRAMA": 46,
            "TITULO_OTORGADO": 36,
            "NIVEL_DE_FORMACIÓN": 24,
            "MODALIDAD": 14,
            "NOMBRE_INSTITUCIÓN": 34,
            "CARÁCTER_ACADÉMICO": 18,
            "SECTOR": 12,
            "DEPARTAMENTO_OFERTA_PROGRAMA": 22,
            "MUNICIPIO_OFERTA_PROGRAMA": 22,
            "ÁREA_DE_CONOCIMIENTO": 24,
            "NÚCLEO_BÁSICO_DEL_CONOCIMIENTO": 28,
            "ESTADO_PROGRAMA": 14,
            "CATEGORIA_FINAL": 32,
            "FUENTE_CATEGORIA": 18,
            "PROBABILIDAD": 14,
            "CATEGORIA_ALTERNATIVA": 32,
            "PROBABILIDAD_ALTERNATIVA": 14,
            "REQUIERE_REVISION": 16,
        }
        for ci, col in enumerate(df_h.columns, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = ANCHOS.get(col, 16)

        # Filas de datos
        col_fuente = list(df_h.columns).index("FUENTE_CATEGORIA") if "FUENTE_CATEGORIA" in df_h.columns else None
        col_rev_i = list(df_h.columns).index("REQUIERE_REVISION") if "REQUIERE_REVISION" in df_h.columns else None
        col_prob_i = list(df_h.columns).index("PROBABILIDAD") if "PROBABILIDAD" in df_h.columns else None

        for ri, (_, row) in enumerate(df_h.iterrows(), start=3):
            fuente = str(row.iloc[col_fuente]) if col_fuente is not None else ""
            req_val = str(row.iloc[col_rev_i]).lower() in ("true", "1", "yes", "sí") if col_rev_i is not None else False
            try:
                prob_f = float(row.iloc[col_prob_i]) if col_prob_i is not None and pd.notna(row.iloc[col_prob_i]) else None
            except (TypeError, ValueError):
                prob_f = None

            bg = _color_fila(fuente, req_val, prob_f)
            fill = (
                PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                if bg != BLANCO
                else (
                    PatternFill(start_color=GRIS_ALT, end_color=GRIS_ALT, fill_type="solid")
                    if ri % 2 == 0
                    else None
                )
            )

            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci)
                # Serializar valores Python nativos
                try:
                    import math as _math

                    if val is None or (isinstance(val, float) and (_math.isnan(val) or _math.isinf(val))):
                        cell.value = ""
                    elif hasattr(val, "item"):
                        # numpy scalar → Python nativo
                        native = val.item()
                        cell.value = "Sí" if isinstance(native, bool) and native else (
                            "No" if isinstance(native, bool) else native
                        )
                    elif isinstance(val, bool):
                        cell.value = "Sí" if val else "No"
                    elif pd.isna(val):
                        cell.value = ""
                    else:
                        cell.value = val
                except Exception:
                    cell.value = str(val) if val is not None else ""
                if fill:
                    cell.fill = fill
                cell.border = borde
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                # Formato número para columnas de probabilidad
                col_name = df_h.columns[ci - 1]
                if col_name in ("PROBABILIDAD", "PROBABILIDAD_ALTERNATIVA") and isinstance(
                    cell.value, (int, float)
                ):
                    cell.number_format = "0.0%"

        # Freeze + autofiltro
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(df_h.columns))}2"

    # ── Escribir Excel con openpyxl puro (sin ExcelWriter para evitar corrupción) ──
    import openpyxl as _opxl

    wb = _opxl.Workbook()

    # Hoja principal
    ws_main = wb.active
    ws_main.title = "Programas_Categorizados"
    _escribir_hoja(
        ws_main,
        df_export,
        f"Programas Esp+Maestría con Categorías de Mercado — {len(df_export):,} programas",
    )

    # Hoja revisión requerida
    if len(df_revision) > 0:
        ws_rev = wb.create_sheet("Revision_Requerida")
        _escribir_hoja(
            ws_rev,
            df_revision,
            f"Programas que Requieren Revisión Manual — {len(df_revision):,} registros (confianza KNN < 50%)",
        )

    # ── Hoja Leyenda ─────────────────────────────────────────────────────────
    ws_ley = wb.create_sheet("Leyenda")

    FILAS_METODOLOGIA = [
        ("METODOLOGÍA", "Definición", "Detalle"),
        (
            "Primer curso",
            "Nuevos estudiantes matriculados",
            "Suma de primer_curso semestral del SNIES. Es el indicador principal "
            "de demanda real de un mercado. Equivale a los nuevos matriculados en "
            "el año (S1 + S2). Diferente de matrícula total (que incluye "
            "estudiantes de todos los semestres).",
        ),
        (
            "AAGR primer curso (robusto)",
            "Crecimiento anual histórico del mercado",
            "Average Annual Growth Rate calculado sobre primer_curso 2019-2024. "
            "Formula: promedio de las variaciones interanuales "
            "(2019-2020, 2020-2021, ..., 2023-2024). "
            "Para categorias nuevas (sin dato en 2019), se calcula desde el "
            "primer año con dato. El pipeline calcula este indicador "
            "correctamente. El AAGR del archivo manual de referencia tenia un "
            "error de formula (AVERAGE/5) que producia valores 5 veces menores "
            "al valor real.",
        ),
        (
            "Señal tendencia actual",
            "Momento actual del mercado (ultimo año)",
            "Basada en la variacion primer_curso 2023-2024 (YoY). "
            "ACELERANDO: crecimiento mayor a 10% en el ultimo año. "
            "ESTABLE: variacion entre -10% y +10%. "
            "EN DECLIVE: caida en el ultimo año. "
            "CONTRACCION: caida fuerte. "
            "COMPLEMENTA el AAGR historico: una categoria puede tener buen AAGR "
            "(crecio 5 años) pero señal EN DECLIVE (cayo este año), o viceversa. "
            "Ambas metricas son validas y complementarias.",
        ),
        (
            "Registros con matriculas vs Programas con dato",
            "Dos conteos distintos de programas",
            "Registros con matriculas: filas unicas en SNIES con matricula mayor "
            "a 0 en 2024. Un mismo programa SNIES puede generar multiples "
            "registros si se ofrece en varias ciudades o modalidades. "
            "Programas con dato 2024: codigos SNIES unicos con primer_curso "
            "mayor a 0. Es normal que Registros sea mayor a Programas unicos.",
        ),
        (
            "Prom. primer curso (por registro)",
            "Promedio de primer_curso sobre registros individuales",
            "suma_primer_curso dividido por n_registros_con_dato. El denominador "
            "incluye cada combinacion programa, modalidad, municipio y semestre. "
            "No es equivalente a suma dividido por n_programas_unicos. "
            "El scoring S. Primer curso usa esta cifra de forma consistente "
            "en todas las categorias.",
        ),
        (
            "Distancia vs benchmark EAFIT (%)",
            "Cuanto cuesta el mercado vs lo que cobra EAFIT",
            "Formula: (Costo promedio del mercado menos Benchmark EAFIT) "
            "dividido por Benchmark EAFIT, por 100. "
            "Negativo: el mercado es mas barato que EAFIT (EAFIT tiene margen "
            "de precio). Positivo: el mercado cobra mas que EAFIT. "
            "Benchmarks: ESP universitaria 11910000, "
            "ESP Medico-Quirurgica 31895490, Maestria 13686800.",
        ),
        (
            "Pct Registros con matricula",
            "Fraccion de registros SNIES activos con matricula real",
            "Registros con matricula dividido por la suma de Registros con "
            "matricula mas Registros sin matricula. Siempre entre 0% y 100%. "
            "Un valor bajo indica que muchos programas registrados no tienen "
            "estudiantes activos.",
        ),
    ]

    FILAS_FUENTES = [
        ("Color", "Fuente de categoria", "Significado"),
        ("Verde", "CRUCE_SNIES", "Categoria por cruce exacto de codigo SNIES. Certeza 100%."),
        ("Verde", "MATCH_NOMBRE", "Categoria por coincidencia exacta de nombre. Certeza 100%."),
        ("Verde", "MATCH_CATEGORIA", "Nombre coincide directamente con una categoria. Certeza 100%."),
        ("Amarillo", "KNN_TFIDF (>=50%)", "Categoria por KNN+TF-IDF. Confianza >= 50%. Aceptable."),
        ("Rojo", "KNN_TFIDF (<50%)", "Categoria por KNN. Confianza < 50%. Requiere revision manual."),
    ]

    def _celda_texto(ws, row: int, col: int, texto: object):
        """
        Escribe una celda como texto puro (data_type='s'), nunca como fórmula.
        Evita corrupción XML cuando el texto contiene =, /, flechas u otros símbolos.
        """
        cell = ws.cell(row=row, column=col)
        valor_limpio = str(texto).strip() if texto is not None else ""
        if valor_limpio.startswith("="):
            valor_limpio = valor_limpio.lstrip("=").strip()
        cell.value = valor_limpio
        cell.data_type = "s"
        return cell

    fila_actual = 1
    for tupla in FILAS_METODOLOGIA:
        nombre, definicion, detalle = tupla
        c_nom = _celda_texto(ws_ley, fila_actual, 1, nombre)
        c_def = _celda_texto(ws_ley, fila_actual, 2, definicion)
        c_det = _celda_texto(ws_ley, fila_actual, 3, detalle)

        if fila_actual == 1:
            for cell in (c_nom, c_def, c_det):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            for cell in (c_nom, c_def, c_det):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            c_nom.font = Font(bold=True)

        fila_actual += 1

    fila_actual += 1  # fila en blanco de separación

    BG_FUENTE = {"Verde": "C6EFCE", "Amarillo": "FFF2CC", "Rojo": "FFC7CE"}
    emoji_map = {"Verde": "🟢 Verde", "Amarillo": "🟡 Amarillo", "Rojo": "🔴 Rojo"}

    for tupla in FILAS_FUENTES:
        color_key, fuente, significado = tupla
        c_col = _celda_texto(ws_ley, fila_actual, 1, color_key)
        c_fue = _celda_texto(ws_ley, fila_actual, 2, fuente)
        c_sig = _celda_texto(ws_ley, fila_actual, 3, significado)

        if color_key == "Color":
            for cell in (c_col, c_fue, c_sig):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
        else:
            bg = BG_FUENTE.get(color_key, BLANCO)
            if bg != BLANCO:
                for cell in (c_col, c_fue, c_sig):
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            if color_key in emoji_map:
                c_col.value = emoji_map[color_key]
                c_col.data_type = "s"

        fila_actual += 1

    ws_ley.column_dimensions["A"].width = 36
    ws_ley.column_dimensions["B"].width = 28
    ws_ley.column_dimensions["C"].width = 80

    wb.save(str(ruta_salida))

    n_verdes = (
        int(df_export["FUENTE_CATEGORIA"].astype(str).str.upper().isin(["CRUCE_SNIES", "MATCH_NOMBRE", "MATCH_CATEGORIA"]).sum())
        if "FUENTE_CATEGORIA" in df_export.columns
        else 0
    )
    n_knn = (
        int(df_export["FUENTE_CATEGORIA"].astype(str).str.upper().eq("KNN_TFIDF").sum())
        if "FUENTE_CATEGORIA" in df_export.columns
        else 0
    )
    n_rev = len(df_revision)
    log_resultado(f"[Exportar F1] Excel generado: {ruta_salida.name}")
    log_resultado(f"  Programas exportados : {len(df_export):,}")
    log_resultado(f"  Certeza 100%         : {n_verdes:,}")
    log_resultado(f"  KNN_TFIDF            : {n_knn:,}  (de los cuales {n_rev:,} requieren revisión)")
    return ruta_salida


def run_pipeline(
    ask_reuse_checkpoint: Callable[[str], bool] | None = None,
) -> None:
    """
    Orquesta el pipeline completo de estudio de mercado.
    Opcionalmente pregunta si reusar base_maestra y/o sabana_consolidada.
    """
    import time
    from etl.config import ARCHIVO_ESTUDIO_MERCADO

    t0 = time.perf_counter()
    log_etapa_iniciada("Pipeline estudio de mercado Colombia")

    ok, errores = validar_archivos_entrada()
    if not ok:
        log_error(f"Pipeline detenido: {len(errores)} error(es) en archivos de entrada.")
        return

    def _ask(name: str) -> bool:
        if ask_reuse_checkpoint is not None:
            return ask_reuse_checkpoint(name)
        try:
            return input(f"¿Reusar checkpoint '{name}'? (s/n): ").strip().lower() == "s"
        except (EOFError, OSError):
            return False

    if not CHECKPOINT_BASE_MAESTRA.exists():
        run_fase1()
    else:
        if _ask("base_maestra.parquet"):
            log_info("Reusando base_maestra.parquet")
        else:
            run_fase1()

    run_fase2()

    sabana_path = CHECKPOINT_BASE_MAESTRA.parent / "sabana_consolidada.parquet"
    if not sabana_path.exists():
        run_fase3()
    else:
        if _ask("sabana_consolidada.parquet"):
            log_info("Reusando sabana_consolidada.parquet")
        else:
            run_fase3()

    ag_pos, ag_pre = run_fase4()
    run_fase5(ag_pos, ag_pre)

    elapsed = time.perf_counter() - t0
    log_resultado(f"Tiempo total: {elapsed:.1f}s")
    log_info(f"Salida: {ARCHIVO_ESTUDIO_MERCADO}")

    try:
        n_prog = len(pd.read_parquet(sabana_path)) if sabana_path.exists() else 0
    except Exception:
        n_prog = 0
    n_cat = len(ag_pos) if ag_pos is not None else 0
    if ag_pos is not None and "calificacion_final" in ag_pos.columns:
        _cal = ag_pos["calificacion_final"]
        verdes = int((_cal >= 4.0).sum())
        amarillos = int(((_cal >= 3.0) & (_cal < 4.0)).sum())
        rojos = int((_cal < 3.0).sum())
    else:
        verdes = amarillos = rojos = 0
    n_cat_pre = len(ag_pre) if ag_pre is not None else 0
    log_resultado(
        f"Programas: {n_prog}, Categorías posgrado: {n_cat}, "
        f"Categorías pregrado: {n_cat_pre}, "
        f"Verde(>=4): {verdes}, Amarillo(>=3): {amarillos}, Rojo(<3): {rojos}"
    )
    log_etapa_completada("Pipeline estudio de mercado Colombia", f"{elapsed:.1f}s")


def run_pipeline_mercado() -> None:
    """
    Ejecuta todas las fases del pipeline de mercado (sin preguntar checkpoints).
    """
    log_etapa_iniciada("Pipeline estudio de mercado (modo automático)")
    ok, errores = validar_archivos_entrada()
    if not ok:
        log_error(f"Pipeline detenido: {len(errores)} error(es) en archivos de entrada.")
        return

    run_fase1()
    run_fase2()
    run_fase3()
    ag_pos, ag_pre = run_fase4()
    run_fase5(ag_pos, ag_pre)


if __name__ == "__main__":
    run_fase1()
