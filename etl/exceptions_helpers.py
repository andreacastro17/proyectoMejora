"""
Funciones auxiliares para manejo robusto de excepciones comunes.

Proporciona wrappers y helpers para operaciones que pueden fallar de manera predecible,
con mensajes claros y sugerencias de solución.
"""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile

from etl.pipeline_logger import log_error, log_warning


def leer_excel_con_reintentos(
    archivo: Path,
    sheet_name: str = "Programas",
    max_intentos: int = 3,
    delay_segundos: float = 2.0,
    **kwargs
) -> pd.DataFrame:
    """
    Lee un archivo Excel con manejo robusto de errores comunes.
    
    Maneja:
    - PermissionError: archivo abierto en Excel
    - BadZipFile/InvalidFileException: archivo corrupto
    - FileNotFoundError: archivo no existe
    
    Args:
        archivo: Ruta al archivo Excel
        sheet_name: Nombre de la hoja a leer
        max_intentos: Número máximo de reintentos si hay PermissionError
        delay_segundos: Segundos de espera entre reintentos
        **kwargs: Argumentos adicionales para pd.read_excel
        
    Returns:
        DataFrame con los datos
        
    Raises:
        FileNotFoundError: Si el archivo no existe
        PermissionError: Si después de reintentos sigue bloqueado
        ValueError: Si el archivo está corrupto o no es un Excel válido
    """
    if not archivo.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo: {archivo}\n\n"
            "Verifica que la ruta sea correcta y que el archivo exista."
        )
    
    # Validar que sea un Excel válido antes de intentar leer
    try:
        # Intento rápido de validación con openpyxl
        wb = load_workbook(archivo, read_only=True, data_only=True)
        wb.close()
    except BadZipFile:
        raise ValueError(
            f"El archivo {archivo.name} está corrupto o no es un archivo Excel válido.\n\n"
            "Verifica que el archivo no esté dañado. Si es necesario, descárgalo nuevamente."
        ) from None
    except InvalidFileException:
        raise ValueError(
            f"El archivo {archivo.name} no es un archivo Excel válido.\n\n"
            "Verifica que el archivo tenga extensión .xlsx o .xls y que no esté corrupto."
        ) from None
    except PermissionError:
        # Si falla la validación por permisos, intentaremos leer de todas formas
        pass
    except Exception as e:
        log_warning(f"Advertencia al validar Excel {archivo.name}: {e}")
    
    # Intentar leer con reintentos si hay PermissionError
    ultimo_error: Exception | None = None
    for intento in range(1, max_intentos + 1):
        try:
            df = pd.read_excel(archivo, sheet_name=sheet_name, **kwargs)
            if intento > 1:
                log_warning(f"Archivo {archivo.name} leído exitosamente en intento {intento}")
            return df
        except PermissionError as e:
            ultimo_error = e
            if intento < max_intentos:
                mensaje = (
                    f"El archivo {archivo.name} está abierto en otro programa (Excel, Power BI, etc.).\n\n"
                    f"Intento {intento}/{max_intentos}. Esperando {delay_segundos}s antes de reintentar...\n"
                    "Por favor, cierra el archivo y vuelve a intentar."
                )
                log_warning(mensaje)
                time.sleep(delay_segundos)
            else:
                raise PermissionError(
                    f"No se pudo leer {archivo.name} después de {max_intentos} intentos.\n\n"
                    "El archivo está abierto en otro programa (Excel, Power BI, etc.).\n"
                    "Por favor:\n"
                    "1. Cierra Excel / Power BI\n"
                    "2. Cierra cualquier visor del archivo\n"
                    "3. Vuelve a ejecutar la operación"
                ) from e
        except (BadZipFile, InvalidFileException) as e:
            raise ValueError(
                f"El archivo {archivo.name} está corrupto o no es un Excel válido.\n\n"
                "Verifica que el archivo no esté dañado. Si es necesario, descárgalo nuevamente."
            ) from e
        except Exception as e:
            # Otros errores se propagan con contexto
            raise RuntimeError(
                f"Error al leer {archivo.name}: {e}\n\n"
                "Verifica que el archivo sea un Excel válido y que tenga la hoja '{sheet_name}'."
            ) from e
    
    # No debería llegar aquí, pero por seguridad
    if ultimo_error:
        raise ultimo_error
    raise RuntimeError(f"No se pudo leer {archivo.name} por razones desconocidas.")


def escribir_excel_con_reintentos(
    archivo: Path,
    df: pd.DataFrame,
    sheet_name: str = "Programas",
    max_intentos: int = 3,
    delay_segundos: float = 2.0,
    **kwargs
) -> None:
    """
    Escribe un DataFrame a un archivo Excel con manejo robusto de errores comunes.
    
    Maneja:
    - PermissionError: archivo abierto en Excel
    - OSError: problemas de permisos o espacio en disco
    
    Args:
        archivo: Ruta al archivo Excel a escribir
        df: DataFrame a escribir
        sheet_name: Nombre de la hoja
        max_intentos: Número máximo de reintentos si hay PermissionError
        delay_segundos: Segundos de espera entre reintentos
        **kwargs: Argumentos adicionales para pd.ExcelWriter
        
    Raises:
        PermissionError: Si después de reintentos sigue bloqueado
        OSError: Si hay problemas de permisos o espacio en disco
    """
    ultimo_error: Exception | None = None
    for intento in range(1, max_intentos + 1):
        try:
            with pd.ExcelWriter(
                archivo,
                mode="a" if archivo.exists() else "w",
                if_sheet_exists="replace",
                engine="openpyxl",
                **kwargs
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            if intento > 1:
                log_warning(f"Archivo {archivo.name} escrito exitosamente en intento {intento}")
            return
        except PermissionError as e:
            ultimo_error = e
            if intento < max_intentos:
                mensaje = (
                    f"El archivo {archivo.name} está abierto en otro programa.\n\n"
                    f"Intento {intento}/{max_intentos}. Esperando {delay_segundos}s antes de reintentar...\n"
                    "Por favor, cierra el archivo y vuelve a intentar."
                )
                log_warning(mensaje)
                time.sleep(delay_segundos)
            else:
                raise PermissionError(
                    f"No se pudo escribir {archivo.name} después de {max_intentos} intentos.\n\n"
                    "El archivo está abierto en otro programa (Excel, Power BI, etc.).\n"
                    "Por favor:\n"
                    "1. Cierra Excel / Power BI\n"
                    "2. Cierra cualquier visor del archivo\n"
                    "3. Vuelve a ejecutar la operación"
                ) from e
        except OSError as e:
            # Problemas de permisos o espacio en disco
            raise OSError(
                f"No se pudo escribir {archivo.name}: {e}\n\n"
                "Verifica:\n"
                "1. Que tengas permisos de escritura en la carpeta\n"
                "2. Que haya espacio suficiente en disco\n"
                "3. Que el archivo no esté abierto en otro programa"
            ) from e
    
    if ultimo_error:
        raise ultimo_error
    raise RuntimeError(f"No se pudo escribir {archivo.name} por razones desconocidas.")


def validar_excel_basico(archivo: Path) -> tuple[bool, str]:
    """
    Valida que un archivo sea un Excel válido sin leerlo completamente.
    
    Args:
        archivo: Ruta al archivo
        
    Returns:
        Tupla (es_valido, mensaje_error)
    """
    if not archivo.exists():
        return False, f"El archivo no existe: {archivo}"
    
    if archivo.suffix.lower() not in ('.xlsx', '.xls'):
        return False, f"El archivo no tiene extensión de Excel: {archivo.suffix}"
    
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        wb.close()
        return True, ""
    except BadZipFile:
        return False, f"El archivo {archivo.name} está corrupto (no es un ZIP válido)"
    except InvalidFileException:
        return False, f"El archivo {archivo.name} no es un Excel válido"
    except Exception as e:
        return False, f"Error al validar {archivo.name}: {e}"


def explicar_error_archivo_abierto(archivo: Path, operacion: str = "acceder") -> str:
    """
    Genera un mensaje claro cuando un archivo está abierto.
    
    Args:
        archivo: Ruta al archivo
        operacion: Operación que se intentaba realizar (leer, escribir, etc.)
        
    Returns:
        Mensaje descriptivo con instrucciones
    """
    return (
        f"No se puede {operacion} el archivo {archivo.name} porque está abierto en otro programa.\n\n"
        "Por favor:\n"
        "1. Cierra Excel si lo tienes abierto\n"
        "2. Cierra Power BI si lo tienes abierto\n"
        "3. Cierra cualquier otro visor del archivo\n"
        "4. Vuelve a intentar la operación"
    )
