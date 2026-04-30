"""
Scraper de matrículas SNIES (Fase 2 pipeline mercado).

Lectura local desde ref/backup/: matriculas/, inscritos/, primer_curso/, graduados/.
Los Excels oficiales se colocan en esas carpetas; Fase 2 genera CSVs en outputs/historico/raw/.
"""

from __future__ import annotations

from pathlib import Path
import unicodedata

import pandas as pd

from etl.config import RAW_HISTORIC_DIR, REF_DIR
from etl.pipeline_logger import log_info, log_warning


def _empty_matriculados() -> pd.DataFrame:
    return pd.DataFrame(columns=["CÓDIGO_SNIES_DEL_PROGRAMA", "MATRICULADOS", "SEMESTRE"])


def _empty_inscritos() -> pd.DataFrame:
    return pd.DataFrame(columns=["CÓDIGO_SNIES_DEL_PROGRAMA", "INSCRITOS", "INSCRITOS_S1", "INSCRITOS_S2"])


def _empty_primer_curso() -> pd.DataFrame:
    return pd.DataFrame(columns=["CÓDIGO_SNIES_DEL_PROGRAMA", "PRIMER_CURSO", "SEMESTRE"])


def _empty_graduados() -> pd.DataFrame:
    return pd.DataFrame(columns=["CÓDIGO_SNIES_DEL_PROGRAMA", "GRADUADOS", "SEMESTRE"])


def _strip_accents_upper(s: str) -> str:
    s = str(s).replace("\n", " ").replace("\r", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.split()).upper()


def _detectar_hoja_datos(sheet_names: list[str]) -> str:
    """
    Devuelve el nombre de la primera hoja que no sea índice.
    Normaliza tildes para comparar: ÍNDICE → INDICE.
    """
    for s in sheet_names:
        s_norm = s.strip().upper().replace("Í", "I").replace("É", "E")
        if "INDICE" not in s_norm and "INDEX" not in s_norm:
            return s
    return sheet_names[0]


def _leer_inscritos_snies(path: Path, year: int) -> pd.DataFrame:
    """
    Lee un Excel oficial SNIES de inscritos con header dinámico.

    Retorna un DataFrame con:
      - CÓDIGO SNIES DEL PROGRAMA (str)
      - INSCRITOS (int)  ← suma anual S1+S2 por programa
      - INSCRITOS_S1 (int)
      - INSCRITOS_S2 (int)

    Suma ambos semestres para maximizar cobertura: ~1,500 SNIESs/año
    solo reportan en S2 (especialmente Especializaciones con cohorte única).
    """
    try:
        import openpyxl

        # Encontrar hoja de datos (descartar hojas de índice)
        wb_tmp = openpyxl.load_workbook(path, read_only=True)
        sheet_names = wb_tmp.sheetnames
        wb_tmp.close()

        sheet_name = _detectar_hoja_datos(sheet_names)

        # Detectar fila header: buscar la fila que contenga "CÓDIGO" y "SNIES"
        preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=30, dtype=str)
        header_idx = None
        for idx in range(len(preview)):
            row = preview.iloc[idx]
            vals = [v for v in row.tolist() if pd.notna(v) and str(v).strip() != ""]
            if not vals:
                continue
            first = _strip_accents_upper(vals[0])
            any_has_snies = any("SNIES" in _strip_accents_upper(v) for v in vals)
            if "CODIGO" in first and any_has_snies:
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError("No se pudo detectar fila de encabezado (CÓDIGO SNIES DEL PROGRAMA).")

        df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, dtype=str)
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=["CÓDIGO SNIES DEL PROGRAMA", "INSCRITOS"])

        # Normalizar nombres de columnas
        df.columns = [str(c).strip() for c in df.columns]
        cols_norm = {c: _strip_accents_upper(c) for c in df.columns}

        col_snies = None
        for c, cn in cols_norm.items():
            if "SNIES" in cn and "PROGRAMA" in cn:
                col_snies = c
                break
        if col_snies is None:
            raise ValueError(f"No se encontró columna SNIES. Columnas: {list(df.columns)[:15]}")

        col_ins = None
        for c, cn in cols_norm.items():
            if "INSCRITOS" in cn:
                col_ins = c
                break
        if col_ins is None:
            raise ValueError(f"No se encontró columna INSCRITOS. Columnas: {list(df.columns)[:15]}")

        col_sem = None
        for c, cn in cols_norm.items():
            if cn == "SEMESTRE":
                col_sem = c
                break
        if col_sem is None:
            raise ValueError("No se encontró columna SEMESTRE.")

        df = df[[col_snies, col_sem, col_ins]].copy()
        df = df.rename(columns={col_snies: "CÓDIGO SNIES DEL PROGRAMA", col_sem: "SEMESTRE", col_ins: "INSCRITOS"})

        df["CÓDIGO SNIES DEL PROGRAMA"] = (
            df["CÓDIGO SNIES DEL PROGRAMA"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )
        df["SEMESTRE"] = pd.to_numeric(df["SEMESTRE"], errors="coerce")
        df["INSCRITOS"] = pd.to_numeric(df["INSCRITOS"], errors="coerce").fillna(0).astype("int64")

        # Agregar por SNIES y semestre para obtener subtotales S1 y S2
        by_sem = (
            df.groupby(["CÓDIGO SNIES DEL PROGRAMA", "SEMESTRE"], as_index=False)["INSCRITOS"]
            .sum()
        )
        s1 = (
            by_sem[by_sem["SEMESTRE"] == 1]
            .set_index("CÓDIGO SNIES DEL PROGRAMA")["INSCRITOS"]
            .rename("INSCRITOS_S1")
        )
        s2 = (
            by_sem[by_sem["SEMESTRE"] == 2]
            .set_index("CÓDIGO SNIES DEL PROGRAMA")["INSCRITOS"]
            .rename("INSCRITOS_S2")
        )

        # Unir S1+S2 — todos los SNIESs, incluyendo los que solo reportan en un semestre
        out = (
            pd.concat([s1, s2], axis=1)
            .fillna(0)
            .astype("int64")
            .reset_index()
        )
        out.columns = ["CÓDIGO SNIES DEL PROGRAMA", "INSCRITOS_S1", "INSCRITOS_S2"]
        out["INSCRITOS"] = out["INSCRITOS_S1"] + out["INSCRITOS_S2"]
        out = out.sort_values("INSCRITOS", ascending=False).reset_index(drop=True)

        n_solo_s1 = int((out["INSCRITOS_S2"] == 0).sum())
        n_solo_s2 = int((out["INSCRITOS_S1"] == 0).sum())
        log_info(
            f"[Fase 2] Inscritos SNIES {year}: {len(out):,} programas "
            f"(solo_S1={n_solo_s1:,} | solo_S2={n_solo_s2:,} | ambos={len(out) - n_solo_s1 - n_solo_s2:,})"
        )
        return out.reset_index(drop=True)
    except Exception as e:
        log_warning(f"[Fase 2] Inscritos {year}: no se pudo leer {path.name}: {e}.")
        return pd.DataFrame(columns=["CÓDIGO SNIES DEL PROGRAMA", "INSCRITOS", "INSCRITOS_S1", "INSCRITOS_S2"])


def _leer_primer_curso_snies(path: Path, year: int, semestre: int) -> pd.DataFrame:
    """
    Lee Excel SNIES de primer curso (header dinámico, misma lógica que inscritos).
    Retorna columnas CÓDIGO SNIES DEL PROGRAMA, PRIMER_CURSO (agregado por programa).
    """
    try:
        import openpyxl

        wb_tmp = openpyxl.load_workbook(path, read_only=True)
        sheet_names = wb_tmp.sheetnames
        wb_tmp.close()
        sheet_name = _detectar_hoja_datos(sheet_names)

        preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=30, dtype=str)
        header_idx = None
        for idx in range(len(preview)):
            row = preview.iloc[idx]
            vals = [v for v in row.tolist() if pd.notna(v) and str(v).strip() != ""]
            if not vals:
                continue
            first = _strip_accents_upper(vals[0])
            any_has_snies = any("SNIES" in _strip_accents_upper(v) for v in vals)
            if "CODIGO" in first and any_has_snies:
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError("No se pudo detectar fila de encabezado (CÓDIGO SNIES DEL PROGRAMA).")

        df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, dtype=str)
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=["CÓDIGO SNIES DEL PROGRAMA", "PRIMER_CURSO"])

        df.columns = [str(c).strip() for c in df.columns]
        cols_norm = {c: _strip_accents_upper(c) for c in df.columns}

        col_snies = None
        for c, cn in cols_norm.items():
            if "SNIES" in cn and "PROGRAMA" in cn:
                col_snies = c
                break
        if col_snies is None:
            raise ValueError(f"No se encontró columna SNIES. Columnas: {list(df.columns)[:15]}")

        col_pc = None
        for c, cn in cols_norm.items():
            if "PRIMER" in cn and "CURSO" in cn:
                col_pc = c
                break
        if col_pc is None:
            raise ValueError(f"No se encontró columna primer curso. Columnas: {list(df.columns)[:15]}")

        col_sem = None
        for c, cn in cols_norm.items():
            if cn == "SEMESTRE":
                col_sem = c
                break
        if col_sem is None:
            raise ValueError("No se encontró columna SEMESTRE.")

        df = df[[col_snies, col_sem, col_pc]].copy()
        df = df.rename(
            columns={col_snies: "CÓDIGO SNIES DEL PROGRAMA", col_sem: "SEMESTRE", col_pc: "PRIMER_CURSO"}
        )

        df["SEMESTRE"] = df["SEMESTRE"].astype(str).str.strip()
        df = df[df["SEMESTRE"] == str(int(semestre))].copy()

        df["CÓDIGO SNIES DEL PROGRAMA"] = (
            df["CÓDIGO SNIES DEL PROGRAMA"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )
        df["PRIMER_CURSO"] = pd.to_numeric(df["PRIMER_CURSO"], errors="coerce").fillna(0).astype("int64")

        out = (
            df.groupby("CÓDIGO SNIES DEL PROGRAMA", as_index=False)["PRIMER_CURSO"]
            .sum()
            .sort_values("PRIMER_CURSO", ascending=False)
        )
        log_info(f"[Fase 2] Primer curso SNIES {year}-S{semestre}: {len(out):,} programas leídos desde {path.name}")
        return out.reset_index(drop=True)
    except Exception as e:
        log_warning(f"[Fase 2] Primer curso {year}-S{semestre}: no se pudo leer {path.name}: {e}.")
        return pd.DataFrame(columns=["CÓDIGO SNIES DEL PROGRAMA", "PRIMER_CURSO"])


def _leer_graduados_snies(path: Path, year: int, semestre: int) -> pd.DataFrame:
    """
    Lee Excel SNIES de graduados (header dinámico, misma lógica que inscritos).
    Retorna columnas CÓDIGO SNIES DEL PROGRAMA, GRADUADOS (agregado por programa).
    """
    try:
        import openpyxl

        wb_tmp = openpyxl.load_workbook(path, read_only=True)
        sheet_names = wb_tmp.sheetnames
        wb_tmp.close()
        sheet_name = _detectar_hoja_datos(sheet_names)

        preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=30, dtype=str)
        header_idx = None
        for idx in range(len(preview)):
            row = preview.iloc[idx]
            vals = [v for v in row.tolist() if pd.notna(v) and str(v).strip() != ""]
            if not vals:
                continue
            first = _strip_accents_upper(vals[0])
            any_has_snies = any("SNIES" in _strip_accents_upper(v) for v in vals)
            if "CODIGO" in first and any_has_snies:
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError("No se pudo detectar fila de encabezado (CÓDIGO SNIES DEL PROGRAMA).")

        df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, dtype=str)
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=["CÓDIGO SNIES DEL PROGRAMA", "GRADUADOS"])

        df.columns = [str(c).strip() for c in df.columns]
        cols_norm = {c: _strip_accents_upper(c) for c in df.columns}

        col_snies = None
        for c, cn in cols_norm.items():
            if "SNIES" in cn and "PROGRAMA" in cn:
                col_snies = c
                break
        if col_snies is None:
            raise ValueError(f"No se encontró columna SNIES. Columnas: {list(df.columns)[:15]}")

        col_grad = None
        for c, cn in cols_norm.items():
            if cn == "GRADUADOS":
                col_grad = c
                break
        if col_grad is None:
            raise ValueError(f"No se encontró columna GRADUADOS. Columnas: {list(df.columns)[:15]}")

        col_sem = None
        for c, cn in cols_norm.items():
            if cn == "SEMESTRE":
                col_sem = c
                break
        if col_sem is None:
            raise ValueError("No se encontró columna SEMESTRE.")

        df = df[[col_snies, col_sem, col_grad]].copy()
        df = df.rename(
            columns={col_snies: "CÓDIGO SNIES DEL PROGRAMA", col_sem: "SEMESTRE", col_grad: "GRADUADOS"}
        )

        df["SEMESTRE"] = df["SEMESTRE"].astype(str).str.strip()
        df = df[df["SEMESTRE"] == str(int(semestre))].copy()

        df["CÓDIGO SNIES DEL PROGRAMA"] = (
            df["CÓDIGO SNIES DEL PROGRAMA"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )
        df["GRADUADOS"] = pd.to_numeric(df["GRADUADOS"], errors="coerce").fillna(0).astype("int64")

        out = (
            df.groupby("CÓDIGO SNIES DEL PROGRAMA", as_index=False)["GRADUADOS"]
            .sum()
            .sort_values("GRADUADOS", ascending=False)
        )
        log_info(f"[Fase 2] Graduados SNIES {year}-S{semestre}: {len(out):,} programas leídos desde {path.name}")
        return out.reset_index(drop=True)
    except Exception as e:
        log_warning(f"[Fase 2] Graduados {year}-S{semestre}: no se pudo leer {path.name}: {e}.")
        return pd.DataFrame(columns=["CÓDIGO SNIES DEL PROGRAMA", "GRADUADOS"])


class SNIESMatriculasScraper:
    """
    Lee Excels de matrículas desde ref/backup/matriculas/ (descarga manual del usuario).
    Detecta dinámicamente la fila de encabezados (palabra 'snies' en las primeras 20 filas),
    divide por semestre y guarda CSVs en raw (outputs/historico/raw/).
    Inscritos, primer curso y graduados se leen desde ref/backup/ cuando existan los Excels.
    """

    def __init__(self, raw_dir: Path | None = None) -> None:
        self.raw_dir = raw_dir or RAW_HISTORIC_DIR
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.manual_dir = REF_DIR / "backup" / "matriculas"

    @staticmethod
    def _norm_col_name(name: str) -> str:
        s = str(name).replace("\n", " ").replace("\r", " ").strip()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = " ".join(s.split())
        return s.upper()

    @classmethod
    def _normalize_cached_columns(cls, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or len(df.columns) == 0:
            return df
        rename_map = {}
        for c in df.columns:
            cn = cls._norm_col_name(c)
            if "SNIES" in cn and "PROGRAMA" in cn:
                rename_map[c] = "CÓDIGO_SNIES_DEL_PROGRAMA"
        if rename_map:
            df = df.rename(columns=rename_map)
        return df

    def _load_cached(self, path: Path, required_cols: set[str]) -> pd.DataFrame | None:
        if not path.exists():
            return None
        try:
            df = pd.read_csv(path, dtype={"CÓDIGO_SNIES_DEL_PROGRAMA": str}, encoding="utf-8-sig")
            df = self._normalize_cached_columns(df)
            if required_cols.issubset(set(df.columns)):
                return df
        except Exception as e:
            log_warning(f"[Fase 2] Error leyendo {path.name}: {e}.")
        return None

    def _find_manual_excel_for_year(self, year: int) -> Path | None:
        """Busca en self.manual_dir un archivo cuyo nombre contenga el año (ej. 2019)."""
        if not self.manual_dir.exists():
            return None
        year_str = str(year)
        for ext in ("*.xlsx", "*.xls"):
            for f in self.manual_dir.glob(ext):
                if year_str in f.name:
                    return f
        return None

    def _read_excel_local_dynamic_header(self, filepath: Path) -> pd.DataFrame:
        """
        Lee Excel SNIES con detección robusta de hoja y fila de encabezado.

        Estrategia hoja:
          - Intenta primero la hoja '1.' (presente en 2023 y 2024)
          - Si no existe, usa la última hoja (robustez para formatos futuros)
          - Si falla todo, usa sheet[0]

        Estrategia header:
          - Busca una CELDA INDIVIDUAL cuyo valor normalizado contenga
            'snies' Y 'programa' y sea corto (< 80 chars), para evitar
            falsos positivos en filas de notas/texto libre.
        """
        xl = pd.ExcelFile(filepath)
        sheet_names = xl.sheet_names

        candidates = []
        if "1." in sheet_names:
            candidates.append("1.")
        for s in reversed(sheet_names):
            if s not in candidates:
                candidates.append(s)

        last_error = None
        for sheet in candidates:
            try:
                df = pd.read_excel(filepath, sheet_name=sheet, header=None)
            except Exception as e:
                last_error = e
                continue

            if df is None or len(df) < 5:
                continue

            header_idx = None
            for idx in range(min(25, len(df))):
                row = df.iloc[idx]
                for cell_val in row:
                    if pd.isna(cell_val):
                        continue
                    cell_str = (
                        str(cell_val)
                        .replace("\n", " ")
                        .strip()
                        .lower()
                    )
                    if "snies" in cell_str and "program" in cell_str and len(cell_str) < 80:
                        header_idx = idx
                        break
                if header_idx is not None:
                    break

            if header_idx is None:
                continue

            df.columns = df.iloc[header_idx]
            df = df.iloc[header_idx + 1 :].reset_index(drop=True)
            log_info(
                f"[Fase 2] {filepath.name}: hoja='{sheet}', "
                f"header en fila {header_idx}, {len(df):,} filas de datos"
            )
            return df

        raise ValueError(
            f"No se encontró hoja con datos SNIES en {filepath.name}. "
            f"Hojas disponibles: {sheet_names}. Último error: {last_error}"
        )

    @staticmethod
    def _normalize_cols(df: pd.DataFrame) -> tuple[pd.DataFrame, str, str]:
        cols = [str(c) for c in df.columns]
        cols_norm = {c: str(c).strip().replace("\n", " ").replace("\r", " ").lower() for c in cols}

        col_snies = None
        for c, cn in cols_norm.items():
            if "snies" in cn and "program" in cn:
                col_snies = c
                break
        if col_snies is None:
            for c, cn in cols_norm.items():
                if "snies" in cn:
                    col_snies = c
                    break
        col_sem = None
        for c, cn in cols_norm.items():
            if "semestre" in cn:
                col_sem = c
                break

        if not col_snies:
            raise ValueError(f"No se pudo detectar columna SNIES. Columnas: {list(df.columns)[:15]}")

        if not col_sem:
            df["SEMESTRE"] = 1
            col_sem = "SEMESTRE"

        df = df.rename(columns={col_snies: "CÓDIGO_SNIES_DEL_PROGRAMA", col_sem: "SEMESTRE"})
        df["CÓDIGO_SNIES_DEL_PROGRAMA"] = (
            df["CÓDIGO_SNIES_DEL_PROGRAMA"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )
        df["SEMESTRE"] = pd.to_numeric(df["SEMESTRE"], errors="coerce").astype("Int64")

        for c in list(df.columns):
            cn = str(c).strip().replace("\n", " ").replace("\r", " ").lower()
            if "matriculad" in cn:
                df = df.rename(columns={c: "MATRICULADOS"})
                break
        for c in list(df.columns):
            cn = str(c).strip().replace("\n", " ").replace("\r", " ").lower()
            if "inscrit" in cn:
                df = df.rename(columns={c: "INSCRITOS"})
                break

        return df, "CÓDIGO_SNIES_DEL_PROGRAMA", "SEMESTRE"

    def _split_y_cache_por_semestre(
        self,
        df: pd.DataFrame,
        year: int,
        prefix: str,
        value_col_candidates: list[str],
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        df, _, _ = self._normalize_cols(df)

        value_col = None
        cols_norm = {c: str(c).strip().replace("\n", " ").replace("\r", " ").lower() for c in df.columns}
        for cand in value_col_candidates:
            cand_l = cand.lower()
            for c, cn in cols_norm.items():
                if cand_l in cn:
                    value_col = c
                    break
            if value_col:
                break
        if not value_col:
            numeric_cols = [c for c in df.columns if c not in ("CÓDIGO_SNIES_DEL_PROGRAMA", "SEMESTRE")]
            for c in numeric_cols:
                try:
                    s = pd.to_numeric(df[c], errors="coerce")
                    if s.notna().sum() > 0:
                        value_col = c
                        break
                except Exception:
                    continue
        if not value_col:
            raise ValueError("No se pudo detectar columna de valor (matriculados/inscritos).")

        out_value_col = "MATRICULADOS" if prefix == "matriculados" else "INSCRITOS"
        df[out_value_col] = pd.to_numeric(df[value_col], errors="coerce").fillna(0).astype(int)
        df_out = df[["CÓDIGO_SNIES_DEL_PROGRAMA", out_value_col, "SEMESTRE"]].copy()

        df_1 = df_out[df_out["SEMESTRE"] == 1].copy()
        df_2 = df_out[df_out["SEMESTRE"] == 2].copy()

        p1 = self.raw_dir / f"{prefix}_{year}_1.csv"
        p2 = self.raw_dir / f"{prefix}_{year}_2.csv"
        df_1.to_csv(p1, index=False, encoding="utf-8-sig")
        df_2.to_csv(p2, index=False, encoding="utf-8-sig")
        log_info(f"[Fase 2] {prefix} {year}: guardados {p1.name} ({len(df_1):,}) y {p2.name} ({len(df_2):,})")

        return df_1, df_2

    def download_matriculados(self, year: int, semestre: int) -> pd.DataFrame:
        """
        Lee matrículas desde ref/backup/matriculas/ (archivo cuyo nombre contenga el año).
        Si ya existen los CSVs en raw, los carga desde disco.
        Lectura dinámica: header=None y búsqueda de fila con 'snies' en las primeras 20 filas.
        """
        archivo = self.raw_dir / f"matriculados_{year}_{semestre}.csv"
        cached = self._load_cached(archivo, {"CÓDIGO_SNIES_DEL_PROGRAMA", "MATRICULADOS"})
        if cached is not None:
            # Si el Excel fuente es más nuevo, invalidar caché y reconstruir ambos semestres
            try:
                filepath = self._find_manual_excel_for_year(year)
                if filepath is not None:
                    csv_mtime = archivo.stat().st_mtime
                    xlsx_mtime = filepath.stat().st_mtime
                    if xlsx_mtime > csv_mtime:
                        log_info(
                            f"[Fase 2] Excel {filepath.name} fue modificado después del CSV en caché. "
                            f"Reconstruyendo CSVs para {year}..."
                        )
                        for sem in (1, 2):
                            old = self.raw_dir / f"matriculados_{year}_{sem}.csv"
                            if old.exists():
                                old.unlink()
                        cached = None
            except Exception:
                pass
            if cached is not None:
                log_info(f"[Fase 2] Matriculados {year}-{semestre}: cargado desde disco ({len(cached):,} filas)")
                return cached

        filepath = self._find_manual_excel_for_year(year)
        if filepath is None:
            log_warning(
                f"[Fase 2] Matriculados {year}: no hay archivo en {self.manual_dir} con el año en el nombre. "
                "Coloque el Excel de matriculados (ej. matriculados_2019.xlsx) en ref/backup/matriculas/."
            )
            return _empty_matriculados()

        try:
            df_raw = self._read_excel_local_dynamic_header(filepath)
            if df_raw is None or len(df_raw) == 0:
                return _empty_matriculados()
            df_1, df_2 = self._split_y_cache_por_semestre(
                df_raw,
                year=year,
                prefix="matriculados",
                value_col_candidates=["MATRICULADOS", "MATRICULA"],
            )
            return df_1 if int(semestre) == 1 else df_2
        except Exception as e:
            log_warning(f"[Fase 2] Matriculados {year}-{semestre}: {e}. Continuando con vacío.")
            return _empty_matriculados()

    def download_inscritos(self, year: int) -> pd.DataFrame:
        """
        Lee inscritos desde Excel oficial en ref/backup/inscritos/inscritos_{year}.xlsx.
        Si no existe o falla, retorna DataFrame vacío con columnas esperadas (no bloquea el pipeline).
        """
        archivo = self.raw_dir / f"inscritos_{year}.csv"
        cached = self._load_cached(
            archivo,
            {"CÓDIGO_SNIES_DEL_PROGRAMA", "INSCRITOS", "INSCRITOS_S1", "INSCRITOS_S2"},
        )
        if cached is not None:
            # Si el Excel fuente es más nuevo, invalidar caché
            try:
                src = REF_DIR / "backup" / "inscritos" / f"inscritos_{year}.xlsx"
                if src.exists():
                    if src.stat().st_mtime > archivo.stat().st_mtime:
                        log_info(
                            f"[Fase 2] Excel {src.name} fue modificado después del CSV en caché. "
                            f"Reconstruyendo inscritos para {year}..."
                        )
                        cached = None
            except Exception:
                pass
            if cached is not None:
                log_info(f"[Fase 2] Inscritos {year}: cargado desde disco ({len(cached):,} filas)")
                return cached

        src = REF_DIR / "backup" / "inscritos" / f"inscritos_{year}.xlsx"
        if not src.exists():
            log_warning(
                f"[Fase 2] Inscritos {year}: no existe {src}. "
                "Coloque el Excel oficial (inscritos_YYYY.xlsx) en ref/backup/inscritos/."
            )
            return _empty_inscritos()

        df_in = _leer_inscritos_snies(src, year=year)
        if df_in is None or len(df_in) == 0:
            return _empty_inscritos()

        out = df_in.rename(columns={"CÓDIGO SNIES DEL PROGRAMA": "CÓDIGO_SNIES_DEL_PROGRAMA"}).copy()
        out_path = self.raw_dir / f"inscritos_{year}.csv"
        try:
            out[
                ["CÓDIGO_SNIES_DEL_PROGRAMA", "INSCRITOS", "INSCRITOS_S1", "INSCRITOS_S2"]
            ].to_csv(out_path, index=False, encoding="utf-8-sig")
            log_info(f"[Fase 2] Inscritos {year}: guardado {out_path.name} ({len(out):,} filas)")
        except Exception as e:
            log_warning(f"[Fase 2] Inscritos {year}: no se pudo guardar CSV: {e}.")
        return out[["CÓDIGO_SNIES_DEL_PROGRAMA", "INSCRITOS", "INSCRITOS_S1", "INSCRITOS_S2"]].copy()

    def download_primer_curso(self, year: int, semestre: int) -> pd.DataFrame:
        """
        Lee primer curso desde ref/backup/primer_curso/primer_curso_{year}.xlsx.
        Si no existe o falla, retorna DataFrame vacío (no bloquea el pipeline).
        """
        archivo = self.raw_dir / f"primer_curso_{year}_{semestre}.csv"
        cached = self._load_cached(archivo, {"CÓDIGO_SNIES_DEL_PROGRAMA", "PRIMER_CURSO"})
        if cached is not None:
            try:
                src = REF_DIR / "backup" / "matriculas primer curso" / f"primer_curso_{year}.xlsx"
                if src.exists() and src.stat().st_mtime > archivo.stat().st_mtime:
                    log_info(
                        f"[Fase 2] Excel {src.name} fue modificado después del CSV en caché. "
                        f"Reconstruyendo primer curso para {year}-S{semestre}..."
                    )
                    cached = None
            except Exception:
                pass
            if cached is not None:
                log_info(f"[Fase 2] Primer curso {year}-{semestre}: cargado desde disco ({len(cached):,} filas)")
                return cached

        src = REF_DIR / "backup" / "matriculas primer curso" / f"primer_curso_{year}.xlsx"
        if not src.exists():
            log_warning(
                f"[Fase 2] Primer curso {year}: no existe {src}. "
                "Coloque el Excel en ref/backup/primer_curso/."
            )
            return _empty_primer_curso()

        df_pc = _leer_primer_curso_snies(src, year=year, semestre=int(semestre))
        if df_pc is None or len(df_pc) == 0:
            return _empty_primer_curso()

        out = df_pc.rename(columns={"CÓDIGO SNIES DEL PROGRAMA": "CÓDIGO_SNIES_DEL_PROGRAMA"}).copy()
        out["SEMESTRE"] = int(semestre)
        out_path = self.raw_dir / f"primer_curso_{year}_{semestre}.csv"
        try:
            out[["CÓDIGO_SNIES_DEL_PROGRAMA", "PRIMER_CURSO", "SEMESTRE"]].to_csv(
                out_path, index=False, encoding="utf-8-sig"
            )
            log_info(f"[Fase 2] Primer curso {year}-{semestre}: guardado {out_path.name} ({len(out):,} filas)")
        except Exception as e:
            log_warning(f"[Fase 2] Primer curso {year}-{semestre}: no se pudo guardar CSV: {e}.")
        return out[["CÓDIGO_SNIES_DEL_PROGRAMA", "PRIMER_CURSO", "SEMESTRE"]].copy()

    def download_graduados(self, year: int, semestre: int) -> pd.DataFrame:
        """
        Lee graduados desde ref/backup/graduados/graduados_{year}.xlsx.
        Si no existe o falla, retorna DataFrame vacío (no bloquea el pipeline).
        """
        archivo = self.raw_dir / f"graduados_{year}_{semestre}.csv"
        cached = self._load_cached(archivo, {"CÓDIGO_SNIES_DEL_PROGRAMA", "GRADUADOS"})
        if cached is not None:
            try:
                src = REF_DIR / "backup" / "graduados" / f"graduados_{year}.xlsx"
                if src.exists() and src.stat().st_mtime > archivo.stat().st_mtime:
                    log_info(
                        f"[Fase 2] Excel {src.name} fue modificado después del CSV en caché. "
                        f"Reconstruyendo graduados para {year}-S{semestre}..."
                    )
                    cached = None
            except Exception:
                pass
            if cached is not None:
                log_info(f"[Fase 2] Graduados {year}-{semestre}: cargado desde disco ({len(cached):,} filas)")
                return cached

        src = REF_DIR / "backup" / "graduados" / f"graduados_{year}.xlsx"
        if not src.exists():
            log_warning(
                f"[Fase 2] Graduados {year}: no existe {src}. Coloque el Excel en ref/backup/graduados/."
            )
            return _empty_graduados()

        df_g = _leer_graduados_snies(src, year=year, semestre=int(semestre))
        if df_g is None or len(df_g) == 0:
            return _empty_graduados()

        out = df_g.rename(columns={"CÓDIGO SNIES DEL PROGRAMA": "CÓDIGO_SNIES_DEL_PROGRAMA"}).copy()
        out["SEMESTRE"] = int(semestre)
        out_path = self.raw_dir / f"graduados_{year}_{semestre}.csv"
        try:
            out[["CÓDIGO_SNIES_DEL_PROGRAMA", "GRADUADOS", "SEMESTRE"]].to_csv(
                out_path, index=False, encoding="utf-8-sig"
            )
            log_info(f"[Fase 2] Graduados {year}-{semestre}: guardado {out_path.name} ({len(out):,} filas)")
        except Exception as e:
            log_warning(f"[Fase 2] Graduados {year}-{semestre}: no se pudo guardar CSV: {e}.")
        return out[["CÓDIGO_SNIES_DEL_PROGRAMA", "GRADUADOS", "SEMESTRE"]].copy()
