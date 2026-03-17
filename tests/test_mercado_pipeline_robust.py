"""
Suite robusta de tests para el pipeline de Estudio de Mercado (Fases 1–5).

Objetivo:
- Verificar el flujo completo con sandbox temporal (sin tocar outputs/ref reales).
- Validar contratos de datos por fase (columnas, flags, artefactos).
- Probar casos borde (feedback HITL, ausencia de columnas, etc.).
- Probar Fase 2 sin red usando descargas simuladas (requests monkeypatch).

Notas:
- Para Parquet: requiere pyarrow (o fastparquet). El test lo marca como requisito.
- Tests "network" (opcionales) pueden añadirse aparte si se desea pegarle al portal real.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import pytest


def _make_excel_bytes(df: pd.DataFrame, sheet_name: str = "1.") -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Simular 5 filas basura con un DF vacío y luego el DF real empezando en fila 6
        pd.DataFrame({"x": []}).to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
    return bio.getvalue()


def _bootstrap_inputs(tmp_path: Path):
    """
    Crea entradas mínimas y coherentes para ejecutar Fases 1–5 en sandbox.
    Retorna un dict con paths y valores útiles.
    """
    pytest.importorskip("pyarrow")
    from etl import config

    # Pre-crear stubs ANTES de update_paths_for_base_dir para evitar prints "[ERROR]" ruidosos
    # durante la resolución de archivos de referencia (referentesUnificados, catalogoOfertasEAFIT).
    ref_backup = tmp_path / "ref" / "backup"
    ref_backup.mkdir(parents=True, exist_ok=True)
    for nombre in ("referentesUnificados.xlsx", "catalogoOfertasEAFIT.xlsx", "Referente_Categorias.xlsx"):
        p = ref_backup / nombre
        if not p.exists():
            with pd.ExcelWriter(p, engine="openpyxl") as writer:
                pd.DataFrame({"_stub": [1]}).to_excel(writer, sheet_name="stub", index=False)

    config.update_paths_for_base_dir(tmp_path)
    from etl.config import (
        ARCHIVO_PROGRAMAS,
        REF_DIR,
        RAW_HISTORIC_DIR,
    )

    # Programas
    n_programas = 30
    rows = []
    for i in range(1, n_programas + 1):
        rows.append(
            {
                "CÓDIGO_SNIES_DEL_PROGRAMA": str(20_000 + i),
                "NOMBRE_DEL_PROGRAMA": f"Programa {i} Economía",
                "TITULO_OTORGADO": "Economista",
                "CINE_F_2013_AC_CAMPO_DETALLADO": "Economía",
                "CINE_F_2013_AC_CAMPO_ESPECÍFIC": "Finanzas",
                "ÁREA_DE_CONOCIMIENTO": "Economía, administración, contaduría y afines",
                "NOMBRE_INSTITUCIÓN": "IES Test",
                "NIVEL_DE_FORMACIÓN": "Pregrado",
                "ESTADO_PROGRAMA": "Activo",
                "FECHA_DE_REGISTRO_EN_SNIES": "2023-01-01",
                "COSTO_MATRÍCULA_ESTUD_NUEVOS": pd.NA,
            }
        )
    df_prog = pd.DataFrame(rows)
    df_cob = pd.DataFrame(
        [
            {
                "CÓDIGO_SNIES_DEL_PROGRAMA": str(20_000 + i),
                "TIPO_CUBRIMIENTO": "PRINCIPAL",
                "VALOR_MATRICULA": 7_500_000 + (i * 15_000),
            }
            for i in range(1, 16)
        ]
    )
    ARCHIVO_PROGRAMAS.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(ARCHIVO_PROGRAMAS, engine="openpyxl") as writer:
        df_prog.to_excel(writer, sheet_name="Programas", index=False)
        df_cob.to_excel(writer, sheet_name="Cobertura", index=False)

    # Referente (>=5 por clase en y_train, con SNIES únicos)
    (REF_DIR / "backup").mkdir(parents=True, exist_ok=True)
    cats = ["CAT_X", "CAT_Y", "CAT_Z"]
    ref_rows = []
    for ci, cat in enumerate(cats):
        for j in range(1, 11):  # 10 por clase
            snies = str(90_000 + (ci * 1000 + j))
            ref_rows.append(
                {
                    "SNIES": snies,
                    "CATEGORIA_FINAL": cat,
                    # salario en SMLMV
                    "SALARIO_OLE": 3.0 + (ci * 0.5) + (j * 0.05),
                    "TASA_COTIZANTES": 0.60 + ci * 0.05,
                    "INSCRITOS_2023": 80 + j,
                    "INSCRITOS_2024": 90 + j,
                    "NOMBRE_DEL_PROGRAMA": f"Ref {cat} {j}",
                    "TITULO_OTORGADO": "Economista",
                    "ÁREA_DE_CONOCIMIENTO": "Economía",
                    "CINE_F_2013_AC_CAMPO_DETALLADO": "Economía",
                    "CINE_F_2013_AC_CAMPO_ESPECÍFIC": "Finanzas",
                }
            )
    df_ref = pd.DataFrame(ref_rows)
    for p in (REF_DIR / "backup" / "Referente_Categorias.xlsx", REF_DIR / "Referente_Categorias.xlsx"):
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            df_ref.to_excel(writer, sheet_name="1_Consolidado", index=False)

    # OLE backup local (cabeceras sucias)
    df_ole = pd.DataFrame(
        {
            "CÓDIGO_SNIES_DEL_PROGRAMA": [str(20_001), str(20_002), str(20_003)],
            "TASA COTIZANTES\n(0.0-1.0)": [0.70, 0.65, 0.55],
            "SALARIO OLE": [3.2, 3.6, 4.1],
        }
    )
    df_ole.to_csv(REF_DIR / "backup" / "ole_indicadores.csv", index=False, encoding="utf-8-sig")

    # Cache históricos (matrícula/inscritos) para evitar red
    RAW_HISTORIC_DIR.mkdir(parents=True, exist_ok=True)
    for year in range(2019, 2025):
        for sem in (1, 2):
            pd.DataFrame(
                {
                    "CÓDIGO_SNIES_DEL_PROGRAMA": [str(20_001), str(20_002), str(20_003)],
                    "MATRICULADOS": [10 + sem, 20 + sem, 30 + sem],
                    "SEMESTRE": [sem, sem, sem],
                }
            ).to_csv(RAW_HISTORIC_DIR / f"matriculados_{year}_{sem}.csv", index=False, encoding="utf-8-sig")
            pd.DataFrame(
                {
                    "CÓDIGO_SNIES_DEL_PROGRAMA": [str(20_001), str(20_002), str(20_003)],
                    "INSCRITOS": [40 + sem, 50 + sem, 60 + sem],
                    "SEMESTRE": [sem, sem, sem],
                }
            ).to_csv(RAW_HISTORIC_DIR / f"inscritos_{year}_{sem}.csv", index=False, encoding="utf-8-sig")

    # Importante: etl.mercado_pipeline importa rutas desde etl.config por binding.
    # Si otros tests ya importaron el módulo, hay que recargarlo para que tome las rutas del tmp_path actual.
    import importlib
    import etl.mercado_pipeline as mp  # noqa: F401
    import etl.scraper_ole as so  # noqa: F401
    import etl.scraper_matriculas as sm  # noqa: F401
    importlib.reload(mp)
    importlib.reload(so)
    importlib.reload(sm)

    return {"n_programas": n_programas}


@pytest.mark.slow
def test_fase1_aplica_feedback_manual(tmp_path: Path):
    pytest.importorskip("pyarrow")
    from etl import config

    _bootstrap_inputs(tmp_path)

    # Crear feedback_manual que corrija una categoría en el referente
    from etl.config import REF_DIR

    fb = pd.DataFrame(
        [
            {"SNIES": "90001", "CATEGORIA_FINAL": "CAT_OVERRIDE"},
            {"SNIES": "90001", "CATEGORIA_FINAL": "CAT_OVERRIDE_FINAL"},  # última gana
        ]
    )
    fb.to_csv(REF_DIR / "feedback_manual.csv", index=False, encoding="utf-8-sig")

    from etl.mercado_pipeline import run_fase1

    df_base = run_fase1()
    assert df_base is not None and len(df_base) > 0
    # No podemos observar df_referente directo, pero sí asegurar que el pipeline no revienta aplicando feedback.


@pytest.mark.slow
def test_fase2_snies_split_por_semestre_sin_red(tmp_path: Path, monkeypatch):
    """
    Prueba que el scraper de matrículas:
    - descarta API
    - usa portal HTML
    - descarga un excel anual simulado
    - detecta columnas flexibles
    - guarda ambos CSVs (sem1 y sem2)
    """
    pytest.importorskip("pyarrow")
    from etl import config

    # Evitar ruido de config: pre-crear stubs antes de update_paths_for_base_dir
    ref_backup = tmp_path / "ref" / "backup"
    ref_backup.mkdir(parents=True, exist_ok=True)
    for nombre in ("referentesUnificados.xlsx", "catalogoOfertasEAFIT.xlsx", "Referente_Categorias.xlsx"):
        p = ref_backup / nombre
        if not p.exists():
            with pd.ExcelWriter(p, engine="openpyxl") as writer:
                pd.DataFrame({"_stub": [1]}).to_excel(writer, sheet_name="stub", index=False)

    config.update_paths_for_base_dir(tmp_path)
    from etl.config import RAW_HISTORIC_DIR
    from etl.scraper_matriculas import SNIESMatriculasScraper

    # Excel anual simulado con columnas variantes
    df_excel = pd.DataFrame(
        {
            "CÓDIGO SNIES DEL PROGRAMA": ["1", "2", "3", "1"],
            "SEMESTRE": [1, 1, 2, 2],
            "MATRICULADOS": [100, 200, 300, 150],
        }
    )
    excel_bytes = _make_excel_bytes(df_excel, sheet_name="1.")

    class _Resp:
        def __init__(self, status_code: int, text: str = "", content: bytes = b""):
            self.status_code = status_code
            self.text = text
            self.content = content

        def raise_for_status(self):
            if self.status_code >= 400:
                raise RuntimeError(f"HTTP {self.status_code}")

        def json(self):
            return {}

    # API POST falla => fallback a portal HTML
    def fake_post(url, json=None, timeout=None):
        return _Resp(404)

    portal_html = """
    <html><body>
    <a href="articles-401908_recurso.xlsx">Estudiantes matriculados 2019</a>
    </body></html>
    """

    def fake_get(url, headers=None, timeout=None):
        if url.endswith("/portal/ESTADISTICAS/Bases-consolidadas/") or url.endswith("/portal/ESTADISTICAS/Bases-consolidadas"):
            return _Resp(200, text=portal_html)
        if url.endswith("articles-401908_recurso.xlsx"):
            return _Resp(200, content=excel_bytes)
        return _Resp(404)

    import requests

    monkeypatch.setattr(requests, "post", fake_post)
    monkeypatch.setattr(requests, "get", fake_get)

    s = SNIESMatriculasScraper(raw_dir=RAW_HISTORIC_DIR)
    df_sem1 = s.download_matriculados(2019, 1)
    df_sem2 = s.download_matriculados(2019, 2)

    assert (RAW_HISTORIC_DIR / "matriculados_2019_1.csv").exists()
    assert (RAW_HISTORIC_DIR / "matriculados_2019_2.csv").exists()
    assert df_sem1["SEMESTRE"].dropna().unique().tolist() == [1]
    assert df_sem2["SEMESTRE"].dropna().unique().tolist() == [2]


@pytest.mark.slow
def test_pipeline_mercado_robusto_artifactos_y_excel(tmp_path: Path):
    """
    Corre Fase 1–5 y valida:
    - checkpoints existen
    - columnas esenciales existen
    - Excel exportado contiene la columna salario_proyectado_pesos_hoy
    - formato moneda en esa columna (openpyxl number_format)
    """
    pytest.importorskip("pyarrow")
    from etl import config

    _bootstrap_inputs(tmp_path)

    from etl.mercado_pipeline import run_fase1, run_fase2, run_fase3, run_fase4, run_fase5
    from etl.config import (
        ARCHIVO_ESTUDIO_MERCADO,
        CHECKPOINT_BASE_MAESTRA,
        MODELO_CLASIFICADOR_MERCADO,
        RAW_HISTORIC_DIR,
    )

    df_base = run_fase1()
    assert CHECKPOINT_BASE_MAESTRA.exists()
    assert MODELO_CLASIFICADOR_MERCADO.exists()
    assert "CATEGORIA_FINAL" in df_base.columns
    assert "FUENTE_CATEGORIA" in df_base.columns

    run_fase2()
    assert (RAW_HISTORIC_DIR / "ole_indicadores.csv").exists()

    run_fase3()
    sabana_path = CHECKPOINT_BASE_MAESTRA.parent / "sabana_consolidada.parquet"
    assert sabana_path.exists()

    ag = run_fase4()
    assert ag is not None and len(ag) > 0
    assert "calificacion_final" in ag.columns
    assert "salario_proyectado_pesos_hoy" in ag.columns

    run_fase5(ag)
    assert ARCHIVO_ESTUDIO_MERCADO.exists()

    # Validar que hoja total tiene la columna en el encabezado fila 2
    import openpyxl

    wb = openpyxl.load_workbook(ARCHIVO_ESTUDIO_MERCADO)
    assert "total" in wb.sheetnames
    ws = wb["total"]
    header_row = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert "salario_proyectado_pesos_hoy" in header_row

    # Validar formato moneda en una celda de esa columna (fila 3)
    col_idx = header_row.index("salario_proyectado_pesos_hoy") + 1
    fmt = ws.cell(row=3, column=col_idx).number_format
    assert fmt in ("#,##0",)  # moneda_fmt actual

